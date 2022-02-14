#!/usr/bin/python3

import shutil
import json
import sys
import glob
import os
import re
import mpu.io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
from openpyxl.styles import numbers
from openpyxl.chart import BarChart, Reference, Series, LineChart
from openpyxl.chart.label import DataLabelList
import timeit
from datetime import datetime
from datetime import timedelta
import pandas as pd
import math
import numpy
from os import listdir
import time
from logdef import KEY_WORD
import modules
import ssh_modules
import socket

now = datetime.now() # current date and time
date_time = now.strftime("%Y%m%d_%H%M%S")
print("date and time:",date_time)

#TEST on KEY_WORD
#print(KEY_WORD.NIC_DIAG_TEST_RSLT_RE)

scriptname = os.path.basename(__file__)
scriptname = scriptname.replace('.py','')

def main():

    get_lock(scriptname)

    start=datetime.now()
    pr = dict()
    pr['modules'] = modules.modules()

    startdate = checkstartdate()
    ARGV = dict()
    inputconfig = checkconfigfile(ARGV)
    print("FIND startdate input: {}".format(startdate))
    print(json.dumps(ARGV, indent = 4 ))
    #sys.exit()
    print(json.dumps(inputconfig, indent = 4))
    print(inputconfig['FILE']['datebasesjsonfile'])
    pr['modules'].checkDirectoryexist(inputconfig['DIR'])

    datebase_DB = pr['modules'].readjsonfile(inputconfig['FILE']['datebasesjsonfile'])
    # if os.path.exists(inputconfig['FILE']['datebasesjsonfile']):
    #     datebase_DB = mpu.io.read(inputconfig['FILE']['datebasesjsonfile'])    
    DATA = datebase_DB

    

    snmaclist = getsnmaclist(DATA,inputconfig)
    macsnlist = getmacsnlist(DATA,inputconfig)

    pr['SNlist'] = processFailureSNFlexFlowCheck(DATA,inputconfig,pr)

    reportdir = inputconfig['DIR']["reportpath"]

    dest_filename = "{}EXECL_{}_DATA.xlsx".format(reportdir,date_time)
    filenameheader = "{}EXECL_{}_DATA".format(reportdir,date_time)
    if "NAME" in inputconfig:
        dest_filename = "{}{}_{}_DATA.xlsx".format(reportdir,inputconfig["NAME"],date_time)
        filenameheader = "{}{}_{}_DATA".format(reportdir,inputconfig["NAME"],date_time)

    if len(snmaclist):
        dest_filenameformac = dest_filename.replace('DATA','SNandMAClistwithFlexflow')
        wsandc = Workbook()
        generateexeclmacandsnreport(pr,DATA,wsandc,snmaclist)
        if len(macsnlist):
            generateexeclmacandsnreport2(pr,DATA,wsandc,macsnlist)
        wsandc.save(filename = dest_filenameformac)

    difftime = datetime.now()-start
    print('Done Time: ', difftime)       
    print("How many seconds use?: {} seconds".format(difftime.total_seconds()))       
    return None

def processFailureSNFlexFlowCheck(DATA,inputconfig,pr):
    start=datetime.now()
    FailureSNlist = dict()
    #FailureSNlist["DATA"] = GetFailureSNList(DATA,inputconfig,pr)
    FailureSNlist["FailureSN"] = list()
    FailureSNlist["FailureSNFlexflow"] = dict()
    FailureSNlist["FlexflowType"] = dict()

    pr['modules'].print_anyinformation(FailureSNlist)

    workingSNlist = list()
    for SN in DATA['SN']['LIST']:
        if "FPN" in SN:
            if not SN in workingSNlist:
                workingSNlist.append(SN)    

    if "FLEXFLOW" in inputconfig:
        mkssh = ssh_modules.ssh_modules(inputconfig["FLEXFLOW"])
        mkssh.ssh_login()
        mkssh.getuptimeinfo()
        count = 0
        for sn in workingSNlist:
            FailureSNlist["FailureSN"].append(sn)
            flexflowdata = mkssh.getflowflex(sn)
            FailureSNlist["FailureSNFlexflow"][sn] = flexflowdata
            if not flexflowdata in FailureSNlist["FlexflowType"]:
                FailureSNlist["FlexflowType"][flexflowdata] = list()
            FailureSNlist["FlexflowType"][flexflowdata].append(sn)
            count += 1
            print("{} of {}".format(count, len(workingSNlist)))
            #if count > 100:
                #break
        mkssh.ssh_logout()

    else:
        return None

    #pr['modules'].print_anyinformation(FailureSNlist)
    difftime = datetime.now()-start
    print('Done Time: ', difftime)     
    #sys.exit()

    return FailureSNlist

def processtocreatedailyreport(pr,DATA,inputconfig,date_time,startdate):

    if "historypath" in inputconfig["DIR"]:
        movereporttohistorydir(inputconfig)

    if "MTP_STATUS" in inputconfig:
        generateMTPreport(pr,DATA,inputconfig,startdate=None)
        generateMTPreport(pr,DATA,inputconfig,startdate=getbeforedayinformation(checkday=31))

    if "DIE_ID" in inputconfig:
        wb2 = Workbook()
        dest_filename2 = "{}ELBA_DIE_ID_{}_DATA.xlsx".format(inputconfig['DIR']["reportpath"],date_time)
        print(dest_filename2)

        generateexecldieidreport(DATA,wb2,inputconfig,start=startdate)

        wb2.save(filename = dest_filename2) 

    #sys.exit()

    createteststatusreport(pr,DATA,inputconfig)

    if pr['FailureSNlist']:
        createfailureteststatusreport(pr,DATA,inputconfig)

    startdate = getbeforedayinformation(checkday=31)
    createteststatusreport(pr,DATA,inputconfig,startdate=startdate)

    startdate = getbeforedayinformation(checkday=15)
    createteststatusreport(pr,DATA,inputconfig,startdate=startdate)

    # if "NAME" in inputconfig:
    #     if "ORTANO2" == inputconfig["NAME"].upper():
    #         workingonSNlist = getsnlistafteestartdate(DATA,inputconfig,startdate=None)
    #         SNbyPN = GetSNlistbyPNfromDLtest(workingonSNlist,DATA["teststep"]['DL'],teststep='DL')
    #         for PN in SNbyPN:
    #             #,listofsn=[],specpn=None
    #             createteststatusreport(pr,DATA,inputconfig,startdate=None, listofsn=SNbyPN[PN], specpn=PN)

    return None

def generateMTPreport(pr,DATA,inputconfig,startdate=None):
    wb3 = Workbook()
    dest_filename3 = "{}MTP_STATUS_{}_DATA.xlsx".format(inputconfig['DIR']["reportpath"],date_time)
    if startdate:
        dest_filename3 = "{}MTP_STATUS_{}_DATA_StartWith_{}.xlsx".format(inputconfig['DIR']["reportpath"],date_time,startdate)
    print(dest_filename3)

    byTestdata = generateexeclmtpstatussummaryreport(DATA,wb3,inputconfig,startdate)

    for mtp in DATA["MTPCHASSIS"]:
        generateexeclmtpstatusbyeachmtpreport(DATA,wb3,inputconfig,mtp,startdate)

    try:
        if "mtpdatabasejsonfile" in inputconfig["FILE"]:
            mtpdatabase = pr['modules'].readjsonfile(inputconfig["FILE"]["mtpdatabasejsonfile"])
            mtpdatabase[inputconfig["NAME"]] = DATA["MTPCHASSIS"]
            pr['modules'].wirtejsonfile(inputconfig["FILE"]["mtpdatabasejsonfile"],mtpdatabase)
    except:
        print("mtpdatabasejsonfile issue!")

    wb3.save(filename = dest_filename3) 

    return None

def getbeforedayinformation(checkday=30):
    from datetime import date, timedelta

    current_date = date.today().isoformat()   
    days_before = (date.today()-timedelta(days=checkday)).isoformat()
    days_after = (date.today()+timedelta(days=checkday)).isoformat()  

    # print("\nCurrent Date: ",current_date)
    # print("30 days before current date: ",days_before)
    # print("30 days after current date : ",days_after)
    current_date = str(current_date)
    days_before = str(days_before)
    days_after = str(days_after)

    # print("\nCurrent Date: ",current_date)
    # print("30 days before current date: ",days_before)
    # print("30 days after current date : ",days_after)

    return days_before

def createfailureteststatusreport(pr,DATA,inputconfig,startdate=None):
    
    #workingonSNlist = DATA['SN']['LIST']
    
    print("START DATE: {}".format(startdate))
    workingonSNlist = pr['FailureSNlist']["FailureSN"]
    workingonSNlist.sort(reverse=True)
    print("COUNT SN: {}".format(len(workingonSNlist)))
    #sys.exit()    
    wb = Workbook()
    
    reportdir = inputconfig['DIR']["reportpath"]

    dest_filename = "{}FAILURE_{}_FAILURE.xlsx".format(reportdir,date_time)
    filenameheader = "{}FAILURE_{}_FAILURE".format(reportdir,date_time)
    if "NAME" in inputconfig:
        dest_filename = "{}{}_{}_FAILURE.xlsx".format(reportdir,inputconfig["NAME"],date_time)
        filenameheader = "{}{}_{}_FAILURE".format(reportdir,inputconfig["NAME"],date_time)
    if startdate:
        dest_filename = "{}_withStartDate_{}.xlsx".format(filenameheader,startdate)

    print('OUTPUT FILE NAME: ' + dest_filename)

    generateexeclfailurereport(DATA,wb,inputconfig,workingonSNlist,pr,start=startdate)

    generateexeclsnfailurestatusalldata(DATA,pr, workingonSNlist,'LAST',wb,inputconfig,Withallerror=True)
    generateexeclsnTopFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig,byweek=False)

    for eachteststep in DATA['SN']['TEST']:
        generateexecltest(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)
        generateexeclerrdata(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)
        generateexeclerrdata2(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)

    wb.save(filename = dest_filename)
    
    print("OUTPUT FILE: {}".format(dest_filename))

    return None    

def createteststatusreport(pr,DATA,inputconfig,startdate=None,listofsn=[],specpn=None):
    
    #workingonSNlist = DATA['SN']['LIST']
    
    print("START DATE: {}".format(startdate))
    workingonSNlist = getsnlistafteestartdate(DATA,inputconfig,startdate=None)
    if len(listofsn):
        workingonSNlist = listofsn
    workingonSNlist.sort(reverse=True)
    print("COUNT SN: {}".format(len(workingonSNlist)))
    #sys.exit()    
    wb = Workbook()
    
    reportdir = inputconfig['DIR']["reportpath"]
    if specpn:
        reportdir = reportdir + specpn + '/'
        pr['modules'].createdirinserver(reportdir)

    dest_filename = "{}EXECL_{}_DATA.xlsx".format(reportdir,date_time)
    filenameheader = "{}EXECL_{}_DATA".format(reportdir,date_time)
    if "NAME" in inputconfig:
        dest_filename = "{}{}_{}_DATA.xlsx".format(reportdir,inputconfig["NAME"],date_time)
        filenameheader = "{}{}_{}_DATA".format(reportdir,inputconfig["NAME"],date_time)
    if startdate:
        dest_filename = "{}_withStartDate_{}.xlsx".format(filenameheader,startdate)

    dest_chartfilename = dest_filename.replace('DATA','CHART')
    print('OUTPUT FILE NAME: ' + dest_filename)
    print('OUTPUT CHART FILE NAME: ' + dest_chartfilename)
    
    #sys.exit()
    chartdata = dict()

    generateexeclreport(DATA,wb,inputconfig,workingonSNlist,chartdata,pr,start=startdate)
    #pr['modules'].print_anyinformation(chartdata)


    if 'CM' in inputconfig:
        if not startdate:
            if inputconfig['CM'] == "FSJ" or inputconfig['CM'] == "FLEX":

                generateexeclsnFSJstatus(DATA, workingonSNlist,'LAST',wb)
            
            wb.save(filename = dest_filename)
            createlastfailurelogfolder(pr,DATA,workingonSNlist,inputconfig)

    else:
        if not startdate:
            wg = Workbook()
            generateexeclchart(DATA,wg,inputconfig,workingonSNlist,chartdata,pr,start=startdate)
            wg.save(filename = dest_chartfilename)

        # if "NAME" in inputconfig:
        #     if "ORTANO2" == inputconfig["NAME"].upper():
        #         SNbyPN = GetSNlistbyPNfromDLtest(workingonSNlist,DATA["teststep"]['DL'],teststep='DL')
        #         for PN in SNbyPN:
        #             checkSNstartdate = getbeforedayinformation(checkday=30)
        #             if len(checksnlistafteestartdate(DATA,SNbyPN[PN],startdate=checkSNstartdate)) > 0:
        #                 generateexeclDatabyPN(DATA,wb,inputconfig,SNbyPN[PN],PN,chartdata,pr,start=None)

        print("START DATE: {}".format(startdate))
        workingonSNlist = getsnlistafteestartdate(DATA,inputconfig,startdate=startdate)
        if len(listofsn):
            workingonSNlist = listofsn
        workingonSNlist.sort(reverse=True)
        print("COUNT SN: {}".format(len(workingonSNlist)))

        if not len(workingonSNlist):
            return None

        snmaclist = getsnmaclist(DATA,inputconfig)
        macsnlist = getmacsnlist(DATA,inputconfig)

        if not startdate:
            if len(snmaclist):
                dest_filenameformac = dest_filename.replace('DATA','SNandMAClist')
                wsandc = Workbook()
                generateexeclmacandsnreport(DATA,wsandc,snmaclist)
                if len(macsnlist):
                    generateexeclmacandsnreport2(DATA,wsandc,macsnlist)
                wsandc.save(filename = dest_filenameformac)
                try:
                    if "snandmacjsonfile" in inputconfig["FILE"]:
                        snmacdatabase = pr['modules'].readjsonfile(inputconfig["FILE"]["snandmacjsonfile"])
                        if not "snmaclist" in snmacdatabase:
                            snmacdatabase["snmaclist"] = dict()
                        if not "macsnlist" in snmacdatabase:
                            snmacdatabase["macsnlist"] = dict()
                        snmacdatabase["snmaclist"][inputconfig["NAME"]] = snmaclist
                        snmacdatabase["macsnlist"][inputconfig["NAME"]] = macsnlist
                        pr['modules'].wirtejsonfile(inputconfig["FILE"]["snandmacjsonfile"],snmacdatabase)
                except:
                    print("snandmacjsonfile issue!")
        #sys.exit()
        TempHVLVdata = dict()

        generateexeclsnstatus(DATA, workingonSNlist,'FIRST',wb,inputconfig)
        #generateexeclsnstatus(DATA, workingonSNlist,'LAST',wb,inputconfig)
        #generateexeclsnstatus(DATA, workingonSNlist,'LAST',wb,inputconfig,Withallerror=True)
        generateexeclsnstatusalldata(DATA, workingonSNlist,'LAST',wb,inputconfig,Withallerror=True)
        generateexeclsnTopFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig)
        generateexeclsnTopFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig)
        generateexeclsnTopFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig,byweek=False)

        generateexeclby4CChambertemp(workingonSNlist,wb,DATA,pr,start=startdate)

        generateexeclby4CChambertime(workingonSNlist,wb,DATA,pr,start=startdate)
        generateexeclby4CChambertime(workingonSNlist,wb,DATA,pr,start=startdate,totimezone='Asia/Singapore')
        workingonSNlist2 = getsnlistafteestartdate(DATA,inputconfig,startdate=None)
        workingonSNlist2.sort(reverse=True)
        if check2C4CinTestinfunction(DATA['SN']['TEST']):
            for eachteststep in DATA['SN']['TEST']:
                if "4C" in eachteststep:
                    TempHVLVdata[eachteststep] = generateHVLVdata(workingonSNlist2,DATA["teststep"][eachteststep],eachteststep,wb,DATA)

            #print(json.dumps(TempHVLVdata, indent = 4))
            generateexeclHVLVsummary(workingonSNlist2,DATA["teststep"],wb,DATA,TempHVLVdata,pr) 
            generateexeclHVLVdata(workingonSNlist,DATA["teststep"],wb,DATA,TempHVLVdata)           
            #sys.exit()
            #generateexeclsn4CFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig)
            generateexeclsn4CFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig,duplicatesn=False)
            #generateexeclsn4CFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig)
            #generateexeclsn4CFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig,duplicatesn=False)

            #generateexeclsn4CFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig,byweek=True)
            generateexeclsn4CFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig,duplicatesn=False,byweek=True)
            #generateexeclsn4CFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig,byweek=True)
            generateexeclsn4CFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig,duplicatesn=False,byweek=True)
        if not startdate:
            print("OUTPUT FILE: {}".format(dest_filename.replace('DATA', 'SUMMARY')))
            wb.save(filename = dest_filename.replace('DATA', 'SUMMARY'))

        for eachteststep in DATA['SN']['TEST']:
            generateexecltest(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)
            if "4C" in eachteststep:
                generateexecltestby4Ctesttime(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA,pr)
            else:
                generateexecltestbyNon4Ctesttime(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA,pr)
            generateexeclerrdata(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)
            #L1 Sub Test only passed
            generateexeclerrdata2(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)

        wb.save(filename = dest_filename)

        copyallfailurelogfolder(pr,DATA,workingonSNlist,inputconfig)
    
    print("OUTPUT FILE: {}".format(dest_filename))

    return None    
        
def check2C4CinTestinfunction(teststep):
    check2C4CinTest = False
    for eachteststep in teststep:
        if "4C" in eachteststep:
            check2C4CinTest = True
        # if "2C" in eachteststep:
        #     check2C4CinTest = True
    return check2C4CinTest

def movereporttohistorydir(inputconfig):
    filelisttowork = getallfilebyfolder(inputconfig["DIR"]["reportpath"],keyword=None,level=1)
    foldlisttowork = getallfolder(inputconfig["DIR"]["reportpath"],level=2)

    #print(json.dumps(filelisttowork, indent = 4))
    #print(json.dumps(foldlisttowork, indent = 4))

    workingfileandfoldermovinglist = list()
    for eachfolder in foldlisttowork:
        if eachfolder != inputconfig["DIR"]["reportpath"]: 
            if eachfolder != inputconfig["DIR"]["historypath"]:
                workingfileandfoldermovinglist.append(eachfolder)

    for eachfile in filelisttowork:
        workingfileandfoldermovinglist.append(eachfile)
    #print(json.dumps(workingfileandfoldermovinglist, indent = 4))
    try:
        for eachworkitem in workingfileandfoldermovinglist:
            if os.path.isdir(eachworkitem):
                smallitemslist = eachworkitem.split('/')
                #print(smallitemslist)
                movetopath = "{}/{}".format(inputconfig["DIR"]["historypath"],smallitemslist[-1])
                #print(movetopath)
                dest = shutil.move(eachworkitem, movetopath, copy_function = shutil.copytree)
                #print(dest)
            else:
                smallitemslist = eachworkitem.split('/')
                #print(smallitemslist)
                movetopath = "{}/{}".format(inputconfig["DIR"]["historypath"],smallitemslist[-1])
                #print(movetopath)
                dest = shutil.move(eachworkitem, movetopath)
                #print(dest)    
    except:
        print("Something else went wrong")        
    #sys.exit()
    return None

def workingonLastdata(DATA,inputconfig):

    DATA['SN']["KEEPLASTPASS"] = dict()
    DATA['SN']['LAST'] = dict()

    workingSNlist = list()
    for test in DATA['SN']['TEST']:
        for SN in DATA['SN']['FIRST'][test]:
            if not SN in workingSNlist:
                workingSNlist.append(SN) 

    for test in DATA['SN']['TEST']:
        print("TEST: {}".format(test))

        if not test in DATA['SN']["LAST"]:
            DATA['SN']["LAST"][test] = dict()

        for SN in workingSNlist:
            #print("{}: {}".format(test, SN))
            #print(json.dumps(DATA['SN']['LAST'][test][SN], indent = 4))
            checkdata = None 
            NewSN = False
            if not SN in DATA['SN']['LAST'][test]:
                DATA['SN']['LAST'][test][SN] = dict()
                NewSN = True
            counttesttime = 0
            if SN in DATA['teststep'][test]['SN']:
                for chassis in DATA['teststep'][test]['SN'][SN]:
                    for testdate in DATA['teststep'][test]['SN'][SN][chassis]:
                        for testtime in DATA['teststep'][test]['SN'][SN][chassis][testdate]:
                            if "checktime" in DATA['SN']['LAST'][test][SN]:
                                checkdata = DATA['SN']['LAST'][test][SN]["checktime"]
                            checktime2 = "{}_{}".format(testdate, testtime)
                            if not checkdata:
                                checkdata = checktime2
                            counttesttime += 1
                            parpareData(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST')
                            if NewSN:
                                writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST')
                                NewSN = False
                            if checktime2 >= checkdata:
                                writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST')

            delSNwithnoData(DATA,test,SN,counttesttime,status='LAST')

        if not test in DATA['SN']["KEEPLASTPASS"]:
            DATA['SN']["KEEPLASTPASS"][test] = dict()

        for SN in workingSNlist:
            #print("{}: {}".format(test, SN))
            #print(json.dumps(DATA['SN']['LAST'][test][SN], indent = 4))
            checkdata = None 
            NewSN = False
            if not SN in DATA['SN']['KEEPLASTPASS'][test]:
                DATA['SN']['KEEPLASTPASS'][test][SN] = dict()
                NewSN = True
            counttesttime = 0
            if SN in DATA['teststep'][test]['SN']:
                for chassis in DATA['teststep'][test]['SN'][SN]:
                    for testdate in DATA['teststep'][test]['SN'][SN][chassis]:
                        for testtime in DATA['teststep'][test]['SN'][SN][chassis][testdate]:
                            if "checktime" in DATA['SN']['KEEPLASTPASS'][test][SN]:
                                checkdata = DATA['SN']['KEEPLASTPASS'][test][SN]["checktime"]
                            checktime2 = "{}_{}".format(testdate, testtime)
                            if not checkdata:
                                checkdata = checktime2
                            counttesttime += 1
                            parpareData(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                            if NewSN:
                                writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                                NewSN = False
                            if checktime2 >= checkdata:
                                if "FINALRESULT" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
                                    if not 'PASS' in DATA['SN']['KEEPLASTPASS'][test][SN]["result"] and not "PASS" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]:
                                        writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                                    elif not 'PASS' in DATA['SN']['KEEPLASTPASS'][test][SN]["result"] and "PASS" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]:
                                        writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                                    elif 'PASS' in DATA['SN']['KEEPLASTPASS'][test][SN]["result"] and "PASS" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]:
                                        writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                            else:
                                if "FINALRESULT" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
                                    if not 'PASS' in DATA['SN']['KEEPLASTPASS'][test][SN]["result"] and "PASS" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]:
                                        writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                                    
            delSNwithnoData(DATA,test,SN,counttesttime,status='KEEPLASTPASS')

        # workingSNlist = list()
        # for SN in DATA['SN']['FIRST'][test]:
        #     workingSNlist.append(SN)

        for SN in workingSNlist:

            checkdata = None 
            counttesttime = 0
            if SN in DATA['teststep'][test]['SN']:
                for chassis in DATA['teststep'][test]['SN'][SN]:
                    for testdate in DATA['teststep'][test]['SN'][SN][chassis]:
                        for testtime in DATA['teststep'][test]['SN'][SN][chassis][testdate]:
                            if "checktime" in DATA['SN']['FIRST'][test][SN]:
                                checkdata = DATA['SN']['FIRST'][test][SN]["checktime"]
                            checktime2 = "{}_{}".format(testdate, testtime)
                            if not checkdata:
                                checkdata = checktime2
                            counttesttime += 1
                            parpareData(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='FIRST')
                            if checktime2 <= checkdata:
                                writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='FIRST')

            delSNwithnoData(DATA,test,SN,counttesttime,status='FIRST')

    if "KEEPLASTPASS" in inputconfig:
        DATA['SN']['ORGLAST'] = DATA['SN']['LAST']
        DATA['SN']['LAST'] = DATA['SN']["KEEPLASTPASS"]

    return None

def parpareData(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST'):

    if not "IMAGE" in DATA['SN'][status][test][SN]:
        if "IMAGE" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["IMAGE"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["IMAGE"]
    if not "MTPINFO" in DATA['SN'][status][test][SN]:
        if "MTPINFO" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["MTPINFO"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["MTPINFO"]
    if not "NICINFO" in DATA['SN'][status][test][SN]:
        if "NICINFO" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["NICINFO"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["NICINFO"]

    return None

def writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST'):
    if "IMAGE" in DATA['SN'][status][test][SN]:
        if "IMAGE" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["IMAGE"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["IMAGE"]
    if "MTPINFO" in DATA['SN'][status][test][SN]:
        if "MTPINFO" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["MTPINFO"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["MTPINFO"]
    if "NICINFO" in DATA['SN'][status][test][SN]:
        if "NICINFO" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["NICINFO"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["NICINFO"]
    checkdata = checktime2
    DATA['SN'][status][test][SN]["checktime"] = checktime2
    if "FINALRESULT" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
        DATA['SN'][status][test][SN]["result"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]
    else:
        DATA['SN'][status][test][SN]["result"] = "NONE"
    if "SPECIAL" in inputconfig:
        if SN == inputconfig["SPECIAL"]:
            if "FAIL" in DATA['SN'][status][test][SN]["result"]:
                DATA['SN'][status][test][SN]["result"] = "PASS"
    if not "PASS" in DATA['SN'][status][test][SN]["result"]:
        DATA['SN'][status][test][SN]["result"] = "FAIL"
    DATA['SN'][status][test][SN]["testfile"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["eachfile"]
    DATA['SN'][status][test][SN]["ERROR"] = dict()
    DATA['SN'][status][test][SN]["ERROR"]["DETAIL"] = dict()
    DATA['SN'][status][test][SN]["ERROR"]["LIST"] = list()
    for eachteststepdetail in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["TESTSTEPLIST"]:
        for teststepkey in eachteststepdetail:
            if not eachteststepdetail[teststepkey] == 'PASS':
                teststepkeywork = "{} <{}>".format(teststepkey, eachteststepdetail[teststepkey])
                if not teststepkeywork in DATA['SN'][status][test][SN]["ERROR"]["DETAIL"]:
                    DATA['SN'][status][test][SN]["ERROR"]["DETAIL"][teststepkeywork] = 1
                else:
                    DATA['SN'][status][test][SN]["ERROR"]["DETAIL"][teststepkeywork] += 1
                if not teststepkeywork in DATA['SN'][status][test][SN]["ERROR"]["LIST"]:
                    DATA['SN'][status][test][SN]["ERROR"]["LIST"].append(teststepkeywork)
    return None

def delSNwithnoData(DATA,test,SN,counttesttime,status='LAST'):
    if counttesttime:
        DATA['SN'][status][test][SN]["count"] = counttesttime
    else:
        if SN in DATA['SN'][status][test]:
            del DATA['SN'][status][test][SN]
        if SN in DATA['teststep'][test]["SN"]:
            del DATA['teststep'][test]["SN"][SN]
    return None

def checkconfigfile(ARGV):
    print(sys.argv)
    inputfile = sys.argv[1]
    print('GET INPUT CONFIG FILE: ' + inputfile)
    
    if os.path.exists(inputfile):
        print('CHECK INPUT CONFIG FILE EXIST PASS: ' + inputfile)
        inputconfig = mpu.io.read(inputfile)
    else:
        print('CHECK INPUT CONFIG FILE DOES NOT EXIST: ' + inputfile)
        sys.exit()

    for argvitem in sys.argv:
        #print(argvitem)
        matchcheck = re.findall(r"(.*)=(.*)", argvitem)
        if matchcheck:
            #print(matchcheck)
            ARGV[matchcheck[0][0]] = matchcheck[0][1]

    #print(json.dumps(ARGV, indent = 4 ))

    #sys.exit()

    return inputconfig

def checkstartdate():
    returndate = None
    #startdate
    countkey = 1
    for checkkey in sys.argv:
        print("CHECK KEY({}): {}".format(countkey, checkkey))
        countkey += 1
        if 'start' in checkkey:

            match = re.findall(KEY_WORD.CHECKSTARTDATE, checkkey) 
            #print(match)
            if match:
                returndate = match[0]
                print("FIND startdate input: {}".format(returndate))
            else:
                print("FIND startdate keyword, but date format is not correct by YYYY-MM-DD : {}".format(checkkey))
                sys.exit()
    print("FIND startdate input: {}".format(returndate))
    #sys.exit()     
    return returndate

def generateexeclsnDLstatus(DATA,status,wb):

    #DATA['SN']['LIST'] = list()
    #DATA['SN']['TEST'] = list()
    #DATA['SN']['FIRST'] = dict()
    #DATA['SN']['LAST'] = dict()
    
    ws2 = wb.create_sheet(title=status + '_IMAGE')
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('SLOT')
    wirtedata.append('DATE_TIME')
    wirtedata.append('NAME')
    for eachstep in DATA['SN']['IMAGE']:
         wirtedata.append(eachstep)
    ws2.append(wirtedata)
    
    for sn in DATA['SN']['LIST']:

        if sn in DATA['SN']['DL'][status]:
            for dateANDtime in DATA['SN']['DL'][status][sn]:
                for eachname in DATA['SN']['DL'][status][sn][dateANDtime]["DATA"]:
                    for eachdata in DATA['SN']['DL'][status][sn][dateANDtime]["DATA"][eachname]:
                        wirtedata = list()
                        wirtedata.append(sn)                
                        wirtedata.append(DATA['SN']['DL'][status][sn][dateANDtime]["TESTCHASSIS"])
                        wirtedata.append(DATA['SN']['DL'][status][sn][dateANDtime]["SLOT"])
                        wirtedata.append(dateANDtime)
                        wirtedata.append(eachname)
                        #DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD']
                        for eachstep in DATA['SN']['IMAGE']:
                            wirtedata.append(eachdata[eachstep])
                        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    return 0

def takecareteststepDICT(inputconfig,DATA):

    for teststep in inputconfig["TESTLIST"]:
        CreateNeedDict(DATA,teststep)

    return None

def CreateNeedDict(DATA,teststep):
    if not teststep in DATA['teststep']:
        DATA['teststep'][teststep] = dict()
        DATA['teststep'][teststep]['SN'] = dict()
        DATA['teststep'][teststep]['DETAILTESTSTEP'] = list()
        DATA['teststep'][teststep]['FILELISTS'] = list()
    if not teststep in DATA['SN']['FIRST']:
        DATA['SN']['FIRST'][teststep] = dict()
    if not teststep in DATA['SN']['LAST']:
        DATA['SN']['LAST'][teststep] = dict()
    if not teststep in DATA['SN']['ERROR']:
        DATA['SN']['ERROR'][teststep] = dict()
    return None

def workingoneachtest(pr,inputconfig,DATA,testfolder,redo=False):

    takecareteststepDICT(inputconfig,DATA)
    #print(json.dumps(DATA, indent = 4))

    read_dir_path = "{}/{}".format(inputconfig['DIR']['testlogpath'],testfolder)
    print("FOLDER: {}".format(read_dir_path))

    allfolder = getallfolder(read_dir_path)
    #print(json.dumps(allfolder, indent = 4))
    
    teststeplist = list()

    if not 'MTPCHASSIS' in DATA:
        DATA['MTPCHASSIS'] = dict()

    count = 0
    newcount = 0
    for eachfolder in allfolder:
        #print(eachfolder)
        #print(len(eachfolder))
        #print(len(read_dir_path))
        sn = os.path.basename(eachfolder)
        
        
        if 'unknown' in sn.lower():
            continue
        if len(sn) < 5:
            continue
        count += 1
        if count % 100 == 0:
            print("count: {}".format(count))
        #time.sleep(5)
        # if count > 10:
            # time.sleep(5)
            # break
        # if count % 1000 == 0:
        #     print("COUNT: {} Will save to database FILES: {}".format(count,inputconfig['FILE']['datebasesjsonfile']))
        #     mpu.io.write(inputconfig['FILE']['datebasesjsonfile'], DATA)

        if not len(eachfolder) == len(read_dir_path):
          
            getallfilelist = getallfilebyfolder(eachfolder, keyword=".tar.gz")
            # print(json.dumps(getallfilelist, indent = 4))
            for eachfile in getallfilelist:
                #print(eachfile)
                eachfilenameonly = os.path.basename(eachfile)
                getinformationeachfilenameonly = eachfilenameonly.replace('.tar.gz', '')
                #print(eachfilenameonly)
                filearray = getinformationeachfilenameonly.split('_')
                #print(filearray)
                teststep = filearray[0]
                if 'TESTSUB' in inputconfig:
                    if teststep in inputconfig['TESTSUB']:
                        teststep = inputconfig['TESTSUB'][teststep]
                TESTCHASSIS = filearray[1]
                TESTDATE = filearray[2]
                if not teststep in teststeplist:
                    teststeplist.append(teststep)
                    CreateNeedDict(DATA,teststep)
                #if TESTDATE < '2021-07-09':
                    #continue
                # elif TESTDATE > '2021-07-07':
                    # continue

                if not sn in DATA['teststep'][teststep]['SN']:
                    DATA['teststep'][teststep]['SN'][sn] = dict()
                #print("SN: {}".format(sn))
                if not sn in DATA['SN']['LIST']:
                    DATA['SN']['LIST'].append(sn)
                if not sn in DATA['SN']['FIRST'][teststep]:
                    DATA['SN']['FIRST'][teststep][sn] = dict()
                if not sn in DATA['SN']['LAST'][teststep]:
                    DATA['SN']['LAST'][teststep][sn] = dict()
                if not sn in DATA['SN']['ERROR'][teststep]:
                    DATA['SN']['ERROR'][teststep][sn] = dict()  
                    DATA['SN']['ERROR'][teststep][sn]['LIST'] = list()
                    DATA['SN']['ERROR'][teststep][sn]['DETAIL'] = dict()
                
                if not eachfile in DATA['teststep'][teststep]['FILELISTS']:
                    DATA['teststep'][teststep]['FILELISTS'].append(eachfile)
                    newcount += 1
                    DATA['NEWFILECOUNT'] += 1
                    #print("newcount: {}".format(count))
                    if newcount % 1000 == 0:
                        print("COUNT: {} Will save to database FILES: {}".format(newcount,inputconfig['FILE']['datebasesjsonfile']))
                        #mpu.io.write(inputconfig['FILE']['datebasesjsonfile'], DATA)
                        pr['modules'].wirtejsonfile(inputconfig['FILE']['datebasesjsonfile'], DATA)
                else:
                    if not redo:
                        continue

                if not TESTCHASSIS in DATA['MTPCHASSIS']:
                    DATA['MTPCHASSIS'][TESTCHASSIS] = dict()

                if not teststep in DATA['MTPCHASSIS'][TESTCHASSIS]:
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep] = dict()

                TESTFINISHTIME = filearray[3]
                CHECKTESTTIME = TESTDATE + '_' + TESTFINISHTIME
                #print("CHASSIS: {} DATE: {} TIME: {}".format(TESTCHASSIS,TESTDATE,TESTFINISHTIME))
                if not TESTCHASSIS in DATA['teststep'][teststep]['SN'][sn]:
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS] = dict()
                if not TESTDATE in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS]:
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE] = dict()
                if not TESTFINISHTIME in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE]:
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME] = dict()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DETAILTESTSTEP'] = dict()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TESTSTEPLIST'] = list()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['PN'] = dict()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'] = dict()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'] = list()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['IMAGE'] = dict()

                if not CHECKTESTTIME in DATA['MTPCHASSIS'][TESTCHASSIS][teststep]:
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME] = dict()

                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['eachfile'] = eachfile
                if os.path.isdir(inputconfig['DIR']['TEMPDIR']):
                    try:
                        shutil.rmtree(inputconfig['DIR']['TEMPDIR'])
                    except OSError as e:
                        print("Error: %s : %s" % (inputconfig['DIR']['TEMPDIR'], e.strerror))  
                    pr['modules'].createdirinserver(inputconfig['DIR']['TEMPDIR'])

                if os.path.isfile(eachfile):
                    try:
                        shutil.unpack_archive(eachfile, inputconfig['DIR']['TEMPDIR'])
                    except Exception as e:
                        print("Error: %s".format(eachfile)) 
                # else:
                #     continue 

                #mtp_test, test_dl, test_swi, test_fst
                getallunzipfilelist = getallfilebyfolder(inputconfig['DIR']['TEMPDIR'], 'test')

                #print(getallunzipfilelist)


                for readfilefromzip in getallunzipfilelist:
                    #print(readfilefromzip)
                    tempsavelog = ''
                    f = open(readfilefromzip, 'r', encoding="ISO-8859-1")

                    
                    getsomething = False
                    nic_test_rslt_reg_exp = KEY_WORD.NIC_DIAG_TEST_RSLT_RE.format(sn)
                    nic_test_err_reg_exp = KEY_WORD.NIC_DIAG_TEST_ERR_RE.format(sn)
                    nic_test_rslt_final_exp = KEY_WORD.NIC_DIAG_TEST_FINAL_RE.format(sn)
                    nic_test_rslt_cpld_exp = KEY_WORD.NIC_DIAG_TEST_CPLD_IMAGE
                    nic_test_rslt_qspi_exp = KEY_WORD.NIC_DIAG_TEST_QSPI_IMAGE
                    nic_test_rslt_pn_exp = KEY_WORD.NIC_DIAG_TEST_PN.format(sn)
                    nic_test_failure_reg_exp = KEY_WORD.NIC_DIAG_TEST_FAILURE_RE.format('NIC')
                    if "KEY" in inputconfig:
                        nic_test_failure_reg_exp = KEY_WORD.NIC_DIAG_TEST_FAILURE_RE.format(inputconfig["KEY"])
                    failurelist = list()
                    # print(nic_test_rslt_reg_exp)
                    # print(nic_test_rslt_final_exp)
                    # print(nic_test_rslt_cpld_exp)
                    # print(nic_test_rslt_pn_exp)
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] = None
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CARDTYPE'] = "NO CARD TYPE"
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['FINALRESULT'] = "NO RESULT"
                    # if not 'result' in DATA['SN']['FIRST'][teststep][sn]:
                        # DATA['SN']['FIRST'][teststep][sn]['result'] = "NO RESULT"
                        # DATA['SN']['FIRST'][teststep][sn]['checktime'] = CHECKTESTTIME
                    # if not 'result' in DATA['SN']['LAST'][teststep][sn]:
                        # DATA['SN']['LAST'][teststep][sn]['result'] = "NO RESULT"
                        # DATA['SN']['LAST'][teststep][sn]['checktime'] = CHECKTESTTIME
                    TEMPSAVEERROR = dict()
                    TEMPSAVEERROR['LIST'] = list()
                    TEMPSAVEERROR['DETAIL'] = dict()
                    tempmtpinformation = dict()
                    tempnicinformation = dict()
                    tempnicresult = dict()
                    tempimage = dict()
                    temperrormessagelist = list()
                    temperrormessagedata = ''
                    starttocatcherrormessage = False
                    for x in f:
                        tempsavelog += x

                        sub_match = re.findall(KEY_WORD.ERRORMESSAGESTART, x)
                        if sub_match:
                            starttocatcherrormessage = True

                        if starttocatcherrormessage:
                            temperrormessagedata += x
                            #temperrormessagedata += "\r"

                        sub_match = re.findall(KEY_WORD.ERRORMESSAGEEND, x)
                        if sub_match:
                            starttocatcherrormessage = False
                            temperrormessagelist.append(temperrormessagedata)
                            temperrormessagedata = ''
                            #print(temperrormessagelist)
                            #print(x)
                            #sys.exit()



                        #GETIMAGE
                        sub_match = re.findall(KEY_WORD.GETIMAGE, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            mtpnumber = sub_match[0][0]
                            nicnumber = sub_match[0][1]
                            nickey = sub_match[0][2]
                            nicdata = sub_match[0][3]
                            if "IMAGESUB" in inputconfig:
                                if nickey in inputconfig["IMAGESUB"]:
                                    nickey = inputconfig["IMAGESUB"][nickey]
                            if not nicnumber in tempimage:
                                tempimage[nicnumber] = dict()
                            tempimage[nicnumber][nickey] = nicdata
                            #print(tempimage)
                            #sys.exit()

                        sub_match = re.findall(KEY_WORD.GERNICFINALRESULT, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            mtpnumber = sub_match[0][0]
                            nicnumber = sub_match[0][1]
                            nictype = sub_match[0][2]
                            nicsn = sub_match[0][3]
                            nicresult = sub_match[0][4]
                            if not nicnumber in tempnicresult:
                                tempnicresult[nicnumber] = dict()
                            tempnicresult[nicnumber]['SN'] = nicsn
                            tempnicresult[nicnumber]['TYPE'] = nictype
                            tempnicresult[nicnumber]['RESULT'] = nicresult
                            #print(tempnicresult)
                            #sys.exit()

                        sub_match = re.findall(KEY_WORD.GETNICINFORMATION, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            mtpnumber = sub_match[0][0]
                            nicnumber = sub_match[0][1]
                            nickey = sub_match[0][2]
                            nicdata = sub_match[0][3]
                            if not nicnumber in tempnicinformation:
                                tempnicinformation[nicnumber] = dict()
                            tempnicinformation[nicnumber][nickey] = nicdata
                            #print(tempnicinformation)
                            #sys.exit()

                        sub_match = re.findall(KEY_WORD.GETMTPINFORMATION, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            mtpnumber = sub_match[0][0]
                            mptkey = sub_match[0][1]
                            mtpdata = sub_match[0][2]
                            tempmtpinformation['MTP'] = mtpnumber
                            tempmtpinformation[mptkey] = mtpdata
                            #sys.exit()

                        sub_match = re.findall(nic_test_rslt_reg_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            getsomething = True
                            detailteststep = sub_match[0][0]
                            if detailteststep == 'DL2':
                                detailteststep = 'DL'
                            if detailteststep == 'DL3':
                                detailteststep = 'DL'
                            detailteststep = detailteststep + ' ' + sub_match[0][1]
                            detailteststepresult = sub_match[0][2]
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DETAILTESTSTEP'][detailteststep] = detailteststepresult
                            if not detailteststep in DATA['teststep'][teststep]['DETAILTESTSTEP']:
                                DATA['teststep'][teststep]['DETAILTESTSTEP'].append(detailteststep)
                            #DATA['SN']['ERROR'][teststep][sn]
                            reportdetailstep = "{} <{}>".format(detailteststep, detailteststepresult)
                            if not detailteststepresult == 'PASS':
                                if not reportdetailstep in DATA['SN']['ERROR'][teststep][sn]['DETAIL']:
                                    DATA['SN']['ERROR'][teststep][sn]['DETAIL'][reportdetailstep] = 1
                                else:
                                    DATA['SN']['ERROR'][teststep][sn]['DETAIL'][reportdetailstep] += 1

                                if not reportdetailstep in DATA['SN']['ERROR'][teststep][sn]['LIST']:
                                    DATA['SN']['ERROR'][teststep][sn]['LIST'].append(reportdetailstep)

                                if not reportdetailstep in TEMPSAVEERROR['DETAIL']:
                                    TEMPSAVEERROR['DETAIL'][reportdetailstep] = 1
                                else:
                                    TEMPSAVEERROR['DETAIL'][reportdetailstep] += 1

                                if not reportdetailstep in TEMPSAVEERROR['LIST']:
                                    TEMPSAVEERROR['LIST'].append(reportdetailstep)
                            tempeachtestresult = dict()
                            tempeachtestresult[detailteststep] = detailteststepresult
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TESTSTEPLIST'].append(tempeachtestresult)
                                    
                                    
                        #nic_test_err_reg_exp            
                        sub_match = re.findall(nic_test_err_reg_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            detailerror = sub_match[0]

                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'].append(detailerror)

                        sub_match = re.findall(nic_test_failure_reg_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            detailerror = sub_match[0]

                            failurelist.append(detailerror)                               
                                    
                        sub_match = re.findall(nic_test_rslt_final_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            getsomething = True
                            testslot = sub_match[0][1]
                            carttype = sub_match[0][2]
                            finalresult = sub_match[0][3]
                            
                            #NIC-09 ORTANO2 FLM21100052 NIC_DIAG_REGRESSION_TEST_FAIL

                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] = testslot
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CARDTYPE'] = carttype
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['FINALRESULT'] = finalresult
                            if not 'result' in DATA['SN']['FIRST'][teststep][sn]:
                                DATA['SN']['FIRST'][teststep][sn]['result'] = finalresult
                                DATA['SN']['FIRST'][teststep][sn]['checktime'] = CHECKTESTTIME
                                DATA['SN']['FIRST'][teststep][sn]['testfile'] = eachfile
                                DATA['SN']['FIRST'][teststep][sn]['ERROR'] = TEMPSAVEERROR
                            elif CHECKTESTTIME < DATA['SN']['FIRST'][teststep][sn]['checktime']:
                                    DATA['SN']['FIRST'][teststep][sn]['result'] = finalresult
                                    DATA['SN']['FIRST'][teststep][sn]['checktime'] = CHECKTESTTIME
                                    DATA['SN']['FIRST'][teststep][sn]['testfile'] = eachfile 
                                    DATA['SN']['FIRST'][teststep][sn]['ERROR'] = TEMPSAVEERROR                              
                            if not 'result' in DATA['SN']['LAST'][teststep][sn]:
                                DATA['SN']['LAST'][teststep][sn]['result'] = finalresult
                                DATA['SN']['LAST'][teststep][sn]['checktime'] = CHECKTESTTIME
                                DATA['SN']['LAST'][teststep][sn]['count'] = 1
                                DATA['SN']['LAST'][teststep][sn]['testfile'] = eachfile
                                DATA['SN']['LAST'][teststep][sn]['ERROR'] = TEMPSAVEERROR
                            elif CHECKTESTTIME > DATA['SN']['LAST'][teststep][sn]['checktime']:
                                    DATA['SN']['LAST'][teststep][sn]['result'] = finalresult
                                    DATA['SN']['LAST'][teststep][sn]['checktime'] = CHECKTESTTIME
                                    DATA['SN']['LAST'][teststep][sn]['count'] += 1
                                    DATA['SN']['LAST'][teststep][sn]['testfile'] = eachfile
                                    DATA['SN']['LAST'][teststep][sn]['ERROR'] = TEMPSAVEERROR

                        ## [2021-03-23_16-26-57] LOG: [MTP-207]: [NIC-01]: CPLD1 image: naples200_Ortano2_rev3_2_03112021.bin
                        ## [2021-03-23_16-26-57] LOG: [MTP-207]: [NIC-01]: CPLD2 image: naples200_Ortano2_failsafe_rev3_2_03112021.bin

                        sub_match = re.findall(nic_test_rslt_cpld_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            cpldslot = sub_match[0][0]
                            cpldlocate = sub_match[0][1]
                            if sub_match[0][2]:
                                cpldlocate = cpldlocate + str(sub_match[0][2])
                            version = sub_match[0][3]
                            #DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DETAILTESTSTEP'][detailteststep] = detailteststepresult
                            if not cpldlocate in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD']:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'][cpldlocate] = dict()
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'][cpldlocate][cpldslot] = version
                            #print(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'])
                            #sys.exit()
                            
                        sub_match = re.findall(nic_test_rslt_qspi_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            qspislot = sub_match[0][0]
                            qspilocate = sub_match[0][1]
                            version = sub_match[0][3]
                            #DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DETAILTESTSTEP'][detailteststep] = detailteststepresult
                            if not 'QSPI' in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['QSPI'] = dict()
                                
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['QSPI'][qspislot] = version
                            #print(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'])
                            #sys.exit()
                            
                        ## [2021-04-28_22-09-08] LOG: [MTP-208]: [NIC-04]: SN = FLM211000A6; MAC = 00-AE-CD-F6-0E-C8; PN = 68-0021-02 01
                        
                        sub_match = re.findall(nic_test_rslt_pn_exp, x)
                        #print(nic_test_rslt_pn_exp)
                        #sys.exit()
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            macaddress = sub_match[0][0]
                            pn = sub_match[0][1]
                            dom = sub_match[0][2]
                            #[('00-AE-CD-F6-0F-40', '68-0021-02 01', '042921')] [0][0|1|2|]

                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['MACADDRESS'] = macaddress
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['PN'] = pn
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DOM'] = dom
                            #sys.exit()
                                
                    f.close()

                    if not getsomething:
                        continue
                    #print(json.dumps(tempmtpinformation, indent = 4))
                    #print(json.dumps(tempnicinformation, indent = 4))
                    #tempnicresult
                    #print(json.dumps(tempnicresult, indent = 4))
                    #tempimage
                    #print(json.dumps(tempimage, indent = 4))
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['IMAGE'] = None
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['NICINFO'] = None
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['NICRESULT'] = None

                    if DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] in tempimage:
                        DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['IMAGE'] = tempimage[DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT']]
                    if DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] in tempnicinformation:
                        DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['NICINFO'] = tempnicinformation[DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT']]
                    if DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] in tempnicresult:
                        DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['NICRESULT'] = tempnicresult[DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT']]

                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['MTPINFO'] = tempmtpinformation
                    #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['MTPINFO'] = tempmtpinformation
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['NICINFO'] = tempnicinformation
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['NICRESULT'] = tempnicresult
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['NICIMAGE'] = tempimage
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['LOG'] = eachfile
                    #print(json.dumps(DATA['MTPCHASSIS'], indent = 4))
                    #sys.exit()

                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TESTTIME'] = findtotaltesttime(tempsavelog)
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPERATURE'] = findtesttemperature(tempsavelog)
                    if len(failurelist) > 0:
                        #print(failurelist)
                        checkslot = "NIC-{}".format(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'])
                        if "KEY" in inputconfig:
                            checkslot = "{}-{}".format(inputconfig["KEY"],DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'])
                        #print(checkslot)
                        for eachfailuremessage in failurelist:
                            if checkslot in eachfailuremessage:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'].append(eachfailuremessage)
                        #print(nic_test_failure_reg_exp)
                        #print(nic_test_rslt_final_exp)
                        #print(json.dumps(DATA['teststep'][teststep]['SN'], indent = 4))
                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'], indent = 4))

                    if len(temperrormessagelist) > 0:
                        checkslot = "NIC-{}".format(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'])
                        if "KEY" in inputconfig:
                            checkslot = "{}-{}".format(inputconfig["KEY"],DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'])
                        for eacherrormessage in temperrormessagelist:
                            if checkslot in eacherrormessage:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'].append(eacherrormessage)
                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'], indent = 4))

                        #sys.exit()

                    #print(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'])
                    #sys.exit() 


                if "TYPE" in inputconfig:
                    if inputconfig["TYPE"] == "NIC":
                        getallunzipfilelist = getallfilebyfolder(inputconfig['DIR']['TEMPDIR'], 'NIC')
                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                        #print(json.dumps(getallunzipfilelist, indent = 4))
                        if "SLOT" in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]:
                            if DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]["SLOT"]:
                                cardslot = "NIC-{}" .format(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]["SLOT"])
                                #print("CARD: {}".format(cardslot))
                                for eachniccardlog in getallunzipfilelist:
                                    if cardslot in eachniccardlog:
                                        #print("LOG: {}".format(eachniccardlog))
                                        f = open(eachniccardlog, 'r', encoding="ISO-8859-1")
                                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                                        for x in f:
                                            sub_match = re.findall(KEY_WORD.GETELBAASICDIEID, x)
                                            if sub_match:
                                                #print(x)
                                                #print(sub_match)                                     
                                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]["ELBA_DIE_ID"] = sub_match[0]

                                                
                                                #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                                                #sys.exit()

                                        if not "ELBA_DIE_ID" in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]:
                                            for x in f:
                                                if sub_match:
                                                    #print(x)
                                                    #print(sub_match)                                     
                                                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]["ELBA_DIE_ID"] = sub_match[0]

                                                    
                                                    #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                                                    #sys.exit()
                                                #print(x)
                                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                                        #sys.exit()
                                        f.close()
                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG'] = dict()
                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG']['LIST'] = list()
                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG']['DICT'] = dict()
                getallunzipfilelist = getallfilebyfolder(inputconfig['DIR']['TEMPDIR'], 'mtp_diag.log')

                #pr['modules'].print_anyinformation(getallunzipfilelist)

                for eachniccardlog in getallunzipfilelist:
                    f = open(eachniccardlog, 'r', encoding="ISO-8859-1")
                    #pr['modules'].print_anyinformation(f)
                    temploglist = list()
                    tempgrouplog = ''
                    startrecordtemplog = False
                    count = 0
                    for x in f:
                        sub_match = re.findall(KEY_WORD.TEMPGROUPFROMMFGLOGSTART, x)
                        if sub_match:
                            startrecordtemplog = True

                        if startrecordtemplog:
                            tempgrouplog += x 
                            if count > 1:
                                sub_match = re.findall(KEY_WORD.TEMPGROUPFROMMFGLOGEND, x)
                                if sub_match:
                                    startrecordtemplog = False
                                    temploglist.append(tempgrouplog)
                                    tempgrouplog = ''
                                    count = 0
                            count += 1

                    f.close()
                    #pr['modules'].print_anyinformation(temploglist)
                    for eachstr in temploglist:
                        #print(eachstr)
                        sub_match = re.findall(KEY_WORD.GETEMPINFORMATION, eachstr)
                        if sub_match:
                            #print(sub_match)
                            listofname = list()
                            listofdata = list()
                            recorddate = None 
                            recordtime = None 

                            for eachmatch in sub_match:
                                if not recorddate:
                                    recorddate = eachmatch[0]
                                if not recordtime:
                                    recordtime = eachmatch[1].replace(':','-')

                                if eachmatch[2] == 'NAME':
                                    for eachkey in eachmatch[3].split():
                                        listofname.append(eachkey)
                                if eachmatch[2] == 'FAN':
                                    for eachvalue in eachmatch[3].split():
                                        listofdata.append(eachvalue)

                            recorddateandtime = "{}_{}".format(recorddate,recordtime)
                            #print(recorddateandtime)
                            #print(listofname)
                            #print(listofdata)
                            if not recorddateandtime in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG']['LIST']:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG']['LIST'].append(recorddateandtime)
                            if not recorddateandtime in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG']['DICT']:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG']['DICT'][recorddateandtime] = dict()
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG']['DICT'][recorddateandtime]['NAME'] = listofname
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TEMPFROMMFGLOG']['DICT'][recorddateandtime]['VALUE'] = listofdata

                        #sys.exit()

                #pr['modules'].print_anyinformation(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME])
                #sys.exit()

                try:
                    shutil.rmtree(inputconfig['DIR']['TEMPDIR'])
                except OSError as e:
                    print("Error: %s : %s" % (inputconfig['DIR']['TEMPDIR'], e.strerror))  

          
    for teststep in teststeplist:   
        jsonfileoutputname = "{}/{}_{}_DATA.json".format(inputconfig['DIR']["TEMPDATABASE"],teststep,date_time)
        #mpu.io.write(jsonfileoutputname, DATA['teststep'][teststep])
        pr['modules'].wirtejsonfile(jsonfileoutputname, DATA['teststep'][teststep])
    
    return None

def findtesttemperature(f):
    
    #print("findtesttemperature")

    listoftemp = list()

    #print(f)

    match = re.findall(KEY_WORD.FINDTEMPWITHTIMESTAMP, f)

    if match:
        #print(match)
        listoftemp = match
        #sys.exit()

    return listoftemp

def findtotaltesttime(f):
    #print(f)
    match = re.findall(KEY_WORD.FINDDATEANDTIME, f) 

    #print(match)

    if match:
        #print(match[0])

        #print(match[-1])

        starttime = "{}_{}".format(match[0][0], match[0][1])
        endtime = "{}_{}".format(match[-1][0], match[-1][1])
        if starttime > endtime:
            #print(match)
            #print(len(match))
            #sys.exit()
            for x in range(len(match)):
                starttime = "{}_{}".format(match[x][0], match[x][1])
                if endtime > starttime:
                    break
        startdatetime_object = datetime.strptime(starttime, '%Y-%m-%d_%H-%M-%S')
        enddatetime_object = datetime.strptime(endtime, '%Y-%m-%d_%H-%M-%S')

        #print(type(startdatetime_object))
        #print(startdatetime_object)  # printed in default format

        #print(type(enddatetime_object))
        #print(enddatetime_object)  # printed in default format

        difftime = enddatetime_object-startdatetime_object
        #print('Done Time: ', difftime)       
        #print("How many seconds use?: {} seconds".format(difftime.total_seconds())) 
        returninseconds = "{}".format(difftime.total_seconds())
        #sys.exit()
        return returninseconds

    else:
        return None

def cal_total_testTime(starttime,endtime):

    startdatetime_object = datetime.strptime(starttime, '%Y-%m-%d_%H-%M-%S')
    enddatetime_object = datetime.strptime(endtime, '%Y-%m-%d_%H-%M-%S')
    difftime = enddatetime_object-startdatetime_object
    returninseconds = int(float("{}".format(difftime.total_seconds())))

    return returninseconds

def generateexeclsnFSJstatus(DATA, workingonSNlist, status,wb,Withallerror=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnFSJstatus: {}".format(status))
    titlename = status
    if Withallerror:
        titlename = "{} All Error".format(status)
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('TEST_WEEK')
    for test in DATA['SN']['TEST']:
         wirtedata.append(test)
    ws2.append(wirtedata)
    
    #print(json.dumps(DATA['SN']['LIST'], indent = 4))
    #time.sleep(2)

    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        overall = dict()
        overall['result'] = 'PASS'
        overall['TESTWEEK'] = 'NODATA'
        overall['LAST'] = dict()
        overall['LAST']['test'] = None
        overall['LAST']['result'] = None  
        overall['LAST']['date'] = None 
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                firsttesttimestamp = DATA['SN'][status][test][sn]["checktime"]
                break
        testcount = 0
        for test in DATA['SN']['TEST']:

            if sn in DATA['SN'][status][test]:

                testcount += 1
                if "result" in DATA['SN'][status][test][sn]:
                    if not overall['LAST']['result']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])
                    elif DATA['SN'][status][test][sn]["checktime"] > overall['LAST']['date']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])                     
                    if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                        overall['result'] = "FAIL"

        if overall['result'] == 'PASS':
            if not testcount == len(DATA['SN']['TEST']):
                overall['result'] = "INCOMPLETE"
        #wirtedata.append(overall['result'])
        wirtedata.append(overall['TESTWEEK'])

        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            if sn in DATA['SN'][status][test]:
                #print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                #time.sleep(5)
                if "result" in DATA['SN'][status][test][sn]:
                    #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                    dateandtime = DATA['SN'][status][test][sn]["checktime"]
                    datetimearray = dateandtime.split('_')
                    reportresut = "{} ({})".format(DATA['SN'][status][test][sn]["result"],datetimearray[0])
                       
                    wirtedata.append(reportresut)
                else:
                    wirtedata.append('NO TEST RESULT')
            else:
                wirtedata.append('NO TEST DATA')
            #time.sleep(2)
        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')
    freezePosition(ws2,'C2')
    
    
    return 0

def getsnlistafteestartdate(DATA,inputconfig,startdate=None):

    returnSNlist = list()

    if startdate:
        for sn in DATA['SN']['LIST']:
            for test in DATA['SN']['TEST']:
                if sn in DATA['SN']['LAST'][test]:
                    if "result" in DATA['SN']['LAST'][test][sn]:
                        #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                        dateandtime = DATA['SN']['LAST'][test][sn]["checktime"]
                        datetimearray = dateandtime.split('_')
                        #print("SN: {} -> {} vs {}".format(sn, datetimearray[0], startdate))
                        if datetimearray[0] >= startdate:
                            if not sn in returnSNlist:
                                returnSNlist.append(sn)
                                returnSNlist.sort()
                                #sys.exit()
                break

    returnSNlist2 = list()

    if len(returnSNlist) == 0:
        returnSNlist = DATA['SN']['LIST']

    for SN in returnSNlist:
        recordSN = True
        if "SKIPPrefix" in inputconfig:
            if inputconfig["SKIPPrefix"] in SN:
                recordSN = False
        if "STARTCountSN" in inputconfig:
            if SN < inputconfig["STARTCountSN"]:
                recordSN = False
        if recordSN:
            if not SN in returnSNlist2:
                returnSNlist2.append(SN)  
                returnSNlist2.sort()

    return returnSNlist2

def getsnmaclist(DATA,inputconfig):

    returnSNlist = dict()

    statuslist = ['FIRST', 'LAST']

    for sn in DATA['SN']['LIST']:
        if not sn in returnSNlist:
            returnSNlist[sn] = dict()
            returnSNlist[sn]['MAC'] = list()
            returnSNlist[sn]['PN'] = list()
        for test in DATA['SN']['TEST']:
            for status in statuslist:
                if sn in DATA['SN'][status][test]:
                    if "NICINFO" in DATA['SN'][status][test][sn]:
                        try:
                            if not "FRU" in DATA['SN'][status][test][sn]["NICINFO"]:
                                continue
                            frudata = DATA['SN'][status][test][sn]["NICINFO"]["FRU"]
                            #print("FRU: {}".format(frudata))
                            frudatalist = frudata.split(', ')
                            #print(frudatalist)
                            if not frudatalist[1] in returnSNlist[sn]['MAC']:
                                returnSNlist[sn]['MAC'].append(frudatalist[1])
                            if not frudatalist[2] in returnSNlist[sn]['PN']:
                                returnSNlist[sn]['PN'].append(frudatalist[2])
                            #print(returnSNlist)
                            #sys.exit()
                        except:
                            print("An exception occurred in SN {}, {} test: {}".format(sn,status,test))

    return returnSNlist

def getmacsnlist(DATA,inputconfig):

    returnSNlist = dict()

    statuslist = ['FIRST', 'LAST']

    for sn in DATA['SN']['LIST']:
        for test in DATA['SN']['TEST']:
            for status in statuslist:
                if sn in DATA['SN'][status][test]:
                    if "NICINFO" in DATA['SN'][status][test][sn]:
                        try:
                            if not "FRU" in DATA['SN'][status][test][sn]["NICINFO"]:
                                continue
                            frudata = DATA['SN'][status][test][sn]["NICINFO"]["FRU"]
                            #print("FRU: {}".format(frudata))
                            frudatalist = frudata.split(', ')
                            #print(frudatalist)
                            if not frudatalist[1] in returnSNlist:
                                returnSNlist[frudatalist[1]] = dict()
                                returnSNlist[frudatalist[1]]['SN'] = list()
                                returnSNlist[frudatalist[1]]['PN'] = list()
                            if not frudatalist[0] in returnSNlist[frudatalist[1]]['SN']:
                                returnSNlist[frudatalist[1]]['SN'].append(frudatalist[0])
                            if not frudatalist[2] in returnSNlist[frudatalist[1]]['PN']:
                                returnSNlist[frudatalist[1]]['PN'].append(frudatalist[2])
                            #print(returnSNlist)
                            
                        except:
                            print("An exception occurred in SN {}, {} test: {}".format(sn,status,test))
                        #sys.exit()

    return returnSNlist

def checksnlistafteestartdate(DATA,snlist,startdate=None):

    returnSNlist = list()
    print("{} Count SN with start date {}: {}".format("checksnlistafteestartdate",startdate,len(snlist)))

    if startdate:
        for sn in DATA['SN']['LIST']:
            for test in DATA['SN']['TEST']:
                if sn in DATA['SN']['LAST'][test]:
                    if "result" in DATA['SN']['LAST'][test][sn]:
                        #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                        dateandtime = DATA['SN']['LAST'][test][sn]["checktime"]
                        datetimearray = dateandtime.split('_')
                        #print("SN: {} -> {} vs {}".format(sn, datetimearray[0], startdate))
                        if datetimearray[0] >= startdate:
                            if not sn in returnSNlist:
                                returnSNlist.append(sn)
                                returnSNlist.sort()
                                #sys.exit()
                break

    returnSNlist2 = snlist

    if len(returnSNlist) > 0:
        for sn in snlist:
            if not sn in returnSNlist:
                returnSNlist2.remove(sn)
    
    print("{} return Count SN with start date {}: {}".format("checksnlistafteestartdate",startdate,len(snlist)))

    return returnSNlist2

def generateexeclsn4CFailurestatus(DATA, workingonSNlist,status,wb,inputconfig,Withallerror=False,duplicatesn=True,byweek=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsn4CFailurestatus: {}".format(status))
    titlename = "Top4C_F_By{}".format(status)
    if not duplicatesn:
        titlename = "{}_NoDupSN".format(titlename)
    print("TITLE: {}".format(titlename))
    topfailuredata = dict()
    topfailuredata['week'] = list()
    topfailuredata['weekday'] = dict()
    topfailuredata['TEST'] = DATA['SN']['TEST']
    topfailuredata['DATA'] = dict()
    topfailuredata['OVERALL'] = dict()
    topfailuredata['OVERALL']['TEST'] = dict()
    topfailuredata['OVERALL']['DATA'] = dict()
    topfailuredata['OVERALL']['TEST_HVLV'] = dict()
    topfailuredata['OVERALL']['TEST_HVLV']["DATA"] = dict()
    topfailuredata['OVERALL']['TEST_HVLV']["TEST"] = list()
    topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"] = dict()
    topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'] = dict()
    topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list'] = list()
    topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'] = dict()

    for test in DATA['SN']['TEST']:
        print("{}: {}".format("generateexeclsn4CFailurestatus",test))
    for sn in workingonSNlist:
        for test in DATA['SN']['TEST']:
            #print("{}: {}".format("generateexeclsn4CFailurestatus",test))
            if not "4C" in test:
                continue
            if not test in topfailuredata['OVERALL']['TEST']:
                topfailuredata['OVERALL']['TEST'][test] = dict()
            if not test in topfailuredata['OVERALL']['DATA']:
                topfailuredata['OVERALL']['DATA'][test] = dict()
            if sn in DATA['SN'][status][test]:
                testweek = findworkweek(DATA['SN'][status][test][sn]["checktime"])
                if not testweek in topfailuredata['weekday']:
                    topfailuredata['weekday'][testweek] = list()
                testdayandtime = DATA['SN'][status][test][sn]["checktime"].split("_")
                testday = testdayandtime[0]
                if not testday in topfailuredata['weekday'][testweek]:
                    topfailuredata['weekday'][testweek].append(testday)
                    topfailuredata['weekday'][testweek].sort()
                if not testweek in topfailuredata['DATA']:
                    topfailuredata['DATA'][testweek] = dict()
                if not test in topfailuredata['DATA'][testweek]:
                    topfailuredata['DATA'][testweek][test] = dict()
                if not testweek in topfailuredata['week']:
                    topfailuredata['week'].append(testweek)
                    topfailuredata['week'].sort(reverse=True)

                testresult = DATA['SN'][status][test][sn]["result"]
                if 'FAIL' in testresult:
                    #print(json.dumps(DATA['SN'][status][test][sn]["ERROR"], indent = 4))
                    if len(DATA['SN'][status][test][sn]["ERROR"]["LIST"]):
                        if not DATA['SN'][status][test][sn]["ERROR"]["LIST"][0] in topfailuredata['DATA'][testweek][test]:
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]] = dict()
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] = 1
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"] = list()
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        else:
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] += 1
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                    else:
                        if not "UNKNOWN" in topfailuredata['DATA'][testweek][test]:
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"] = dict()
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["count"] = 1
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"] = list()
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"].append(sn)
                        else:
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["count"] += 1
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"].append(sn)   
                    if len(DATA['SN'][status][test][sn]["ERROR"]["LIST"]):
                        if not DATA['SN'][status][test][sn]["ERROR"]["LIST"][0] in topfailuredata['OVERALL']['TEST'][test]:
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]] = dict()
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] = 1
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"] = list()
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        else:
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] += 1
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        test_HVLV = DATA['SN'][status][test][sn]["ERROR"]["LIST"][0][:2]
                        testdetail = DATA['SN'][status][test][sn]["ERROR"]["LIST"][0][3:]
                        testdetail = testdetail.replace("FAILED", "FAIL")
                        testdetail = testdetail.replace("FAILURE", "FAIL")
                        if "HV" == test_HVLV or "LV" == test_HVLV:
                            newtest = "{} ({})".format(test,test_HVLV)
                            if not newtest in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
                                topfailuredata['OVERALL']['TEST_HVLV']["TEST"].append(newtest)
                                topfailuredata['OVERALL']['TEST_HVLV']["TEST"].sort()
                            if not testdetail in topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"]:
                                topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"][testdetail] = 1
                            else:
                                topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"][testdetail] += 1
                            #print("{} | {}".format(newtest,testdetail))
                            if not newtest in topfailuredata['OVERALL']['TEST_HVLV']["DATA"]:
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest] = dict()
                            if not testdetail in topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest]:
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail] = dict()
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["count"] = 1
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["SN"] = list()
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["SN"].append(sn)
                            else:
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["count"] += 1
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["SN"].append(sn)

                            if not testweek in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list']:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list'].append(testweek)
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list'].sort(reverse=True)

                            if not testweek in topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY']:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek] = list()

                            if not testday in topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek].append(testday)
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek].sort()

                            if not testweek in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK']:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek] = dict()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"] = dict()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"] = list()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"] = dict()                                

                            if not newtest in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"].append(newtest)
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"].sort()
                            if not testdetail in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"][testdetail] = 1
                            else:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"][testdetail] += 1
                            #print("{} | {}".format(newtest,testdetail))
                            if not newtest in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest] = dict()
                            if not testdetail in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail] = dict()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["count"] = 1
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["SN"] = list()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["SN"].append(sn)
                            else:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["count"] += 1
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["SN"].append(sn)

                    else:
                        if not "UNKNOWN" in topfailuredata['OVERALL']['TEST'][test]:
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"] = dict()
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["count"] = 1
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["SN"] = list()
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["SN"].append(sn)
                        else:
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["count"] += 1
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["SN"].append(sn)  


                    if not duplicatesn:
                        break
    #print(json.dumps(topfailuredata, indent = 4))                    
    #sys.exit()



    if "4CFLOW" in inputconfig:
        backuplist = topfailuredata['OVERALL']['TEST_HVLV']["TEST"]
        topfailuredata['OVERALL']['TEST_HVLV']["TEST"] = inputconfig["4CFLOW"]
        for teststep in backuplist:
            if not teststep in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
                topfailuredata['OVERALL']['TEST_HVLV']["TEST"].append(teststep)

    if not byweek:
        ws2 = wb.create_sheet(title=titlename)
        

        wirtedata = list()
        wirtedata.append("4C Process sequence")
        count = 1
        for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
            wirtedata.append(count)
            count += 1
        ws2.append(wirtedata)
        wirtedata = list()
        wirtedata.append("FAILURE TYPE")
        for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
             wirtedata.append(test)
        wirtedata.append("TOTAL")
        ws2.append(wirtedata)
        testtotal = dict()
        for failuretype, number in sorted(topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"].items(), key=lambda item: item[1],reverse=True):
            print("Failure: {} - {}".format(failuretype,number))
            wirtedata = list()
            wirtedata.append(failuretype)
            
            for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
                if not test in testtotal:
                    testtotal[test] = 0
                if test in topfailuredata['OVERALL']['TEST_HVLV']["DATA"]:
                    if failuretype in topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test]:
                        wirtedata.append(topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test][failuretype]["count"])
                        testtotal[test] += topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test][failuretype]["count"]
                    else:
                        wirtedata.append("")
                else:
                    wirtedata.append("")
            wirtedata.append(number)
            ws2.append(wirtedata)

        wirtedata = list()
        wirtedata.append("TOTAL")
        alltotal = 0
        for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
            if test in testtotal:
                wirtedata.append(testtotal[test])
                alltotal += testtotal[test]
        wirtedata.append(alltotal)
        ws2.append(wirtedata)
        wirtedata = list()
        ws2.append(wirtedata)
        wirtedata = list()
        wirtedata.append("TEST")
        wirtedata.append("SN")
        wirtedata.append("FAILURE TYPE")
        wirtedata.append("RE-TEST COUNT")
        wirtedata.append("LAST RESULT")
        ws2.append(wirtedata)
        for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
            if test in topfailuredata['OVERALL']['TEST_HVLV']["DATA"]:
                for failuretype in topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test]:
                    for SN in topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test][failuretype]["SN"]:
                        wirtedata = list()
                        wirtedata.append(test)
                        wirtedata.append(SN)
                        wirtedata.append(failuretype)
                        testnamelist = test.split(' ')
                        wirtedata.append(DATA['SN']["LAST"][testnamelist[0]][SN]["count"])
                        wirtedata.append(DATA['SN']["LAST"][testnamelist[0]][SN]["result"])
                        ws2.append(wirtedata)
        wirtedata = list()
        ws2.append(wirtedata)
        ws2.append(wirtedata)
        ws2.append(wirtedata)

    else:
        titlename = "{}_byWeek".format(titlename)
        ws2 = wb.create_sheet(title=titlename)
        
        for testweek in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list']:
            wirtedata = list()
            wirtedata.append("Week: {}".format(testweek))
            wirtedata.append("START:")
            wirtedata.append(topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek][0])
            wirtedata.append("END:")
            wirtedata.append(topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek][-1])            
            ws2.append(wirtedata)

            wirtedata = list()
            wirtedata.append("4C Process sequence")
            count = 1
            if "4CFLOW" in inputconfig:
                backuplist = topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]
                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"] = inputconfig["4CFLOW"]
                for teststep in backuplist:
                    if not teststep in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                        topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"].append(teststep)

            for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                wirtedata.append(count)
                count += 1
            ws2.append(wirtedata)
            wirtedata = list()
            wirtedata.append("FAILURE TYPE")
            for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                 wirtedata.append(test)
            wirtedata.append("TOTAL")
            ws2.append(wirtedata)        
            testtotal = dict()
            for failuretype, number in sorted(topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"].items(), key=lambda item: item[1],reverse=True):
                print("Failure: {} - {}".format(failuretype,number))
                wirtedata = list()
                wirtedata.append(failuretype)
                
                for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                    if not test in testtotal:
                        testtotal[test] = 0
                    if test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"]:
                        if failuretype in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test]:
                            wirtedata.append(topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test][failuretype]["count"])
                            testtotal[test] += topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test][failuretype]["count"]
                        else:
                            wirtedata.append("")
                    else:
                        wirtedata.append("")
                wirtedata.append(number)
                ws2.append(wirtedata)

            wirtedata = list()
            wirtedata.append("TOTAL")
            alltotal = 0
            for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                wirtedata.append(testtotal[test])
                alltotal += testtotal[test]
            wirtedata.append(alltotal)
            ws2.append(wirtedata)
            wirtedata = list()
            ws2.append(wirtedata)
            wirtedata = list()
            wirtedata.append("TEST")
            wirtedata.append("SN")
            wirtedata.append("FAILURE TYPE")
            wirtedata.append("RE-TEST COUNT")
            wirtedata.append("LAST RESULT")
            ws2.append(wirtedata)
            for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                if test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"]:
                    for failuretype in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test]:
                        for SN in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test][failuretype]["SN"]:
                            wirtedata = list()
                            wirtedata.append(test)
                            wirtedata.append(SN)
                            wirtedata.append(failuretype)
                            testnamelist = test.split(' ')
                            wirtedata.append(DATA['SN']["LAST"][testnamelist[0]][SN]["count"])
                            wirtedata.append(DATA['SN']["LAST"][testnamelist[0]][SN]["result"])
                            ws2.append(wirtedata)
            wirtedata = list()
            ws2.append(wirtedata)
            ws2.append(wirtedata)
            ws2.append(wirtedata)

    fixcolumnssize(ws2,False)
    highlightnumberincell(ws2)
    #highlightinyellow(ws2,'TIMEOUT')
    #highlightinyellow(ws2,'NO TEST DATA')
    #highlightinred(ws2, 'FAIL')
    #highlightinred(ws2, 'FAILED')
    #highlightingreen(ws2,'PASS')
    #highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    #freezePosition(ws2,'C2')
    
    return 0

def generateexeclsnTopFailurestatus(DATA, workingonSNlist,status,wb,inputconfig,Withallerror=False,byweek=True):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnTopFailurestatus: {}".format(status))
    titlename = "TopFailureBy{}".format(status)
    print("TITLE: {}".format(titlename))
    topfailuredata = dict()
    topfailuredata['week'] = list()
    topfailuredata['weekday'] = dict()
    topfailuredata['TEST'] = DATA['SN']['TEST']
    topfailuredata['DATA'] = dict()
    topfailuredata['TESTSNLIST'] = dict()
    alltopfailuredata = dict()
    alltopfailuredata['TEST'] = DATA['SN']['TEST']
    alltopfailuredata['DATA'] = dict()
    alltopfailuredata['DATA2'] = dict()
    alltopfailuredata['TESTSNLIST'] = dict()

    for sn in workingonSNlist:
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                testweek = findworkweek(DATA['SN'][status][test][sn]["checktime"])
                if not testweek in topfailuredata['weekday']:
                    topfailuredata['weekday'][testweek] = list()
                if not testweek in topfailuredata['TESTSNLIST']:
                    topfailuredata['TESTSNLIST'][testweek] = dict()
                if not test in topfailuredata['TESTSNLIST'][testweek]:
                    topfailuredata['TESTSNLIST'][testweek][test] = list()
                if not test in alltopfailuredata['TESTSNLIST']:
                    alltopfailuredata['TESTSNLIST'][test] = list()
                if not sn in topfailuredata['TESTSNLIST'][testweek][test]:
                    topfailuredata['TESTSNLIST'][testweek][test].append(sn)
                if not sn in alltopfailuredata['TESTSNLIST'][test]:
                    alltopfailuredata['TESTSNLIST'][test].append(sn)
                testdayandtime = DATA['SN'][status][test][sn]["checktime"].split("_")
                testday = testdayandtime[0]
                if not testday in topfailuredata['weekday'][testweek]:
                    topfailuredata['weekday'][testweek].append(testday)
                    topfailuredata['weekday'][testweek].sort()
                if not testweek in topfailuredata['DATA']:
                    topfailuredata['DATA'][testweek] = dict()
                if not test in topfailuredata['DATA'][testweek]:
                    topfailuredata['DATA'][testweek][test] = dict()
                if not test in alltopfailuredata['DATA']:
                    alltopfailuredata['DATA'][test] = dict()
                if not test in alltopfailuredata['DATA2']:
                    alltopfailuredata['DATA2'][test] = dict()
                if not testweek in topfailuredata['week']:
                    topfailuredata['week'].append(testweek)
                    topfailuredata['week'].sort(reverse=True)

                testresult = DATA['SN'][status][test][sn]["result"]
                if 'FAIL' in testresult:
                    #print(json.dumps(DATA['SN'][status][test][sn]["ERROR"], indent = 4))
                    if len(DATA['SN'][status][test][sn]["ERROR"]["LIST"]):
                        if not DATA['SN'][status][test][sn]["ERROR"]["LIST"][0] in topfailuredata['DATA'][testweek][test]:
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]] = dict()
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] = 1
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"] = list()
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        else:
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] += 1
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        failurestep = DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]
                        failurestep = getsimpleteststep(failurestep)
                        if not failurestep in alltopfailuredata['DATA2'][test]:
                            alltopfailuredata['DATA2'][test][failurestep] = dict()
                            alltopfailuredata['DATA2'][test][failurestep]["count"] = 1 
                            alltopfailuredata['DATA2'][test][failurestep]["SN"] = list() 
                            alltopfailuredata['DATA2'][test][failurestep]["SN"].append(sn)
                        else:
                            alltopfailuredata['DATA2'][test][failurestep]["count"] += 1
                            alltopfailuredata['DATA2'][test][failurestep]["SN"].append(sn)

                        if not DATA['SN'][status][test][sn]["ERROR"]["LIST"][0] in alltopfailuredata['DATA'][test]:
                            alltopfailuredata['DATA'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]] = dict()
                            alltopfailuredata['DATA'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] = 1 
                            alltopfailuredata['DATA'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"] = list() 
                            alltopfailuredata['DATA'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        else:
                            alltopfailuredata['DATA'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] += 1
                            alltopfailuredata['DATA'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                    else:
                        if not "UNKNOWN" in topfailuredata['DATA'][testweek][test]:
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"] = dict()
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["count"] = 1
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"] = list()
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"].append(sn)
                        else:
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["count"] += 1
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"].append(sn)
                        if not "UNKNOWN" in alltopfailuredata['DATA'][test]:
                            alltopfailuredata['DATA'][test]["UNKNOWN"] = dict()
                            alltopfailuredata['DATA'][test]["UNKNOWN"]["count"] = 1
                            alltopfailuredata['DATA'][test]["UNKNOWN"]["SN"] = list()
                            alltopfailuredata['DATA'][test]["UNKNOWN"]["SN"].append(sn)
                        else:
                            alltopfailuredata['DATA'][test]["UNKNOWN"]["count"] += 1
                            alltopfailuredata['DATA'][test]["UNKNOWN"]["SN"].append(sn)    
                    #print(json.dumps(topfailuredata, indent = 4))                    
                    #sys.exit()

    topfailuredata['OVERALL'] = dict()
    for testweek in topfailuredata['week']:
        for test in topfailuredata['TEST']:
            if not test in topfailuredata['DATA'][testweek]:
                continue
            for failuretype in topfailuredata['DATA'][testweek][test]:
                if not failuretype in topfailuredata['OVERALL']:
                    topfailuredata['OVERALL'][failuretype] = topfailuredata['DATA'][testweek][test][failuretype]["count"]
                else:
                    topfailuredata['OVERALL'][failuretype] += topfailuredata['DATA'][testweek][test][failuretype]["count"]

    alltopfailuredata['OVERALL'] = dict()
    for test in alltopfailuredata['TEST']:
        if not test in alltopfailuredata['DATA2']:
            continue
        for failuretype in alltopfailuredata['DATA2'][test]:
            if not failuretype in alltopfailuredata['OVERALL']:
                alltopfailuredata['OVERALL'][failuretype] = alltopfailuredata['DATA2'][test][failuretype]["count"]
            else:
                alltopfailuredata['OVERALL'][failuretype] += alltopfailuredata['DATA2'][test][failuretype]["count"]

    topfailuredata['OVERALLbyWeek'] = dict()
    for testweek in topfailuredata['week']:
        if not testweek in topfailuredata['OVERALLbyWeek']:
            topfailuredata['OVERALLbyWeek'][testweek] = dict()
        for test in topfailuredata['TEST']:
            if not test in topfailuredata['DATA'][testweek]:
                continue
            for failuretype in topfailuredata['DATA'][testweek][test]:
                if not failuretype in topfailuredata['OVERALLbyWeek'][testweek]:
                    topfailuredata['OVERALLbyWeek'][testweek][failuretype] = topfailuredata['DATA'][testweek][test][failuretype]["count"]
                else:
                    topfailuredata['OVERALLbyWeek'][testweek][failuretype] += topfailuredata['DATA'][testweek][test][failuretype]["count"]
    #print(json.dumps(topfailuredata, indent = 4))
    for testweek in sorted(topfailuredata['OVERALLbyWeek'],reverse=True):
        print("TEST Week: {}".format(testweek))
        for failuretype, number in sorted(topfailuredata['OVERALLbyWeek'][testweek].items(), key=lambda item: item[1],reverse=True):
            print("Failure: {} - {}".format(failuretype,number))

    #sys.exit()
    if byweek:
        ws2 = wb.create_sheet(title="{}_byWeek".format(titlename))
        
        for testweek in topfailuredata['week']:
            wirtedata = list()
            wirtedata.append("WEEK: {} {} Top Failure Type".format(testweek,status))
            wirtedata.append("START:")
            wirtedata.append(topfailuredata['weekday'][testweek][0])
            wirtedata.append("END:")
            wirtedata.append(topfailuredata['weekday'][testweek][-1])
            ws2.append(wirtedata)
            wirtedata = list()
            wirtedata.append("FAILURE TYPE")
            eachweektotallist = list()
            eachweektotoaldetail = dict()
            for test in topfailuredata['TEST']:
                writetestinformation = "{} <0>".format(test)
                if test in topfailuredata['TESTSNLIST'][testweek]:
                    writetestinformation = "{} <{}>".format(test,len(topfailuredata['TESTSNLIST'][testweek][test]))
                wirtedata.append(writetestinformation)
                eachweektotallist.append(test)
                eachweektotoaldetail[test] = 0
            wirtedata.append("TOTAL")
            eachweektotallist.append("TOTAL")
            eachweektotoaldetail["TOTAL"] = 0
            
            ws2.append(wirtedata)

            for failuretype, number in sorted(topfailuredata['OVERALLbyWeek'][testweek].items(), key=lambda item: item[1],reverse=True):
                print("Failure: {} - {}".format(failuretype,number))
                wirtedata = list()
                wirtedata.append(failuretype)
                for test in topfailuredata['TEST']:
                    if testweek in topfailuredata['DATA']:
                        if test in topfailuredata['DATA'][testweek]:
                            if failuretype in topfailuredata['DATA'][testweek][test]:
                                wirtedata.append(topfailuredata['DATA'][testweek][test][failuretype]["count"])
                                eachweektotoaldetail[test] += topfailuredata['DATA'][testweek][test][failuretype]["count"]
                            else:
                                wirtedata.append("")
                        else:
                            wirtedata.append("")
                    else:
                        wirtedata.append("")
                wirtedata.append(number)
                eachweektotoaldetail["TOTAL"] += number
                ws2.append(wirtedata)

            wirtedata = list()
            wirtedata.append("TOTAL")
            for test in eachweektotallist:
                wirtedata.append(eachweektotoaldetail[test])
            ws2.append(wirtedata)

            wirtedata = list()
            ws2.append(wirtedata)
            wirtedata = list()
            wirtedata.append("TEST")
            wirtedata.append("SN")
            wirtedata.append("FAILURE TYPE")
            wirtedata.append("RE-TEST COUNT")
            wirtedata.append("LAST RESULT")
            ws2.append(wirtedata)
            for test in topfailuredata['TEST']:
                if testweek in topfailuredata['DATA']:
                    if test in topfailuredata['DATA'][testweek]:
                        for failuretype in topfailuredata['DATA'][testweek][test]:
                            for SN in topfailuredata['DATA'][testweek][test][failuretype]["SN"]:
                                wirtedata = list()
                                wirtedata.append(test)
                                wirtedata.append(SN)
                                wirtedata.append(failuretype)
                                wirtedata.append(DATA['SN']["LAST"][test][SN]["count"])
                                wirtedata.append(DATA['SN']["LAST"][test][SN]["result"])
                                ws2.append(wirtedata)
            wirtedata = list()
            ws2.append(wirtedata)
            ws2.append(wirtedata)
            ws2.append(wirtedata)
    else:
        ws2 = wb.create_sheet(title=titlename)
        
        wirtedata = list()
        wirtedata.append("FAILURE TYPE")
        eachweektotallist = list()
        eachweektotoaldetail = dict()
        for test in topfailuredata['TEST']:
            writetestinformation = "{} <0>".format(test)
            if test in alltopfailuredata['TESTSNLIST']:
                writetestinformation = "{} <{}>".format(test,len(alltopfailuredata['TESTSNLIST'][test]))
            wirtedata.append(writetestinformation)
            eachweektotallist.append(test)
            eachweektotoaldetail[test] = 0
        wirtedata.append("TOTAL")
        eachweektotallist.append("TOTAL")
        eachweektotoaldetail["TOTAL"] = 0
        
        ws2.append(wirtedata)

        for failuretype, number in sorted(alltopfailuredata['OVERALL'].items(), key=lambda item: item[1],reverse=True):
            print("Failure: {} - {}".format(failuretype,number))
            wirtedata = list()
            wirtedata.append(failuretype)
            for test in alltopfailuredata['TEST']:
                if test in alltopfailuredata['DATA2']:
                    if failuretype in alltopfailuredata['DATA2'][test]:
                        wirtedata.append(alltopfailuredata['DATA2'][test][failuretype]["count"])
                        eachweektotoaldetail[test] += alltopfailuredata['DATA2'][test][failuretype]["count"]
                    else:
                        wirtedata.append("")
                else:
                    wirtedata.append("")
            wirtedata.append(number)
            eachweektotoaldetail["TOTAL"] += number
            ws2.append(wirtedata)

        wirtedata = list()
        wirtedata.append("TOTAL")
        for test in eachweektotallist:
            wirtedata.append(eachweektotoaldetail[test])
        ws2.append(wirtedata)

        wirtedata = list()
        ws2.append(wirtedata)
        wirtedata = list()
        wirtedata.append("TEST")
        wirtedata.append("SN")
        wirtedata.append("FAILURE TYPE")
        wirtedata.append("RE-TEST COUNT")
        wirtedata.append("LAST RESULT")
        ws2.append(wirtedata)
        for test in alltopfailuredata['TEST']:
            if test in alltopfailuredata['DATA']:
                for failuretype in alltopfailuredata['DATA'][test]:
                    for SN in alltopfailuredata['DATA'][test][failuretype]["SN"]:
                        wirtedata = list()
                        wirtedata.append(test)
                        wirtedata.append(SN)
                        wirtedata.append(failuretype)
                        wirtedata.append(DATA['SN']["LAST"][test][SN]["count"])
                        wirtedata.append(DATA['SN']["LAST"][test][SN]["result"])
                        ws2.append(wirtedata)
        wirtedata = list()
        ws2.append(wirtedata)
        ws2.append(wirtedata)
        ws2.append(wirtedata)


    fixcolumnssize(ws2,False)
    highlightnumberincell(ws2)
    #highlightinyellow(ws2,'TIMEOUT')
    #highlightinyellow(ws2,'NO TEST DATA')
    #highlightinred(ws2, 'FAIL')
    #highlightinred(ws2, 'FAILED')
    #highlightingreen(ws2,'PASS')
    #highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    #freezePosition(ws2,'C2')
    
    return 0

def getsimpleteststep(failureteststep):

    failureteststep = failureteststep.replace('LV_','')
    failureteststep = failureteststep.replace('HV_','')
    failureteststep = failureteststep.replace('DL ','')
    failureteststep = failureteststep.replace('FST ','')
    failureteststep = failureteststep.replace('SWI ','')
    #<FAIL>
    failureteststep = failureteststep.replace('<FAIL> ','<FAILED>')

    return failureteststep

def generateexeclsnstatus(DATA, workingonSNlist,status,wb,inputconfig,Withallerror=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnstatus: {}".format(status))
    titlename = status
    if Withallerror:
        titlename = "{} All Error".format(status)
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('OVERALL')
    wirtedata.append('TEST_WEEK')
    # if status == 'LAST':
    #     wirtedata.append('LAST_TESTDATE')
    #     wirtedata.append('LAST_STATION(RESULT)')
    
    #     wirtedata.append('QSPI')
    #     wirtedata.append('CPLD1')
    if "DISPLAYINDO" in inputconfig:
        for displaykey in inputconfig["DISPLAYINDO"]:
            wirtedata.append(displaykey)

    #wirtedata.append('')
    for test in DATA['SN']['TEST']:
         wirtedata.append(test)
    ws2.append(wirtedata)

    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        overall = dict()
        overall['result'] = 'PASS'
        overall['TESTWEEK'] = 'NODATA'
        overall['LAST'] = dict()
        overall['LAST']['test'] = None
        overall['LAST']['result'] = None  
        overall['LAST']['date'] = None 
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                firsttesttimestamp = DATA['SN'][status][test][sn]["checktime"]
                break

        testcount = 0
        for test in DATA['SN']['TEST']:

            if sn in DATA['SN'][status][test]:

                testcount += 1
                if "result" in DATA['SN'][status][test][sn]:
                    if not overall['LAST']['result']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])
                    elif DATA['SN'][status][test][sn]["checktime"] > overall['LAST']['date']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])                     
                    if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                        overall['result'] = "FAIL"

        if overall['result'] == 'PASS':
            if not testcount == len(DATA['SN']['TEST']):
                overall['result'] = "INCOMPLETE"
        wirtedata.append(overall['result'])
        wirtedata.append(overall['TESTWEEK'])
        if "DISPLAYINDO" in inputconfig:
            for displaykey in inputconfig["DISPLAYINDO"]:
                #finddisplay = False
                displaydict = dict()
                for test in DATA['SN']['TEST']:
                    if sn in DATA['SN'][status][test]:
                        for checkkey in inputconfig["CHECKKEYS"]:
                            if checkkey in DATA['SN'][status][test][sn]:
                                if isinstance(DATA['SN'][status][test][sn][checkkey], dict):
                                    if displaykey in DATA['SN'][status][test][sn][checkkey]:
                                        if not "data" in displaydict:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]
                                        elif DATA['SN'][status][test][sn]["checktime"] > displaydict["checktime"]:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]                                            
                                        #finddisplay = True
                                        #wirtedata.append(DATA['SN'][status][test][sn][checkkey][displaykey])
                                        #break
                if "data" in displaydict:
                    wirtedata.append(displaydict["data"])
                else:
                    wirtedata.append('NODATA')
        # if status == 'LAST':
        #     last_dateandtime = overall['LAST']['date']

        #     last_datetimearray = last_dateandtime.split('_')
        #     reporttestandresult = "{} ({})".format(overall['LAST']['test'],overall['LAST']['result'])
        #     wirtedata.append(last_datetimearray[0])
        #     wirtedata.append(reporttestandresult)
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'QSPI' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['QSPI'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #     #'CPLD1'
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'CPLD1' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['CPLD1'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #wirtedata.append('')

        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            #sys.exit()
            if sn in DATA['SN'][status][test]:
                # print("TEST: {} | STATUS: {}".format(test, status))
                # print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                # sys.exit()
                #time.sleep(5)

                if "result" in DATA['SN'][status][test][sn]:
                    #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                    dateandtime = DATA['SN'][status][test][sn]["checktime"]
                    datetimearray = dateandtime.split('_')
                    reportresut = "{} ({})".format(DATA['SN'][status][test][sn]["result"],datetimearray[0])
                    #DATA['SN']['LAST'][teststep][sn]['count']
                    if "count" in DATA['SN'][status][test][sn]:
                        reportresut = "{} <{}> ".format(reportresut,DATA['SN'][status][test][sn]["count"])
                    if Withallerror:
                        if sn in DATA['SN']['ERROR'][test]:
                            for eachfailstep in DATA['SN']['ERROR'][test][sn]['LIST']:
                                reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                    else:
                        if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                            for eachfailstep in DATA['SN'][status][test][sn]['ERROR']['LIST']:
                                reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                    wirtedata.append(reportresut)
                else:
                    wirtedata.append('NO TEST DATA')

            else:
                wirtedata.append('NO TEST DATA')
            #time.sleep(2)
        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')
    highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    freezePosition(ws2,'C2')
    
    return 0

def generateexeclsnfailurestatusalldata(DATA,pr, workingonSNlist,status,wb,inputconfig,Withallerror=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnfailurestatusalldata: {}".format(status))
    #titlename = "{} ALL DATE".format(status)
    titlename = "{}".format(status)
    if Withallerror:
        titlename = "{} ALL ERROR".format(titlename)
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('FLEXFLOWSTATUS')
    wirtedata.append('OVERALL')
    wirtedata.append('TEST_WEEK')
    # if status == 'LAST':
    #     wirtedata.append('LAST_TESTDATE')
    #     wirtedata.append('LAST_STATION(RESULT)')
    
    #     wirtedata.append('QSPI')
    #     wirtedata.append('CPLD1')
    if "DISPLAYINDO" in inputconfig:
        for displaykey in inputconfig["DISPLAYINDO"]:
            wirtedata.append(displaykey)

    #wirtedata.append('')
    for test in DATA['SN']['TEST']:
         wirtedata.append(test)
    ws2.append(wirtedata)

    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        wirtedata.append(pr["FailureSNlist"]["FailureSNFlexflow"][sn])
        overall = dict()
        overall['result'] = 'PASS'
        overall['TESTWEEK'] = 'NODATA'
        overall['LAST'] = dict()
        overall['LAST']['test'] = None
        overall['LAST']['result'] = None  
        overall['LAST']['date'] = None 
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                firsttesttimestamp = DATA['SN'][status][test][sn]["checktime"]
                break

        testcount = 0
        for test in DATA['SN']['TEST']:

            if sn in DATA['SN'][status][test]:

                testcount += 1
                if "result" in DATA['SN'][status][test][sn]:
                    if not overall['LAST']['result']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])
                    elif DATA['SN'][status][test][sn]["checktime"] > overall['LAST']['date']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])                     
                    if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                        overall['result'] = "FAIL"

        if overall['result'] == 'PASS':
            if not testcount == len(DATA['SN']['TEST']):
                overall['result'] = "INCOMPLETE"
        wirtedata.append(overall['result'])
        wirtedata.append(overall['TESTWEEK'])
        if "DISPLAYINDO" in inputconfig:
            for displaykey in inputconfig["DISPLAYINDO"]:
                #finddisplay = False
                displaydict = dict()
                for test in DATA['SN']['TEST']:
                    if sn in DATA['SN'][status][test]:
                        for checkkey in inputconfig["CHECKKEYS"]:
                            if checkkey in DATA['SN'][status][test][sn]:
                                if isinstance(DATA['SN'][status][test][sn][checkkey], dict):
                                    if displaykey in DATA['SN'][status][test][sn][checkkey]:
                                        if not "data" in displaydict:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]
                                        elif DATA['SN'][status][test][sn]["checktime"] > displaydict["checktime"]:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]                                            
                                        #finddisplay = True
                                        #wirtedata.append(DATA['SN'][status][test][sn][checkkey][displaykey])
                                        #break
                if "data" in displaydict:
                    wirtedata.append(displaydict["data"])
                else:
                    wirtedata.append('NODATA')
        # if status == 'LAST':
        #     last_dateandtime = overall['LAST']['date']

        #     last_datetimearray = last_dateandtime.split('_')
        #     reporttestandresult = "{} ({})".format(overall['LAST']['test'],overall['LAST']['result'])
        #     wirtedata.append(last_datetimearray[0])
        #     wirtedata.append(reporttestandresult)
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'QSPI' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['QSPI'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #     #'CPLD1'
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'CPLD1' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['CPLD1'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #wirtedata.append('')

        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            #sys.exit()
            if sn in DATA['SN'][status][test]:
                # print("TEST: {} | STATUS: {}".format(test, status))
                # print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                # sys.exit()
                #time.sleep(5)

                if "result" in DATA['SN'][status][test][sn]:
                    #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                    dateandtime = DATA['SN'][status][test][sn]["checktime"]
                    datetimearray = dateandtime.split('_')
                    reportresut = "{} ({})".format(DATA['SN'][status][test][sn]["result"],datetimearray[0])
                    #DATA['SN']['LAST'][teststep][sn]['count']
                    if "count" in DATA['SN'][status][test][sn]:
                        reportresut = "{} <{}> ".format(reportresut,DATA['SN'][status][test][sn]["count"])
                    if Withallerror:
                        if sn in DATA['SN']['ERROR'][test]:
                            for eachfailstep in DATA['SN']['ERROR'][test][sn]['LIST']:
                                reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                    else:
                        if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                            for eachfailstep in DATA['SN'][status][test][sn]['ERROR']['LIST']:
                                reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                    wirtedata.append(reportresut)
                else:
                    wirtedata.append('NO TEST DATA')
            else:
                wirtedata.append('NO TEST DATA')
            #time.sleep(2)
        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')
    highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    freezePosition(ws2,'C2')
    
    return 0

def generateexeclsnstatusalldata(DATA, workingonSNlist,status,wb,inputconfig,Withallerror=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnstatusalldata: {}".format(status))
    #titlename = "{} ALL DATE".format(status)
    titlename = "{}".format(status)
    if Withallerror:
        titlename = "{} ALL ERROR".format(titlename)
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('OVERALL')
    wirtedata.append('TEST_WEEK')
    # if status == 'LAST':
    #     wirtedata.append('LAST_TESTDATE')
    #     wirtedata.append('LAST_STATION(RESULT)')
    
    #     wirtedata.append('QSPI')
    #     wirtedata.append('CPLD1')
    if "DISPLAYINDO" in inputconfig:
        for displaykey in inputconfig["DISPLAYINDO"]:
            wirtedata.append(displaykey)

    #wirtedata.append('')
    for test in DATA['SN']['TEST']:
         wirtedata.append(test)
    ws2.append(wirtedata)

    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        overall = dict()
        overall['result'] = 'PASS'
        overall['TESTWEEK'] = 'NODATA'
        overall['LAST'] = dict()
        overall['LAST']['test'] = None
        overall['LAST']['result'] = None  
        overall['LAST']['date'] = None 
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                firsttesttimestamp = DATA['SN'][status][test][sn]["checktime"]
                break

        testcount = 0
        for test in DATA['SN']['TEST']:

            if sn in DATA['SN'][status][test]:

                testcount += 1
                if "result" in DATA['SN'][status][test][sn]:
                    if not overall['LAST']['result']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])
                    elif DATA['SN'][status][test][sn]["checktime"] > overall['LAST']['date']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])                     
                    if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                        overall['result'] = "FAIL"

        if overall['result'] == 'PASS':
            if not testcount == len(DATA['SN']['TEST']):
                overall['result'] = "INCOMPLETE"
        wirtedata.append(overall['result'])
        wirtedata.append(overall['TESTWEEK'])
        if "DISPLAYINDO" in inputconfig:
            for displaykey in inputconfig["DISPLAYINDO"]:
                #finddisplay = False
                displaydict = dict()
                for test in DATA['SN']['TEST']:
                    if sn in DATA['SN'][status][test]:
                        for checkkey in inputconfig["CHECKKEYS"]:
                            if checkkey in DATA['SN'][status][test][sn]:
                                if isinstance(DATA['SN'][status][test][sn][checkkey], dict):
                                    if displaykey in DATA['SN'][status][test][sn][checkkey]:
                                        if not "data" in displaydict:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]
                                        elif DATA['SN'][status][test][sn]["checktime"] > displaydict["checktime"]:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]                                            
                                        #finddisplay = True
                                        #wirtedata.append(DATA['SN'][status][test][sn][checkkey][displaykey])
                                        #break
                if "data" in displaydict:
                    wirtedata.append(displaydict["data"])
                else:
                    wirtedata.append('NODATA')
        # if status == 'LAST':
        #     last_dateandtime = overall['LAST']['date']

        #     last_datetimearray = last_dateandtime.split('_')
        #     reporttestandresult = "{} ({})".format(overall['LAST']['test'],overall['LAST']['result'])
        #     wirtedata.append(last_datetimearray[0])
        #     wirtedata.append(reporttestandresult)
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'QSPI' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['QSPI'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #     #'CPLD1'
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'CPLD1' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['CPLD1'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #wirtedata.append('')

        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            #sys.exit()
            if sn in DATA['SN'][status][test]:
                # print("TEST: {} | STATUS: {}".format(test, status))
                # print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                # sys.exit()
                #time.sleep(5)

                if "result" in DATA['SN'][status][test][sn]:
                    #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                    dateandtime = DATA['SN'][status][test][sn]["checktime"]
                    datetimearray = dateandtime.split('_')
                    reportresut = "{} ({})".format(DATA['SN'][status][test][sn]["result"],datetimearray[0])
                    #DATA['SN']['LAST'][teststep][sn]['count']
                    if "count" in DATA['SN'][status][test][sn]:
                        reportresut = "{} <{}> ".format(reportresut,DATA['SN'][status][test][sn]["count"])
                    if Withallerror:
                        if sn in DATA['SN']['ERROR'][test]:
                            for eachfailstep in DATA['SN']['ERROR'][test][sn]['LIST']:
                                reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                    else:
                        if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                            for eachfailstep in DATA['SN'][status][test][sn]['ERROR']['LIST']:
                                reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                    wirtedata.append(reportresut)
                else:
                    wirtedata.append('NO TEST DATA')
            else:
                wirtedata.append('NO TEST DATA')
            #time.sleep(2)
        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')
    highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    freezePosition(ws2,'C2')
    
    return 0

def createlastfailurelogfolder(pr,DATA,workingonSNlist,inputconfig):
    lastfailurefolder = "{}LastFail_{}".format(inputconfig['DIR']["reportpath"],date_time)
    pr['modules'].createdirinserver(lastfailurefolder)
    status = 'LAST'
    countforcreateunzipfile = 0
    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            if sn in DATA['SN'][status][test]:
                #print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                #time.sleep(5)
                if "result" in DATA['SN'][status][test][sn]:
                    if "FAIL" in DATA['SN'][status][test][sn]["result"].upper():
                        dateandtime = DATA['SN'][status][test][sn]["checktime"]
                        datetimearray = dateandtime.split('_')
                        eachfile = DATA['SN'][status][test][sn]["testfile"]
                        #shutil.copy2(eachfile, lastfailurefolder)
                        unzipfolder = "{}/{}_{}_{}".format(lastfailurefolder,test,sn,datetimearray[0])
                        pr['modules'].createdirinserver(unzipfolder)
                        shutil.unpack_archive(eachfile, unzipfolder)
                        countforcreateunzipfile += 1

    print("createlastfailurelogfolder: <{}>".format(countforcreateunzipfile))
    return 0

def copyallfailurelogfolder(pr,DATA,workingonSNlist,inputconfig):
    lastfailurefolder = "{}LastFail_{}".format(inputconfig['DIR']["reportpath"],date_time)
    pr['modules'].createdirinserver(lastfailurefolder)
    status = 'LAST'
    countforcreateunzipfile = 0
    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            if sn in DATA['SN'][status][test]:
                if "result" in DATA['SN'][status][test][sn]:
                    if "FAIL" in DATA['SN'][status][test][sn]["result"].upper():
                        if sn in DATA['SN'][status][test]:
                            copyfolder = "{}/{}_{}".format(lastfailurefolder,test,sn)
                            pr['modules'].createdirinserver(copyfolder)
                            for eachfile in DATA['teststep'][test]['FILELISTS']:
                                if sn.upper() in eachfile.upper():

                                    shutil.copy2(eachfile, copyfolder)
                                    countforcreateunzipfile += 1
                            change_permissions_recursive(copyfolder, 0o777)
    linux_cmd_change_permissions = "chmod -R 777 {}".format(lastfailurefolder)
    os.system(linux_cmd_change_permissions)

    print("copyallfailurelogfolder: <{}>".format(countforcreateunzipfile))
    return 0

def change_permissions_recursive(path, mode):
    for root, dirs, files in os.walk(path, topdown=False):
        for dir in [os.path.join(root,d) for d in dirs]:
            os.chmod(dir, mode)
    for file in [os.path.join(root, f) for f in files]:
            os.chmod(file, mode)

def generatedailydataintxtfile(DATA,teststep):
    
    ## [2021-04-01_03-51-47] LOG: [MTP-213]: NIC-10 ORTANO2 FLM2110004F NIC_DIAG_REGRESSION_TEST_PASS
    ## [2021-04-01_03-51-47] ERR: [MTP-213]: NIC-07 ORTANO2 FLM2110005B NIC_DIAG_REGRESSION_TEST_FAIL (LV_ASIC L1)
    ## [2021-04-16_12-30-18] LOG: [MTPS-402]: NIC-02 ORTANO2 FLM21100003 NIC_DIAG_REGRESSION_TEST_PASS
    print("")
    print("teststep: {}".format(teststep))
    print("")
    print("")
    checkdatasetup = "2021-07-11"
    BYCHASSIS = dict()
    for sn in DATA["SN"]:
        for chassis in DATA["SN"][sn]:
            if not chassis in BYCHASSIS:
                BYCHASSIS[chassis] = list()
            for testdate in DATA["SN"][sn][chassis]:
                if not testdate == checkdatasetup:
                    continue
                for testetime in DATA["SN"][sn][chassis][testdate]:
                    
                    wirteline = "## [{}_{}] ".format(testdate, testetime)
                    if 'PASS' in DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']:
                        wirteline += "LOG: "
                    else:
                        wirteline += "ERR: "
                    wirteline += "[{}]: NIC-{} {} {} NIC_DIAG_REGRESSION_TEST_{}".format(chassis,DATA["SN"][sn][chassis][testdate][testetime]['SLOT'],DATA["SN"][sn][chassis][testdate][testetime]['CARDTYPE'],sn,DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                    if 'FAIL' in DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']:
                        wirteline += " ("
                        for eachdeatilteststep in DATA["DETAILTESTSTEP"]:
                            if eachdeatilteststep in DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]:
                                if 'FAIL' in DATA["SN"][sn][chassis][testdate][testetime]['DETAILTESTSTEP'][eachdeatilteststep]:
                                    wirteline += "{},".format(eachdeatilteststep)
                        if wirteline[-1] == ",":
                            wirteline = wirteline[:-1]
                        wirteline += ")"
                            
                    BYCHASSIS[chassis].append(wirteline)
                    print(wirteline)
                      
                            
                    
    
    #print(json.dumps(BYCHASSIS, indent = 4))
    print("teststep: {}".format(teststep))
    for eachchassis in BYCHASSIS:
        print("CHASSIS: {}".format(eachchassis))
        for eachline in BYCHASSIS[eachchassis]:
            print(eachline)
    print("")
    print("")
    print("")
    print("")
    print("")
    print("")
    print("")
    print("")
    print("")

    
    
    return 0

def GetTestTimedictbyweek(workingonSNlist,DATA,teststep,FULLDATA,TimeData,pr):

    DatabyWeek = dict()
    DatabyWeek['WEEKLIST'] = list()
    DatabyWeek['DATA'] = dict()

    #TimeData = dict()
    if not "DATA" in TimeData:
        TimeData["DATA"] = dict()
        TimeData["LIST"] = list()
        TimeData["CHAMBER"] = dict()

    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                chambername = pr['modules'].where_is_my_chamber(chassis)
                if chambername:
                    if not chambername in TimeData["CHAMBER"]:
                        TimeData["CHAMBER"][chambername] = dict()
                        TimeData["CHAMBER"][chambername]["DATA"] = dict()
                        TimeData["CHAMBER"][chambername]["LIST"] = list()
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        endtime = "{}_{}".format(testdate,testetime)
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            unittesttime = str(int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME'])))
                            #print("{}: {}".format(unittesttime,endtime))
                            #sys.exit()
                            begintime = findbegintesttime(endtime,unittesttime)
                            recordchassistime = "{}_{}".format(chassis,endtime)
                            if not int(unittesttime) in TimeData["LIST"]:
                                TimeData["LIST"].append(int(unittesttime))
                                TimeData["LIST"].sort(reverse=True)
                            if not unittesttime in TimeData["DATA"]:
                                TimeData["DATA"][unittesttime] = dict()
                                TimeData["DATA"][unittesttime]["MTP"] = dict()
                                TimeData["DATA"][unittesttime]["TESTTIME"] = int(unittesttime)
                            if not recordchassistime in TimeData["DATA"][unittesttime]["MTP"]:
                                TimeData["DATA"][unittesttime]["MTP"][recordchassistime] = dict()
                            TimeData["DATA"][unittesttime]["MTP"][recordchassistime]["start"] = begintime
                            TimeData["DATA"][unittesttime]["MTP"][recordchassistime]["end"] = endtime
                            if chambername:
                                if not int(unittesttime) in TimeData["CHAMBER"][chambername]["LIST"]:
                                    TimeData["CHAMBER"][chambername]["LIST"].append(int(unittesttime))
                                    TimeData["CHAMBER"][chambername]["LIST"].sort(reverse=True)
                                if not unittesttime in TimeData["CHAMBER"][chambername]["DATA"]:
                                    TimeData["CHAMBER"][chambername]["DATA"][unittesttime] = dict()
                                    TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"] = dict()
                                    TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["TESTTIME"] = int(unittesttime)
                                if not recordchassistime in TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"]:
                                    TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"][recordchassistime] = dict()
                                TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"][recordchassistime]["start"] = begintime
                                TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"][recordchassistime]["end"] = endtime

    if len(TimeData["CHAMBER"].keys()):
        for chambername in TimeData["CHAMBER"]:
            pr['modules'].print_anyinformation(chambername)
            TimeData["CHAMBER"][chambername] = findmaxchamberUsingtimeinChamber(TimeData["CHAMBER"][chambername])

    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                chambername = pr['modules'].where_is_my_chamber(chassis)
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        weeknumber = findworkweek("{}_{}".format(testdate,testetime))
                        if not weeknumber in DatabyWeek['WEEKLIST']:
                            DatabyWeek['WEEKLIST'].append(weeknumber)
                            DatabyWeek['WEEKLIST'].sort()
                        if not weeknumber in DatabyWeek['DATA']:
                            DatabyWeek['DATA'][weeknumber] = dict()
                            DatabyWeek['DATA'][weeknumber]["numberofSLOTlist"] = list()
                            DatabyWeek['DATA'][weeknumber]["DATAofNumber"] = dict()
                        timestamp = "{}_{}".format(testdate,testetime)
                        numberofslotfortest = findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp)
                        if isinstance(numberofslotfortest, int):
                            numberofslotfortest = "{:02d}".format(numberofslotfortest)
                        if not numberofslotfortest in DatabyWeek['DATA'][weeknumber]["numberofSLOTlist"]:
                            if not 'NO' in numberofslotfortest.upper():
                                DatabyWeek['DATA'][weeknumber]["numberofSLOTlist"].append(numberofslotfortest)
                                DatabyWeek['DATA'][weeknumber]["numberofSLOTlist"].sort()
                        if not numberofslotfortest in DatabyWeek['DATA'][weeknumber]["DATAofNumber"]:
                            DatabyWeek['DATA'][weeknumber]["DATAofNumber"][numberofslotfortest] = list()

                        if '4C' in teststep:
                            endtestime = "{}_{}".format(testdate,testetime)
                            if chambername:
                                MaxTestTime = findmaxtesttimeinChamber(endtestime,TimeData["CHAMBER"][chambername])
                                DatabyWeek['DATA'][weeknumber]["DATAofNumber"][numberofslotfortest].append(int(MaxTestTime))
                            else:
                                MaxTestTime = findmaxtesttimeinChamber(endtestime,TimeData)
                                DatabyWeek['DATA'][weeknumber]["DATAofNumber"][numberofslotfortest].append(int(MaxTestTime))
                        else:
                            if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                                DatabyWeek['DATA'][weeknumber]["DATAofNumber"][numberofslotfortest].append(int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME'])))
    
    return DatabyWeek

def generateexecltestbyNon4Ctesttime(workingonSNlist,DATA,teststep,wb,FULLDATA,pr):

    titlename = "{}_{}".format("Testtime", teststep)
    ws2 = wb.create_sheet(title=titlename)
    print("{}: {}".format("generateexecltestbyNon4Ctesttime", titlename))
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('WEEK')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('TESTTIME(HH:MM:SS)')
    wirtedata.append('TIME_RANGE')
    wirtedata.append('TESTSTATUS')
    wirtedata.append('# of Card in MTP')
    TimeData = dict()
    WeekTimedata = GetTestTimedictbyweek(workingonSNlist,DATA,teststep,FULLDATA,TimeData,pr)
    #TimeData["DATA"] = dict()
    #TimeData["LIST"] = list()

    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        wirtedata = list()
                        wirtedata.append(teststep)
                        wirtedata.append(sn)
                        wirtedata.append(chassis)
                        timestamp = "{}_{}".format(testdate,testetime)

                        wirtedata.append(findworkweek("{}_{}".format(testdate,testetime)))
                        endtestime = "{}_{}".format(testdate,testetime)
                        wirtedata.append(testdate)
                        wirtedata.append(testetime.replace("-",":"))
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(timedelta(seconds=int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                        else:
                            wirtedata.append('NO TESTTIME')
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(converttoHourtoHour(int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                        else:
                            wirtedata.append("No TESTTIME")
                        #sys.exit()
                        if 'FINALRESULT' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                        else:
                            wirtedata.append("NONE")
                        timestamp = "{}_{}".format(testdate,testetime)
                        wirtedata.append(findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp))
                        ws2.append(wirtedata)
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'K2')
    return 0

def generateexeclby4CChambertemp(workingonSNlist,wb,FULLDATA,pr,start=None,totimezone=None):

    print("{}!".format("generateexeclby4CChambertime"))

    Chamberref = dict()

    for eachteststep in FULLDATA['SN']['TEST']:
        if '4C' in eachteststep:
            TimeData = dict()
            DATA = FULLDATA["teststep"][eachteststep]
            WeekTimedata = GetTestTimedictbyweek(workingonSNlist,DATA,eachteststep,FULLDATA,TimeData,pr)
            
            for sn in workingonSNlist:
                if sn in DATA["SN"]:
                    for chassis in DATA["SN"][sn]:
                        for testdate in DATA["SN"][sn][chassis]:
                            for testetime in DATA["SN"][sn][chassis][testdate]:
                                chambername = pr['modules'].where_is_my_chamber(chassis)
                                if chambername:
                                    if not chambername in Chamberref:
                                        Chamberref[chambername] = dict()
                                        Chamberref[chambername]['LIST'] = list()
                                        Chamberref[chambername]['DICT'] = dict()

                                    endtestime = "{}_{}".format(testdate,testetime)
                                    MaxTestTime = findmaxtesttimeinChamber(endtestime,TimeData["CHAMBER"][chambername])
                                    chamberstarttime,chamberendtime = findmaxtestchambertimeinChamber(endtestime,TimeData["CHAMBER"][chambername])
                                    teststatus = DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']

                                    if not chamberendtime in Chamberref[chambername]['LIST']:
                                        Chamberref[chambername]['LIST'].append(chamberendtime)
                                        Chamberref[chambername]['LIST'].sort(reverse=True)
                                    if not chamberendtime in Chamberref[chambername]['DICT']:
                                        Chamberref[chambername]['DICT'][chamberendtime] = dict()
                                        Chamberref[chambername]['DICT'][chamberendtime]['SN'] = list()
                                        Chamberref[chambername]['DICT'][chamberendtime]['PASS'] = list()
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'] = dict()

                                    Chamberref[chambername]['DICT'][chamberendtime]['SN'].append(sn)
                                    if 'PASS' in teststatus.upper():
                                        Chamberref[chambername]['DICT'][chamberendtime]['PASS'].append(sn)
                                    Chamberref[chambername]['DICT'][chamberendtime]['START'] = chamberstarttime
                                    Chamberref[chambername]['DICT'][chamberendtime]['END'] = chamberendtime
                                    Chamberref[chambername]['DICT'][chamberendtime]['TESTTIME'] = MaxTestTime
                                    Chamberref[chambername]['DICT'][chamberendtime]['TEST'] = eachteststep

                                    #pr['modules'].print_anyinformation(DATA["SN"][sn][chassis][testdate][testetime])
                                    if not chassis in Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS']:
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis] = dict()
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['SN'] = list()
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['PASS'] = list()
                                    
                                    if "TEMPERATURE" in DATA["SN"][sn][chassis][testdate][testetime]:
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["TEMPERATURE"] = getlistoftemp(DATA["SN"][sn][chassis][testdate][testetime]["TEMPERATURE"])
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["TEMPERATUREDETAIL"] = getlistoftempdetail(DATA["SN"][sn][chassis][testdate][testetime]["TEMPERATURE"])
                                    else:
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["TEMPERATURE"] = []
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["TEMPERATUREDETAIL"] = []
                                    Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['SN'].append(sn)
                                    if 'PASS' in teststatus.upper():
                                        Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['PASS'].append(sn)

                                    #pr['modules'].print_anyinformation(DATA["SN"][sn][chassis][testdate][testetime])
                                    Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["ADDITION_TEMP"] = getchamberTempbymfglog(pr, DATA["SN"][sn][chassis][testdate][testetime])
                                    #sys.exit()

    #pr['modules'].print_anyinformation(Chamberref)
    #sys.exit()
    
    #pr['modules'].print_anyinformation(Chamberref)
    #sys.exit()

    additionTemp = ['Local', 'Outlet', 'Inlet-1', 'Inlet-2']
    additionTemp2 = ['Local', 'Remote-1', 'Remote-2', 'Remote-3']

    titlename = "{}_{}".format("Chamber", "TEMP")
    if totimezone:
        titlename = "{}_{}".format(titlename,totimezone.replace('/','_'))
    ws2 = wb.create_sheet(title=titlename)
    
    wirtedata = list()
    wirtedata.append('#')
    wirtedata.append('CHAMBER')
    wirtedata.append('FLOOR')
    wirtedata.append('TEST')
    wirtedata.append('START_DATE')
    wirtedata.append('START_TIME')
    wirtedata.append('END_DATE')
    wirtedata.append('END_TIME')
    #wirtedata.append('WAIT_TO_START_TIME')
    wirtedata.append('RUNNING_TIME')
    wirtedata.append('MTP')
    wirtedata.append('TEST_UNITS')
    wirtedata.append('PASS_UNITS')
    wirtedata.append('YIELD')
    wirtedata.append('HIGH-TEMP')
    wirtedata.append('LOW-TEMP')
    wirtedata.append('AVG-TEMP')
    for additionkey in additionTemp:
        wirtedata.append("{}_AVG-TEMP".format(additionkey))
    ws2.append(wirtedata)
    count = 0
    for chambername in sorted(Chamberref.keys()):
        countnumber = 1
        eachchamberdata = pr['modules'].getchamberlistorderbyfloor(chambername)
        for chamberendtime in Chamberref[chambername]['LIST'][:10]:
            count += 1
            for eachfloor in eachchamberdata['LIST']:
                for chassis in eachchamberdata['DATA'][eachfloor]:
                    if chassis in sorted(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'].keys()):
                        wirtedata = list()
                        if start:
                            if start > chamberendtime.split('_')[0]:
                                break
                        wirtedata.append(count)
                        wirtedata.append(chambername)
                        wirtedata.append(pr['modules'].which_floor_is_for_my_mtp(chassis))
                        wirtedata.append(Chamberref[chambername]['DICT'][chamberendtime]['TEST'])
                        chamberstarttimefromrecord = Chamberref[chambername]['DICT'][chamberendtime]['START']
                        chamberendtimefromrecord = Chamberref[chambername]['DICT'][chamberendtime]['END']
                        if totimezone:
                            chamberstarttimefromrecord = converttimetonewtimezone(chamberstarttimefromrecord,totimezone)
                            chamberendtimefromrecord = converttimetonewtimezone(chamberendtimefromrecord,totimezone)

                        wirtedata.append(chamberstarttimefromrecord.split('_')[0])
                        wirtedata.append(chamberstarttimefromrecord.split('_')[1].replace('-',':'))
                        wirtedata.append(chamberendtimefromrecord.split('_')[0])
                        wirtedata.append(chamberendtimefromrecord.split('_')[1].replace('-',':'))
                        #wirtedata.append(Chamberref[chambername]['DICT'][chamberendtime]['TESTTIME'])
                        # if countnumber < len(Chamberref[chambername]['LIST']):
                        #     starttime = Chamberref[chambername]['DICT'][chamberendtime]['START']
                        #     lastendtime = Chamberref[chambername]['DICT'][Chamberref[chambername]['LIST'][countnumber]]['END']
                        #     newtesttime = cal_total_testTime(lastendtime,starttime)
                        #     wirtedata.append(timedelta(seconds=int(float(newtesttime))))
                        # else:
                        #     wirtedata.append('NULL')
                        wirtedata.append(timedelta(seconds=int(float(Chamberref[chambername]['DICT'][chamberendtime]['TESTTIME']))))
                        wirtedata.append(chassis)
                        wirtedata.append(len(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['SN']))
                        wirtedata.append(len(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['PASS']))
                        wirtedata.append("{:.2f}%".format(len(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['PASS'])/len(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['SN']) * 100))
                        if len(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['TEMPERATURE']):
                            wirtedata.append(max(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['TEMPERATURE']))
                            wirtedata.append(min(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['TEMPERATURE']))
                            from statistics import mean
                            wirtedata.append(round(mean(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]['TEMPERATURE']),2))
                            for additionkey in additionTemp:
                                from statistics import mean
                                if additionkey in Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["ADDITION_TEMP"]["DATA"]:
                                    if len(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["ADDITION_TEMP"]["DATA"][additionkey]):
                                        wirtedata.append(round(mean(Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["ADDITION_TEMP"]["DATA"][additionkey]),2))
                                    else:
                                        wirtedata.append("NO DATA")
                                else:
                                    wirtedata.append("NO DATA")
                        else:
                            wirtedata.append("NO DATA")
                            wirtedata.append("NO DATA")
                            wirtedata.append("NO DATA")
                            for additionkey in additionTemp:
                                wirtedata.append("NO DATA")


                        wirtedata.append("|||")
                        for eachdata in Chamberref[chambername]['DICT'][chamberendtime]['CHASSIS'][chassis]["TEMPERATUREDETAIL"]:
                            wirtedata.append(eachdata)
                        ws2.append(wirtedata)
                        countnumber += 1

                    else: 
                        wirtedata = list()
                        wirtedata.append(count)
                        wirtedata.append(chambername)
                        wirtedata.append(pr['modules'].which_floor_is_for_my_mtp(chassis))
                        wirtedata.append(Chamberref[chambername]['DICT'][chamberendtime]['TEST'])
                        for i in range(16):
                            wirtedata.append('NULL')
                        ws2.append(wirtedata)

            wirtedata = list()
            for i in range(20):
                wirtedata.append('---')
            ws2.append(wirtedata)
    fixcolumnssize(ws2)
    #highlightinyellow(ws2,'TIMEOUT')
    #highlightinyellow(ws2,'NO TEST DATA')
    #highlightingreen(ws2,'PASS')
    #highlightinred(ws2, 'FAIL')
    #highlightinred(ws2, 'FAILED')
    highlightinAqua(ws2,'NULL')
    highlightinyellow(ws2,'---')
    freezePosition(ws2,'E2')
    converttoPERCENTAGEnumber(ws2)
    return 0

def getchamberTempbymfglog(pr, onedata):
    #pr['modules'].print_anyinformation(onedata)
    if "TEMPERATURE" in onedata:
        if len(onedata["TEMPERATURE"]):
            counttime = onedata["TEMPERATURE"][0][0]
        else:
            return None
    else:
        return None
    #print("START: {}".format(counttime))
    listofdatewillcheck = list()
    for eachdatetime in onedata["TEMPFROMMFGLOG"]["LIST"]:
        #print(eachdatetime)
        if eachdatetime >= counttime:
            #print("COUNT: {}".format(eachdatetime))
            listofdatewillcheck.append(eachdatetime)

    #print(eachdatetime)

    # [5]                     Local   <TYPE:<class 'str'>>
    # [5]                     Remote-1    <TYPE:<class 'str'>>
    # [5]                     Remote-2    <TYPE:<class 'str'>>
    # [5]                     Remote-3    <TYPE:<class 'str'>>    

    dataforreturn = dict()
    dataforreturn["LIST1"] = ['Local', 'Remote-1', 'Remote-2', 'Remote-3']
    dataforreturn["LIST"] = ['Local', 'Outlet', 'Inlet-1', 'Inlet-2']
    refdata = dict(zip(dataforreturn["LIST"], dataforreturn["LIST1"]))
    dataforreturn["DATA"] = dict()
    dataforreturn["DATA"]["CHECKPOINT"] = listofdatewillcheck
    for eachcheckpoint in listofdatewillcheck:
        res = dict(zip(onedata["TEMPFROMMFGLOG"]["DICT"][eachcheckpoint]["NAME"], onedata["TEMPFROMMFGLOG"]["DICT"][eachcheckpoint]["VALUE"]))
        #pr['modules'].print_anyinformation(res)
        for eachkey in dataforreturn["LIST"]:
            if not eachkey in dataforreturn["DATA"]:
                dataforreturn["DATA"][eachkey] = list()
            if eachkey in res:
                dataforreturn["DATA"][eachkey].append(float(res[eachkey]))
            elif refdata[eachkey] in res:
                dataforreturn["DATA"][eachkey].append(float(res[refdata[eachkey]]))

        #sys.exit()
    #pr['modules'].print_anyinformation(dataforreturn)
    #sys.exit()
    return dataforreturn

def getlistoftemp(timewithtemplist):
    timelist = list()
    templist = list()
    for eachdata in timewithtemplist:
        timelist.append(eachdata[0])
        templist.append(float(eachdata[1]))
    return templist

def getlistoftempdetail(timewithtemplist):
    timelist = list()
    templist = list()
    tempdetail = list()
    starttime = None
    for eachdata in timewithtemplist:
        timelist.append(eachdata[0])
        templist.append(float(eachdata[1]))
        if not starttime:
            starttime = eachdata[0]
        runtesttime = cal_total_testTime(starttime,eachdata[0])
        recordtempwithtime = "{} ({})".format(float(eachdata[1]),timedelta(seconds=int(float(runtesttime))))
        tempdetail.append(recordtempwithtime)

    return tempdetail

def generateexeclby4CChambertime(workingonSNlist,wb,FULLDATA,pr,start=None,totimezone=None):

    print("{}!".format("generateexeclby4CChambertime"))

    Chamberref = dict()

    for eachteststep in FULLDATA['SN']['TEST']:
        if '4C' in eachteststep:
            TimeData = dict()
            DATA = FULLDATA["teststep"][eachteststep]
            WeekTimedata = GetTestTimedictbyweek(workingonSNlist,DATA,eachteststep,FULLDATA,TimeData,pr)
            
            for sn in workingonSNlist:
                if sn in DATA["SN"]:
                    for chassis in DATA["SN"][sn]:
                        for testdate in DATA["SN"][sn][chassis]:
                            for testetime in DATA["SN"][sn][chassis][testdate]:
                                chambername = pr['modules'].where_is_my_chamber(chassis)
                                if chambername:
                                    if not chambername in Chamberref:
                                        Chamberref[chambername] = dict()
                                        Chamberref[chambername]['LIST'] = list()
                                        Chamberref[chambername]['DICT'] = dict()

                                    endtestime = "{}_{}".format(testdate,testetime)
                                    MaxTestTime = findmaxtesttimeinChamber(endtestime,TimeData["CHAMBER"][chambername])
                                    chamberstarttime,chamberendtime = findmaxtestchambertimeinChamber(endtestime,TimeData["CHAMBER"][chambername])
                                    teststatus = DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']

                                    if not chamberendtime in Chamberref[chambername]['LIST']:
                                        Chamberref[chambername]['LIST'].append(chamberendtime)
                                        Chamberref[chambername]['LIST'].sort(reverse=True)
                                    if not chamberendtime in Chamberref[chambername]['DICT']:
                                        Chamberref[chambername]['DICT'][chamberendtime] = dict()
                                        Chamberref[chambername]['DICT'][chamberendtime]['SN'] = list()
                                        Chamberref[chambername]['DICT'][chamberendtime]['PASS'] = list()
                                    Chamberref[chambername]['DICT'][chamberendtime]['SN'].append(sn)
                                    if 'PASS' in teststatus.upper():
                                        Chamberref[chambername]['DICT'][chamberendtime]['PASS'].append(sn)
                                    Chamberref[chambername]['DICT'][chamberendtime]['START'] = chamberstarttime
                                    Chamberref[chambername]['DICT'][chamberendtime]['END'] = chamberendtime
                                    Chamberref[chambername]['DICT'][chamberendtime]['TESTTIME'] = MaxTestTime
                                    Chamberref[chambername]['DICT'][chamberendtime]['TEST'] = eachteststep
    
    #pr['modules'].print_anyinformation(Chamberref)
    #sys.exit()
    titlename = "{}_{}".format("Chamber", "Time")
    if totimezone:
        titlename = "{}_{}".format(titlename,totimezone.replace('/','_'))
    ws2 = wb.create_sheet(title=titlename)
    
    wirtedata = list()
    wirtedata.append('CHAMBER')
    wirtedata.append('TEST')
    wirtedata.append('START_DATE')
    wirtedata.append('START_TIME')
    wirtedata.append('END_DATE')
    wirtedata.append('END_TIME')
    wirtedata.append('IDLE_TIME')
    wirtedata.append('TOTAL_TIME')
    wirtedata.append('TEST_UNITS')
    wirtedata.append('PASS_UNITS')
    wirtedata.append('YIELD')
    ws2.append(wirtedata)
    for chambername in sorted(Chamberref.keys()):
        countnumber = 1
        for chamberendtime in Chamberref[chambername]['LIST']:
            wirtedata = list()
            if start:
                if start > chamberendtime.split('_')[0]:
                    break
            wirtedata.append(chambername)
            wirtedata.append(Chamberref[chambername]['DICT'][chamberendtime]['TEST'])
            chamberstarttimefromrecord = Chamberref[chambername]['DICT'][chamberendtime]['START']
            chamberendtimefromrecord = Chamberref[chambername]['DICT'][chamberendtime]['END']
            if totimezone:
                chamberstarttimefromrecord = converttimetonewtimezone(chamberstarttimefromrecord,totimezone)
                chamberendtimefromrecord = converttimetonewtimezone(chamberendtimefromrecord,totimezone)

            wirtedata.append(chamberstarttimefromrecord.split('_')[0])
            wirtedata.append(chamberstarttimefromrecord.split('_')[1].replace('-',':'))
            wirtedata.append(chamberendtimefromrecord.split('_')[0])
            wirtedata.append(chamberendtimefromrecord.split('_')[1].replace('-',':'))
            #wirtedata.append(Chamberref[chambername]['DICT'][chamberendtime]['TESTTIME'])
            if countnumber < len(Chamberref[chambername]['LIST']):
                starttime = Chamberref[chambername]['DICT'][chamberendtime]['START']
                lastendtime = Chamberref[chambername]['DICT'][Chamberref[chambername]['LIST'][countnumber]]['END']
                newtesttime = cal_total_testTime(lastendtime,starttime)
                wirtedata.append(timedelta(seconds=int(float(newtesttime))))
            else:
                wirtedata.append('NULL')
            wirtedata.append(timedelta(seconds=int(float(Chamberref[chambername]['DICT'][chamberendtime]['TESTTIME']))))
            wirtedata.append(len(Chamberref[chambername]['DICT'][chamberendtime]['SN']))
            wirtedata.append(len(Chamberref[chambername]['DICT'][chamberendtime]['PASS']))
            wirtedata.append("{:.2f}%".format(len(Chamberref[chambername]['DICT'][chamberendtime]['PASS'])/len(Chamberref[chambername]['DICT'][chamberendtime]['SN']) * 100))
            ws2.append(wirtedata)
            countnumber += 1


    fixcolumnssize(ws2)
    #highlightinyellow(ws2,'TIMEOUT')
    #highlightinyellow(ws2,'NO TEST DATA')
    #highlightingreen(ws2,'PASS')
    #highlightinred(ws2, 'FAIL')
    #highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'A2')
    converttoPERCENTAGEnumber(ws2)
    return 0

def converttimetonewtimezone(timetoconvert,totimezone):
    print("TIME: {} | TimeZone: {}".format(timetoconvert,totimezone))
    newtime = timetoconvert
    import pytz
    from datetime import datetime, timezone
    datetime_object = datetime.strptime(timetoconvert, '%Y-%m-%d_%H-%M-%S')
    psttimezone = pytz.timezone('US/Pacific')
    converttimezone = pytz.timezone(totimezone)
    datetime_object = psttimezone.localize(datetime_object)
    #print(datetime_object.tzname())
    #print(datetime_object)
    # Convert time zone
    datetime_object = datetime_object.astimezone(converttimezone)
    #print(datetime_object.tzname())
    #print(datetime_object)
    newtime = datetime_object.strftime("%Y-%m-%d_%H-%M-%S")
    #print(newtime)
    #sys.exit()

    return newtime


def generateexecltestby4Ctesttime(workingonSNlist,DATA,teststep,wb,FULLDATA,pr):

    titlename = "{}_{}".format("Testtime", teststep)
    ws2 = wb.create_sheet(title=titlename)
    print("{}: {}".format("generateexecltestby4Ctesttime", titlename))
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('CHAMBER')
    wirtedata.append('WEEK')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('TESTTIME(HH:MM:SS)')
    wirtedata.append('GROUP_TESTTIME(HH:MM:SS)')
    wirtedata.append('CHAMBER_START_DATE')
    wirtedata.append('CHAMBER_START_TIME')
    wirtedata.append('CHAMBER_END_DATE')
    wirtedata.append('CHAMBER_END_TIME')
    wirtedata.append('TIME_RANGE')
    wirtedata.append('TESTSTATUS')
    wirtedata.append('# of Card in MTP')
    TimeData = dict()
    WeekTimedata = GetTestTimedictbyweek(workingonSNlist,DATA,teststep,FULLDATA,TimeData,pr)
    # TimeData["DATA"] = dict()
    # TimeData["LIST"] = list()
    # TimeData["CHAMBER"] = dict()

    # for sn in workingonSNlist:
    #     if sn in DATA["SN"]:
    #         for chassis in DATA["SN"][sn]:
    #             chambername = pr['modules'].where_is_my_chamber(chassis)
    #             if chambername:
    #                 if not chambername in TimeData["CHAMBER"]:
    #                     TimeData["CHAMBER"][chambername] = dict()
    #                     TimeData["CHAMBER"][chambername]["DATA"] = dict()
    #                     TimeData["CHAMBER"][chambername]["LIST"] = list()
    #             for testdate in DATA["SN"][sn][chassis]:
    #                 for testetime in DATA["SN"][sn][chassis][testdate]:
    #                     endtime = "{}_{}".format(testdate,testetime)
    #                     unittesttime = str(int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME'])))
    #                     #print("{}: {}".format(unittesttime,endtime))
    #                     #sys.exit()
    #                     begintime = findbegintesttime(endtime,unittesttime)
    #                     recordchassistime = "{}_{}".format(chassis,endtime)
    #                     if not int(unittesttime) in TimeData["LIST"]:
    #                         TimeData["LIST"].append(int(unittesttime))
    #                         TimeData["LIST"].sort(reverse=True)
    #                     if not unittesttime in TimeData["DATA"]:
    #                         TimeData["DATA"][unittesttime] = dict()
    #                         TimeData["DATA"][unittesttime]["MTP"] = dict()
    #                         TimeData["DATA"][unittesttime]["TESTTIME"] = int(unittesttime)
    #                     if not recordchassistime in TimeData["DATA"][unittesttime]["MTP"]:
    #                         TimeData["DATA"][unittesttime]["MTP"][recordchassistime] = dict()
    #                     TimeData["DATA"][unittesttime]["MTP"][recordchassistime]["start"] = begintime
    #                     TimeData["DATA"][unittesttime]["MTP"][recordchassistime]["end"] = endtime
    #                     if chambername:
    #                         if not int(unittesttime) in TimeData["CHAMBER"][chambername]["LIST"]:
    #                             TimeData["CHAMBER"][chambername]["LIST"].append(int(unittesttime))
    #                             TimeData["CHAMBER"][chambername]["LIST"].sort(reverse=True)
    #                         if not unittesttime in TimeData["CHAMBER"][chambername]["DATA"]:
    #                             TimeData["CHAMBER"][chambername]["DATA"][unittesttime] = dict()
    #                             TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"] = dict()
    #                             TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["TESTTIME"] = int(unittesttime)
    #                         if not recordchassistime in TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"]:
    #                             TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"][recordchassistime] = dict()
    #                         TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"][recordchassistime]["start"] = begintime
    #                         TimeData["CHAMBER"][chambername]["DATA"][unittesttime]["MTP"][recordchassistime]["end"] = endtime                            

    #print(json.dumps(TimeData, indent = 4))
    #sys.exit()
    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        slot = DATA["SN"][sn][chassis][testdate][testetime]['SLOT']
                        wirtedata = list()
                        wirtedata.append(teststep)
                        wirtedata.append(sn)
                        wirtedata.append(chassis)
                        chambername = pr['modules'].where_is_my_chamber(chassis)
                        wirtedata.append(chambername)
                        #sys.exit()
                        if chambername:
                            wirtedata.append(findworkweek("{}_{}".format(testdate,testetime)))
                            endtestime = "{}_{}".format(testdate,testetime)
                            wirtedata.append(testdate)
                            wirtedata.append(testetime.replace("-",":"))
                            if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                                wirtedata.append(timedelta(seconds=int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                            else:
                                wirtedata.append('NO TESTTIME')
                            MaxTestTime = findmaxtesttimeinChamber(endtestime,TimeData["CHAMBER"][chambername])
                            chamberstarttime,chamberendtime = findmaxtestchambertimeinChamber(endtestime,TimeData["CHAMBER"][chambername])
                            wirtedata.append(timedelta(seconds=int(MaxTestTime)))
                            wirtedata.append(chamberstarttime.split('_')[0])
                            wirtedata.append(chamberstarttime.split('_')[1].replace('-',':'))
                            wirtedata.append(chamberendtime.split('_')[0])
                            wirtedata.append(chamberendtime.split('_')[1].replace('-',':'))
                            wirtedata.append(converttoHourtoHour(MaxTestTime))
                            wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                            #sys.exit()
                            timestamp = "{}_{}".format(testdate,testetime)
                            wirtedata.append(findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp))
                        else:

                            wirtedata.append(findworkweek("{}_{}".format(testdate,testetime)))
                            endtestime = "{}_{}".format(testdate,testetime)
                            wirtedata.append(testdate)
                            wirtedata.append(testetime.replace("-",":"))
                            if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                                wirtedata.append(timedelta(seconds=int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                            else:
                                wirtedata.append('NO TESTTIME')
                            MaxTestTime = findmaxtesttimeinChamber(endtestime,TimeData)

                            wirtedata.append(timedelta(seconds=int(MaxTestTime)))
                            wirtedata.append('NO CHAMBER FIND')
                            wirtedata.append('NO CHAMBER FIND')
                            wirtedata.append('NO CHAMBER FIND')
                            wirtedata.append('NO CHAMBER FIND')
                            wirtedata.append(converttoHourtoHour(MaxTestTime))
                            wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                            #sys.exit()
                            timestamp = "{}_{}".format(testdate,testetime)
                            wirtedata.append(findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp))
                        ws2.append(wirtedata)
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'H2')
    return 0

def converttoHourtoHour(MaxTestTime):
    hours = int(MaxTestTime / 3600)
    displaystr = "{:02d} ~ {:02d}".format(hours, hours + 1)
    return displaystr


def findmaxtestchambertimeinChamber(endtesttime,TimeData):
    count = 1
    for unittesttime in TimeData["LIST"]:
        #print("{}: {}".format(count,unittesttime))
        #print(json.dumps(TimeData["DATA"][str(unittesttime)], indent = 4))
        for mtp in TimeData["DATA"][str(unittesttime)]["MTP"]:
            if endtesttime >= TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["start"] and endtesttime <= TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["end"]:
                
                #print("Return {}: {} by END TIME: {}".format(count,unittesttime,endtesttime))
                #sys.exit()
                return TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["start"],TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["end"]
        
        count += 1

    return None, None

def findmaxtesttimeinChamber(endtesttime,TimeData):
    count = 1
    for unittesttime in TimeData["LIST"]:
        #print("{}: {}".format(count,unittesttime))
        #print(json.dumps(TimeData["DATA"][str(unittesttime)], indent = 4))
        for mtp in TimeData["DATA"][str(unittesttime)]["MTP"]:
            if endtesttime >= TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["start"] and endtesttime <= TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["end"]:
                
                #print("Return {}: {} by END TIME: {}".format(count,unittesttime,endtesttime))
                #sys.exit()
                return unittesttime
        
        count += 1

    return None

def findmaxchamberUsingtimeinChamber(TimeData):
    start=datetime.now()
    Keeprun = True
    countforwholerun = 0

    count = 0
    countforwholerun += 1
    workingonlist = TimeData["LIST"]
    for unittesttime in workingonlist:
        #print("{}: {}".format(count,unittesttime))
        #print(json.dumps(TimeData["DATA"][str(unittesttime)], indent = 4))
        for mtp in TimeData["DATA"][str(unittesttime)]["MTP"]:
            starttime = TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["start"]
            endtime = TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["end"]
            newstartime = None 
            newendtime = None 
            for unittesttime2 in TimeData["LIST"]:
                for mtp2 in TimeData["DATA"][str(unittesttime2)]["MTP"]:            

                    if starttime > TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["start"] and starttime < TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["end"]:
                        
                        #print("Return {}: {} by starttime TIME: {}".format(count,unittesttime2,starttime))
                        #print(json.dumps(TimeData["DATA"][str(unittesttime2)], indent = 4))
                        if not newstartime:
                            newstartime = TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["start"]
                        elif newstartime > TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["start"]:
                            newstartime = TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["start"]

                        #sys.exit()
                    if endtime > TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["start"] and endtime < TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["end"]:
                        
                        #print("Return {}: {} by endtime TIME: {}".format(count,unittesttime2,endtime))
                        #print(json.dumps(TimeData["DATA"][str(unittesttime2)], indent = 4))
                        if not newendtime:
                            newendtime = TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["end"]
                        elif newendtime < TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["end"]:
                            newendtime = TimeData["DATA"][str(unittesttime2)]["MTP"][mtp2]["end"]
                        #sys.exit()
            if newstartime or newendtime:
                if not newstartime:
                    newstartime = starttime
                if not newendtime:
                    newendtime = endtime
                newtesttime = cal_total_testTime(newstartime,newendtime)
                #print("Start Time: {} to End Time: {} | TESTTIME: {}".format(newstartime,newendtime,newtesttime))
                if not newtesttime in TimeData["LIST"]:
                    TimeData["LIST"].append(newtesttime)
                    TimeData["LIST"].sort(reverse=True)
                if not str(newtesttime) in TimeData["DATA"]:
                    TimeData["DATA"][str(newtesttime)] = dict()
                    TimeData["DATA"][str(newtesttime)]["MTP"] = dict()
                    TimeData["DATA"][str(newtesttime)]["TESTTIME"] = newtesttime
                if not newendtime in TimeData["DATA"][str(newtesttime)]["MTP"]:
                    TimeData["DATA"][str(newtesttime)]["MTP"][newendtime] = dict()
                TimeData["DATA"][str(newtesttime)]["MTP"][newendtime]["start"] = newstartime
                TimeData["DATA"][str(newtesttime)]["MTP"][newendtime]["end"] = newendtime
                #print(json.dumps(TimeData["DATA"][str(newtesttime)], indent = 4))
                #sys.exit()
                count += 1

    print("countforwholerun: {} with count {}".format(countforwholerun, count))
    difftime = datetime.now()-start
    print("findmaxchamberUsingtimeinChamber: use {} seconds".format(difftime.total_seconds()))
    return TimeData

def GetSNlistbyPNfromDLtest(workingonSNlist,DATA,teststep='DL'):

    SNbyPN = dict()
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        PN = None
                        if 'PN' in DATA["SN"][sn][chassis][testdate][testetime]:
                            PN = str(DATA["SN"][sn][chassis][testdate][testetime]['PN'])
                        else:
                            PN = 'NULL'
                        
                        teststatus = DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']
                        if PN:
                            if len(PN) > 5:
                                if not PN in SNbyPN:
                                    SNbyPN[PN] = list()
                                if not sn in SNbyPN[PN]:
                                    SNbyPN[PN].append(sn)
                                    SNbyPN[PN].sort(reverse=True)

    #print(json.dumps(SNbyPN, indent = 4))

    #sys.exit()
                        
    return SNbyPN

def generateHVLVdata(workingonSNlist,DATA,teststep,wb,FULLDATA):
    start=datetime.now()
    print("{}: {}".format("generateHVLVdata", teststep))
    HLteststeplist = dict()
    LVteststeplist = dict()
    HLteststeplist['MAIN'] = list()
    LVteststeplist['MAIN'] = list()
    for eachdeatilteststep in DATA["DETAILTESTSTEP"]:
        if "HV" in eachdeatilteststep:
            HLteststeplist['MAIN'].append(eachdeatilteststep)
        if "LV" in eachdeatilteststep:
            LVteststeplist['MAIN'].append(eachdeatilteststep)
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        if "PASS" in DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']:
                            if not testdate in HLteststeplist:
                                HLteststeplist[testdate] = list()
                                for eachdeatilteststep in DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]:
                                    if "HV" in eachdeatilteststep:
                                        HLteststeplist[testdate].append(eachdeatilteststep)
                            if not testdate in LVteststeplist:
                                LVteststeplist[testdate] = list()
                                for eachdeatilteststep in DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]:
                                    if "LV" in eachdeatilteststep:
                                        LVteststeplist[testdate].append(eachdeatilteststep)

    #print(json.dumps(HLteststeplist, indent = 4))
    #print(json.dumps(LVteststeplist, indent = 4))

    HVLVData = dict()

    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        timestamp = "{}_{}".format(testdate, testetime)
                        if not sn in HVLVData:
                            HVLVData[sn] = dict()
                            HVLVData[sn]["LIST"] = list()
                            HVLVData[sn]["DATA"] = dict()
                        if not timestamp in HVLVData[sn]["DATA"]:
                            HVLVData[sn]["DATA"][timestamp] = dict()
                        if not timestamp in HVLVData[sn]["LIST"]:
                            HVLVData[sn]["LIST"].append(timestamp)
                            HVLVData[sn]["LIST"].sort()
                        TestStatus = DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']
                        HVstatus = "PASS"
                        HVtestcount = 0
                        HVtestcountinSN = 0
                        HLteststeplistforcheck = list()
                        if testdate in HLteststeplist:
                            HLteststeplistforcheck = HLteststeplist[testdate]
                        else:
                            HLteststeplistforcheck = HLteststeplist['MAIN']
                        for eachdeatilteststep in HLteststeplistforcheck:
                            HVtestcount += 1
                            if eachdeatilteststep in DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]:
                                HVtestcountinSN += 1
                                if not "PASS" in DATA["SN"][sn][chassis][testdate][testetime]['DETAILTESTSTEP'][eachdeatilteststep].upper():
                                    HVstatus = "FAIL"
                        if HVstatus == "PASS":
                            if not HVtestcountinSN:
                                HVstatus = "NO TEST"
                            else:
                                if not "PASS" in TestStatus:
                                    if HVtestcount != HVtestcountinSN:
                                        HVstatus = "FAIL"
                        LVstatus = "PASS"
                        LVtestcount = 0
                        LVtestcountinSN = 0
                        LVteststeplistforcheck = list()
                        if testdate in LVteststeplist:
                            LVteststeplistforcheck = LVteststeplist[testdate]
                        else:
                            LVteststeplistforcheck = LVteststeplist['MAIN']
                        for eachdeatilteststep in LVteststeplistforcheck:
                            LVtestcount += 1
                            if eachdeatilteststep in DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]:
                                LVtestcountinSN += 1
                                if not "PASS" in DATA["SN"][sn][chassis][testdate][testetime]['DETAILTESTSTEP'][eachdeatilteststep].upper():
                                    LVstatus = "FAIL"
                        if LVstatus == "PASS":
                            if not LVtestcountinSN:
                                LVstatus = "NO TEST"
                            else:
                                if not "PASS" in TestStatus:
                                    if LVtestcount != LVtestcountinSN:
                                        LVstatus = "FAIL"

                        HVLVData[sn]["DATA"][timestamp]["HVstatus"] = HVstatus
                        HVLVData[sn]["DATA"][timestamp]["LVstatus"] = LVstatus
                        HVLVData[sn]["DATA"][timestamp]["RESULT"] = DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']
                        HVLVData[sn]["DATA"][timestamp]["DETAILTESTSTEP"] = DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]
                        HVLVData[sn]["DATA"][timestamp]["TESTSTEPLIST"] = DATA["SN"][sn][chassis][testdate][testetime]["TESTSTEPLIST"]
                        #if HVstatus == "PASS" and LVstatus == "FAIL":
                            #print(json.dumps(HVLVData[sn], indent = 4))    
                            #sys.exit()

    #print(json.dumps(HVLVData, indent = 4))  
    difftime = datetime.now()-start
    print("generateHVLVdata: {} use {} seconds".format(teststep,difftime.total_seconds()))
    return HVLVData


def generateexeclHVLVsummary(workingonSNlist,DATA,wb,FULLDATA,TempHVLVdata,pr):
    start=datetime.now()
    sheettitle = "HVLV_summary_By_1st"
    ws2 = wb.create_sheet(title=sheettitle)
    print("{}".format("generateexeclHVLVsummary"))
    FailureSNlist = list()
    teststeplist = list()
    listofweeknumber = list()
    for teststep in TempHVLVdata:
        teststeplist.append(teststep)

    summarydata = dict()
    checksteplist = list()
    for teststep in teststeplist:
        checksteplist.append("{}_HV".format(teststep))
        checksteplist.append("{}_LV".format(teststep))
    columnList = list()
    columnList.append("4C-L_LV_Failwith4C-L_HVand4C-H_LVPass")
    columnList.append("4C-H_HV_Failwith4C-L_HVand4C-H_LVPass")
    for sn in workingonSNlist:
        
        checksnstatus = False
        for teststep in teststeplist:
            if sn in TempHVLVdata[teststep]:
                checksnstatus = True
        if not checksnstatus:
            continue
        checkweeknumber = None

        for teststep in teststeplist:
            if sn in TempHVLVdata[teststep]:
                if not checkweeknumber:
                    checkweeknumber = findworkweek(TempHVLVdata[teststep][sn]["LIST"][0])
        if not checkweeknumber in summarydata:
            summarydata[checkweeknumber] = dict()
            summarydata[checkweeknumber]["total"] = 0
            summarydata[checkweeknumber]["4C-L_LV_Failwith4C-L_HVand4C-H_LVPass"] = 0
            summarydata[checkweeknumber]["4C-H_HV_Failwith4C-L_HVand4C-H_LVPass"] = 0
            summarydata[checkweeknumber]["4C-L_LV_Failwith4C-L_HVand4C-H_LVPassList"] = list()
            summarydata[checkweeknumber]["4C-H_HV_Failwith4C-L_HVand4C-H_LVPassList"] = list()

        if not checkweeknumber in listofweeknumber:
            listofweeknumber.append(checkweeknumber)
            listofweeknumber.sort(reverse=True)

        
        HVLVstatuslist = list()
        havebothtestresult = True
        for teststep in teststeplist:
            if sn in TempHVLVdata[teststep]:
                HVLVstatuslist.append(TempHVLVdata[teststep][sn]["DATA"][TempHVLVdata[teststep][sn]["LIST"][0]]["HVstatus"])
                HVLVstatuslist.append(TempHVLVdata[teststep][sn]["DATA"][TempHVLVdata[teststep][sn]["LIST"][0]]["LVstatus"])
                #wirtedata.append(TempHVLVdata[teststep][sn]["DATA"][TempHVLVdata[teststep][sn]["LIST"][0]]["RESULT"])
            else:
                HVLVstatuslist.append("NULL")
                HVLVstatuslist.append("NULL")
                havebothtestresult = False
        if HVLVstatuslist[0] == "PASS" and HVLVstatuslist[3] == "PASS":
            summarydata[checkweeknumber]["total"] += 1
            if HVLVstatuslist[1] == "FAIL":
                summarydata[checkweeknumber]["4C-L_LV_Failwith4C-L_HVand4C-H_LVPass"] += 1
                summarydata[checkweeknumber]["4C-L_LV_Failwith4C-L_HVand4C-H_LVPassList"].append(sn)
                if not sn in FailureSNlist:
                    FailureSNlist.append(sn)
            if HVLVstatuslist[2] == "FAIL":
                summarydata[checkweeknumber]["4C-H_HV_Failwith4C-L_HVand4C-H_LVPass"] += 1
                summarydata[checkweeknumber]["4C-H_HV_Failwith4C-L_HVand4C-H_LVPassList"].append(sn)
                if not sn in FailureSNlist:
                    FailureSNlist.append(sn)                
        # if summarydata[checkweeknumber]["4C-L_LV_Failwith4C-L_HVand4C-H_LVPass"] or summarydata[checkweeknumber]["4C-H_HV_Failwith4C-L_HVand4C-H_LVPass"]:
        #     for teststep in teststeplist:
        #         if sn in TempHVLVdata[teststep]:
        #             print(teststep)
        #             pr['modules'].print_anyinformation(TempHVLVdata[teststep][sn])
        #     pr['modules'].print_anyinformation(summarydata)
        #     pr['modules'].print_anyinformation(checksteplist)
        #     pr['modules'].print_anyinformation(HVLVstatuslist)

        #     sys.exit()

    #pr['modules'].print_anyinformation(summarydata)
    #pr['modules'].print_anyinformation(FailureSNlist)
    #sys.exit()

    wirtedata = list()
    wirtedata.append("WeekNumber")
    wirtedata.append("TOTALwith4C-L_HVand4C-H_LVPass")
    for culumnName in columnList:
        wirtedata.append(culumnName)
        wirtedata.append("FailureRate")
    
    ws2.append(wirtedata)

    for workonweeknumber in listofweeknumber:

        wirtedata = list()

        wirtedata.append(workonweeknumber)
        wirtedata.append(summarydata[workonweeknumber]['total'])
        for culumnName in columnList:
            wirtedata.append(summarydata[workonweeknumber][culumnName])
            failurerate = 0
            if summarydata[workonweeknumber]['total']:
                failurerate = summarydata[workonweeknumber][culumnName]/summarydata[workonweeknumber]['total']
            wirtedata.append("{:.2f}%".format(failurerate * 100))

        ws2.append(wirtedata)

    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinOrange(ws2,'NULL')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'B2')
    converttoPERCENTAGEnumberbyFailurerange(ws2)
    difftime = datetime.now()-start
    print("generateexeclHVLVsummary: {} use {} seconds".format(teststep,difftime.total_seconds()))
    generateexeclHVLVdata(FailureSNlist,DATA,wb,FULLDATA,TempHVLVdata,failureonly=True)
    return 0

def generateexeclHVLVdata(workingonSNlist,DATA,wb,FULLDATA,TempHVLVdata,failureonly=False):
    start=datetime.now()
    sheettitle = "HVLV_status_By_1st"
    if failureonly:
        sheettitle = "HVLV_FailureData"
    ws2 = wb.create_sheet(title=sheettitle)
    print("{}".format("generateexeclHVLVdata"))
    teststeplist = list()
    for teststep in TempHVLVdata:
        teststeplist.append(teststep)
    wirtedata = list()
    #wirtedata.append('TEST')
    wirtedata.append('SN')
    for teststep in teststeplist:
        wirtedata.append("{}_WeekNumber".format(teststep))  
    for teststep in teststeplist:
        wirtedata.append("{}_HV_RESULT".format(teststep))
        wirtedata.append("{}_LV_RESULT".format(teststep))
    #wirtedata.append('RESULT')
    for teststep in teststeplist:
        for eachdeatilteststep in DATA[teststep]["DETAILTESTSTEP"]:
             wirtedata.append("{}_{}".format(teststep, eachdeatilteststep))
    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        
        checksnstatus = False
        for teststep in teststeplist:
            if sn in TempHVLVdata[teststep]:
                checksnstatus = True
        if not checksnstatus:
            continue

        wirtedata = list()
        #wirtedata.append(teststep)
        wirtedata.append(sn)
        for teststep in teststeplist:
            if sn in TempHVLVdata[teststep]:
                wirtedata.append(findworkweek(TempHVLVdata[teststep][sn]["LIST"][0]))
            else:
                wirtedata.append("NULL")
        for teststep in teststeplist:
            if sn in TempHVLVdata[teststep]:
                wirtedata.append(TempHVLVdata[teststep][sn]["DATA"][TempHVLVdata[teststep][sn]["LIST"][0]]["HVstatus"])
                wirtedata.append(TempHVLVdata[teststep][sn]["DATA"][TempHVLVdata[teststep][sn]["LIST"][0]]["LVstatus"])
                #wirtedata.append(TempHVLVdata[teststep][sn]["DATA"][TempHVLVdata[teststep][sn]["LIST"][0]]["RESULT"])
            else:
                wirtedata.append("NULL")
                wirtedata.append("NULL")
        for teststep in teststeplist:
            for eachdeatilteststep in DATA[teststep]["DETAILTESTSTEP"]:
                if sn in TempHVLVdata[teststep]:
                    if eachdeatilteststep in TempHVLVdata[teststep][sn]["DATA"][TempHVLVdata[teststep][sn]["LIST"][0]]["DETAILTESTSTEP"]:
                        wirtedata.append(TempHVLVdata[teststep][sn]["DATA"][TempHVLVdata[teststep][sn]["LIST"][0]]['DETAILTESTSTEP'][eachdeatilteststep])
                    else:
                        wirtedata.append("NO TEST DATA")
                else:
                    wirtedata.append("NULL")
        
        ws2.append(wirtedata)
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinOrange(ws2,'NULL')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'H2')
    difftime = datetime.now()-start
    print("generateexeclHVLVdata: {} use {} seconds".format(teststep,difftime.total_seconds()))
    return 0

def generateexecltest(workingonSNlist,DATA,teststep,wb,FULLDATA):

    ws2 = wb.create_sheet(title=teststep)
    print("{}: {}".format("generateexecltest", teststep))
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('WEEK')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('TESTTIME(HH:MM:SS)')
    wirtedata.append('# of Card in MTP')
    wirtedata.append('CARDTYPE')
    wirtedata.append('SLOT')
    wirtedata.append('RESULT')
    if teststep == 'DL':
        wirtedata.append('PN')
        wirtedata.append('MACADDRESS')
        #wirtedata.append('CPLD1')
        #wirtedata.append('CPLD2')
        #wirtedata.append('QSPI')
    chassisinfo = list()
    imageinfo = list()
    nicinfo = list()

    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        if "MTPINFO" in DATA["SN"][sn][chassis][testdate][testetime]:
                            if type(DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"]) is dict:
                                for chassiskey in DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"]:
                                    if not chassiskey in chassisinfo:
                                        chassisinfo.append(chassiskey)
                                        chassisinfo.sort()
                        if "IMAGE" in DATA["SN"][sn][chassis][testdate][testetime]:
                            if type(DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"]) is dict:
                                for imagekey in DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"]:
                                    if not imagekey in imageinfo:
                                        imageinfo.append(imagekey)
                                        imageinfo.sort()
                        if "NICINFO" in DATA["SN"][sn][chassis][testdate][testetime]:
                            if type(DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"]) is dict:
                                for nickey in DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"]:
                                    if not nickey in nicinfo:
                                        nicinfo.append(nickey)
                                        nicinfo.sort()

    for chassiskey in chassisinfo:
        wirtedata.append(chassiskey)
    for imagekey in imageinfo:
        wirtedata.append(imagekey)
    for nickey in nicinfo:
        wirtedata.append(nickey)

    for eachdeatilteststep in DATA["DETAILTESTSTEP"]:
         wirtedata.append(eachdeatilteststep)
    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        if 'SLOT' in DATA["SN"][sn][chassis][testdate][testetime]:
                            slot = DATA["SN"][sn][chassis][testdate][testetime]['SLOT']
                        else:
                            slot = None
                        wirtedata = list()
                        wirtedata.append(teststep)
                        wirtedata.append(sn)
                        wirtedata.append(chassis)
                        wirtedata.append(findworkweek("{}_{}".format(testdate,testetime)))
                        wirtedata.append(testdate)
                        wirtedata.append(testetime.replace("-",":"))
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(timedelta(seconds=int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                        else:
                            wirtedata.append('NO TESTTIME')
                        timestamp = "{}_{}".format(testdate,testetime)
                        wirtedata.append(findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp))
                        if 'CARDTYPE' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CARDTYPE'])
                        else:
                            wirtedata.append("NONE")
                        if 'SLOT' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['SLOT'])
                        else:
                            wirtedata.append("NONE")
                        if 'FINALRESULT' in DATA["SN"][sn][chassis][testdate][testetime]:
                            #DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD']
                            wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                        else:
                            wirtedata.append("NONE")
                        if teststep == 'DL':
                            if 'PN' in DATA["SN"][sn][chassis][testdate][testetime]:
                                wirtedata.append(str(DATA["SN"][sn][chassis][testdate][testetime]['PN']))
                            else:
                                wirtedata.append('CANNOT FIND')
                                
                            if 'MACADDRESS' in DATA["SN"][sn][chassis][testdate][testetime]:
                                wirtedata.append(str(DATA["SN"][sn][chassis][testdate][testetime]['MACADDRESS']))
                            else:
                                wirtedata.append('CANNOT FIND')
                            #print(DATA["SN"][sn][chassis][testdate][testetime])
                            #sys.exit()
                            # if 'CPLD1' in DATA["SN"][sn][chassis][testdate][testetime]['CPLD']:
                            #     wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CPLD']['CPLD1'][slot])
                            # else:
                            #     wirtedata.append('CANNOT FIND')
                            # if 'CPLD2' in DATA["SN"][sn][chassis][testdate][testetime]['CPLD']:
                            #     wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CPLD']['CPLD2'][slot])
                            # else:
                            #     wirtedata.append('CANNOT FIND')
                            # if 'QSPI' in DATA["SN"][sn][chassis][testdate][testetime]:
                            #     wirtedata.append(str(DATA["SN"][sn][chassis][testdate][testetime]['QSPI'][slot]))
                            # else:
                            #     wirtedata.append('CANNOT FIND')
                        
                        for chassiskey in chassisinfo:
                            #wirtedata.append(chassiskey)
                            if "MTPINFO" in DATA["SN"][sn][chassis][testdate][testetime]:
                                if type(DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"]) is dict:
                                    if chassiskey in DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"]:
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"][chassiskey])
                                    else:
                                        wirtedata.append('CANNOT FIND')
                                else:
                                    wirtedata.append('CANNOT FIND')
                            else:
                                wirtedata.append('CANNOT FIND')
                        for imagekey in imageinfo:
                            #wirtedata.append(imagekey)
                            if "IMAGE" in DATA["SN"][sn][chassis][testdate][testetime]:
                                if type(DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"]) is dict:
                                    if imagekey in DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"]:
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"][imagekey])
                                    else:
                                        wirtedata.append('CANNOT FIND')
                                else:
                                    wirtedata.append('CANNOT FIND')
                            else:
                                wirtedata.append('CANNOT FIND')
                        for nickey in nicinfo:
                            #wirtedata.append(nickey)
                            if "NICINFO" in DATA["SN"][sn][chassis][testdate][testetime]:
                                if type(DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"]) is dict:
                                    if nickey in DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"]:
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"][nickey])
                                    else:
                                        wirtedata.append('CANNOT FIND')
                                else:
                                    wirtedata.append('CANNOT FIND')
                            else:
                                wirtedata.append('CANNOT FIND')    

                        for eachdeatilteststep in DATA["DETAILTESTSTEP"]:
                            if eachdeatilteststep in DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]:
                                wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['DETAILTESTSTEP'][eachdeatilteststep])
                            else:
                                wirtedata.append("NO TEST DATA")
                        
                        ws2.append(wirtedata)
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'L2')
    return 0

def generateexeclerrdata(workingonSNlist,DATA,teststep,wb,FULLDATA):

    teststeperrortitle = "{}_ERROR".format(teststep)
    print("{}: {}".format("generateexeclerrdata", teststeperrortitle))
    ws2 = wb.create_sheet(title=teststeperrortitle)
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('CARDTYPE')
    wirtedata.append('SLOT')

    wirtedata.append('RESULT')
    wirtedata.append('ERROR_DATA')
    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        if 'FINALRESULT' in DATA["SN"][sn][chassis][testdate][testetime]:
                            if "FAIL" in DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']:
                                slot = DATA["SN"][sn][chassis][testdate][testetime]['SLOT']
                                howmanyerror = len(DATA["SN"][sn][chassis][testdate][testetime]['ERRORDETAIL'])
                                if howmanyerror:
                                    for eacherror in DATA["SN"][sn][chassis][testdate][testetime]['ERRORDETAIL']:
                                        eacherrorlist = eacherror.split('\n')
                                        neweacherror = ''
                                        for checkeacherror in eacherrorlist:

                                            checkeacherror = _removeIllegalCharacterError(checkeacherror)
                                            neweacherror += checkeacherror
                                            neweacherror += "\r"
                                        wirtedata = list()
                                        wirtedata.append(teststep)
                                        wirtedata.append(sn)
                                        wirtedata.append(chassis)
                                        wirtedata.append(testdate)
                                        wirtedata.append(testetime)
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CARDTYPE'])
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['SLOT'])

                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                                        #print(eacherror)
                                        #if len(eacherror) > 200:
                                            #eacherror = eacherror[:200]
                                        wirtedata.append(neweacherror)
                                        #print(wirtedata)
                                        ws2.append(wirtedata)
                                        #print(wirtedata)
    
    fixcolumnssize2(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    wraptest(ws2)
    return 0

def generateexeclerrdata2(workingonSNlist,DATA,teststep,wb,FULLDATA):

    teststeperrortitle = "{}_ERROR_2".format(teststep)
    print("{}: {}".format("generateexeclerrdata", teststeperrortitle))
    ws2 = wb.create_sheet(title=teststeperrortitle)
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('CARDTYPE')
    wirtedata.append('SLOT')

    wirtedata.append('RESULT')
    wirtedata.append('FAILED_STEP')
    wirtedata.append('ERROR_DATA')
    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        if 'FINALRESULT' in DATA["SN"][sn][chassis][testdate][testetime]:
                            if "FAIL" in DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']:
                                slot = DATA["SN"][sn][chassis][testdate][testetime]['SLOT']
                                howmanyerror = len(DATA["SN"][sn][chassis][testdate][testetime]['ERRORDETAIL'])
                                if howmanyerror:
                                    FAILURE_STEPS = list()
                                    FAILURE_DICT = dict()
                                    for eacherror in DATA["SN"][sn][chassis][testdate][testetime]['ERRORDETAIL']:
                                        if "ERR MSG ==" in eacherror:
                                            failurestep = None
                                            if "L1 Sub Test only passed" in eacherror:
                                                continue
                                            if "DIAG TEST" in eacherror:
                                                Needworkonlist = eacherror.split('DIAG TEST')
                                                match = re.findall(r"(\w.*\w),\s+ERR MSG ==\s?(.*)",Needworkonlist[-1])
                                                #print(match)
                                                if match:
                                                    failurestep = match[0][0]
                                                    failurestep = failurestep.strip() 
                                            else:
                                                match = re.findall(r"(\w.*\w),\s+ERR MSG ==\s?(.*)",eacherror)
                                                #print(match)
                                                if match:
                                                    failurestep = match[0][0]
                                                    failurestep = failurestep.strip()                                   
                                            eacherrorlist = eacherror.split('\n')
                                            neweacherror = ''
                                            for checkeacherror in eacherrorlist:
                                                checkeacherror = _removeIllegalCharacterError(checkeacherror)
                                                neweacherror += checkeacherror
                                                neweacherror += "\r"
                                            if not failurestep in FAILURE_STEPS:
                                                FAILURE_STEPS.append(failurestep)
                                            if not failurestep in FAILURE_DICT:
                                                FAILURE_DICT[failurestep] = neweacherror
                                            else:
                                                FAILURE_DICT[failurestep] += neweacherror 

                                    for failurestep in FAILURE_STEPS:
                                        wirtedata = list()
                                        wirtedata.append(teststep)
                                        wirtedata.append(sn)
                                        wirtedata.append(chassis)
                                        wirtedata.append(testdate)
                                        wirtedata.append(testetime)
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CARDTYPE'])
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['SLOT'])
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                                        #print(eacherror)
                                        #if len(eacherror) > 200:
                                            #eacherror = eacherror[:200]
                                        wirtedata.append(failurestep)
                                        wirtedata.append(FAILURE_DICT[failurestep][:-2])
                                        #print(wirtedata)
                                        ws2.append(wirtedata)
                                        
                                        #print(wirtedata)
    
    fixcolumnssize2(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    wraptest(ws2)
    return 0

def fixcolumnssize(ws,enablefilter=True):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
    for col, value in dims.items():
        if value > 30:
            value = 30
        ws.column_dimensions[col].width = value + 2
    if enablefilter:
        ws.auto_filter.ref = ws.dimensions

def fixcolumnssize2(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
    for col, value in dims.items():
        if value > 100:
            value = 100
        ws.column_dimensions[col].width = value + 2

    ws.auto_filter.ref = ws.dimensions

def wraptest(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True,vertical='top') 

def freezePosition(ws,keyword):
    ws.freeze_panes = keyword

def highlightinAqua(ws,keyword):
    from openpyxl.styles import Color, PatternFill, Font, Border
    maxRow = ws.max_row
    maxCol = ws.max_column
    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
    for rowNum in range(1, maxRow + 1):
        fillcolor = 0
        for colNum in range(1, maxCol + 1):
            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
            if keyword in str(ws.cell(row=rowNum, column=colNum).value):
                ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type = 'solid')

def highlightinOrange(ws,keyword):
    from openpyxl.styles import Color, PatternFill, Font, Border
    maxRow = ws.max_row
    maxCol = ws.max_column
    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
    for rowNum in range(1, maxRow + 1):
        fillcolor = 0
        for colNum in range(1, maxCol + 1):
            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
            if keyword in str(ws.cell(row=rowNum, column=colNum).value):
                ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type = 'solid')

def highlightinyellow(ws,keyword):
	from openpyxl.styles import Color, PatternFill, Font, Border
	maxRow = ws.max_row
	maxCol = ws.max_column
	#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	for rowNum in range(1, maxRow + 1):
		fillcolor = 0
		for colNum in range(1, maxCol + 1):
			#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
			if keyword in str(ws.cell(row=rowNum, column=colNum).value):
				ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')

def highlightnumberincell(ws):
    import numbers
    from openpyxl.styles import Color, PatternFill, Font, Border
    maxRow = ws.max_row
    maxCol = ws.max_column
    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
    for rowNum in range(1, maxRow + 1):
        fillcolor = 0
        for colNum in range(1, maxCol + 1):
            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
            #if keyword in str(ws.cell(row=rowNum, column=colNum).value):
            if isinstance(ws.cell(row=rowNum, column=colNum).value, numbers.Number):
                if ws.cell(row=rowNum, column=colNum).value < 3:
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
                elif ws.cell(row=rowNum, column=colNum).value >= 3 and ws.cell(row=rowNum, column=colNum).value < 8:
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFBF00', end_color='FFBF00', fill_type = 'solid')
                else:
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')

def highlightingreen(ws,keyword):
	from openpyxl.styles import Color, PatternFill, Font, Border
	maxRow = ws.max_row
	maxCol = ws.max_column
	#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	for rowNum in range(1, maxRow + 1):
		fillcolor = 0
		for colNum in range(1, maxCol + 1):
			#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
			if keyword in str(ws.cell(row=rowNum, column=colNum).value):
				fillcolor = 1
				ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='66f86a', end_color='66f86a', fill_type = 'solid')
                
def highlightinred(ws,keyword):
	from openpyxl.styles import Color, PatternFill, Font, Border
	maxRow = ws.max_row
	maxCol = ws.max_column
	#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	for rowNum in range(1, maxRow + 1):
		fillcolor = 0
		for colNum in range(1, maxCol + 1):
			#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
			if keyword in str(ws.cell(row=rowNum, column=colNum).value):
				ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')                

def highlightinlightredrow(ws,keyword):
    from openpyxl.styles import Color, PatternFill, Font, Border
    maxRow = ws.max_row
    maxCol = ws.max_column
    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
    for rowNum in range(1, maxRow + 1):
        fillcolor = 0
        for colNum in range(1, maxCol + 1):
            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
            if keyword in str(ws.cell(row=rowNum, column=colNum).value):
                fillcolor = 1
            if fillcolor:
                ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type = 'solid')   

def converttoPERCENTAGEnumber(ws):
    from openpyxl.styles import Color, PatternFill, Font, Border
    maxRow = ws.max_row
    maxCol = ws.max_column
    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
    for rowNum in range(1, maxRow + 1):
        for colNum in range(1, maxCol + 1):
            checkvalue = str(ws.cell(row=rowNum, column=colNum).value)
            if '%' in checkvalue:
                checkvalue = float(checkvalue[:-1])/100
                ws.cell(row=rowNum, column=colNum).value = checkvalue
                ws.cell(row=rowNum, column=colNum).number_format = '0.00%'
                if checkvalue > 0.95:
                    #Green
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='66f86a', end_color='66f86a', fill_type = 'solid')
                elif checkvalue < 0.1:
                    #Red
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                elif checkvalue > 0.75:
                    #Yellow
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
                elif checkvalue < 0.25:
                    #Pink
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type = 'solid')
                elif checkvalue > 0.50:
                    #Arylide Yellow
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='E9D66B', end_color='E9D66B', fill_type = 'solid')
                else:
                    #Orange
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type = 'solid')

def converttoPERCENTAGEnumberbyFailurerange(ws):
    from openpyxl.styles import Color, PatternFill, Font, Border
    maxRow = ws.max_row
    maxCol = ws.max_column
    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
    for rowNum in range(1, maxRow + 1):
        for colNum in range(1, maxCol + 1):
            checkvalue = str(ws.cell(row=rowNum, column=colNum).value)
            if '%' in checkvalue:
                checkvalue = float(checkvalue[:-1])/100
                ws.cell(row=rowNum, column=colNum).value = checkvalue
                ws.cell(row=rowNum, column=colNum).number_format = '0.00%'
                if checkvalue < 0.005:
                    #Green
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type = 'solid')
                elif checkvalue < 0.01:
                    #Sap Green
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='66FF00', end_color='66FF00', fill_type = 'solid')
                elif checkvalue < 0.02:
                    #Antique Bronze
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='BBFF00', end_color='BBFF00', fill_type = 'solid')
                elif checkvalue < 0.03:
                    #Chestnut
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type = 'solid')
                elif checkvalue < 0.04:
                    #Sweet Brown
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFAA00', end_color='FFAA00', fill_type = 'solid')
                elif checkvalue < 0.05:
                    #Cardinal
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF5500', end_color='FF5500', fill_type = 'solid')
                else:
                    #Red
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')

    # FF0000 <-- red
    # FF1100
    # FF2200
    # FF3300
    # FF4400
    # FF5500
    # FF6600
    # FF7700
    # FF8800
    # FF9900
    # FFAA00
    # FFBB00
    # FFCC00
    # FFDD00
    # FFEE00
    # FFFF00 <-- yellow
    # EEFF00
    # DDFF00
    # CCFF00
    # BBFF00
    # AAFF00
    # 99FF00
    # 88FF00
    # 77FF00
    # 66FF00
    # 55FF00
    # 44FF00
    # 33FF00
    # 22FF00
    # 11FF00
    # 00FF00 <-- green

def findhowmanycardinthistestbymtp(DATA,mtp,test,timestamp):
    
    if not mtp in DATA["MTPCHASSIS"]:
        return "No Data"

    if not test in DATA["MTPCHASSIS"][mtp]:
        return "No Data"

    if not timestamp in DATA["MTPCHASSIS"][mtp][test]:
        return "No Data"

    mtpdata = DATA["MTPCHASSIS"][mtp][test][timestamp]

    #print(json.dumps(mtpdata, indent = 4))

    if not "NICRESULT" in mtpdata:
        return "No Data"

    return len(mtpdata["NICRESULT"].keys())

def generateexeclmtpstatusbyeachmtpreport(DATA,wb,inputconfig,mtp,startdate=None):
    print("generateexeclmtpstatusbyeachmtpreport: {}".format(mtp))
    titlename = mtp
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append("TEST")
    wirtedata.append("DATE_TIME")
    for slot in inputconfig["MTP_STATUS"]:
        wirtedata.append(slot)
    wirtedata.append("FAILURE_TYPE")
    for slot in inputconfig["MTP_STATUS"]:
        wirtedata.append(slot)
    ws2.append(wirtedata)

    timelinelist = list()
    for test in DATA["MTPCHASSIS"][mtp]:
        for datetime in sorted(DATA["MTPCHASSIS"][mtp][test].keys(), reverse=True):
            if startdate:
                datelist = datetime.split("_")
                if datelist[0] < startdate:
                    continue
            if not datetime in timelinelist:
                timelinelist.append(datetime)

    timelinelist.sort(reverse=True)
    for datetime in timelinelist:
        for test in DATA["MTPCHASSIS"][mtp]:
            if datetime in DATA["MTPCHASSIS"][mtp][test]:
                wirtedata = list()
                wirtedataFailure = list()
                wirtedata.append(test)
                wirtedata.append(datetime)
                for slot in inputconfig["MTP_STATUS"]:
                    if not "FAILURESTEP" in DATA["MTPCHASSIS"][mtp][test][datetime]:
                        DATA["MTPCHASSIS"][mtp][test][datetime]["FAILURESTEP"] = dict()

                    if slot in DATA["MTPCHASSIS"][mtp][test][datetime]["NICRESULT"]:
                        if "FAIL" in DATA["MTPCHASSIS"][mtp][test][datetime]["NICRESULT"][slot]["RESULT"]:
                            #print(json.dumps(DATA["MTPCHASSIS"][mtp][test][datetime], indent = 4))
                            testdateandtime = datetime.split("_")
                            SN = DATA["MTPCHASSIS"][mtp][test][datetime]["NICRESULT"][slot]["SN"]
                            if SN in DATA["teststep"][test]["SN"]:
                                #print(json.dumps(DATA["teststep"][test]["SN"][SN][mtp][testdateandtime[0]][testdateandtime[1]], indent = 4))
                                firstFailurestep = None 
                                if SN in DATA["teststep"][test]["SN"]:
                                    if mtp in DATA["teststep"][test]["SN"][SN]:
                                        if testdateandtime[0] in DATA["teststep"][test]["SN"][SN][mtp]:
                                            if testdateandtime[1] in DATA["teststep"][test]["SN"][SN][mtp][testdateandtime[0]]:
                                                for teststepdata in DATA["teststep"][test]["SN"][SN][mtp][testdateandtime[0]][testdateandtime[1]]["TESTSTEPLIST"]:
                                                    for teststep in teststepdata:
                                                        if "FAIL" in teststepdata[teststep]:
                                                            firstFailurestep = "{} <{}>".format(teststep,teststepdata[teststep])
                                                    if firstFailurestep:
                                                        break
                                #print(firstFailurestep)
                                if not firstFailurestep:
                                    firstFailurestep = "CANNOT FIND FAILURE STEP"
                                wirtedataFailure.append(firstFailurestep)
                                DATA["MTPCHASSIS"][mtp][test][datetime]["FAILURESTEP"][slot] = firstFailurestep
                                #sys.exit()
                            else:
                                wirtedataFailure.append("CANNOT FIND SN: {}".format(SN))
                                DATA["MTPCHASSIS"][mtp][test][datetime]["FAILURESTEP"][slot] = "CANNOT FIND SN: {}".format(SN)
                        else:
                            wirtedataFailure.append("PASS")
                            DATA["MTPCHASSIS"][mtp][test][datetime]["FAILURESTEP"][slot] = "PASS"
                        wirtedata.append(DATA["MTPCHASSIS"][mtp][test][datetime]["NICRESULT"][slot]["RESULT"])
                    else:
                        wirtedata.append("None")
                        wirtedataFailure.append("None")
                        DATA["MTPCHASSIS"][mtp][test][datetime]["FAILURESTEP"][slot] = "None"
                wirtedata.append("|||")
                for eachdata in wirtedataFailure:
                    wirtedata.append(eachdata)

                ws2.append(wirtedata)

    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')      
    fixcolumnssize(ws2)

    return True

def generateexeclmtpstatussummaryreport(DATA,wb,inputconfig,startdate=None):
    ws1 = wb.active
    ws1.title = "SUMMARY"

    mtpchassisusecountbyslot = dict()
    bytestmtpchassisusecountbyslot = dict()
    for MTP in DATA["MTPCHASSIS"]:
        #print(MTP)
        if not MTP in mtpchassisusecountbyslot:
            mtpchassisusecountbyslot[MTP] = dict()
            mtpchassisusecountbyslot[MTP]["TEST"] = list()
            for slot in inputconfig["MTP_STATUS"]:
                mtpchassisusecountbyslot[MTP][slot] = dict()
                mtpchassisusecountbyslot[MTP][slot]["TOTAL"] = 0
                mtpchassisusecountbyslot[MTP][slot]["PASS"] = 0
                mtpchassisusecountbyslot[MTP][slot]["FAIL"] = 0
        for test in DATA["MTPCHASSIS"][MTP]:
            #print(test)
            if not test in mtpchassisusecountbyslot[MTP]["TEST"]:
                mtpchassisusecountbyslot[MTP]["TEST"].append(test)
            for datetime in DATA["MTPCHASSIS"][MTP][test]:
                #print(datetime)
                #print(json.dumps(DATA["MTPCHASSIS"][MTP][test][datetime], indent = 4))
                if startdate:
                    datelist = datetime.split("_")
                    if datelist[0] < startdate:
                        continue
                for eachslot in DATA["MTPCHASSIS"][MTP][test][datetime]["NICRESULT"]:
                    if "PASS" in DATA["MTPCHASSIS"][MTP][test][datetime]["NICRESULT"][eachslot]["RESULT"]:
                        mtpchassisusecountbyslot[MTP][eachslot]["TOTAL"] += 1
                        mtpchassisusecountbyslot[MTP][eachslot]["PASS"] += 1
                    else:
                        mtpchassisusecountbyslot[MTP][eachslot]["TOTAL"] += 1
                        mtpchassisusecountbyslot[MTP][eachslot]["FAIL"] += 1                        
            #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
            #sys.exit()
        #break

    
    #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
    #sys.exit()
    wirtedata = list()
    wirtedata.append('MTP')
    wirtedata.append('TEST')
    wirtedata.append('STATUS')
    wirtedata.append('TOTAL')
    for slot in inputconfig["MTP_STATUS"]:
        wirtedata.append(slot)
    ws1.append(wirtedata)
    for mtp in mtpchassisusecountbyslot:
        wirtedatatotal = list()
        wirtedatapass = list()
        wirtedatafail = list()
        wirtedatapassrate = list()
        wirtedatatotal.append(mtp)
        wirtedatapass.append(mtp)
        wirtedatafail.append(mtp)
        wirtedatapassrate.append(mtp)
        teststep = mtpchassisusecountbyslot[mtp]["TEST"][0]
        if len(mtpchassisusecountbyslot[mtp]["TEST"]) > 1:
            for eachtest in mtpchassisusecountbyslot[mtp]["TEST"][1:]:
                teststep = "{},{}".format(teststep,eachtest)
        wirtedatatotal.append(teststep)
        wirtedatapass.append(teststep)
        wirtedatafail.append(teststep)
        wirtedatapassrate.append(teststep)
        listoftotal = list()
        listofpass = list()
        listoffail = list()
        for slot in inputconfig["MTP_STATUS"]:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            listofpass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            listoffail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])  
        wirtedatatotal.append("TOTAL")
        wirtedatapass.append("PASS")
        wirtedatafail.append("FAIL")
        wirtedatapassrate.append("PASS_RATE")
        wirtedatatotal.append(sum(listoftotal))
        wirtedatapass.append(sum(listofpass))
        wirtedatafail.append(sum(listoffail))
        wirtedatapassrate.append(calculePass_rate(sum(listoftotal),sum(listofpass)))           
        for slot in inputconfig["MTP_STATUS"]:
            wirtedatatotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            wirtedatapass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            wirtedatafail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])
            wirtedatapassrate.append(calculePass_rate(mtpchassisusecountbyslot[mtp][slot]["TOTAL"],mtpchassisusecountbyslot[mtp][slot]["PASS"]))
        ws1.append(wirtedatatotal)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatafail)
        ws1.append(wirtedatapassrate)          
    fixcolumnssize(ws1)
    converttoPERCENTAGEnumber(ws1)
    highlightinlightredrow(ws1,'FAIL')
    freezePosition(ws1,'D2')

    return bytestmtpchassisusecountbyslot

def calculePass_rate(totalnumber,passnumber):

    calcule_yeild = 1
    if totalnumber:
        calcule_yeild = passnumber / totalnumber

    yeilddisplay = "{:.2f}%".format(calcule_yeild * 100)  

    return yeilddisplay

def generateexeclmacandsnreport(pr,DATA,wb,snmaclist):
    ws1 = wb.active
    ws1.title = "SNvsMAC"

    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('PN')
    wirtedata.append('MAC')
    wirtedata.append('FLEXFLOW')
    ws1.append(wirtedata)

    for SN in snmaclist:
        wirtedata = list()
        wirtedata.append(SN)        
        if len(snmaclist[SN]['PN']):
            wirtedata.append(snmaclist[SN]['PN'][0])
        else:
            wirtedata.append('') 
        if len(snmaclist[SN]['MAC']):
            wirtemacaddress = ''
            for macaddress in snmaclist[SN]['MAC']:
                wirtemacaddress += macaddress.replace('-','')
                wirtemacaddress += '\r'
            wirtedata.append(wirtemacaddress[:-2])
        else:
            wirtedata.append('')    
        if SN in pr['SNlist']["FailureSNFlexflow"]:
            wirtedata.append(pr['SNlist']["FailureSNFlexflow"][SN]) 
        ws1.append(wirtedata)

    wraptest(ws1)
    freezePosition(ws1,'B2')
    fixcolumnssize(ws1)


    return True

def generateexeclmacandsnreport2(pr,DATA,wb,snmaclist):
    ws1 = wb.create_sheet(title="MACvsSN")

    wirtedata = list()
    wirtedata.append('MAC')
    wirtedata.append('PN')
    wirtedata.append('SN')
    ws1.append(wirtedata)

    for mac in snmaclist:
        wirtedata = list()
        wirtedata.append(mac.replace('-',''))        
        if len(snmaclist[mac]['PN']):
            wirtedata.append(snmaclist[mac]['PN'][0])
        if len(snmaclist[mac]['SN']):
            for sn in snmaclist[mac]['SN']:
                wirtedata.append(sn)   
        ws1.append(wirtedata)

    wraptest(ws1)
    freezePosition(ws1,'B2')
    fixcolumnssize(ws1)


    return True

def generateexecldieidreport(DATA,wb,inputconfig,start=None):
    ws1 = wb.active
    ws1.title = "DIE_ID_DATA"

    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('DIE_ID')
    ws1.append(wirtedata)

    alllistofSN = list()
    for test in DATA['SN']['TEST']:
        wirtedata = list()
        wirtedata.append(test)
        listofSN = list()
        for SN in DATA['teststep'][test]["SN"]:
            recordSN = True
            if "SKIPPrefix" in inputconfig:
                if inputconfig["SKIPPrefix"] in SN:
                    recordSN = False
            if "STARTCountSN" in inputconfig:
                if SN < inputconfig["STARTCountSN"]:
                    recordSN = False
            if recordSN:
                listofSN.append(SN)
                if not SN in alllistofSN:
                    alllistofSN.append(SN)   

    for SN in alllistofSN:
        #DATA["SN"][sn][chassis][testdate][testetime]
        die_ids = list()
        for test in DATA['teststep']:
            if SN in DATA['teststep'][test]["SN"]:
                for chassis in DATA['teststep'][test]["SN"][SN]:
                    for testdate in DATA['teststep'][test]["SN"][SN][chassis]:
                        for testetime in DATA['teststep'][test]["SN"][SN][chassis][testdate]:
                            if "ELBA_DIE_ID" in DATA['teststep'][test]["SN"][SN][chassis][testdate][testetime]:
                                die_id = DATA['teststep'][test]["SN"][SN][chassis][testdate][testetime]["ELBA_DIE_ID"]
                                if not die_id in die_ids:
                                    die_ids.append(die_id)
        wirtedata = list()
        wirtedata.append(SN)        
        if len(die_ids):
            for die_id in die_ids:
                wirtedata.append(die_id)
        else:
            wirtedata.append("Cannot find")
        ws1.append(wirtedata)

    fixcolumnssize(ws1)

    return True


def generateexeclDatabyPN(DATA,wb,inputconfig,workingonSNlist,PN,chartdata,pr,start=None):

    titlename = "{}_SUMMARY".format(PN)
    ws2 = wb.create_sheet(title=titlename)
    print("{}: {}".format("generateexeclDatabyPN", titlename))

    chartdata['PN'] = SummaryReportDetail(DATA,ws2,inputconfig,workingonSNlist,pr,start)

    return None    

def generateexeclreport(DATA,wb,inputconfig,workingonSNlist,chartdata,pr,start=None):

    # 1st Sheet keep active
    ws1 = wb.active
    
    # Will creat extra "sheet" using below
    #ws1 = wb.create_sheet(title='SUMMARY')
    ws1.title = "SUMMARY"

    chartdata["SUMMARY"] = SummaryReportDetail(DATA,ws1,inputconfig,workingonSNlist,pr,start=None)

    return None

def generateexeclfailurereport(DATA,wb,inputconfig,workingonSNlist,pr,start=None):

    # 1st Sheet keep active
    ws1 = wb.active
    
    # Will creat extra "sheet" using below
    #ws1 = wb.create_sheet(title='SUMMARY')
    ws1.title = "FailureSUMMARY"

    SummaryfailureReportDetail(DATA,ws1,inputconfig,workingonSNlist,pr,start=None)

    return None

def generateexeclchart(DATA,wb,inputconfig,workingonSNlist,chartdata,pr,start=None):
    # 1st Sheet keep active
    ws1 = wb.active
    
    # Will creat extra "sheet" using below
    #ws1 = wb.create_sheet(title='SUMMARY')
    ws1.title = "SUMMARY"

    SummaryReportDetailchart(DATA,ws1,inputconfig,workingonSNlist,chartdata["SUMMARY"],pr,start)

    for test in DATA['SN']['TEST']:
        if test in chartdata["SUMMARY"]["WEEKTIME"]:
            ws2 = wb.create_sheet(title="{}_TestTimeChart".format(test))
            print("{}: {}".format("generateexeclchart", test))
            SummaryReportDetailTimechart(DATA,ws2,inputconfig,workingonSNlist,chartdata["SUMMARY"]["WEEKTIME"][test],pr,test,start)

    testlist = list()
    testlist.append('DL')
    testlist.append('P2C')

    SummaryReportDetailTimechartwithcoupletest(DATA,wb,inputconfig,workingonSNlist,chartdata["SUMMARY"]["WEEKTIME"],testlist,pr,start)        

    testlist = list()
    testlist.append('4C-L')
    testlist.append('4C-H')

    SummaryReportDetailTimechartwithcoupletest(DATA,wb,inputconfig,workingonSNlist,chartdata["SUMMARY"]["WEEKTIME"],testlist,pr,start)        

    return None

def SummaryReportDetailTimechartwithcoupletest(DATA,wb,inputconfig,workingonSNlist,chartdata,testlist,pr,start):

    #pr['modules'].print_anyinformation(chartdata)

    Havethosetest = True
    checkweeknumber = list()
    testname = ''
    for test in testlist:
        if test in chartdata:
            if len(testname) == 0:
                testname = "{}_".format(test)
            else:
                testname = "{}_{}_".format(testname, test)
            for weeknumber in chartdata[test]["WEEKLIST"]:
                if not weeknumber in checkweeknumber:
                    checkweeknumber.append(weeknumber)
                    checkweeknumber.sort()
        else:
            Havethosetest = False

    if not Havethosetest:
        return False

    ws1 = wb.create_sheet(title="{}TestTimeChart".format(testname))

    countrow = 1
    countcol = 1
    wirtedata = list()
    wirtedata.append('Weeknumber')
    wirtedata.append('Test Units (AVG)')
    wirtedata.append('MAX TestTime')
    wirtedata.append('AVG TestTime')  
    ws1.append(wirtedata)


    wirtedata = list()
    for weeknumber in checkweeknumber:
        countrow += 1
        wirtedata = list()
        wirtedata.append(weeknumber)
        avgnumber = list()
        MaxTimeinData = 0
        AveTimeinData = 0
        for test in testlist:
            if weeknumber in chartdata[test]["DATA"]:
                MaxSlotinData = chartdata[test]["DATA"][weeknumber]["numberofSLOTlist"][-1]
                avgnumber.append(len(chartdata[test]["DATA"][weeknumber]["DATAofNumber"][MaxSlotinData]))
                MaxTimeinData += max(chartdata[test]["DATA"][weeknumber]["DATAofNumber"][MaxSlotinData])
                from statistics import mean
                AveTimeinData += mean(chartdata[test]["DATA"][weeknumber]["DATAofNumber"][MaxSlotinData])
        from statistics import mean
        if len(avgnumber):
            wirtedata.append(int(mean(avgnumber)))
        else:
            wirtedata.append(0)
        wirtedata.append(timedelta(seconds=int(MaxTimeinData)))
        wirtedata.append(timedelta(seconds=int(AveTimeinData)))
        ws1.append(wirtedata)

    fixcolumnssize(ws1)
    converttoPERCENTAGEnumber(ws1)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "{} Test Time by Week".format(testname)
    chart1.y_axis.title = 'Units'
    chart1.x_axis.title = 'WEEK Number'

    data = Reference(ws1, min_col=2, min_row=1, max_row=countrow, max_col=2)
    cats = Reference(ws1, min_col=1, min_row=2, max_row=countrow)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.dataLabels = DataLabelList() 
    #chart1.dataLabels.showVal = True
    chart1.shape = 4

    # Create a second chart
    c2 = LineChart()
    v2 = Reference(ws1, min_col=3, min_row=1, max_row=countrow, max_col=4)
    c2.add_data(v2, titles_from_data=True)
    c2.set_categories(cats)
    c2.y_axis.axId = 200
    c2.y_axis.title = "Test Time in HH:MM:SS"

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    c2.y_axis.crosses = "max"
    chart1 += c2
    chart1.height = 20 # default is 7.5
    chart1.width = 30 # default is 15
    ws1.add_chart(chart1, "E1")     

    return None

def SummaryReportDetailTimechart(DATA,ws1,inputconfig,workingonSNlist,chartdata,pr,test,start):

    #pr['modules'].print_anyinformation(chartdata)

    countrow = 1
    countcol = 1
    wirtedata = list()
    wirtedata.append('Weeknumber')
    wirtedata.append('Test Units')
    wirtedata.append('MAX TestTime')
    wirtedata.append('AVG TestTime')  
    ws1.append(wirtedata)

    wirtedata = list()
    for weeknumber in chartdata["WEEKLIST"]:
        countrow += 1
        wirtedata = list()
        wirtedata.append(weeknumber)
        if weeknumber in chartdata["DATA"]:
            MaxSlotinData = chartdata["DATA"][weeknumber]["numberofSLOTlist"][-1]
            wirtedata.append(len(chartdata["DATA"][weeknumber]["DATAofNumber"][MaxSlotinData]))
            MaxTimeinData = max(chartdata["DATA"][weeknumber]["DATAofNumber"][MaxSlotinData])
            from statistics import mean
            AveTimeinData = mean(chartdata["DATA"][weeknumber]["DATAofNumber"][MaxSlotinData])
            wirtedata.append(timedelta(seconds=int(MaxTimeinData)))
            wirtedata.append(timedelta(seconds=int(AveTimeinData)))
        else:
            wirtedata.append(0)
            wirtedata.append("00:00:00")
            wirtedata.append("00:00:00")

        ws1.append(wirtedata)

    fixcolumnssize(ws1)
    converttoPERCENTAGEnumber(ws1)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "{} Test Time by Week".format(test)
    chart1.y_axis.title = 'Units'
    chart1.x_axis.title = 'WEEK Number'

    data = Reference(ws1, min_col=2, min_row=1, max_row=countrow, max_col=2)
    cats = Reference(ws1, min_col=1, min_row=2, max_row=countrow)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.dataLabels = DataLabelList() 
    #chart1.dataLabels.showVal = True
    chart1.shape = 4

    # Create a second chart
    c2 = LineChart()
    v2 = Reference(ws1, min_col=3, min_row=1, max_row=countrow, max_col=4)
    c2.add_data(v2, titles_from_data=True)
    c2.set_categories(cats)
    c2.y_axis.axId = 200
    c2.y_axis.title = "Test Time in HH:MM:SS"

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    c2.y_axis.crosses = "max"
    chart1 += c2
    chart1.height = 20 # default is 7.5
    chart1.width = 30 # default is 15
    ws1.add_chart(chart1, "E1")     

    return None


def SummaryReportDetailchart(DATA,ws1,inputconfig,workingonSNlist,chartdata,pr,start):

    #pr['modules'].print_anyinformation(chartdata)

    countrow = 1
    countcol = 1
    wirtedata = list()
    wirtedata.append('Weeknumber')
    for test in DATA['SN']['TEST']:
        countcol += 1
        wirtedata.append(test)
    wirtedata.append('EndtoEndYeild')
    ws1.append(wirtedata)

    #pr['modules'].print_anyinformation(chartdata["WEEKFIRST"])

    wirtedata = list()
    for weeknumber in chartdata["WEEKFIRST"]["LIST"]:
        countrow += 1
        wirtedata = list()
        wirtedata.append(weeknumber)
        
        if weeknumber in chartdata["WEEKFIRST"]["DATA"]:
            for test in DATA['SN']['TEST']:
                if "TOTAL" in chartdata["WEEKFIRST"]["DATA"][weeknumber][test]:
                    wirtedata.append(chartdata["WEEKFIRST"]["DATA"][weeknumber][test]["TOTAL"])
                else:
                    wirtedata.append(0)
        else:
            for test in DATA['SN']['TEST']:
                wirtedata.append(0)
        if weeknumber in chartdata["WEEKFIRST"]["DATA"]["testendtoendyeildlist"]:
            endtoendyeild = numpy.prod(chartdata["WEEKFIRST"]["DATA"]["testendtoendyeildlist"][weeknumber])
            endtoendyeilddisplay = "{:.2f}%".format(endtoendyeild * 100)
            wirtedata.append(endtoendyeilddisplay)
        else:
            wirtedata.append("0.0%")
        ws1.append(wirtedata)

    fixcolumnssize(ws1)
    converttoPERCENTAGEnumber(ws1)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "First Pass data by Week"
    chart1.y_axis.title = 'Units'
    chart1.x_axis.title = 'WEEK Number'

    data = Reference(ws1, min_col=2, min_row=1, max_row=countrow, max_col=countcol)
    cats = Reference(ws1, min_col=1, min_row=2, max_row=countrow)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.dataLabels = DataLabelList() 
    #chart1.dataLabels.showVal = True
    chart1.shape = 4

    # Create a second chart
    c2 = LineChart()
    v2 = Reference(ws1, min_col=countcol+1, min_row=1, max_row=countrow, max_col=countcol+1)
    c2.add_data(v2, titles_from_data=True)
    c2.set_categories(cats)
    c2.y_axis.axId = 200
    c2.y_axis.title = "End to End Yeild in %"

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    c2.y_axis.crosses = "max"
    chart1 += c2
    chart1.height = 20 # default is 7.5
    chart1.width = 30 # default is 15
    ws1.add_chart(chart1, "E1")     

    return None

def GetFailureSNList(DATA,inputconfig,pr):

    workingonSNlist = getsnlistafteestartdate(DATA,inputconfig,startdate=None)

    LastfailureremoveDupSN = dict()
    for SN in workingonSNlist:

        for test in DATA['SN']['TEST']:
            if not test in LastfailureremoveDupSN:
                LastfailureremoveDupSN[test] = dict()
                LastfailureremoveDupSN[test]["count"] = 0
                LastfailureremoveDupSN[test]["SN"] = list()
            if SN in DATA['SN']['LAST'][test]:
                if "result" in DATA['SN']['LAST'][test][SN]:
                    LastfailureremoveDupSN[test]["count"] += 1
                    LastfailureremoveDupSN[test]["SN"].append(SN)
                    break

    #print(json.dumps(LastfailureremoveDupSN, indent = 4))

    #sys.exit()

    return LastfailureremoveDupSN

def SummaryfailureReportDetail(DATA,ws1,inputconfig,workingonSNlist,pr,start=None):

    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN_QTY')
    #wirtedata.append('DETAIL_STEP_QTY')
    #wirtedata.append('First_PASS')
    #wirtedata.append('First_FAIL')
    #wirtedata.append('First_YIELD')
    #wirtedata.append('LAST_PASS')
    #wirtedata.append('LAST_FAIL')
    #wirtedata.append('LAST_YIELD')
    wirtedata.append('')
    wirtedata.append('LAST_Failure\rCount (NoDupSN)')
    if pr['FailureSNlist']:
        wirtedata.append('')
        wirtedata.append('FLEXFLOWSTATUS')
    ws1.append(wirtedata)

    LastfailureremoveDupSN = dict()
    for SN in workingonSNlist:

        for test in DATA['SN']['TEST']:
            if not test in LastfailureremoveDupSN:
                LastfailureremoveDupSN[test] = dict()
                LastfailureremoveDupSN[test]["count"] = 0
                LastfailureremoveDupSN[test]["SN"] = list()
            if SN in DATA['SN']['LAST'][test]:
                if "result" in DATA['SN']['LAST'][test][SN]:
                    if "FAIL" in DATA['SN']['LAST'][test][SN]["result"]:
                        LastfailureremoveDupSN[test]["count"] += 1
                        LastfailureremoveDupSN[test]["SN"].append(SN)
                        break
                else:
                    print("LAST <{}> SN: {}".format(test, SN))
                    #print(json.dumps(DATA['SN']['LAST'][test][SN], indent = 4))
                    del DATA['SN']['LAST'][test][SN]
                    #sys.exit()
            if SN in DATA['SN']['FIRST'][test]:
                if not "result" in DATA['SN']['FIRST'][test][SN]:
                    print("FIRST <{}> SN: {}".format(test, SN))
                    #print(json.dumps(DATA['SN']['FIRST'][test][SN], indent = 4))
                    del DATA['SN']['FIRST'][test][SN]

    #print(json.dumps(LastfailureremoveDupSN, indent = 4))
    #sys.exit()
    summarydata = dict()
    resultlist = ["PASS", "FAIL"]
    alllistofSN = list()
    for test in DATA['SN']['TEST']:
        if not test in summarydata:
            summarydata[test] = dict()
        wirtedata = list()
        wirtedata.append(test)
        listofSN = list()
        for SN in DATA['teststep'][test]["SN"]:
            recordSN = True
            if "SKIPPrefix" in inputconfig:
                if inputconfig["SKIPPrefix"] in SN:
                    recordSN = False
            if "STARTCountSN" in inputconfig:
                if SN < inputconfig["STARTCountSN"]:
                    recordSN = False
            if recordSN:
                if SN in workingonSNlist:
                    listofSN.append(SN)
                    if not SN in alllistofSN:
                        alllistofSN.append(SN)    

        wirtedata.append(len(listofSN))
        summarydata[test]["TOTAL"] = len(listofSN)
        #wirtedata.append(len(DATA['teststep'][test]["DETAILTESTSTEP"]))
        summarydata[test]["TOTALTestStep"] = len(DATA['teststep'][test]["DETAILTESTSTEP"])
        countResult = dict()
        countResult["PASS"] = 0
        for sn in listofSN:
            if "result" in DATA['SN']['FIRST'][test][sn]:
                testresult = DATA['SN']['FIRST'][test][sn]["result"]
                if not testresult in countResult:
                    countResult[testresult] = 1
                else:
                    countResult[testresult] += 1
        for testresult in resultlist:
            if testresult in countResult:
                #wirtedata.append(countResult[testresult])
                summarydata[test]["FIRST{}".format(testresult)] = countResult[testresult]
            else:
                 #wirtedata.append(0)
                 summarydata[test]["FIRST{}".format(testresult)] = 0
        if len(listofSN):
            firstyeild = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
            summarydata[test]["FIRSTYEILD"] = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
        else:
            firstyeild = "{:.2f}%".format(0)
            summarydata[test]["FIRSTYEILD"] = "{:.2f}%".format(0)
        #wirtedata.append(firstyeild)
        countResult = dict()
        countResult["PASS"] = 0
        for sn in listofSN:
            testresult = DATA['SN']['LAST'][test][sn]["result"]
            if not testresult in countResult:
                countResult[testresult] = 1
            else:
                countResult[testresult] += 1
        for testresult in resultlist:
            if testresult in countResult:
                #wirtedata.append(countResult[testresult])
                summarydata[test]["LAST{}".format(testresult)] = countResult[testresult]
            else:
                 #wirtedata.append(0)
                 summarydata[test]["LASTYEILD"] = "{:.2f}%".format(0)
        if len(listofSN):
            firstyeild = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
            summarydata[test]["LASTYEILD"] = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
        else:
            firstyeild = "{:.2f}%".format(0)
            summarydata[test]["LASTYEILD"] = "{:.2f}%".format(0)
        #wirtedata.append(firstyeild)
        wirtedata.append('')
        if test in LastfailureremoveDupSN:
            wirtedata.append(LastfailureremoveDupSN[test]["count"])
        else:
            wirtedata.append(0)

        if pr['FailureSNlist']:
            wirtedata.append('')
            if test in pr['FailureSNlist']['DATA']:
                for flexflowstatus in pr['FailureSNlist']['DATA'][test]["FlexflowType"]:
                    wirtedata.append("{} <{}>".format(flexflowstatus,len(pr['FailureSNlist']['DATA'][test]["FlexflowType"][flexflowstatus])))
        ws1.append(wirtedata)

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    wirtedata = list()
    wirtedata.append('FLEXFLOWSTATUS')
    wirtedata.append('SN_QTY')
    ws1.append(wirtedata)

    for flexflowtype in pr["FailureSNlist"]["FlexflowType"]:
        wirtedata = list()
        wirtedata.append(flexflowtype)
        wirtedata.append(len(pr["FailureSNlist"]["FlexflowType"][flexflowtype]))
        ws1.append(wirtedata)

    fixcolumnssize(ws1)
    converttoPERCENTAGEnumber(ws1)
    freezePosition(ws1, "B1")

    return None

def SummaryReportDetail(DATA,ws1,inputconfig,workingonSNlist,pr,start=None):

    #print(json.dumps(inputconfig, indent = 4))
    WeekTimedata = dict()
    for eachteststep in DATA['SN']['TEST']:
        TimeData = dict()
        WeekTimedata[eachteststep] = GetTestTimedictbyweek(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,DATA,TimeData,pr)
        #print(json.dumps(WeekTimedata, indent = 4))
        #sys.exit()

    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN_QTY')
    wirtedata.append('DETAIL_STEP_QTY')
    wirtedata.append('First_PASS')
    wirtedata.append('First_FAIL')
    wirtedata.append('First_YIELD')
    wirtedata.append('LAST_PASS')
    wirtedata.append('LAST_FAIL')
    wirtedata.append('LAST_YIELD')
    wirtedata.append('')
    wirtedata.append('LAST_Failure\rCount (NoDupSN)')
    if pr['FailureSNlist']:
        wirtedata.append('')
        wirtedata.append('FLEXFLOWSTATUS')
    ws1.append(wirtedata)

    LastfailureremoveDupSN = dict()
    for SN in workingonSNlist:

        for test in DATA['SN']['TEST']:
            if not test in LastfailureremoveDupSN:
                LastfailureremoveDupSN[test] = dict()
                LastfailureremoveDupSN[test]["count"] = 0
                LastfailureremoveDupSN[test]["SN"] = list()
            if SN in DATA['SN']['LAST'][test]:
                if "result" in DATA['SN']['LAST'][test][SN]:
                    if "FAIL" in DATA['SN']['LAST'][test][SN]["result"]:
                        LastfailureremoveDupSN[test]["count"] += 1
                        LastfailureremoveDupSN[test]["SN"].append(SN)
                        break
                else:
                    print("LAST <{}> SN: {}".format(test, SN))
                    #print(json.dumps(DATA['SN']['LAST'][test][SN], indent = 4))
                    del DATA['SN']['LAST'][test][SN]
                    #sys.exit()
            if SN in DATA['SN']['FIRST'][test]:
                if not "result" in DATA['SN']['FIRST'][test][SN]:
                    print("FIRST <{}> SN: {}".format(test, SN))
                    #print(json.dumps(DATA['SN']['FIRST'][test][SN], indent = 4))
                    del DATA['SN']['FIRST'][test][SN]

    #print(json.dumps(LastfailureremoveDupSN, indent = 4))
    #sys.exit()
    summarydata = dict()
    resultlist = ["PASS", "FAIL"]
    alllistofSN = list()
    for test in DATA['SN']['TEST']:
        if not test in summarydata:
            summarydata[test] = dict()
        wirtedata = list()
        wirtedata.append(test)
        listofSN = list()
        for SN in DATA['teststep'][test]["SN"]:
            recordSN = True
            if "SKIPPrefix" in inputconfig:
                if inputconfig["SKIPPrefix"] in SN:
                    recordSN = False
            if "STARTCountSN" in inputconfig:
                if SN < inputconfig["STARTCountSN"]:
                    recordSN = False
            if recordSN:
                if SN in workingonSNlist:
                    listofSN.append(SN)
                    if not SN in alllistofSN:
                        alllistofSN.append(SN)    

        wirtedata.append(len(listofSN))
        summarydata[test]["TOTAL"] = len(listofSN)
        wirtedata.append(len(DATA['teststep'][test]["DETAILTESTSTEP"]))
        summarydata[test]["TOTALTestStep"] = len(DATA['teststep'][test]["DETAILTESTSTEP"])
        countResult = dict()
        countResult["PASS"] = 0
        for sn in listofSN:
            if "result" in DATA['SN']['FIRST'][test][sn]:
                testresult = DATA['SN']['FIRST'][test][sn]["result"]
                if not testresult in countResult:
                    countResult[testresult] = 1
                else:
                    countResult[testresult] += 1
        for testresult in resultlist:
            if testresult in countResult:
                wirtedata.append(countResult[testresult])
                summarydata[test]["FIRST{}".format(testresult)] = countResult[testresult]
            else:
                 wirtedata.append(0)
                 summarydata[test]["FIRST{}".format(testresult)] = 0
        if len(listofSN):
            firstyeild = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
            summarydata[test]["FIRSTYEILD"] = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
        else:
            firstyeild = "{:.2f}%".format(0)
            summarydata[test]["FIRSTYEILD"] = "{:.2f}%".format(0)
        wirtedata.append(firstyeild)
        countResult = dict()
        countResult["PASS"] = 0
        for sn in listofSN:
            testresult = DATA['SN']['LAST'][test][sn]["result"]
            if not testresult in countResult:
                countResult[testresult] = 1
            else:
                countResult[testresult] += 1
        for testresult in resultlist:
            if testresult in countResult:
                wirtedata.append(countResult[testresult])
                summarydata[test]["LAST{}".format(testresult)] = countResult[testresult]
            else:
                 wirtedata.append(0)
                 summarydata[test]["LASTYEILD"] = "{:.2f}%".format(0)
        if len(listofSN):
            firstyeild = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
            summarydata[test]["LASTYEILD"] = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
        else:
            firstyeild = "{:.2f}%".format(0)
            summarydata[test]["LASTYEILD"] = "{:.2f}%".format(0)
        wirtedata.append(firstyeild)
        wirtedata.append('')
        if test in LastfailureremoveDupSN:
            wirtedata.append(LastfailureremoveDupSN[test]["count"])
        else:
            wirtedata.append(0)

        if pr['FailureSNlist']:
            wirtedata.append('')
            if test in pr['FailureSNlist']['DATA']:
                for flexflowstatus in pr['FailureSNlist']['DATA'][test]["FlexflowType"]:
                    wirtedata.append("{} <{}>".format(flexflowstatus,len(pr['FailureSNlist']['DATA'][test]["FlexflowType"][flexflowstatus])))
        ws1.append(wirtedata)

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    ws1.append(wirtedata)

    overalldata = dict()
    overalldata['TOP'] = summarydata

    overalldata['WEEKFIRST'] = testyeildbyworkweek(DATA, ws1, alllistofSN, resultlist, status='FIRST')
    overalldata['WEEKLAST'] = testyeildbyworkweek(DATA, ws1, alllistofSN, resultlist, status='LAST')
    overalldata['WEEKTIME'] = testtimebyworkweek(DATA, ws1, alllistofSN, WeekTimedata)

    testyeildbySNperfixworkweek(DATA, ws1, alllistofSN, resultlist, status='FIRST')
    testyeildbySNperfixworkweek(DATA, ws1, alllistofSN, resultlist, status='LAST')

    fixcolumnssize(ws1)
    converttoPERCENTAGEnumber(ws1)
    freezePosition(ws1, "B1")

    return overalldata

def findworkweek(checktime):
    checktime_object = datetime.strptime(checktime, '%Y-%m-%d_%H-%M-%S')
    #print(checktime)
    #print(checktime_object.isocalendar()[:2])
    #hereweekcode = week_from_date(checktime_object)
    hereweekcode = checktime_object.isocalendar()[:2]
    recordweekcode = "{}wk{:02d}".format(hereweekcode[0],hereweekcode[1])
    return recordweekcode

def findbegintesttime(checktime,testime):
    checktime_object = datetime.strptime(checktime, '%Y-%m-%d_%H-%M-%S')
    begintime_object = checktime_object - timedelta(seconds=float(testime))
    #print(checktime_object)
    #print(begintime_object)
    begintimestr = begintime_object.strftime("%Y-%m-%d_%H-%M-%S")
    #print(begintimestr)
    return begintimestr

def testtimebyworkweek(DATA, ws1, alllistofSN, WeekTimedata):
    myweeknumber = dict()
    myweeknumber['LIST'] = list()

    for teststep in WeekTimedata:
        for weeknumber in WeekTimedata[teststep]["WEEKLIST"]:
            if not weeknumber in myweeknumber['LIST']:
                myweeknumber['LIST'].append(weeknumber)
                myweeknumber['LIST'].sort()
    WeekTimedata['OVERALLWEEKLIST'] = myweeknumber['LIST']
    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    wirtedata.append('TEST TIME FOR MAX CARD IN MTP BY WEEKCODE')

    ws1.append(wirtedata)
    wirtedata = list()
    wirtedata.append('TEST')
    for weekcode in myweeknumber['LIST']:
        wirtedata.append(weekcode)
    ws1.append(wirtedata)
    for test in DATA['SN']['TEST']:
        wirtedataslot = list()
        wirtedatatestunits = list()
        wirtedataMAX = list()
        wirtedataAVG = list()
        wirtedataslot.append("{}_UNIT_in_MTP".format(test))
        wirtedatatestunits.append("{}_Units".format(test))
        wirtedataMAX.append("{}_MAX_TIME".format(test))
        wirtedataAVG.append("{}_AVG_TIME".format(test))
        for weekcode in myweeknumber['LIST']:
            if weekcode in WeekTimedata[test]["DATA"]:
                MaxSlotinData = WeekTimedata[test]["DATA"][weekcode]["numberofSLOTlist"][-1]
                TestunitinMaxSlotData = len(WeekTimedata[test]["DATA"][weekcode]["DATAofNumber"][MaxSlotinData])
                MaxTimeinData = max(WeekTimedata[test]["DATA"][weekcode]["DATAofNumber"][MaxSlotinData])
                from statistics import mean
                AveTimeinData = mean(WeekTimedata[test]["DATA"][weekcode]["DATAofNumber"][MaxSlotinData])
                wirtedataslot.append(MaxSlotinData)
                wirtedatatestunits.append(TestunitinMaxSlotData)
                wirtedataMAX.append(timedelta(seconds=int(MaxTimeinData)))
                wirtedataAVG.append(timedelta(seconds=int(AveTimeinData)))
                
            else:
                wirtedataslot.append("None")
                wirtedatatestunits.append("None")
                wirtedataMAX.append("None")
                wirtedataAVG.append("None")
        ws1.append(wirtedataslot)
        ws1.append(wirtedatatestunits)
        ws1.append(wirtedataMAX) 
        ws1.append(wirtedataAVG) 

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)

    return WeekTimedata

def testyeildbyworkweek(DATA, ws1, alllistofSN, resultlist, status='FIRST'):
    myweeknumber = dict()
    myweeknumber['LIST'] = list()
    myweeknumber['DETAIL'] = dict()
    myweeknumber['DATA'] = dict()

    for sn in alllistofSN:
        for test in DATA['SN']['TEST']:
            if not test in myweeknumber['DETAIL']:
                myweeknumber['DETAIL'][test] = dict()
            if sn in DATA['SN'][status][test]:
                checktime = DATA['SN'][status][test][sn]["checktime"]
                recordweekcode = findworkweek(checktime)
                #print(recordweekcode)
                if not recordweekcode in myweeknumber['LIST']:
                    myweeknumber['LIST'].append(recordweekcode)
                    myweeknumber['LIST'].sort()
                if not recordweekcode in myweeknumber['DETAIL'][test]:
                    myweeknumber['DETAIL'][test][recordweekcode] = dict()
                    myweeknumber['DETAIL'][test][recordweekcode]['TOTAL'] = 0
                    myweeknumber['DETAIL'][test][recordweekcode]['PASS'] = 0
                myweeknumber['DETAIL'][test][recordweekcode]['TOTAL'] += 1
                if 'PASS' == DATA['SN'][status][test][sn]["result"]:
                    myweeknumber['DETAIL'][test][recordweekcode]['PASS'] += 1

    #print(json.dumps(myweeknumber, indent = 4))

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    if status == 'FIRST':
        wirtedata.append('FIRST PASS YIELD BY WEEKCODE')
    else:
        wirtedata.append('LAST YIELD BY WEEKCODE')
    ws1.append(wirtedata)
    wirtedata = list()
    wirtedata.append('TEST')
    testendtoendyeildlist = dict()
    wirtedataendtoendyeild = list()
    wirtedataendtoendyeild.append("TEST_ENDtoEND_YIELD")
    #print(myweeknumber['LIST'])
    for weekcode in myweeknumber['LIST']:
        #wirtedata.append("{}_PASS".format(SNPerfix))
        #wirtedata.append("{}_FAIL".format(SNPerfix))
        #wirtedata.append("{}_%".format(SNPerfix))
        wirtedata.append(weekcode)
    ws1.append(wirtedata)
    for test in DATA['SN']['TEST']:
        wirtedata = list()
        wirtedatapass = list()
        wirtedatayield = list()
        #wirtedata.append(test)
        wirtedata.append("{}_TOTAL".format(test))
        wirtedatapass.append("{}_PASS".format(test))
        wirtedatayield.append("{}_YIELD".format(test))
        for weekcode in myweeknumber['LIST']:
            if not weekcode in myweeknumber['DATA']:
                myweeknumber['DATA'][weekcode] = dict()
            if not test in myweeknumber['DATA'][weekcode]:
                myweeknumber['DATA'][weekcode][test] = dict()
            if not weekcode in testendtoendyeildlist:
                testendtoendyeildlist[weekcode] = list()
            if weekcode in myweeknumber['DETAIL'][test]:
                wirtedata.append(myweeknumber['DETAIL'][test][weekcode]['TOTAL'])
                myweeknumber['DATA'][weekcode][test]['TOTAL'] = myweeknumber['DETAIL'][test][weekcode]['TOTAL']
                if 'PASS' in myweeknumber['DETAIL'][test][weekcode]:
                    wirtedatapass.append(myweeknumber['DETAIL'][test][weekcode]['PASS'])
                    myweeknumber['DATA'][weekcode][test]['PASS'] = myweeknumber['DETAIL'][test][weekcode]['PASS']
                    firstyield = "{:.2f}%".format(myweeknumber['DETAIL'][test][weekcode]['PASS']/myweeknumber['DETAIL'][test][weekcode]['TOTAL'] * 100)
                    firstyieldsave = myweeknumber['DETAIL'][test][weekcode]['PASS']/myweeknumber['DETAIL'][test][weekcode]['TOTAL']
                else:
                    wirtedatapass.append(0)
                    firstyield = "{:.2f}%".format(0)
                    firstyieldsave = 0                    
                wirtedatayield.append(firstyield)
                myweeknumber['DATA'][weekcode][test]['YIELD'] = firstyield
                if myweeknumber['DETAIL'][test][weekcode]['TOTAL']:
                    testendtoendyeildlist[weekcode].append(firstyieldsave)
                else:
                    testendtoendyeildlist[weekcode].append(1)
            else:
                wirtedata.append(0)
                wirtedatapass.append(0)
                wirtedatayield.append("0%")  
                testendtoendyeildlist[weekcode].append(1)
        ws1.append(wirtedata)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatayield) 

    wirtedata = list()
    ws1.append(wirtedata)    
    for weekcode in myweeknumber['LIST']:
        endtoendyeild = numpy.prod(testendtoendyeildlist[weekcode])
        endtoendyeilddisplay = "{:.2f}%".format(endtoendyeild * 100)
        wirtedataendtoendyeild.append(endtoendyeilddisplay)    
    ws1.append(wirtedataendtoendyeild)
    myweeknumber['DATA']['testendtoendyeildlist'] = testendtoendyeildlist
    myweeknumber['DATA']['wirtedataendtoendyeild'] = wirtedataendtoendyeild

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)

    return myweeknumber

def testyeildbySNperfixworkweek(DATA, ws1, alllistofSN, resultlist, status='FIRST'):
    checkeachwkcodeSNlist = dict()
    checkeachwkcodeSNlist['LIST'] = list()
    checkeachwkcodeSNlist['DETAIL'] = dict()
    for sn in alllistofSN:
        if not sn[:7] in checkeachwkcodeSNlist['LIST']:
            checkeachwkcodeSNlist['LIST'].append(sn[:7])
            checkeachwkcodeSNlist['LIST'].sort()
            checkeachwkcodeSNlist['DETAIL'][sn[:7]] = list()

        if not sn in checkeachwkcodeSNlist['DETAIL'][sn[:7]]:
            checkeachwkcodeSNlist['DETAIL'][sn[:7]].append(sn)
    wirtedata = list()
    if status == 'FIRST':
        wirtedata.append('FIRST PASS YIELD BY SN WEEKCODE')
    else:
        wirtedata.append('LAST YIELD BY SN WEEKCODE')
    ws1.append(wirtedata)
    wirtedata = list()
    wirtedata.append('TEST')
    print(checkeachwkcodeSNlist['LIST'])
    for SNPerfix in checkeachwkcodeSNlist['LIST']:
        #wirtedata.append("{}_PASS".format(SNPerfix))
        #wirtedata.append("{}_FAIL".format(SNPerfix))
        #wirtedata.append("{}_%".format(SNPerfix))
        wirtedata.append(SNPerfix)
    ws1.append(wirtedata)
    testendtoendyeildlist = dict()
    wirtedataendtoendyeild = list()
    wirtedataendtoendyeild.append("TEST_ENDtoEND_YIELD")
    for test in DATA['SN']['TEST']:
        wirtedata = list()
        wirtedatapass = list()
        wirtedatayield = list()
        #wirtedata.append(test)
        wirtedata.append("{}_TOTAL".format(test))
        wirtedatapass.append("{}_PASS".format(test))
        wirtedatayield.append("{}_YIELD".format(test))
        for SNPerfix in checkeachwkcodeSNlist['LIST']:
            if not SNPerfix in testendtoendyeildlist:
                testendtoendyeildlist[SNPerfix] = list()
            countResult = dict()
            for sn in checkeachwkcodeSNlist['DETAIL'][SNPerfix]:
                if sn in DATA['SN'][status][test]:
                    testresult = DATA['SN'][status][test][sn]["result"]
                    if not testresult in countResult:
                        countResult[testresult] = 1
                    else:
                        countResult[testresult] += 1
            total = 0
            for testresult in resultlist:
                if testresult in countResult:
                    #wirtedata.append(countResult[testresult])
                    total += countResult[testresult]
            print(countResult)
            if 'PASS' in countResult:
                wirtedatapass.append(countResult["PASS"])
                firstyieldsave = countResult["PASS"]/total
                firstyield = "{:.2f}%".format(countResult["PASS"]/total * 100)
            else:
                wirtedatapass.append(0)
                firstyieldsave = 0
                firstyield = "{:.2f}%".format(0)
            testendtoendyeildlist[SNPerfix].append(firstyieldsave)
            wirtedatayield.append(firstyield)
            wirtedata.append(total)

        ws1.append(wirtedata)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatayield)
    for SNPerfix in checkeachwkcodeSNlist['LIST']:
        endtoendyeild = numpy.prod(testendtoendyeildlist[SNPerfix])
        endtoendyeilddisplay = "{:.2f}%".format(endtoendyeild * 100)
        wirtedataendtoendyeild.append(endtoendyeilddisplay)    
    ws1.append(wirtedataendtoendyeild)

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)


def week_from_date(date_object):
    date_ordinal = date_object.toordinal()
    year = date_object.year
    week = ((date_ordinal - _week1_start_ordinal(year)) // 7) + 1
    if week >= 52:
        if date_ordinal >= _week1_start_ordinal(year + 1):
            year += 1
            week = 1
    return year, week

def _week1_start_ordinal(year):
    from datetime import date
    jan1 = date(year, 1, 1)
    jan1_ordinal = jan1.toordinal()
    jan1_weekday = jan1.weekday()
    week1_start_ordinal = jan1_ordinal - ((jan1_weekday + 1) % 7)
    return week1_start_ordinal

def getallfilebyfolder(rootdir,keyword=None,level=None):

    listofdirs = getallfolder(rootdir,keyword=keyword,level=level)

    listoffiles = list()

    for eachdir in listofdirs:
        somelists = glob.glob(eachdir + '/*')
        for something in somelists:
            if os.path.isfile(something):  
                basename = os.path.basename(something)
                if keyword:
                    if not keyword.upper() in basename.upper():
                        continue
                if not something in listoffiles:
                    listoffiles.append(something)

    listoffiles.sort()

    #print(json.dumps(listoffiles, indent = 4))

    return listoffiles

def getallfolder(rootdir,keyword=None,level=None):

    listofdirs = list()

    listofdirs.append(rootdir)

    stillhavedir = True

    count = 1
    #print(level)
    #sys.exit()

    while stillhavedir:
        
        if level:
            #print(json.dumps(listofdirs, indent = 4))
            if count >= level:
                break
        stillhavedir = False
        resultlist = list()         

        for checkdir in listofdirs:
            somelists = glob.glob(checkdir + '/*')
            #print("COUNT: {} DIR: {} STILLHAVE: {}".format(count,checkdir,stillhavedir))
            for something in somelists:
                if os.path.isdir(something):
                    if not something in listofdirs:
                        stillhavedir = True
                        resultlist.append(something)
        #print("COUNT: {} TOTAL resultlist: {} STILLHAVE: {}".format(count,len(resultlist), stillhavedir))
        for eachdir in resultlist:
            if not eachdir in listofdirs:
                listofdirs.append(eachdir)

        #print("COUNT: {} TOTAL listofdirs: {} STILLHAVE: {}".format(count,len(listofdirs), stillhavedir))
        count += 1


    #print("LEVEL: {}".format(level))

    return listofdirs

def _removeIllegalCharacterError(s):
    s = re.sub('[^A-Za-z0-9 =\[\]!\(\)_,-:\.]+', '', s)
    return s    

def _removeNonAscii(s): return "".join(i for i in s if ord(i)<128)

def _removespace(s): return "".join(i for i in s if ord(i)!=32)

def _onlykeepword(s): return "".join(i for i in s if (ord(i)>48 and ord(i)<57) or (ord(i)>97 and ord(i)<122) or (ord(i)>65 and ord(i)<90))

def _remove47(s): return "".join(i for i in s if ord(i)!=47 and ord(i)!=43 and ord(i)!=58 and ord(i)!=33)

def get_lock(process_name):
    # Without holding a reference to our socket somewhere it gets garbage
    # collected when the function exits
    get_lock._lock_socket = socket.socket(socket.AF_UNIX, socket.SOCK_DGRAM)

    try:
        # The null byte (\0) means the socket is created 
        # in the abstract namespace instead of being created 
        # on the file system itself.
        # Works only in Linux
        get_lock._lock_socket.bind('\0' + process_name)
        print('{} got the lock'.format(process_name))
    except socket.error:
        print('{} lock exists, another jobs running...'.format(process_name))
        sys.exit()

if __name__ == "__main__":
    sys.exit(main())