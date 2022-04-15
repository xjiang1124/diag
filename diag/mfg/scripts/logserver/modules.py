#!/usr/bin/python3

import warnings
import openpyxl
import os
import sys
import time
import glob
import os.path
from os import path
import datetime
import re
import json
import mpu.io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
from openpyxl.styles import numbers
from openpyxl.chart import BarChart, Reference, Series, LineChart
from openpyxl.chart.label import DataLabelList
import timeit
from datetime import datetime
import dateutil.relativedelta
from statistics import mean 
import statistics
import pandas as pd
#import mysql_202009_db
import math
import csv
import operator
from collections import OrderedDict
#import opsfs_mail
#import modules
import socket
#import script_logging
import gzip
import shutil
import yaml

from datetime import date

today = date.today()
# dd/mm/YY
todayday = today.strftime("%Y/%m/%d")
print("todayday =", todayday)

from datetime import datetime

now = datetime.now() # current date and time
date_time = now.strftime("%Y%m%d_%H%M%S")
print("date and time:",date_time)

class modules(object):
	def __init__(self, logfile=None):
		self.debug = True
		self.logfile = None
		if logfile:
			self.logfile = logfile
		self.debug_print("Start modules...")
		self.chamber_config = self.get_chamber_config()
		#self.getchamberlistorderbyfloor("Chamber 1")
		#self.print_anyinformation(self.chamber_config)
		#print(self.where_is_my_chamber('MTP-637'))
		#sys.exit()

	def getchamberlistorderbyfloor(self,chamberName):
		#self.debug_print("where_is_my_chamber")
		floorlist = list()
		floordict = dict()
		if chamberName in self.chamber_config:
			for MTP in self.chamber_config[chamberName]:
				floorname = self.which_floor_is_for_my_mtp(MTP)
				if not floorname in floorlist:
					floorlist.append(floorname)
					floorlist.sort(reverse=True)
				if not floorname in floordict:
					floordict[floorname] = list()
				if not MTP in floordict[floorname]:
					floordict[floorname].append(MTP)

			returndict = dict()
			returndict['LIST'] = floorlist
			returndict['DATA'] = floordict
			#self.print_anyinformation(returndict)
			#sys.exit()
		else:
			returndict = dict()
			returndict['LIST'] = floorlist
			returndict['DATA'] = floordict			
		return returndict

	def where_is_my_chamber(self,MTP):
		#self.debug_print("where_is_my_chamber")
		for chamberName in self.chamber_config:
			if chamberName.upper() == "FLOOR":
				continue
			if MTP in self.chamber_config[chamberName]:
				return chamberName
		if not "NoChamberinfo" in self.chamber_config:
			self.chamber_config["NoChamberinfo"] = list()
		if not MTP in self.chamber_config["NoChamberinfo"]:
			self.chamber_config["NoChamberinfo"].append(MTP)
		
		return None

	def which_floor_is_for_my_mtp(self,MTP):
		#self.debug_print("where_is_my_chamber")
		for floorName in self.chamber_config["FLOOR"]:
			if MTP in self.chamber_config["FLOOR"][floorName]:
				return floorName
		return None

	def get_chamber_config(self):
		chamberFile = 'Penang_Chamber_Config.xlsx'
		chamber_config = dict()
		from openpyxl import load_workbook
		#self.debug_print("FILE: {}".format(chamberFile))
		wb = load_workbook(filename = chamberFile)
		for tabname in wb.sheetnames:
			#self.debug_print("tabname: {}".format(tabname))
			eachsheet = wb[tabname]
			headers = [c.value for c in next(eachsheet.iter_rows(min_row=1, max_row=1))]
			#print(headers)
			for chambername in headers:
				if chambername.upper() == "FLOOR":
					continue
				if not chambername in chamber_config:
					chamber_config[chambername] = list()
				chamberraw, chambercolunm = self.findposition(eachsheet,chambername)

				#self.debug_print("{}: Raw {} Colunm {} ({})".format(chambername,chamberraw, chambercolunm, get_column_letter(chambercolunm)))
				colunm = eachsheet[get_column_letter(chambercolunm)]
				for countnumberinrow in range(2,len(colunm)+1):
					MTP = str(eachsheet.cell(row=countnumberinrow, column=chambercolunm).value)
					self.debug_print("MTP: {}".format(MTP))
					if len(MTP) and MTP.upper() != "NONE":
						if not MTP in chamber_config[chambername]:
							chamber_config[chambername].append(MTP)

			floorraw, floorcolunm = self.findposition(eachsheet,"FLOOR")
			colunmnumber = eachsheet[get_column_letter(floorcolunm)]
			rownumber = eachsheet[floorraw]
			#self.debug_print("colunmnumber: {} | rownumber: {} | {} vs {}".format(colunmnumber,rownumber,floorraw,floorcolunm))
			if not "FLOOR" in chamber_config:
				chamber_config["FLOOR"] = dict()
			for countnumberinrow in range(2,len(colunmnumber)+1):
				FLOORNAME = str(eachsheet.cell(row=countnumberinrow, column=floorcolunm).value)
				if not FLOORNAME in chamber_config["FLOOR"]:
					chamber_config["FLOOR"][FLOORNAME] = list()
				for countnumberincolnum in range(2,len(rownumber)+1):
					MTP = str(eachsheet.cell(row=countnumberinrow, column=countnumberincolnum).value)
					self.debug_print("MTP: {}".format(MTP))
					if len(MTP) and MTP.upper() != "NONE":
						if not MTP in chamber_config["FLOOR"][FLOORNAME]:
							chamber_config["FLOOR"][FLOORNAME].append(MTP)					
			#sys.exit()


		#self.print_anyinformation(chamber_config)

		#sys.exit()

		return chamber_config

	def get_snlist_from_xlsx(self,xlsxfilename):
		sn_list = list()
		from openpyxl import load_workbook
		wb = load_workbook(filename = xlsxfilename)
		for tabname in wb.sheetnames:
			#self.debug_print("tabname: {}".format(tabname))
			eachsheet = wb[tabname]

			snraw, sncolunm = self.findposition(eachsheet,'SN')

			colunm = eachsheet[get_column_letter(sncolunm)]
			for countnumberinrow in range(2,len(colunm)+1):
				SN = str(eachsheet.cell(row=countnumberinrow, column=sncolunm).value)
				self.debug_print("SN: {}".format(SN))
				if len(SN) and SN.upper() != "NONE":
					if not SN in sn_list:
						sn_list.append(SN)
						
		return sn_list

	def rsync_and_delete_by_os_system_at_locate(self,source,dest):
		rsync_cmd = "rsync -avrz {} {}".format(source,dest)
		os.system(rsync_cmd)
		try:
			shutil.rmtree(source)
		except:
			return False
		return None

	def copy_file(self,sourcefile,target):
		from shutil import copyfile
		from sys import exit
		try:
		    copyfile(sourcefile, target)
		except IOError as e:
		    self.debug_print("Unable to copy file. %s" % e)
		    return False
		except:
		    self.debug_print("Unexpected error:", sys.exc_info())
		    return False

		self.debug_print("\nFile copy done!\n")
		return True

	def convert_unit(self, size_in_bytes, unit):
	   """ Convert the size from bytes to other units like KB, MB or GB"""
	   if unit.upper() == 'KB':
	       return size_in_bytes/1024
	   elif unit.upper() == 'MB':
	       return size_in_bytes/(1024*1024)
	   elif unit.upper() == 'GB':
	       return size_in_bytes/(1024*1024*1024)
	   else:
	       return size_in_bytes

	def findfilecreatedate(self,checkfile):
		import os.path, time 
		return time.ctime(os.path.getmtime(checkfile))

	def get_file_size(self, file_name, size_type = 'MB' ):
	   """ Get file in size in given unit like KB, MB or GB"""
	   size = os.path.getsize(file_name)
	   return self.convert_unit(size, size_type)

	def getmd5sumoffile(self,checkfile):
		start=datetime.now()
		import hashlib
		if os.path.exists(checkfile):
			md5sumoffile = hashlib.md5(open(checkfile,'rb').read()).hexdigest()
			filesizeinMB = self.get_file_size(checkfile)
			filecreatedate = self.findfilecreatedate(checkfile)
			difftime = datetime.now()-start
			self.debug_print("FILE: {} SIZE: {} MB Create: {} ==> MD5SUM: {} RUNNING TIME: {} seconds".format(checkfile,round(filesizeinMB,3),filecreatedate,md5sumoffile,difftime.total_seconds()))
			return md5sumoffile
		else:
			return None

	def readjsonfile(self,jsonfile):
		start=datetime.now()
		outputdict = dict()

		if os.path.exists(jsonfile):
			outputdict = mpu.io.read(jsonfile)

		difftime = datetime.now()-start
		self.debug_print("readjsonfile: {} use {} seconds".format(jsonfile,difftime.total_seconds()))

		return outputdict

	def wirtejsonfile(self,jsonfile,DATA):
		start=datetime.now()
		outputdict = dict()

		jsonbasedir = os.path.dirname(jsonfile)
		self.createdirinserver(jsonbasedir)
		mpu.io.write(jsonfile, DATA)
		difftime = datetime.now()-start
		self.debug_print("wirtejsonfile: {} use {} seconds".format(jsonfile,difftime.total_seconds()))

		return outputdict

	def readyamltodict(self, yamlfile):
		with open(yamlfile) as file:
		    # The FullLoader parameter handles the conversion from YAML
		    # scalar values to Python the dictionary format
		    yamldict = yaml.load(file, Loader=yaml.FullLoader)

		    #self.print_anyinformation(yamldict)

		return yamldict

	def wirtedicttoyaml(self, yamldict, yamlfile):
		with open(yamlfile, 'w') as file:
		    documents = yaml.dump(yamldict, file)

		    #self.print_anyinformation(documents)
		return None

	def checkDirectoryexist(self, checkDir):
	    for DirName in checkDir.keys():
	        #print(DirName)
	        #print(checkDir[DirName])
	        #print(os.path.isdir(checkDir[DirName]))
	        if "testlogpath" in DirName:
	            continue
	        elif "imagedir" in DirName:
	            continue
	        self.createdirinserver(checkDir[DirName])

	    return None

	def createdirinserver(self, createdir):
	    mode = 0o777
	    if not os.path.isdir(createdir):
	        #os.mkdir(createdir, mode)
	        os.makedirs(createdir, exist_ok=True)
	        self.debug_print("Directory '% s' created" % createdir)

	    return None


	def movereporttohistorydir(self,reportpath,historyreportpath):
	    filelisttowork = self.getallfilebyfolder(reportpath,keyword=None,level=1)
	    foldlisttowork = self.getallfolder(reportpath,level=2)

	    #print(json.dumps(filelisttowork, indent = 4))
	    #print(json.dumps(foldlisttowork, indent = 4))

	    workingfileandfoldermovinglist = list()
	    for eachfolder in foldlisttowork:
	        if eachfolder != reportpath: 
	            if eachfolder != historyreportpath:
	                workingfileandfoldermovinglist.append(eachfolder)

	    for eachfile in filelisttowork:
	        workingfileandfoldermovinglist.append(eachfile)
	    #print(json.dumps(workingfileandfoldermovinglist, indent = 4))
	    try:
	        for eachworkitem in workingfileandfoldermovinglist:
	            if os.path.isdir(eachworkitem):
	                smallitemslist = eachworkitem.split('/')
	                #print(smallitemslist)
	                movetopath = "{}/{}".format(historyreportpath,smallitemslist[-1])
	                #print(movetopath)
	                dest = shutil.move(eachworkitem, movetopath, copy_function = shutil.copytree)
	                #print(dest)
	            else:
	                smallitemslist = eachworkitem.split('/')
	                #print(smallitemslist)
	                movetopath = "{}/{}".format(historyreportpath,smallitemslist[-1])
	                #print(movetopath)
	                dest = shutil.move(eachworkitem, movetopath)
	                #print(dest)    
	    except:
	        print("Something else went wrong")        
	    #sys.exit()
	    return None

	def getallfilebyfolder(self,rootdir,keyword=None,level=None):
	    start=datetime.now()
	    listofdirs = self.getallfolder(rootdir,keyword=keyword,level=level)

	    listoffiles = list()

	    for eachdir in listofdirs:
	        somelists = glob.glob(eachdir + '/*')
	        for something in somelists:
	            if os.path.isfile(something):  
	                basename = os.path.basename(something)
	                if keyword:
	                    if not keyword.upper() in basename.upper():
	                        continue
	                if not something in listoffiles:
	                    listoffiles.append(something)

	    listoffiles.sort()

	    #print(json.dumps(listoffiles, indent = 4))
	    difftime = datetime.now()-start
	    #self.debug_print("DIR: {} have {} files and running {} seconds".format(rootdir,len(listoffiles),difftime.total_seconds()))
	    return listoffiles

	def getallfolder(self,rootdir,keyword=None,level=None):
		start=datetime.now()

		listofdirs = list()

		listofdirs.append(rootdir)

		stillhavedir = True

		count = 1
		#print(level)

		if not level:
		    for checkdir in listofdirs:
		        somelists = glob.glob(checkdir + '/*')
		        for something in somelists:
		            if os.path.isdir(something):
		                if not something in listofdirs:
		                    stillhavedir = True
		                    listofdirs.append(something)
		else:
		    while stillhavedir:
		        
		        if level:
		            #print(json.dumps(listofdirs, indent = 4))
		            if count >= level:
		                break
		        stillhavedir = False
		        resultlist = list()         

		        for checkdir in listofdirs:
		            somelists = glob.glob(checkdir + '/*')
		            #print("COUNT: {} DIR: {} STILLHAVE: {}".format(count,checkdir,stillhavedir))
		            for something in somelists:
		                if os.path.isdir(something):
		                    if not something in listofdirs:
		                        stillhavedir = True
		                        resultlist.append(something)
		        #print("COUNT: {} TOTAL resultlist: {} STILLHAVE: {}".format(count,len(resultlist), stillhavedir))
		        for eachdir in resultlist:
		            if not eachdir in listofdirs:
		                listofdirs.append(eachdir)

		        #print("COUNT: {} TOTAL listofdirs: {} STILLHAVE: {}".format(count,len(listofdirs), stillhavedir))
		        count += 1

		#print("LEVEL: {}".format(level))
		#self.print_anyinformation(listofdirs)
		difftime = datetime.now()-start
		self.debug_print("DIR: {} have {} folder and running {} seconds".format(rootdir,len(listofdirs),difftime.total_seconds()))

		return listofdirs

	def getallfileinfolderbyoswalk(self,rootdir,keyword=None,level=None):
	    listOfFiles = list()
	    for (dirpath, dirnames, filenames) in os.walk(rootdir):
	        listOfFiles += [os.path.join(dirpath, file) for file in filenames]

	    listOfFiles.sort()

	    #print(json.dumps(listOfFiles, indent = 4 ))

	    return listOfFiles

	def fixcolumnssize(self,ws):
		dims = {}
		for row in ws.rows:
			for cell in row:
				if cell.value:
					dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
		for col, value in dims.items():
			usevalue = value
			if value > 100:
				usevalue = 100
			ws.column_dimensions[col].width = usevalue + 2

		ws.auto_filter.ref = ws.dimensions
		return None

	def fixcolumnssize3(self,ws,enablefilter=True):
	    dims = {}
	    for row in ws.rows:
	        for cell in row:
	            if cell.value:
	                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
	    for col, value in dims.items():
	        if value > 30:
	            value = 30
	        ws.column_dimensions[col].width = value + 2
	    if enablefilter:
	        ws.auto_filter.ref = ws.dimensions
	def highlightinlightredrow(self,ws,keyword):
	    from openpyxl.styles import Color, PatternFill, Font, Border
	    maxRow = ws.max_row
	    maxCol = ws.max_column
	    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	    for rowNum in range(1, maxRow + 1):
	        fillcolor = 0
	        for colNum in range(1, maxCol + 1):
	            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
	            if keyword in str(ws.cell(row=rowNum, column=colNum).value):
	                fillcolor = 1
	            if fillcolor:
	                ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type = 'solid')   

	def highlightinlightgreenrow(self,ws,keyword):
	    from openpyxl.styles import Color, PatternFill, Font, Border
	    maxRow = ws.max_row
	    maxCol = ws.max_column
	    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	    for rowNum in range(1, maxRow + 1):
	        fillcolor = 0
	        for colNum in range(1, maxCol + 1):
	            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
	            if keyword in str(ws.cell(row=rowNum, column=colNum).value):
	                fillcolor = 1
	            if fillcolor:
	                ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='66FF00', end_color='66FF00', fill_type = 'solid')   

	def freezePosition(self,ws,keyword):
	    ws.freeze_panes = keyword

	def converttoPERCENTAGEnumberbyFailurerange(self,ws):
	    from openpyxl.styles import Color, PatternFill, Font, Border
	    maxRow = ws.max_row
	    maxCol = ws.max_column
	    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	    for rowNum in range(1, maxRow + 1):
	        for colNum in range(1, maxCol + 1):
	            checkvalue = str(ws.cell(row=rowNum, column=colNum).value)
	            if '%' in checkvalue:
	                checkvalue = float(checkvalue[:-1])/100
	                ws.cell(row=rowNum, column=colNum).value = checkvalue
	                ws.cell(row=rowNum, column=colNum).number_format = '0.00%'
	                if checkvalue < 0.005:
	                    #Green
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type = 'solid')
	                elif checkvalue < 0.01:
	                    #Sap Green
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='66FF00', end_color='66FF00', fill_type = 'solid')
	                elif checkvalue < 0.02:
	                    #Antique Bronze
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='BBFF00', end_color='BBFF00', fill_type = 'solid')
	                elif checkvalue < 0.03:
	                    #Chestnut
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type = 'solid')
	                elif checkvalue < 0.04:
	                    #Sweet Brown
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFAA00', end_color='FFAA00', fill_type = 'solid')
	                elif checkvalue < 0.05:
	                    #Cardinal
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF5500', end_color='FF5500', fill_type = 'solid')
	                else:
	                    #Red
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')

	def converttoPERCENTAGEnumber(self,ws):
	    from openpyxl.styles import Color, PatternFill, Font, Border
	    maxRow = ws.max_row
	    maxCol = ws.max_column
	    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	    for rowNum in range(1, maxRow + 1):
	        for colNum in range(1, maxCol + 1):
	            checkvalue = str(ws.cell(row=rowNum, column=colNum).value)
	            if '%' in checkvalue:
	                checkvalue = float(checkvalue[:-1])/100
	                ws.cell(row=rowNum, column=colNum).value = checkvalue
	                ws.cell(row=rowNum, column=colNum).number_format = '0.00%'
	                if checkvalue > 0.95:
	                    #Green
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='66f86a', end_color='66f86a', fill_type = 'solid')
	                elif checkvalue < 0.1:
	                    #Red
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
	                elif checkvalue > 0.75:
	                    #Yellow
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
	                elif checkvalue < 0.25:
	                    #Pink
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type = 'solid')
	                elif checkvalue > 0.50:
	                    #Arylide Yellow
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='E9D66B', end_color='E9D66B', fill_type = 'solid')
	                else:
	                    #Orange
	                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type = 'solid')

	def highlightinyellownumber(self,ws):
		from openpyxl.styles import Color, PatternFill, Font, Border
		maxRow = ws.max_row
		maxCol = ws.max_column
		#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
		for rowNum in range(1, maxRow + 1):
			fillcolor = 0
			for colNum in range(1, maxCol + 1):
				#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
				if str(ws.cell(row=rowNum, column=colNum).value).isnumeric():
					if int(ws.cell(row=rowNum, column=colNum).value) > 0:
						if colNum < 3:
							continue
						ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
		return 0

	def readMACvsOLDSNvsNEWSNxlsfile(self,xlsxfile,MACvsOLDSNvsNEWSN):
		from openpyxl import load_workbook
		self.debug_print("FILE: {}".format(xlsxfile))
		wb_sn = load_workbook(filename = xlsxfile)
		for tabname in wb_sn.sheetnames:
			self.debug_print("tabname: {}".format(tabname))
			eachsheet = wb_sn[tabname]
			OldSNraw, OldSNcolunm = self.findposition(eachsheet,'OldSN')
			NewSNraw, NewSNcolunm = self.findposition(eachsheet,'NewSN')
			MacAddraw, MacAddcolunm = self.findposition(eachsheet,'MacAdd')

			self.debug_print("OldSN: Raw {} Colunm {} | NewSN: Raw {} Colunm {} | MacAdd: Raw {} Colunm {}".format(OldSNraw, OldSNcolunm, NewSNraw, NewSNcolunm,MacAddraw,MacAddcolunm))
			#sys.exit()
			if OldSNraw and NewSNraw and MacAddraw:
				colunm = eachsheet['A']
				for countnumberinrow in range(2,len(colunm)+1):
					MAC = str(eachsheet.cell(row=countnumberinrow, column=MacAddcolunm).value)
					OLDSN = str(eachsheet.cell(row=countnumberinrow, column=OldSNcolunm).value)
					NEWSN = str(eachsheet.cell(row=countnumberinrow, column=NewSNcolunm).value)
					self.debug_print("MAC: {} | OLDSN: {} | NEWSN: {}".format(MAC, OLDSN, NEWSN))
					MACvsOLDSNvsNEWSN[MAC] = dict()
					MACvsOLDSNvsNEWSN[MAC]["OldSN"] = OLDSN
					MACvsOLDSNvsNEWSN[MAC]["NewSN"] = NEWSN

		return None

	def calculePass_rate(self,totalnumber,passnumber):

	    calcule_yeild = 1
	    if totalnumber:
	        calcule_yeild = passnumber / totalnumber

	    yeilddisplay = "{:.2f}%".format(calcule_yeild * 100)  

	    return yeilddisplay

	def calculeFail_rate(self,totalnumber,failnumber):

	    calcule_yeild = 0
	    if totalnumber:
	        calcule_yeild = failnumber / totalnumber

	    yeilddisplay = "{:.2f}%".format(calcule_yeild * 100)  

	    return yeilddisplay
	    
	def highlightingreen(self,ws,keyword):
		from openpyxl.styles import Color, PatternFill, Font, Border
		maxRow = ws.max_row
		maxCol = ws.max_column
		#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
		for rowNum in range(1, maxRow + 1):
			fillcolor = 0
			for colNum in range(1, maxCol + 1):
				#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
				if keyword in str(ws.cell(row=rowNum, column=colNum).value):
					fillcolor = 1
					ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='66f86a', end_color='66f86a', fill_type = 'solid')
	                
	def highlightinred(self,ws,keyword):
		from openpyxl.styles import Color, PatternFill, Font, Border
		maxRow = ws.max_row
		maxCol = ws.max_column
		#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
		for rowNum in range(1, maxRow + 1):
			fillcolor = 0
			for colNum in range(1, maxCol + 1):
				#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
				if keyword in str(ws.cell(row=rowNum, column=colNum).value):
					ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')                

	def highlightinlightredrow(self,ws,keyword):
	    from openpyxl.styles import Color, PatternFill, Font, Border
	    maxRow = ws.max_row
	    maxCol = ws.max_column
	    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	    for rowNum in range(1, maxRow + 1):
	        fillcolor = 0
	        for colNum in range(1, maxCol + 1):
	            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
	            if keyword in str(ws.cell(row=rowNum, column=colNum).value):
	                fillcolor = 1
	            if fillcolor:
	                ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type = 'solid')   

	def getbeforedayinformation(self,checkday=30):
	    from datetime import date, timedelta

	    current_date = date.today().isoformat()   
	    days_before = (date.today()-timedelta(days=checkday)).isoformat()
	    days_after = (date.today()+timedelta(days=checkday)).isoformat()  

	    # print("\nCurrent Date: ",current_date)
	    # print("30 days before current date: ",days_before)
	    # print("30 days after current date : ",days_after)
	    current_date = str(current_date)
	    days_before = str(days_before)
	    days_after = str(days_after)

	    # print("\nCurrent Date: ",current_date)
	    # print("30 days before current date: ",days_before)
	    # print("30 days after current date : ",days_after)

	    return days_before

	def readSNxlsxfile(self,xlsxfile,SNvsPCBA):
		from openpyxl import load_workbook
		self.debug_print("FILE: {}".format(xlsxfile))
		wb_sn = load_workbook(filename = xlsxfile)
		for tabname in wb_sn.sheetnames:
			self.debug_print("tabname: {}".format(tabname))
			eachsheet = wb_sn[tabname]
			SNraw, SNcolunm = self.findposition(eachsheet,'SN')
			PCBAsnraw, PCBAsncolunm = self.findposition(eachsheet,'PCBA_SN')

			self.debug_print("SN: Raw {} Colunm {} | PCBA_SN: Raw {} Colunm {}".format(SNraw, SNcolunm, PCBAsnraw, PCBAsncolunm))

			if SNraw and PCBAsnraw:
				colunm = eachsheet['A']
				for countnumberinrow in range(2,len(colunm)+1):
					SN = str(eachsheet.cell(row=countnumberinrow, column=SNcolunm).value).strip()
					PCBA_SN = str(eachsheet.cell(row=countnumberinrow, column=PCBAsncolunm).value).strip()
					self.debug_print("SN: {} | PCBA_SN: {}".format(SN, PCBA_SN))
					SNvsPCBA[SN] = PCBA_SN		

		return None

	def wraptest(self,ws):
	    for row in ws.iter_rows():
	        for cell in row:
	            cell.alignment = Alignment(wrap_text=True,vertical='top') 

	def findposition(self,sheet,keyword):

		colunm = sheet[1]
		raw = sheet['A']

		for rawnumber in range(1,len(raw)+1):
			for colunmnumber in range(1,len(colunm)+1):
				if keyword == sheet.cell(row=rawnumber, column=colunmnumber).value:
					return rawnumber, colunmnumber
		return None, None

	def gzip_to_file(self, sourcefile):

		self.debug_print("Current {} FileSize : {}".format(sourcefile,os.path.getsize(sourcefile)))
		returnfile = "{}.gz".format(sourcefile)
		with open(sourcefile, 'rb') as f_in:               	# open original file for reading
			with gzip.open(returnfile, 'wb') as f_out: 		# open gz file for writing
				shutil.copyfileobj(f_in, f_out)           	# write/copy text file into gz file

		if os.path.exists(returnfile):
			self.debug_print("ZIPFILE {} FileSize : {}".format(returnfile,os.path.getsize(returnfile)))
			if os.path.exists(sourcefile):
				self.debug_print("Process to remove FILE : {}".format(sourcefile))
				os.remove(sourcefile)
			else:
				self.debug_print("The file {} does not exist".format(sourcefile))

		return returnfile

	def debug_print(self,msg):
		if self.debug:
			
			if self.logfile:
				self.logfile.WriteLine(str(msg))
			else:
				print(msg)
		return None

	def check_is_module_class(self,inputdata):
		m = re.findall(r'\'(.*)\'', str(type(inputdata)))
		if m:
			#print(m)
			mlist = m[0].lower().split('.')
			#print(mlist)
			#print(len(mlist))
			if len(mlist) == 2:
				if mlist[0] == mlist[1]:
					return True
		if 'object at' in str(inputdata):
			return True

		return False

	def print_anyinformation(self,inputdata):
		countlevel=0
		self.debug_print("")
		specenumber = "\t"
		if countlevel:
			for n in range(countlevel):
				specenumber += "\t"
		if isinstance(inputdata, dict):
			self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <DICT>".format(countlevel,specenumber,inputdata,type(inputdata)))
			self.print_dict(inputdata,countlevel+1)
		elif isinstance(inputdata, list):
			self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <LIST>".format(countlevel,specenumber,inputdata,type(inputdata)))
			self.print_list(inputdata,countlevel+1)
		elif self.check_is_module_class(inputdata):
			self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <CLASS>".format(countlevel,specenumber,inputdata,type(inputdata)))
			self.print_dict(inputdata.__dict__,countlevel+1)				
		else:
			self.debug_print("[{}]{}\t{}\t<TYPE:{}>".format(countlevel,specenumber,inputdata,type(inputdata)))		

	def print_dict(self,dictdata,countlevel=0):
		specenumber = "\t"
		if countlevel:
			countlevelspacenumber = countlevel
			if countlevel > 5:
				countlevelspacenumber = 5
			for n in range(countlevelspacenumber):
				specenumber += "\t"

		for x, y in dictdata.items():
			if countlevel > 100:
				self.debug_print("[{}]{}{}\t=\t{}\t<TYPE:{}>".format(countlevel,specenumber,x,y,type(y)))
				return None
			if isinstance(y, dict):
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <DICT>".format(countlevel,specenumber,x,type(x)))
				self.print_dict(y,countlevel+1)
			elif isinstance(y, list):
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <LIST>".format(countlevel,specenumber,x,type(x)))
				self.print_list(y,countlevel+1)
			elif self.check_is_module_class(y):
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <CLASS>".format(countlevel,specenumber,x,type(x)))
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <CLASS>".format(countlevel,specenumber,y,type(y)))
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <CLASS>".format(countlevel,specenumber,y.__class__,type(y.__class__)))
				self.print_dict(y.__dict__,countlevel+1)				
			else:
				self.debug_print("[{}]{}{}\t=\t{}\t<TYPE:{}>".format(countlevel,specenumber,x,y,type(y)))
		return None

	def print_list(self,listdata,countlevel=0):
		specenumber = "\t"
		if countlevel:
			countlevelspacenumber = countlevel
			if countlevel > 5:
				countlevelspacenumber = 5
			for n in range(countlevelspacenumber):
				specenumber += "\t"

		for x in listdata:
			if countlevel > 100:
				self.debug_print("[{}]{}{}\t<TYPE:{}>".format(countlevel,specenumber,x,type(x)))
				return None
			if isinstance(x, list):
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <LIST>".format(countlevel,specenumber,x,type(x)))
				self.print_list(x,countlevel+1)
			elif isinstance(x, dict):
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <DICT>".format(countlevel,specenumber,x,type(x)))
				self.print_dict(x,countlevel+1)
			elif self.check_is_module_class(x):
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <CLASS>".format(countlevel,specenumber,x,type(x)))
				self.debug_print("[{}]{}\"{}\"\t<TYPE:{}> : <CLASS>".format(countlevel,specenumber,x.__class__,type(x.__class__)))
				self.print_dict(x.__dict__,countlevel+1)
			else:
				self.debug_print("[{}]{}{}\t<TYPE:{}>".format(countlevel,specenumber,x,type(x)))

	def print_json(self,printoutput):

		self.debug_print(json.dumps(printoutput, indent = 4))
		return None

	def print_json_str_only(self,printoutput):
		for eachkey in printoutput:
			recordline = '[{}] : {} ({})'.format(eachkey,printoutput[eachkey],type(printoutput[eachkey]))
			self.debug_print(recordline)
		return None	

	def return_json(self,printoutput):
		return json.dumps(printoutput, indent = 4)

	def getbasename(self,path):
		return os.path.basename(path)

	def checkfilelistkeywork(self,filelist,keyword=None):
		newfilelist = list()
		for eachfile in filelist:
			basename = os.path.basename(eachfile)
			if keyword:
				if not keyword.upper() in basename.upper():
					continue
				if not eachfile in newfilelist:
					newfilelist.append(eachfile)
		newfilelist.sort()
		self.filelist = newfilelist

		self.debug_print(json.dumps(self.filelist, indent = 4))

		return newfilelist		

	def checkandcreatedit(self,pathdir):
		from pathlib import Path
		Path(pathdir).mkdir(parents=True, exist_ok=True)

	def _removeNonAscii(self,s): return "".join(i for i in s if ord(i)<128)

	def _keepnumberandsomestamps(self,s): return "".join(i for i in s if ord(i)<65)

	def _removespace(self,s): return "".join(i for i in s if ord(i)!=32)

	def _removeenter(self,s): return "".join(i for i in s if ord(i)!=12)

	def _onlykeepword(self,s): return "".join(i for i in s if (ord(i)>=48 and ord(i)<=57) or (ord(i)>=97 and ord(i)<=122) or (ord(i)>=65 and ord(i)<=90) or ord(i)==32 or ord(i)==45)

	def _remove47(self,s): return "".join(i for i in s if ord(i)!=47 and ord(i)!=43 and ord(i)!=58 and ord(i)!=33)