#!/usr/bin/python3

import shutil
import json
import sys
import glob
import os
import re
import mpu.io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
import timeit
from datetime import datetime
from datetime import timedelta
import pandas as pd
import math
import numpy
from os import listdir
import time
from logdef import KEY_WORD
import modules

now = datetime.now() # current date and time
date_time = now.strftime("%Y%m%d_%H%M%S")
print("date and time:",date_time)

#TEST on KEY_WORD
#print(KEY_WORD.NIC_DIAG_TEST_RSLT_RE)

def main():

    start=datetime.now()
    pr = dict()
    pr['modules'] = modules.modules()

    startdate = checkstartdate()
    ARGV = dict()
    inputconfig = checkconfigfile(ARGV)
    print("FIND startdate input: {}".format(startdate))
    print(json.dumps(ARGV, indent = 4 ))
    #sys.exit()
    print(json.dumps(inputconfig, indent = 4))
    print(inputconfig['FILE']['datebasesjsonfile'])
    pr['modules'].checkDirectoryexist(inputconfig['DIR'])

    datebase_DB = dict()
    if os.path.exists(inputconfig['FILE']['datebasesjsonfile']):
        datebase_DB = mpu.io.read(inputconfig['FILE']['datebasesjsonfile'])    
    DATA = datebase_DB
    if not 'teststep' in DATA:
        DATA['teststep'] = dict()
        DATA['SN'] = dict()
        DATA['SN']['LIST'] = list()
        DATA['SN']['FIRST'] = dict()
        DATA['SN']['LAST'] = dict()
        DATA['SN']['ERROR'] = dict()
        DATA['SN']['DL'] = dict()

    DATA['SN']['TEST'] = inputconfig["TESTLIST"]
    #print(json.dumps(DATA['SN']['TEST'], indent = 4))
    DATA['NEWFILECOUNT'] = 0
    testfolderlist = inputconfig["TESTFOLDER"]
    #print(testfolderlist)

    if "skipfoldercheck" in ARGV:
        pass 
    else:
        for testfolder in testfolderlist:
            # if "FST" in testfolder:
            #     workingoneachtest(inputconfig,DATA,testfolder,redo=True)
            # else:
            #     workingoneachtest(inputconfig,DATA,testfolder)
            #pass
            workingoneachtest(pr,inputconfig,DATA,testfolder)

    workingonLastdata(DATA,inputconfig)
    #sys.exit()

    DATA['SN']['LIST'].sort()
    jsonfileoutputname = "{}/{}_{}_DATA.json".format(inputconfig['DIR']["TEMPDATABASE"],'OVERALL',date_time)
    mpu.io.write(jsonfileoutputname, DATA)
    mpu.io.write(inputconfig['FILE']['datebasesjsonfile'], DATA)

    if DATA['NEWFILECOUNT']:
        processtocreatedailyreport(pr,DATA,inputconfig,date_time,startdate)
    elif "report" in ARGV:
        processtocreatedailyreport(pr,DATA,inputconfig,date_time,startdate)
 
    if startdate:
        createteststatusreport(pr,DATA,inputconfig,startdate=startdate)

    difftime = datetime.now()-start
    print('Done Time: ', difftime)       
    print("How many seconds use?: {} seconds".format(difftime.total_seconds()))       
    return None

def processtocreatedailyreport(pr,DATA,inputconfig,date_time,startdate):

    if "historypath" in inputconfig["DIR"]:
        movereporttohistorydir(inputconfig)

    if "MTP_STATUS" in inputconfig:
        wb3 = Workbook()
        dest_filename3 = "{}MTP_STATUS_{}_DATA.xlsx".format(inputconfig['DIR']["reportpath"],date_time)
        print(dest_filename3)

        generateexeclmtpstatussummaryreport(DATA,wb3,inputconfig)

        for mtp in DATA["MTPCHASSIS"]:
            generateexeclmtpstatusbyeachmtpreport(DATA,wb3,inputconfig,mtp)

        wb3.save(filename = dest_filename3) 

    if "DIE_ID" in inputconfig:
        wb2 = Workbook()
        dest_filename2 = "{}ELBA_DIE_ID_{}_DATA.xlsx".format(inputconfig['DIR']["reportpath"],date_time)
        print(dest_filename2)

        generateexecldieidreport(DATA,wb2,inputconfig,start=startdate)

        wb2.save(filename = dest_filename2) 

    #sys.exit()

    createteststatusreport(pr,DATA,inputconfig)

    startdate = getbeforedayinformation(checkday=31)
    createteststatusreport(pr,DATA,inputconfig,startdate=startdate)

    return None

def getbeforedayinformation(checkday=30):
    from datetime import date, timedelta

    current_date = date.today().isoformat()   
    days_before = (date.today()-timedelta(days=checkday)).isoformat()
    days_after = (date.today()+timedelta(days=checkday)).isoformat()  

    # print("\nCurrent Date: ",current_date)
    # print("30 days before current date: ",days_before)
    # print("30 days after current date : ",days_after)
    current_date = str(current_date)
    days_before = str(days_before)
    days_after = str(days_after)

    # print("\nCurrent Date: ",current_date)
    # print("30 days before current date: ",days_before)
    # print("30 days after current date : ",days_after)

    return days_before

def createteststatusreport(pr,DATA,inputconfig,startdate=None):
    workingonSNlist = DATA['SN']['LIST']
    
    print("START DATE: {}".format(startdate))
    workingonSNlist = getsnlistafteestartdate(DATA,inputconfig,startdate=None)
    workingonSNlist.sort(reverse=True)
    print("COUNT SN: {}".format(len(workingonSNlist)))
    #sys.exit()    
    wb = Workbook()
    dest_filename = "{}EXECL_{}_DATA.xlsx".format(inputconfig['DIR']["reportpath"],date_time)
    filenameheader = "{}EXECL_{}_DATA".format(inputconfig['DIR']["reportpath"],date_time)
    if "NAME" in inputconfig:
        dest_filename = "{}{}_{}_DATA.xlsx".format(inputconfig['DIR']["reportpath"],inputconfig["NAME"],date_time)
        filenameheader = "{}{}_{}_DATA".format(inputconfig['DIR']["reportpath"],inputconfig["NAME"],date_time)
    if startdate:
        dest_filename = "{}_withStartDate_{}.xlsx".format(filenameheader,startdate)
    print('OUTPUT FILE NAME: ' + dest_filename)
    
    generateexeclreport(DATA,wb,inputconfig,workingonSNlist,start=startdate)
    if "NAME" in inputconfig:
        if "ORTANO2" == inputconfig["NAME"].upper():
            SNbyPN = GetSNlistbyPNfromDLtest(workingonSNlist,DATA["teststep"]['DL'],teststep='DL')
            for PN in SNbyPN:
                generateexeclDatabyPN(DATA,wb,inputconfig,SNbyPN[PN],PN,start=None)


    print("START DATE: {}".format(startdate))
    workingonSNlist = getsnlistafteestartdate(DATA,inputconfig,startdate=startdate)
    workingonSNlist.sort(reverse=True)
    print("COUNT SN: {}".format(len(workingonSNlist)))

    if not len(workingonSNlist):
        return None

    #sys.exit()
    
    if 'CM' in inputconfig:
        if inputconfig['CM'] == "FSJ":

            generateexeclsnFSJstatus(DATA, workingonSNlist,'LAST',wb)
        
        wb.save(filename = dest_filename)
        createlastfailurelogfolder(pr,DATA,workingonSNlist,inputconfig)

    else:

        generateexeclsnstatus(DATA, workingonSNlist,'FIRST',wb,inputconfig)
        #generateexeclsnstatus(DATA, workingonSNlist,'LAST',wb,inputconfig)
        #generateexeclsnstatus(DATA, workingonSNlist,'LAST',wb,inputconfig,Withallerror=True)
        generateexeclsnstatusalldata(DATA, workingonSNlist,'LAST',wb,inputconfig)
        generateexeclsnTopFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig)
        generateexeclsnTopFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig)

        generateexeclsn4CFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig)
        generateexeclsn4CFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig,duplicatesn=False)
        generateexeclsn4CFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig)
        generateexeclsn4CFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig,duplicatesn=False)

        generateexeclsn4CFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig,byweek=True)
        generateexeclsn4CFailurestatus(DATA, workingonSNlist,'FIRST',wb,inputconfig,duplicatesn=False,byweek=True)
        generateexeclsn4CFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig,byweek=True)
        generateexeclsn4CFailurestatus(DATA, workingonSNlist,'LAST',wb,inputconfig,duplicatesn=False,byweek=True)
        #generateexeclsn4Cstatus(DATA,'4C-H',wb)
        #generateexeclsn4Cstatus(DATA,'4C-L',wb)
        #
        for eachteststep in DATA['SN']['TEST']:
            generateexecltest(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)
            if "4C" in eachteststep:
                generateexecltestby4Ctesttime(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)
            else:
                generateexecltestbyNon4Ctesttime(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)
            generateexeclerrdata(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,wb,DATA)

        wb.save(filename = dest_filename)

        copyallfailurelogfolder(pr,DATA,workingonSNlist,inputconfig)
    
    

    print("OUTPUT FILE: {}".format(dest_filename))

    return None    
        

def movereporttohistorydir(inputconfig):
    filelisttowork = getallfilebyfolder(inputconfig["DIR"]["reportpath"],keyword=None,level=1)
    foldlisttowork = getallfolder(inputconfig["DIR"]["reportpath"],level=2)

    #print(json.dumps(filelisttowork, indent = 4))
    #print(json.dumps(foldlisttowork, indent = 4))

    workingfileandfoldermovinglist = list()
    for eachfolder in foldlisttowork:
        if eachfolder != inputconfig["DIR"]["reportpath"]: 
            if eachfolder != inputconfig["DIR"]["historypath"]:
                workingfileandfoldermovinglist.append(eachfolder)

    for eachfile in filelisttowork:
        workingfileandfoldermovinglist.append(eachfile)
    #print(json.dumps(workingfileandfoldermovinglist, indent = 4))

    for eachworkitem in workingfileandfoldermovinglist:
        if os.path.isdir(eachworkitem):
            smallitemslist = eachworkitem.split('/')
            #print(smallitemslist)
            movetopath = "{}/{}".format(inputconfig["DIR"]["historypath"],smallitemslist[-1])
            #print(movetopath)
            dest = shutil.move(eachworkitem, movetopath, copy_function = shutil.copytree)
            #print(dest)
        else:
            smallitemslist = eachworkitem.split('/')
            #print(smallitemslist)
            movetopath = "{}/{}".format(inputconfig["DIR"]["historypath"],smallitemslist[-1])
            #print(movetopath)
            dest = shutil.move(eachworkitem, movetopath)
            #print(dest)            
    #sys.exit()
    return None

def workingonLastdata(DATA,inputconfig):

    DATA['SN']["KEEPLASTPASS"] = dict()
    DATA['SN']['LAST'] = dict()

    for test in DATA['SN']['TEST']:
        print("TEST: {}".format(test))

        if not test in DATA['SN']["LAST"]:
            DATA['SN']["LAST"][test] = dict()

        workingSNlist = list()
        for SN in DATA['SN']['FIRST'][test]:
            workingSNlist.append(SN)

        for SN in workingSNlist:
            #print("{}: {}".format(test, SN))
            #print(json.dumps(DATA['SN']['LAST'][test][SN], indent = 4))
            checkdata = None 
            NewSN = False
            if not SN in DATA['SN']['LAST'][test]:
                DATA['SN']['LAST'][test][SN] = dict()
                NewSN = True
            counttesttime = 0
            if SN in DATA['teststep'][test]['SN']:
                for chassis in DATA['teststep'][test]['SN'][SN]:
                    for testdate in DATA['teststep'][test]['SN'][SN][chassis]:
                        for testtime in DATA['teststep'][test]['SN'][SN][chassis][testdate]:
                            if "checktime" in DATA['SN']['LAST'][test][SN]:
                                checkdata = DATA['SN']['LAST'][test][SN]["checktime"]
                            checktime2 = "{}_{}".format(testdate, testtime)
                            if not checkdata:
                                checkdata = checktime2
                            counttesttime += 1
                            parpareData(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST')
                            if NewSN:
                                writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST')
                                NewSN = False
                            if checktime2 >= checkdata:
                                writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST')

            delSNwithnoData(DATA,test,SN,counttesttime,status='LAST')

        if not test in DATA['SN']["KEEPLASTPASS"]:
            DATA['SN']["KEEPLASTPASS"][test] = dict()

        for SN in workingSNlist:
            #print("{}: {}".format(test, SN))
            #print(json.dumps(DATA['SN']['LAST'][test][SN], indent = 4))
            checkdata = None 
            NewSN = False
            if not SN in DATA['SN']['KEEPLASTPASS'][test]:
                DATA['SN']['KEEPLASTPASS'][test][SN] = dict()
                NewSN = True
            counttesttime = 0
            if SN in DATA['teststep'][test]['SN']:
                for chassis in DATA['teststep'][test]['SN'][SN]:
                    for testdate in DATA['teststep'][test]['SN'][SN][chassis]:
                        for testtime in DATA['teststep'][test]['SN'][SN][chassis][testdate]:
                            if "checktime" in DATA['SN']['KEEPLASTPASS'][test][SN]:
                                checkdata = DATA['SN']['KEEPLASTPASS'][test][SN]["checktime"]
                            checktime2 = "{}_{}".format(testdate, testtime)
                            if not checkdata:
                                checkdata = checktime2
                            counttesttime += 1
                            parpareData(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                            if NewSN:
                                writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                                NewSN = False
                            if checktime2 >= checkdata:
                                if not 'PASS' in DATA['SN']['KEEPLASTPASS'][test][SN]["result"] and not "PASS" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]:
                                    writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                                elif not 'PASS' in DATA['SN']['KEEPLASTPASS'][test][SN]["result"] and "PASS" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]:
                                    writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                                elif 'PASS' in DATA['SN']['KEEPLASTPASS'][test][SN]["result"] and "PASS" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]:
                                    writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                            else:
                                if not 'PASS' in DATA['SN']['KEEPLASTPASS'][test][SN]["result"] and "PASS" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]:
                                    writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='KEEPLASTPASS')
                                    
            delSNwithnoData(DATA,test,SN,counttesttime,status='KEEPLASTPASS')

        workingSNlist = list()
        for SN in DATA['SN']['FIRST'][test]:
            workingSNlist.append(SN)

        for SN in workingSNlist:
            #print("{}: {}".format(test, SN))
            checkdata = None 
            counttesttime = 0
            if SN in DATA['teststep'][test]['SN']:
                for chassis in DATA['teststep'][test]['SN'][SN]:
                    for testdate in DATA['teststep'][test]['SN'][SN][chassis]:
                        for testtime in DATA['teststep'][test]['SN'][SN][chassis][testdate]:
                            if "checktime" in DATA['SN']['FIRST'][test][SN]:
                                checkdata = DATA['SN']['FIRST'][test][SN]["checktime"]
                            checktime2 = "{}_{}".format(testdate, testtime)
                            if not checkdata:
                                checkdata = checktime2
                            counttesttime += 1
                            parpareData(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='FIRST')
                            if checktime2 <= checkdata:
                                writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='FIRST')

            delSNwithnoData(DATA,test,SN,counttesttime,status='FIRST')

    DATA['SN']['ORGLAST'] = DATA['SN']['LAST']
    DATA['SN']['LAST'] = DATA['SN']["KEEPLASTPASS"]

    return None

def parpareData(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST'):

    if not "IMAGE" in DATA['SN'][status][test][SN]:
        if "IMAGE" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["IMAGE"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["IMAGE"]
    if not "MTPINFO" in DATA['SN'][status][test][SN]:
        if "MTPINFO" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["MTPINFO"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["MTPINFO"]
    if not "NICINFO" in DATA['SN'][status][test][SN]:
        if "NICINFO" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["NICINFO"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["NICINFO"]

    return None

def writedatatolastdict(DATA,inputconfig,test,SN,chassis,testdate,testtime,checktime2,status='LAST'):
    if "IMAGE" in DATA['SN'][status][test][SN]:
        if "IMAGE" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["IMAGE"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["IMAGE"]
    if "MTPINFO" in DATA['SN'][status][test][SN]:
        if "MTPINFO" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["MTPINFO"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["MTPINFO"]
    if "NICINFO" in DATA['SN'][status][test][SN]:
        if "NICINFO" in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]:
            DATA['SN'][status][test][SN]["NICINFO"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["NICINFO"]
    checkdata = checktime2
    DATA['SN'][status][test][SN]["checktime"] = checktime2
    DATA['SN'][status][test][SN]["result"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["FINALRESULT"]
    if "SPECIAL" in inputconfig:
        if SN == inputconfig["SPECIAL"]:
            if "FAIL" in DATA['SN'][status][test][SN]["result"]:
                DATA['SN'][status][test][SN]["result"] = "PASS"
    if not "PASS" in DATA['SN'][status][test][SN]["result"]:
        DATA['SN'][status][test][SN]["result"] = "FAIL"
    DATA['SN'][status][test][SN]["testfile"] = DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["eachfile"]
    DATA['SN'][status][test][SN]["ERROR"] = dict()
    DATA['SN'][status][test][SN]["ERROR"]["DETAIL"] = dict()
    DATA['SN'][status][test][SN]["ERROR"]["LIST"] = list()
    for eachteststepdetail in DATA['teststep'][test]['SN'][SN][chassis][testdate][testtime]["TESTSTEPLIST"]:
        for teststepkey in eachteststepdetail:
            if not eachteststepdetail[teststepkey] == 'PASS':
                teststepkeywork = "{} <{}>".format(teststepkey, eachteststepdetail[teststepkey])
                if not teststepkeywork in DATA['SN'][status][test][SN]["ERROR"]["DETAIL"]:
                    DATA['SN'][status][test][SN]["ERROR"]["DETAIL"][teststepkeywork] = 1
                else:
                    DATA['SN'][status][test][SN]["ERROR"]["DETAIL"][teststepkeywork] += 1
                if not teststepkeywork in DATA['SN'][status][test][SN]["ERROR"]["LIST"]:
                    DATA['SN'][status][test][SN]["ERROR"]["LIST"].append(teststepkeywork)
    return None

def delSNwithnoData(DATA,test,SN,counttesttime,status='LAST'):
    if counttesttime:
        DATA['SN'][status][test][SN]["count"] = counttesttime
    else:
        if SN in DATA['SN'][status][test]:
            del DATA['SN'][status][test][SN]
        if SN in DATA['teststep'][test]["SN"]:
            del DATA['teststep'][test]["SN"][SN]
    return None

def checkconfigfile(ARGV):
    print(sys.argv)
    inputfile = sys.argv[1]
    print('GET INPUT CONFIG FILE: ' + inputfile)
    
    if os.path.exists(inputfile):
        print('CHECK INPUT CONFIG FILE EXIST PASS: ' + inputfile)
        inputconfig = mpu.io.read(inputfile)
    else:
        print('CHECK INPUT CONFIG FILE DOES NOT EXIST: ' + inputfile)
        sys.exit()

    for argvitem in sys.argv:
        #print(argvitem)
        matchcheck = re.findall(r"(.*)=(.*)", argvitem)
        if matchcheck:
            #print(matchcheck)
            ARGV[matchcheck[0][0]] = matchcheck[0][1]

    #print(json.dumps(ARGV, indent = 4 ))

    #sys.exit()

    return inputconfig

def checkstartdate():
    returndate = None
    #startdate
    countkey = 1
    for checkkey in sys.argv:
        print("CHECK KEY({}): {}".format(countkey, checkkey))
        countkey += 1
        if 'start' in checkkey:

            match = re.findall(KEY_WORD.CHECKSTARTDATE, checkkey) 
            #print(match)
            if match:
                returndate = match[0]
                print("FIND startdate input: {}".format(returndate))
            else:
                print("FIND startdate keyword, but date format is not correct by YYYY-MM-DD : {}".format(checkkey))
                sys.exit()
    print("FIND startdate input: {}".format(returndate))
    #sys.exit()     
    return returndate

def generateexeclsnDLstatus(DATA,status,wb):

    #DATA['SN']['LIST'] = list()
    #DATA['SN']['TEST'] = list()
    #DATA['SN']['FIRST'] = dict()
    #DATA['SN']['LAST'] = dict()
    
    ws2 = wb.create_sheet(title=status + '_IMAGE')
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('SLOT')
    wirtedata.append('DATE_TIME')
    wirtedata.append('NAME')
    for eachstep in DATA['SN']['IMAGE']:
         wirtedata.append(eachstep)
    ws2.append(wirtedata)
    
    for sn in DATA['SN']['LIST']:

        if sn in DATA['SN']['DL'][status]:
            for dateANDtime in DATA['SN']['DL'][status][sn]:
                for eachname in DATA['SN']['DL'][status][sn][dateANDtime]["DATA"]:
                    for eachdata in DATA['SN']['DL'][status][sn][dateANDtime]["DATA"][eachname]:
                        wirtedata = list()
                        wirtedata.append(sn)                
                        wirtedata.append(DATA['SN']['DL'][status][sn][dateANDtime]["TESTCHASSIS"])
                        wirtedata.append(DATA['SN']['DL'][status][sn][dateANDtime]["SLOT"])
                        wirtedata.append(dateANDtime)
                        wirtedata.append(eachname)
                        #DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD']
                        for eachstep in DATA['SN']['IMAGE']:
                            wirtedata.append(eachdata[eachstep])
                        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    return 0

def takecareteststepDICT(inputconfig,DATA):

    for teststep in inputconfig["TESTLIST"]:
        CreateNeedDict(DATA,teststep)

    return None

def CreateNeedDict(DATA,teststep):
    if not teststep in DATA['teststep']:
        DATA['teststep'][teststep] = dict()
        DATA['teststep'][teststep]['SN'] = dict()
        DATA['teststep'][teststep]['DETAILTESTSTEP'] = list()
        DATA['teststep'][teststep]['FILELISTS'] = list()
    if not teststep in DATA['SN']['FIRST']:
        DATA['SN']['FIRST'][teststep] = dict()
    if not teststep in DATA['SN']['LAST']:
        DATA['SN']['LAST'][teststep] = dict()
    if not teststep in DATA['SN']['ERROR']:
        DATA['SN']['ERROR'][teststep] = dict()
    return None

def workingoneachtest(pr,inputconfig,DATA,testfolder,redo=False):

    takecareteststepDICT(inputconfig,DATA)
    #print(json.dumps(DATA, indent = 4))

    read_dir_path = "{}/{}".format(inputconfig['DIR']['testlogpath'],testfolder)
    print("FOLDER: {}".format(read_dir_path))

    allfolder = getallfolder(read_dir_path)
    #print(json.dumps(allfolder, indent = 4))
    
    teststeplist = list()

    if not 'MTPCHASSIS' in DATA:
        DATA['MTPCHASSIS'] = dict()

    count = 0
    newcount = 0
    for eachfolder in allfolder:
        #print(eachfolder)
        #print(len(eachfolder))
        #print(len(read_dir_path))
        sn = os.path.basename(eachfolder)
        
        
        if 'unknown' in sn.lower():
            continue
        if len(sn) < 5:
            continue
        count += 1
        if count % 100 == 0:
            print("count: {}".format(count))
        #time.sleep(5)
        # if count > 10:
            # time.sleep(5)
            # break
        # if count % 1000 == 0:
        #     print("COUNT: {} Will save to database FILES: {}".format(count,inputconfig['FILE']['datebasesjsonfile']))
        #     mpu.io.write(inputconfig['FILE']['datebasesjsonfile'], DATA)

        if not len(eachfolder) == len(read_dir_path):
          
            getallfilelist = getallfilebyfolder(eachfolder, keyword=".tar.gz")
            # print(json.dumps(getallfilelist, indent = 4))
            for eachfile in getallfilelist:
                #print(eachfile)
                eachfilenameonly = os.path.basename(eachfile)
                getinformationeachfilenameonly = eachfilenameonly.replace('.tar.gz', '')
                #print(eachfilenameonly)
                filearray = getinformationeachfilenameonly.split('_')
                #print(filearray)
                teststep = filearray[0]
                if 'TESTSUB' in inputconfig:
                    if teststep in inputconfig['TESTSUB']:
                        teststep = inputconfig['TESTSUB'][teststep]
                TESTCHASSIS = filearray[1]
                TESTDATE = filearray[2]
                if not teststep in teststeplist:
                    teststeplist.append(teststep)
                    CreateNeedDict(DATA,teststep)
                #if TESTDATE < '2021-07-09':
                    #continue
                # elif TESTDATE > '2021-07-07':
                    # continue

                if not sn in DATA['teststep'][teststep]['SN']:
                    DATA['teststep'][teststep]['SN'][sn] = dict()
                #print("SN: {}".format(sn))
                if not sn in DATA['SN']['LIST']:
                    DATA['SN']['LIST'].append(sn)
                if not sn in DATA['SN']['FIRST'][teststep]:
                    DATA['SN']['FIRST'][teststep][sn] = dict()
                if not sn in DATA['SN']['LAST'][teststep]:
                    DATA['SN']['LAST'][teststep][sn] = dict()
                if not sn in DATA['SN']['ERROR'][teststep]:
                    DATA['SN']['ERROR'][teststep][sn] = dict()  
                    DATA['SN']['ERROR'][teststep][sn]['LIST'] = list()
                    DATA['SN']['ERROR'][teststep][sn]['DETAIL'] = dict()
                
                if not eachfile in DATA['teststep'][teststep]['FILELISTS']:
                    DATA['teststep'][teststep]['FILELISTS'].append(eachfile)
                    newcount += 1
                    DATA['NEWFILECOUNT'] += 1
                    #print("newcount: {}".format(count))
                    if newcount % 1000 == 0:
                        print("COUNT: {} Will save to database FILES: {}".format(newcount,inputconfig['FILE']['datebasesjsonfile']))
                        mpu.io.write(inputconfig['FILE']['datebasesjsonfile'], DATA)
                else:
                    if not redo:
                        continue

                if not TESTCHASSIS in DATA['MTPCHASSIS']:
                    DATA['MTPCHASSIS'][TESTCHASSIS] = dict()

                if not teststep in DATA['MTPCHASSIS'][TESTCHASSIS]:
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep] = dict()

                TESTFINISHTIME = filearray[3]
                CHECKTESTTIME = TESTDATE + '_' + TESTFINISHTIME
                #print("CHASSIS: {} DATE: {} TIME: {}".format(TESTCHASSIS,TESTDATE,TESTFINISHTIME))
                if not TESTCHASSIS in DATA['teststep'][teststep]['SN'][sn]:
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS] = dict()
                if not TESTDATE in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS]:
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE] = dict()
                if not TESTFINISHTIME in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE]:
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME] = dict()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DETAILTESTSTEP'] = dict()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TESTSTEPLIST'] = list()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['PN'] = dict()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'] = dict()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'] = list()
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['IMAGE'] = dict()

                if not CHECKTESTTIME in DATA['MTPCHASSIS'][TESTCHASSIS][teststep]:
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME] = dict()

                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['eachfile'] = eachfile
                if os.path.isdir(inputconfig['DIR']['TEMPDIR']):
                    try:
                        shutil.rmtree(inputconfig['DIR']['TEMPDIR'])
                    except OSError as e:
                        print("Error: %s : %s" % (inputconfig['DIR']['TEMPDIR'], e.strerror))  
                    pr['modules'].createdirinserver(inputconfig['DIR']['TEMPDIR'])

                if os.path.isfile(eachfile):
                    try:
                        shutil.unpack_archive(eachfile, inputconfig['DIR']['TEMPDIR'])
                    except Exception as e:
                        print("Error: %s".format(eachfile)) 
                # else:
                #     continue 

                #mtp_test, test_dl, test_swi, test_fst
                getallunzipfilelist = getallfilebyfolder(inputconfig['DIR']['TEMPDIR'], 'test')

                #print(getallunzipfilelist)


                for readfilefromzip in getallunzipfilelist:
                    #print(readfilefromzip)
                    tempsavelog = ''
                    f = open(readfilefromzip, 'r', encoding="ISO-8859-1")

                    
                    getsomething = False
                    nic_test_rslt_reg_exp = KEY_WORD.NIC_DIAG_TEST_RSLT_RE.format(sn)
                    nic_test_err_reg_exp = KEY_WORD.NIC_DIAG_TEST_ERR_RE.format(sn)
                    nic_test_rslt_final_exp = KEY_WORD.NIC_DIAG_TEST_FINAL_RE.format(sn)
                    nic_test_rslt_cpld_exp = KEY_WORD.NIC_DIAG_TEST_CPLD_IMAGE
                    nic_test_rslt_qspi_exp = KEY_WORD.NIC_DIAG_TEST_QSPI_IMAGE
                    nic_test_rslt_pn_exp = KEY_WORD.NIC_DIAG_TEST_PN.format(sn)
                    nic_test_failure_reg_exp = KEY_WORD.NIC_DIAG_TEST_FAILURE_RE.format('NIC')
                    if "KEY" in inputconfig:
                        nic_test_failure_reg_exp = KEY_WORD.NIC_DIAG_TEST_FAILURE_RE.format(inputconfig["KEY"])
                    failurelist = list()
                    # print(nic_test_rslt_reg_exp)
                    # print(nic_test_rslt_final_exp)
                    # print(nic_test_rslt_cpld_exp)
                    # print(nic_test_rslt_pn_exp)
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] = None
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CARDTYPE'] = "NO CARD TYPE"
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['FINALRESULT'] = "NO RESULT"
                    # if not 'result' in DATA['SN']['FIRST'][teststep][sn]:
                        # DATA['SN']['FIRST'][teststep][sn]['result'] = "NO RESULT"
                        # DATA['SN']['FIRST'][teststep][sn]['checktime'] = CHECKTESTTIME
                    # if not 'result' in DATA['SN']['LAST'][teststep][sn]:
                        # DATA['SN']['LAST'][teststep][sn]['result'] = "NO RESULT"
                        # DATA['SN']['LAST'][teststep][sn]['checktime'] = CHECKTESTTIME
                    TEMPSAVEERROR = dict()
                    TEMPSAVEERROR['LIST'] = list()
                    TEMPSAVEERROR['DETAIL'] = dict()
                    tempmtpinformation = dict()
                    tempnicinformation = dict()
                    tempnicresult = dict()
                    tempimage = dict()
                    temperrormessagelist = list()
                    temperrormessagedata = ''
                    starttocatcherrormessage = False
                    for x in f:
                        tempsavelog += x

                        sub_match = re.findall(KEY_WORD.ERRORMESSAGESTART, x)
                        if sub_match:
                            starttocatcherrormessage = True

                        if starttocatcherrormessage:
                            temperrormessagedata += x
                            #temperrormessagedata += "\r"

                        sub_match = re.findall(KEY_WORD.ERRORMESSAGEEND, x)
                        if sub_match:
                            starttocatcherrormessage = False
                            temperrormessagelist.append(temperrormessagedata)
                            temperrormessagedata = ''
                            #print(temperrormessagelist)
                            #print(x)
                            #sys.exit()



                        #GETIMAGE
                        sub_match = re.findall(KEY_WORD.GETIMAGE, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            mtpnumber = sub_match[0][0]
                            nicnumber = sub_match[0][1]
                            nickey = sub_match[0][2]
                            nicdata = sub_match[0][3]
                            if "IMAGESUB" in inputconfig:
                                if nickey in inputconfig["IMAGESUB"]:
                                    nickey = inputconfig["IMAGESUB"][nickey]
                            if not nicnumber in tempimage:
                                tempimage[nicnumber] = dict()
                            tempimage[nicnumber][nickey] = nicdata
                            #print(tempimage)
                            #sys.exit()

                        sub_match = re.findall(KEY_WORD.GERNICFINALRESULT, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            mtpnumber = sub_match[0][0]
                            nicnumber = sub_match[0][1]
                            nictype = sub_match[0][2]
                            nicsn = sub_match[0][3]
                            nicresult = sub_match[0][4]
                            if not nicnumber in tempnicresult:
                                tempnicresult[nicnumber] = dict()
                            tempnicresult[nicnumber]['SN'] = nicsn
                            tempnicresult[nicnumber]['TYPE'] = nictype
                            tempnicresult[nicnumber]['RESULT'] = nicresult
                            #print(tempnicresult)
                            #sys.exit()

                        sub_match = re.findall(KEY_WORD.GETNICINFORMATION, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            mtpnumber = sub_match[0][0]
                            nicnumber = sub_match[0][1]
                            nickey = sub_match[0][2]
                            nicdata = sub_match[0][3]
                            if not nicnumber in tempnicinformation:
                                tempnicinformation[nicnumber] = dict()
                            tempnicinformation[nicnumber][nickey] = nicdata
                            #print(tempnicinformation)
                            #sys.exit()

                        sub_match = re.findall(KEY_WORD.GETMTPINFORMATION, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            mtpnumber = sub_match[0][0]
                            mptkey = sub_match[0][1]
                            mtpdata = sub_match[0][2]
                            tempmtpinformation['MTP'] = mtpnumber
                            tempmtpinformation[mptkey] = mtpdata
                            #sys.exit()

                        sub_match = re.findall(nic_test_rslt_reg_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            getsomething = True
                            detailteststep = sub_match[0][0]
                            if detailteststep == 'DL2':
                                detailteststep = 'DL'
                            if detailteststep == 'DL3':
                                detailteststep = 'DL'
                            detailteststep = detailteststep + ' ' + sub_match[0][1]
                            detailteststepresult = sub_match[0][2]
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DETAILTESTSTEP'][detailteststep] = detailteststepresult
                            if not detailteststep in DATA['teststep'][teststep]['DETAILTESTSTEP']:
                                DATA['teststep'][teststep]['DETAILTESTSTEP'].append(detailteststep)
                            #DATA['SN']['ERROR'][teststep][sn]
                            reportdetailstep = "{} <{}>".format(detailteststep, detailteststepresult)
                            if not detailteststepresult == 'PASS':
                                if not reportdetailstep in DATA['SN']['ERROR'][teststep][sn]['DETAIL']:
                                    DATA['SN']['ERROR'][teststep][sn]['DETAIL'][reportdetailstep] = 1
                                else:
                                    DATA['SN']['ERROR'][teststep][sn]['DETAIL'][reportdetailstep] += 1

                                if not reportdetailstep in DATA['SN']['ERROR'][teststep][sn]['LIST']:
                                    DATA['SN']['ERROR'][teststep][sn]['LIST'].append(reportdetailstep)

                                if not reportdetailstep in TEMPSAVEERROR['DETAIL']:
                                    TEMPSAVEERROR['DETAIL'][reportdetailstep] = 1
                                else:
                                    TEMPSAVEERROR['DETAIL'][reportdetailstep] += 1

                                if not reportdetailstep in TEMPSAVEERROR['LIST']:
                                    TEMPSAVEERROR['LIST'].append(reportdetailstep)
                            tempeachtestresult = dict()
                            tempeachtestresult[detailteststep] = detailteststepresult
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TESTSTEPLIST'].append(tempeachtestresult)
                                    
                                    
                        #nic_test_err_reg_exp            
                        sub_match = re.findall(nic_test_err_reg_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            detailerror = sub_match[0]

                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'].append(detailerror)

                        sub_match = re.findall(nic_test_failure_reg_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            detailerror = sub_match[0]

                            failurelist.append(detailerror)                               
                                    
                        sub_match = re.findall(nic_test_rslt_final_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            getsomething = True
                            testslot = sub_match[0][1]
                            carttype = sub_match[0][2]
                            finalresult = sub_match[0][3]
                            
                            #NIC-09 ORTANO2 FLM21100052 NIC_DIAG_REGRESSION_TEST_FAIL

                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] = testslot
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CARDTYPE'] = carttype
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['FINALRESULT'] = finalresult
                            if not 'result' in DATA['SN']['FIRST'][teststep][sn]:
                                DATA['SN']['FIRST'][teststep][sn]['result'] = finalresult
                                DATA['SN']['FIRST'][teststep][sn]['checktime'] = CHECKTESTTIME
                                DATA['SN']['FIRST'][teststep][sn]['testfile'] = eachfile
                                DATA['SN']['FIRST'][teststep][sn]['ERROR'] = TEMPSAVEERROR
                            elif CHECKTESTTIME < DATA['SN']['FIRST'][teststep][sn]['checktime']:
                                    DATA['SN']['FIRST'][teststep][sn]['result'] = finalresult
                                    DATA['SN']['FIRST'][teststep][sn]['checktime'] = CHECKTESTTIME
                                    DATA['SN']['FIRST'][teststep][sn]['testfile'] = eachfile 
                                    DATA['SN']['FIRST'][teststep][sn]['ERROR'] = TEMPSAVEERROR                              
                            if not 'result' in DATA['SN']['LAST'][teststep][sn]:
                                DATA['SN']['LAST'][teststep][sn]['result'] = finalresult
                                DATA['SN']['LAST'][teststep][sn]['checktime'] = CHECKTESTTIME
                                DATA['SN']['LAST'][teststep][sn]['count'] = 1
                                DATA['SN']['LAST'][teststep][sn]['testfile'] = eachfile
                                DATA['SN']['LAST'][teststep][sn]['ERROR'] = TEMPSAVEERROR
                            elif CHECKTESTTIME > DATA['SN']['LAST'][teststep][sn]['checktime']:
                                    DATA['SN']['LAST'][teststep][sn]['result'] = finalresult
                                    DATA['SN']['LAST'][teststep][sn]['checktime'] = CHECKTESTTIME
                                    DATA['SN']['LAST'][teststep][sn]['count'] += 1
                                    DATA['SN']['LAST'][teststep][sn]['testfile'] = eachfile
                                    DATA['SN']['LAST'][teststep][sn]['ERROR'] = TEMPSAVEERROR

                        ## [2021-03-23_16-26-57] LOG: [MTP-207]: [NIC-01]: CPLD1 image: naples200_Ortano2_rev3_2_03112021.bin
                        ## [2021-03-23_16-26-57] LOG: [MTP-207]: [NIC-01]: CPLD2 image: naples200_Ortano2_failsafe_rev3_2_03112021.bin

                        sub_match = re.findall(nic_test_rslt_cpld_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            cpldslot = sub_match[0][0]
                            cpldlocate = sub_match[0][1]
                            if sub_match[0][2]:
                                cpldlocate = cpldlocate + str(sub_match[0][2])
                            version = sub_match[0][3]
                            #DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DETAILTESTSTEP'][detailteststep] = detailteststepresult
                            if not cpldlocate in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD']:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'][cpldlocate] = dict()
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'][cpldlocate][cpldslot] = version
                            #print(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'])
                            #sys.exit()
                            
                        sub_match = re.findall(nic_test_rslt_qspi_exp, x)
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            qspislot = sub_match[0][0]
                            qspilocate = sub_match[0][1]
                            version = sub_match[0][3]
                            #DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DETAILTESTSTEP'][detailteststep] = detailteststepresult
                            if not 'QSPI' in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['QSPI'] = dict()
                                
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['QSPI'][qspislot] = version
                            #print(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'])
                            #sys.exit()
                            
                        ## [2021-04-28_22-09-08] LOG: [MTP-208]: [NIC-04]: SN = FLM211000A6; MAC = 00-AE-CD-F6-0E-C8; PN = 68-0021-02 01
                        
                        sub_match = re.findall(nic_test_rslt_pn_exp, x)
                        #print(nic_test_rslt_pn_exp)
                        #sys.exit()
                        if sub_match:
                            #print(x)
                            #print(sub_match)
                            macaddress = sub_match[0][0]
                            pn = sub_match[0][1]
                            dom = sub_match[0][2]
                            #[('00-AE-CD-F6-0F-40', '68-0021-02 01', '042921')] [0][0|1|2|]

                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['MACADDRESS'] = macaddress
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['PN'] = pn
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['DOM'] = dom
                            #sys.exit()
                                
                    f.close()

                    if not getsomething:
                        continue
                    #print(json.dumps(tempmtpinformation, indent = 4))
                    #print(json.dumps(tempnicinformation, indent = 4))
                    #tempnicresult
                    #print(json.dumps(tempnicresult, indent = 4))
                    #tempimage
                    #print(json.dumps(tempimage, indent = 4))
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['IMAGE'] = None
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['NICINFO'] = None
                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['NICRESULT'] = None

                    if DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] in tempimage:
                        DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['IMAGE'] = tempimage[DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT']]
                    if DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] in tempnicinformation:
                        DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['NICINFO'] = tempnicinformation[DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT']]
                    if DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'] in tempnicresult:
                        DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['NICRESULT'] = tempnicresult[DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT']]

                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['MTPINFO'] = tempmtpinformation
                    #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['MTPINFO'] = tempmtpinformation
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['NICINFO'] = tempnicinformation
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['NICRESULT'] = tempnicresult
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['NICIMAGE'] = tempimage
                    DATA['MTPCHASSIS'][TESTCHASSIS][teststep][CHECKTESTTIME]['LOG'] = eachfile
                    #print(json.dumps(DATA['MTPCHASSIS'], indent = 4))
                    #sys.exit()

                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['TESTTIME'] = findtotaltesttime(tempsavelog)
                    if len(failurelist) > 0:
                        #print(failurelist)
                        checkslot = "NIC-{}".format(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'])
                        if "KEY" in inputconfig:
                            checkslot = "{}-{}".format(inputconfig["KEY"],DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['SLOT'])
                        #print(checkslot)
                        for eachfailuremessage in failurelist:
                            if checkslot in eachfailuremessage:
                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'].append(eachfailuremessage)
                        #print(nic_test_failure_reg_exp)
                        #print(nic_test_rslt_final_exp)
                        #print(json.dumps(DATA['teststep'][teststep]['SN'], indent = 4))
                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'], indent = 4))

                    if len(temperrormessagelist) > 0:
                        for eacherrormessage in temperrormessagelist:
                            DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'].append(eacherrormessage)
                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['ERRORDETAIL'], indent = 4))

                        #sys.exit()

                    #print(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD'])
                    #sys.exit() 


                if "TYPE" in inputconfig:
                    if inputconfig["TYPE"] == "NIC":
                        getallunzipfilelist = getallfilebyfolder(inputconfig['DIR']['TEMPDIR'], 'NIC')
                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                        #print(json.dumps(getallunzipfilelist, indent = 4))
                        if "SLOT" in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]:
                            if DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]["SLOT"]:
                                cardslot = "NIC-{}" .format(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]["SLOT"])
                                #print("CARD: {}".format(cardslot))
                                for eachniccardlog in getallunzipfilelist:
                                    if cardslot in eachniccardlog:
                                        #print("LOG: {}".format(eachniccardlog))
                                        f = open(eachniccardlog, 'r', encoding="ISO-8859-1")
                                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                                        for x in f:
                                            sub_match = re.findall(KEY_WORD.GETELBAASICDIEID, x)
                                            if sub_match:
                                                #print(x)
                                                #print(sub_match)                                     
                                                DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]["ELBA_DIE_ID"] = sub_match[0]

                                                
                                                #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                                                #sys.exit()

                                        if not "ELBA_DIE_ID" in DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]:
                                            for x in f:
                                                if sub_match:
                                                    #print(x)
                                                    #print(sub_match)                                     
                                                    DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]["ELBA_DIE_ID"] = sub_match[0]

                                                    
                                                    #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                                                    #sys.exit()
                                                #print(x)
                                        #print(json.dumps(DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME], indent = 4))
                                        #sys.exit()
                                        f.close()


                try:
                    shutil.rmtree(inputconfig['DIR']['TEMPDIR'])
                except OSError as e:
                    print("Error: %s : %s" % (inputconfig['DIR']['TEMPDIR'], e.strerror))  

          
    for teststep in teststeplist:   
        jsonfileoutputname = "{}/{}_{}_DATA.json".format(inputconfig['DIR']["TEMPDATABASE"],teststep,date_time)
        mpu.io.write(jsonfileoutputname, DATA['teststep'][teststep])
    
    return None

def findtotaltesttime(f):
    #print(f)
    match = re.findall(KEY_WORD.FINDDATEANDTIME, f) 

    #print(match)

    if match:
        #print(match[0])

        #print(match[-1])

        starttime = "{}_{}".format(match[0][0], match[0][1])
        endtime = "{}_{}".format(match[-1][0], match[-1][1])
        if starttime > endtime:
            #print(match)
            #print(len(match))
            #sys.exit()
            for x in range(len(match)):
                starttime = "{}_{}".format(match[x][0], match[x][1])
                if endtime > starttime:
                    break
        startdatetime_object = datetime.strptime(starttime, '%Y-%m-%d_%H-%M-%S')
        enddatetime_object = datetime.strptime(endtime, '%Y-%m-%d_%H-%M-%S')

        #print(type(startdatetime_object))
        #print(startdatetime_object)  # printed in default format

        #print(type(enddatetime_object))
        #print(enddatetime_object)  # printed in default format

        difftime = enddatetime_object-startdatetime_object
        #print('Done Time: ', difftime)       
        #print("How many seconds use?: {} seconds".format(difftime.total_seconds())) 
        returninseconds = "{}".format(difftime.total_seconds())
        #sys.exit()
        return returninseconds

    else:
        return None

def generateexeclsnFSJstatus(DATA, workingonSNlist, status,wb,Withallerror=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnFSJstatus: {}".format(status))
    titlename = status
    if Withallerror:
        titlename = "{} All Error".format(status)
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('TEST_WEEK')
    for test in DATA['SN']['TEST']:
         wirtedata.append(test)
    ws2.append(wirtedata)
    
    #print(json.dumps(DATA['SN']['LIST'], indent = 4))
    #time.sleep(2)

    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        overall = dict()
        overall['result'] = 'PASS'
        overall['TESTWEEK'] = 'NODATA'
        overall['LAST'] = dict()
        overall['LAST']['test'] = None
        overall['LAST']['result'] = None  
        overall['LAST']['date'] = None 
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                firsttesttimestamp = DATA['SN'][status][test][sn]["checktime"]
                break
        testcount = 0
        for test in DATA['SN']['TEST']:

            if sn in DATA['SN'][status][test]:

                testcount += 1
                if "result" in DATA['SN'][status][test][sn]:
                    if not overall['LAST']['result']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])
                    elif DATA['SN'][status][test][sn]["checktime"] > overall['LAST']['date']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])                     
                    if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                        overall['result'] = "FAIL"

        if overall['result'] == 'PASS':
            if not testcount == len(DATA['SN']['TEST']):
                overall['result'] = "INCOMPLETE"
        #wirtedata.append(overall['result'])
        wirtedata.append(overall['TESTWEEK'])

        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            if sn in DATA['SN'][status][test]:
                #print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                #time.sleep(5)
                if "result" in DATA['SN'][status][test][sn]:
                    #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                    dateandtime = DATA['SN'][status][test][sn]["checktime"]
                    datetimearray = dateandtime.split('_')
                    reportresut = "{} ({})".format(DATA['SN'][status][test][sn]["result"],datetimearray[0])
                       
                    wirtedata.append(reportresut)
                else:
                    wirtedata.append('NO TEST RESULT')
            else:
                wirtedata.append('NO TEST DATA')
            #time.sleep(2)
        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')
    freezePosition(ws2,'C2')
    
    
    return 0

def getsnlistafteestartdate(DATA,inputconfig,startdate=None):

    returnSNlist = list()

    if startdate:
        for sn in DATA['SN']['LIST']:
            for test in DATA['SN']['TEST']:
                if sn in DATA['SN']['LAST'][test]:
                    if "result" in DATA['SN']['LAST'][test][sn]:
                        #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                        dateandtime = DATA['SN']['LAST'][test][sn]["checktime"]
                        datetimearray = dateandtime.split('_')
                        #print("SN: {} -> {} vs {}".format(sn, datetimearray[0], startdate))
                        if datetimearray[0] >= startdate:
                            if not sn in returnSNlist:
                                returnSNlist.append(sn)
                                returnSNlist.sort()
                                #sys.exit()
                break

    returnSNlist2 = list()

    if len(returnSNlist) == 0:
        returnSNlist = DATA['SN']['LIST']

    for SN in returnSNlist:
        recordSN = True
        if "SKIPPrefix" in inputconfig:
            if inputconfig["SKIPPrefix"] in SN:
                recordSN = False
        if "STARTCountSN" in inputconfig:
            if SN < inputconfig["STARTCountSN"]:
                recordSN = False
        if recordSN:
            if not SN in returnSNlist2:
                returnSNlist2.append(SN)  
                returnSNlist2.sort()

    return returnSNlist2

def generateexeclsn4CFailurestatus(DATA, workingonSNlist,status,wb,inputconfig,Withallerror=False,duplicatesn=True,byweek=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsn4CFailurestatus: {}".format(status))
    titlename = "Top4C_F_By{}".format(status)
    if not duplicatesn:
        titlename = "{}_NoDupSN".format(titlename)
    print("TITLE: {}".format(titlename))
    topfailuredata = dict()
    topfailuredata['week'] = list()
    topfailuredata['weekday'] = dict()
    topfailuredata['TEST'] = DATA['SN']['TEST']
    topfailuredata['DATA'] = dict()
    topfailuredata['OVERALL'] = dict()
    topfailuredata['OVERALL']['TEST'] = dict()
    topfailuredata['OVERALL']['DATA'] = dict()
    topfailuredata['OVERALL']['TEST_HVLV'] = dict()
    topfailuredata['OVERALL']['TEST_HVLV']["DATA"] = dict()
    topfailuredata['OVERALL']['TEST_HVLV']["TEST"] = list()
    topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"] = dict()
    topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'] = dict()
    topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list'] = list()
    topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'] = dict()

    for test in DATA['SN']['TEST']:
        print("{}: {}".format("generateexeclsn4CFailurestatus",test))
    for sn in workingonSNlist:
        for test in DATA['SN']['TEST']:
            #print("{}: {}".format("generateexeclsn4CFailurestatus",test))
            if not "4C" in test:
                continue
            if not test in topfailuredata['OVERALL']['TEST']:
                topfailuredata['OVERALL']['TEST'][test] = dict()
            if not test in topfailuredata['OVERALL']['DATA']:
                topfailuredata['OVERALL']['DATA'][test] = dict()
            if sn in DATA['SN'][status][test]:
                testweek = findworkweek(DATA['SN'][status][test][sn]["checktime"])
                if not testweek in topfailuredata['weekday']:
                    topfailuredata['weekday'][testweek] = list()
                testdayandtime = DATA['SN'][status][test][sn]["checktime"].split("_")
                testday = testdayandtime[0]
                if not testday in topfailuredata['weekday'][testweek]:
                    topfailuredata['weekday'][testweek].append(testday)
                    topfailuredata['weekday'][testweek].sort()
                if not testweek in topfailuredata['DATA']:
                    topfailuredata['DATA'][testweek] = dict()
                if not test in topfailuredata['DATA'][testweek]:
                    topfailuredata['DATA'][testweek][test] = dict()
                if not testweek in topfailuredata['week']:
                    topfailuredata['week'].append(testweek)
                    topfailuredata['week'].sort(reverse=True)

                testresult = DATA['SN'][status][test][sn]["result"]
                if 'FAIL' in testresult:
                    #print(json.dumps(DATA['SN'][status][test][sn]["ERROR"], indent = 4))
                    if len(DATA['SN'][status][test][sn]["ERROR"]["LIST"]):
                        if not DATA['SN'][status][test][sn]["ERROR"]["LIST"][0] in topfailuredata['DATA'][testweek][test]:
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]] = dict()
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] = 1
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"] = list()
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        else:
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] += 1
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                    else:
                        if not "UNKNOWN" in topfailuredata['DATA'][testweek][test]:
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"] = dict()
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["count"] = 1
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"] = list()
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"].append(sn)
                        else:
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["count"] += 1
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"].append(sn)   
                    if len(DATA['SN'][status][test][sn]["ERROR"]["LIST"]):
                        if not DATA['SN'][status][test][sn]["ERROR"]["LIST"][0] in topfailuredata['OVERALL']['TEST'][test]:
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]] = dict()
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] = 1
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"] = list()
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        else:
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] += 1
                            topfailuredata['OVERALL']['TEST'][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        test_HVLV = DATA['SN'][status][test][sn]["ERROR"]["LIST"][0][:2]
                        testdetail = DATA['SN'][status][test][sn]["ERROR"]["LIST"][0][3:]
                        testdetail = testdetail.replace("FAILED", "FAIL")
                        testdetail = testdetail.replace("FAILURE", "FAIL")
                        if "HV" == test_HVLV or "LV" == test_HVLV:
                            newtest = "{} ({})".format(test,test_HVLV)
                            if not newtest in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
                                topfailuredata['OVERALL']['TEST_HVLV']["TEST"].append(newtest)
                                topfailuredata['OVERALL']['TEST_HVLV']["TEST"].sort()
                            if not testdetail in topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"]:
                                topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"][testdetail] = 1
                            else:
                                topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"][testdetail] += 1
                            #print("{} | {}".format(newtest,testdetail))
                            if not newtest in topfailuredata['OVERALL']['TEST_HVLV']["DATA"]:
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest] = dict()
                            if not testdetail in topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest]:
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail] = dict()
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["count"] = 1
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["SN"] = list()
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["SN"].append(sn)
                            else:
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["count"] += 1
                                topfailuredata['OVERALL']['TEST_HVLV']["DATA"][newtest][testdetail]["SN"].append(sn)

                            if not testweek in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list']:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list'].append(testweek)
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list'].sort(reverse=True)

                            if not testweek in topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY']:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek] = list()

                            if not testday in topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek].append(testday)
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek].sort()

                            if not testweek in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK']:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek] = dict()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"] = dict()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"] = list()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"] = dict()                                

                            if not newtest in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"].append(newtest)
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"].sort()
                            if not testdetail in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"][testdetail] = 1
                            else:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"][testdetail] += 1
                            #print("{} | {}".format(newtest,testdetail))
                            if not newtest in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest] = dict()
                            if not testdetail in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest]:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail] = dict()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["count"] = 1
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["SN"] = list()
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["SN"].append(sn)
                            else:
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["count"] += 1
                                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][newtest][testdetail]["SN"].append(sn)

                    else:
                        if not "UNKNOWN" in topfailuredata['OVERALL']['TEST'][test]:
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"] = dict()
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["count"] = 1
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["SN"] = list()
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["SN"].append(sn)
                        else:
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["count"] += 1
                            topfailuredata['OVERALL']['TEST'][test]["UNKNOWN"]["SN"].append(sn)  


                    if not duplicatesn:
                        break
    #print(json.dumps(topfailuredata, indent = 4))                    
    #sys.exit()



    if "4CFLOW" in inputconfig:
        backuplist = topfailuredata['OVERALL']['TEST_HVLV']["TEST"]
        topfailuredata['OVERALL']['TEST_HVLV']["TEST"] = inputconfig["4CFLOW"]
        for teststep in backuplist:
            if not teststep in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
                topfailuredata['OVERALL']['TEST_HVLV']["TEST"].append(teststep)

    if not byweek:
        ws2 = wb.create_sheet(title=titlename)
        

        wirtedata = list()
        wirtedata.append("4C Process sequence")
        count = 1
        for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
            wirtedata.append(count)
            count += 1
        ws2.append(wirtedata)
        wirtedata = list()
        wirtedata.append("FAILURE TYPE")
        for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
             wirtedata.append(test)
        wirtedata.append("TOTAL")
        ws2.append(wirtedata)
        testtotal = dict()
        for failuretype, number in sorted(topfailuredata['OVERALL']['TEST_HVLV']["FAILURESTEP"].items(), key=lambda item: item[1],reverse=True):
            print("Failure: {} - {}".format(failuretype,number))
            wirtedata = list()
            wirtedata.append(failuretype)
            
            for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
                if not test in testtotal:
                    testtotal[test] = 0
                if test in topfailuredata['OVERALL']['TEST_HVLV']["DATA"]:
                    if failuretype in topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test]:
                        wirtedata.append(topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test][failuretype]["count"])
                        testtotal[test] += topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test][failuretype]["count"]
                    else:
                        wirtedata.append("")
                else:
                    wirtedata.append("")
            wirtedata.append(number)
            ws2.append(wirtedata)

        wirtedata = list()
        wirtedata.append("TOTAL")
        alltotal = 0
        for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
            wirtedata.append(testtotal[test])
            alltotal += testtotal[test]
        wirtedata.append(alltotal)
        ws2.append(wirtedata)
        wirtedata = list()
        ws2.append(wirtedata)
        wirtedata = list()
        wirtedata.append("TEST")
        wirtedata.append("SN")
        wirtedata.append("FAILURE TYPE")
        ws2.append(wirtedata)
        for test in topfailuredata['OVERALL']['TEST_HVLV']["TEST"]:
            if test in topfailuredata['OVERALL']['TEST_HVLV']["DATA"]:
                for failuretype in topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test]:
                    for SN in topfailuredata['OVERALL']['TEST_HVLV']["DATA"][test][failuretype]["SN"]:
                        wirtedata = list()
                        wirtedata.append(test)
                        wirtedata.append(SN)
                        wirtedata.append(failuretype)
                        ws2.append(wirtedata)
        wirtedata = list()
        ws2.append(wirtedata)
        ws2.append(wirtedata)
        ws2.append(wirtedata)

    else:
        titlename = "{}_byWeek".format(titlename)
        ws2 = wb.create_sheet(title=titlename)
        
        for testweek in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK_list']:
            wirtedata = list()
            wirtedata.append("Week: {}".format(testweek))
            wirtedata.append("START:")
            wirtedata.append(topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek][0])
            wirtedata.append("END:")
            wirtedata.append(topfailuredata['OVERALL']['TEST_HVLV_by_WEEKDAY'][testweek][-1])            
            ws2.append(wirtedata)

            wirtedata = list()
            wirtedata.append("4C Process sequence")
            count = 1
            if "4CFLOW" in inputconfig:
                backuplist = topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]
                topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"] = inputconfig["4CFLOW"]
                for teststep in backuplist:
                    if not teststep in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                        topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"].append(teststep)

            for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                wirtedata.append(count)
                count += 1
            ws2.append(wirtedata)
            wirtedata = list()
            wirtedata.append("FAILURE TYPE")
            for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                 wirtedata.append(test)
            wirtedata.append("TOTAL")
            ws2.append(wirtedata)        
            testtotal = dict()
            for failuretype, number in sorted(topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["FAILURESTEP"].items(), key=lambda item: item[1],reverse=True):
                print("Failure: {} - {}".format(failuretype,number))
                wirtedata = list()
                wirtedata.append(failuretype)
                
                for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                    if not test in testtotal:
                        testtotal[test] = 0
                    if test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"]:
                        if failuretype in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test]:
                            wirtedata.append(topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test][failuretype]["count"])
                            testtotal[test] += topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test][failuretype]["count"]
                        else:
                            wirtedata.append("")
                    else:
                        wirtedata.append("")
                wirtedata.append(number)
                ws2.append(wirtedata)

            wirtedata = list()
            wirtedata.append("TOTAL")
            alltotal = 0
            for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                wirtedata.append(testtotal[test])
                alltotal += testtotal[test]
            wirtedata.append(alltotal)
            ws2.append(wirtedata)
            wirtedata = list()
            ws2.append(wirtedata)
            wirtedata = list()
            wirtedata.append("TEST")
            wirtedata.append("SN")
            wirtedata.append("FAILURE TYPE")
            ws2.append(wirtedata)
            for test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["TEST"]:
                if test in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"]:
                    for failuretype in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test]:
                        for SN in topfailuredata['OVERALL']['TEST_HVLV_by_WEEK'][testweek]["DATA"][test][failuretype]["SN"]:
                            wirtedata = list()
                            wirtedata.append(test)
                            wirtedata.append(SN)
                            wirtedata.append(failuretype)
                            ws2.append(wirtedata)
            wirtedata = list()
            ws2.append(wirtedata)
            ws2.append(wirtedata)
            ws2.append(wirtedata)

    fixcolumnssize(ws2,False)
    highlightnumberincell(ws2)
    #highlightinyellow(ws2,'TIMEOUT')
    #highlightinyellow(ws2,'NO TEST DATA')
    #highlightinred(ws2, 'FAIL')
    #highlightinred(ws2, 'FAILED')
    #highlightingreen(ws2,'PASS')
    #highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    #freezePosition(ws2,'C2')
    
    return 0

def generateexeclsnTopFailurestatus(DATA, workingonSNlist,status,wb,inputconfig,Withallerror=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnTopFailurestatus: {}".format(status))
    titlename = "TopFailureBy{}".format(status)
    print("TITLE: {}".format(titlename))
    topfailuredata = dict()
    topfailuredata['week'] = list()
    topfailuredata['weekday'] = dict()
    topfailuredata['TEST'] = DATA['SN']['TEST']
    topfailuredata['DATA'] = dict()
    topfailuredata['TESTSNLIST'] = dict()
    for sn in workingonSNlist:
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                testweek = findworkweek(DATA['SN'][status][test][sn]["checktime"])
                if not testweek in topfailuredata['weekday']:
                    topfailuredata['weekday'][testweek] = list()
                if not testweek in topfailuredata['TESTSNLIST']:
                    topfailuredata['TESTSNLIST'][testweek] = dict()
                if not test in topfailuredata['TESTSNLIST'][testweek]:
                    topfailuredata['TESTSNLIST'][testweek][test] = list()
                if not sn in topfailuredata['TESTSNLIST'][testweek][test]:
                    topfailuredata['TESTSNLIST'][testweek][test].append(sn)
                testdayandtime = DATA['SN'][status][test][sn]["checktime"].split("_")
                testday = testdayandtime[0]
                if not testday in topfailuredata['weekday'][testweek]:
                    topfailuredata['weekday'][testweek].append(testday)
                    topfailuredata['weekday'][testweek].sort()
                if not testweek in topfailuredata['DATA']:
                    topfailuredata['DATA'][testweek] = dict()
                if not test in topfailuredata['DATA'][testweek]:
                    topfailuredata['DATA'][testweek][test] = dict()
                if not testweek in topfailuredata['week']:
                    topfailuredata['week'].append(testweek)
                    topfailuredata['week'].sort(reverse=True)

                testresult = DATA['SN'][status][test][sn]["result"]
                if 'FAIL' in testresult:
                    #print(json.dumps(DATA['SN'][status][test][sn]["ERROR"], indent = 4))
                    if len(DATA['SN'][status][test][sn]["ERROR"]["LIST"]):
                        if not DATA['SN'][status][test][sn]["ERROR"]["LIST"][0] in topfailuredata['DATA'][testweek][test]:
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]] = dict()
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] = 1
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"] = list()
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                        else:
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["count"] += 1
                            topfailuredata['DATA'][testweek][test][DATA['SN'][status][test][sn]["ERROR"]["LIST"][0]]["SN"].append(sn)
                    else:
                        if not "UNKNOWN" in topfailuredata['DATA'][testweek][test]:
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"] = dict()
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["count"] = 1
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"] = list()
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"].append(sn)
                        else:
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["count"] += 1
                            topfailuredata['DATA'][testweek][test]["UNKNOWN"]["SN"].append(sn)    
                    #print(json.dumps(topfailuredata, indent = 4))                    
                    #sys.exit()

    topfailuredata['OVERALL'] = dict()
    for testweek in topfailuredata['week']:
        for test in topfailuredata['TEST']:
            if not test in topfailuredata['DATA'][testweek]:
                continue
            for failuretype in topfailuredata['DATA'][testweek][test]:
                if not failuretype in topfailuredata['OVERALL']:
                    topfailuredata['OVERALL'][failuretype] = topfailuredata['DATA'][testweek][test][failuretype]["count"]
                else:
                    topfailuredata['OVERALL'][failuretype] += topfailuredata['DATA'][testweek][test][failuretype]["count"]

    topfailuredata['OVERALLbyWeek'] = dict()
    for testweek in topfailuredata['week']:
        if not testweek in topfailuredata['OVERALLbyWeek']:
            topfailuredata['OVERALLbyWeek'][testweek] = dict()
        for test in topfailuredata['TEST']:
            if not test in topfailuredata['DATA'][testweek]:
                continue
            for failuretype in topfailuredata['DATA'][testweek][test]:
                if not failuretype in topfailuredata['OVERALLbyWeek'][testweek]:
                    topfailuredata['OVERALLbyWeek'][testweek][failuretype] = topfailuredata['DATA'][testweek][test][failuretype]["count"]
                else:
                    topfailuredata['OVERALLbyWeek'][testweek][failuretype] += topfailuredata['DATA'][testweek][test][failuretype]["count"]
    #print(json.dumps(topfailuredata, indent = 4))
    for testweek in sorted(topfailuredata['OVERALLbyWeek'],reverse=True):
        print("TEST Week: {}".format(testweek))
        for failuretype, number in sorted(topfailuredata['OVERALLbyWeek'][testweek].items(), key=lambda item: item[1],reverse=True):
            print("Failure: {} - {}".format(failuretype,number))

    #sys.exit()

    ws2 = wb.create_sheet(title=titlename)
    
    for testweek in topfailuredata['week']:
        wirtedata = list()
        wirtedata.append("WEEK: {} {} Top Failure Type".format(testweek,status))
        wirtedata.append("START:")
        wirtedata.append(topfailuredata['weekday'][testweek][0])
        wirtedata.append("END:")
        wirtedata.append(topfailuredata['weekday'][testweek][-1])
        ws2.append(wirtedata)
        wirtedata = list()
        wirtedata.append("FAILURE TYPE")
        eachweektotallist = list()
        eachweektotoaldetail = dict()
        for test in topfailuredata['TEST']:
            writetestinformation = "{} <0>".format(test)
            if test in topfailuredata['TESTSNLIST'][testweek]:
                writetestinformation = "{} <{}>".format(test,len(topfailuredata['TESTSNLIST'][testweek][test]))
            wirtedata.append(writetestinformation)
            eachweektotallist.append(test)
            eachweektotoaldetail[test] = 0
        wirtedata.append("TOTAL")
        eachweektotallist.append("TOTAL")
        eachweektotoaldetail["TOTAL"] = 0
        
        ws2.append(wirtedata)

        for failuretype, number in sorted(topfailuredata['OVERALLbyWeek'][testweek].items(), key=lambda item: item[1],reverse=True):
            print("Failure: {} - {}".format(failuretype,number))
            wirtedata = list()
            wirtedata.append(failuretype)
            for test in topfailuredata['TEST']:
                if testweek in topfailuredata['DATA']:
                    if test in topfailuredata['DATA'][testweek]:
                        if failuretype in topfailuredata['DATA'][testweek][test]:
                            wirtedata.append(topfailuredata['DATA'][testweek][test][failuretype]["count"])
                            eachweektotoaldetail[test] += topfailuredata['DATA'][testweek][test][failuretype]["count"]
                        else:
                            wirtedata.append("")
                    else:
                        wirtedata.append("")
                else:
                    wirtedata.append("")
            wirtedata.append(number)
            eachweektotoaldetail["TOTAL"] += number
            ws2.append(wirtedata)

        wirtedata = list()
        wirtedata.append("TOTAL")
        for test in eachweektotallist:
            wirtedata.append(eachweektotoaldetail[test])
        ws2.append(wirtedata)

        wirtedata = list()
        ws2.append(wirtedata)
        wirtedata = list()
        wirtedata.append("TEST")
        wirtedata.append("SN")
        wirtedata.append("FAILURE TYPE")
        wirtedata.append("RE-TEST COUNT")
        wirtedata.append("LAST RESULT")
        ws2.append(wirtedata)
        for test in topfailuredata['TEST']:
            if testweek in topfailuredata['DATA']:
                if test in topfailuredata['DATA'][testweek]:
                    for failuretype in topfailuredata['DATA'][testweek][test]:
                        for SN in topfailuredata['DATA'][testweek][test][failuretype]["SN"]:
                            wirtedata = list()
                            wirtedata.append(test)
                            wirtedata.append(SN)
                            wirtedata.append(failuretype)
                            wirtedata.append(DATA['SN']["LAST"][test][SN]["count"])
                            wirtedata.append(DATA['SN']["LAST"][test][SN]["result"])
                            ws2.append(wirtedata)
        wirtedata = list()
        ws2.append(wirtedata)
        ws2.append(wirtedata)
        ws2.append(wirtedata)



    fixcolumnssize(ws2,False)
    highlightnumberincell(ws2)
    #highlightinyellow(ws2,'TIMEOUT')
    #highlightinyellow(ws2,'NO TEST DATA')
    #highlightinred(ws2, 'FAIL')
    #highlightinred(ws2, 'FAILED')
    #highlightingreen(ws2,'PASS')
    #highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    #freezePosition(ws2,'C2')
    
    return 0

def generateexeclsnstatus(DATA, workingonSNlist,status,wb,inputconfig,Withallerror=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnstatus: {}".format(status))
    titlename = status
    if Withallerror:
        titlename = "{} All Error".format(status)
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('OVERALL')
    wirtedata.append('TEST_WEEK')
    # if status == 'LAST':
    #     wirtedata.append('LAST_TESTDATE')
    #     wirtedata.append('LAST_STATION(RESULT)')
    
    #     wirtedata.append('QSPI')
    #     wirtedata.append('CPLD1')
    if "DISPLAYINDO" in inputconfig:
        for displaykey in inputconfig["DISPLAYINDO"]:
            wirtedata.append(displaykey)

    #wirtedata.append('')
    for test in DATA['SN']['TEST']:
         wirtedata.append(test)
    ws2.append(wirtedata)

    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        overall = dict()
        overall['result'] = 'PASS'
        overall['TESTWEEK'] = 'NODATA'
        overall['LAST'] = dict()
        overall['LAST']['test'] = None
        overall['LAST']['result'] = None  
        overall['LAST']['date'] = None 
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                firsttesttimestamp = DATA['SN'][status][test][sn]["checktime"]
                break

        testcount = 0
        for test in DATA['SN']['TEST']:

            if sn in DATA['SN'][status][test]:

                testcount += 1
                if "result" in DATA['SN'][status][test][sn]:
                    if not overall['LAST']['result']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])
                    elif DATA['SN'][status][test][sn]["checktime"] > overall['LAST']['date']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])                     
                    if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                        overall['result'] = "FAIL"

        if overall['result'] == 'PASS':
            if not testcount == len(DATA['SN']['TEST']):
                overall['result'] = "INCOMPLETE"
        wirtedata.append(overall['result'])
        wirtedata.append(overall['TESTWEEK'])
        if "DISPLAYINDO" in inputconfig:
            for displaykey in inputconfig["DISPLAYINDO"]:
                #finddisplay = False
                displaydict = dict()
                for test in DATA['SN']['TEST']:
                    if sn in DATA['SN'][status][test]:
                        for checkkey in inputconfig["CHECKKEYS"]:
                            if checkkey in DATA['SN'][status][test][sn]:
                                if isinstance(DATA['SN'][status][test][sn][checkkey], dict):
                                    if displaykey in DATA['SN'][status][test][sn][checkkey]:
                                        if not "data" in displaydict:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]
                                        elif DATA['SN'][status][test][sn]["checktime"] > displaydict["checktime"]:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]                                            
                                        #finddisplay = True
                                        #wirtedata.append(DATA['SN'][status][test][sn][checkkey][displaykey])
                                        #break
                if "data" in displaydict:
                    wirtedata.append(displaydict["data"])
                else:
                    wirtedata.append('NODATA')
        # if status == 'LAST':
        #     last_dateandtime = overall['LAST']['date']

        #     last_datetimearray = last_dateandtime.split('_')
        #     reporttestandresult = "{} ({})".format(overall['LAST']['test'],overall['LAST']['result'])
        #     wirtedata.append(last_datetimearray[0])
        #     wirtedata.append(reporttestandresult)
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'QSPI' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['QSPI'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #     #'CPLD1'
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'CPLD1' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['CPLD1'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #wirtedata.append('')

        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            #sys.exit()
            if sn in DATA['SN'][status][test]:
                # print("TEST: {} | STATUS: {}".format(test, status))
                # print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                # sys.exit()
                #time.sleep(5)
                if DATA['SN'][status][test][sn]["checktime"] >= firsttesttimestamp:
                    if "result" in DATA['SN'][status][test][sn]:
                        #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                        dateandtime = DATA['SN'][status][test][sn]["checktime"]
                        datetimearray = dateandtime.split('_')
                        reportresut = "{} ({})".format(DATA['SN'][status][test][sn]["result"],datetimearray[0])
                        #DATA['SN']['LAST'][teststep][sn]['count']
                        if "count" in DATA['SN'][status][test][sn]:
                            reportresut = "{} <{}> ".format(reportresut,DATA['SN'][status][test][sn]["count"])
                        if Withallerror:
                            if sn in DATA['SN']['ERROR'][test]:
                                for eachfailstep in DATA['SN']['ERROR'][test][sn]['LIST']:
                                    reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                        else:
                            if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                                for eachfailstep in DATA['SN'][status][test][sn]['ERROR']['LIST']:
                                    reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                        wirtedata.append(reportresut)
                    else:
                        wirtedata.append('NO TEST DATA')
                else:
                    wirtedata.append('NO TEST DATA')
            else:
                wirtedata.append('NO TEST DATA')
            #time.sleep(2)
        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')
    highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    freezePosition(ws2,'C2')
    
    return 0

def generateexeclsnstatusalldata(DATA, workingonSNlist,status,wb,inputconfig,Withallerror=False):

    # DATA['SN']['LIST'] = list()
    # DATA['SN']['TEST'] = list()
    # DATA['SN']['FIRST'] = dict()
    # DATA['SN']['LAST'] = dict()
    
    print("generateexeclsnstatusalldata: {}".format(status))
    #titlename = "{} ALL DATE".format(status)
    titlename = "{}".format(status)
    if Withallerror:
        titlename = "{} ALL ERROR".format(titlename)
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('OVERALL')
    wirtedata.append('TEST_WEEK')
    # if status == 'LAST':
    #     wirtedata.append('LAST_TESTDATE')
    #     wirtedata.append('LAST_STATION(RESULT)')
    
    #     wirtedata.append('QSPI')
    #     wirtedata.append('CPLD1')
    if "DISPLAYINDO" in inputconfig:
        for displaykey in inputconfig["DISPLAYINDO"]:
            wirtedata.append(displaykey)

    #wirtedata.append('')
    for test in DATA['SN']['TEST']:
         wirtedata.append(test)
    ws2.append(wirtedata)

    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        overall = dict()
        overall['result'] = 'PASS'
        overall['TESTWEEK'] = 'NODATA'
        overall['LAST'] = dict()
        overall['LAST']['test'] = None
        overall['LAST']['result'] = None  
        overall['LAST']['date'] = None 
        for test in DATA['SN']['TEST']:
            if sn in DATA['SN'][status][test]:
                firsttesttimestamp = DATA['SN'][status][test][sn]["checktime"]
                break

        testcount = 0
        for test in DATA['SN']['TEST']:

            if sn in DATA['SN'][status][test]:

                testcount += 1
                if "result" in DATA['SN'][status][test][sn]:
                    if not overall['LAST']['result']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])
                    elif DATA['SN'][status][test][sn]["checktime"] > overall['LAST']['date']:
                        overall['LAST']['test'] = test
                        overall['LAST']['result'] = DATA['SN'][status][test][sn]["result"]
                        overall['LAST']['date'] = DATA['SN'][status][test][sn]["checktime"]
                        overall['TESTWEEK'] = findworkweek(overall['LAST']['date'])                     
                    if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                        overall['result'] = "FAIL"

        if overall['result'] == 'PASS':
            if not testcount == len(DATA['SN']['TEST']):
                overall['result'] = "INCOMPLETE"
        wirtedata.append(overall['result'])
        wirtedata.append(overall['TESTWEEK'])
        if "DISPLAYINDO" in inputconfig:
            for displaykey in inputconfig["DISPLAYINDO"]:
                #finddisplay = False
                displaydict = dict()
                for test in DATA['SN']['TEST']:
                    if sn in DATA['SN'][status][test]:
                        for checkkey in inputconfig["CHECKKEYS"]:
                            if checkkey in DATA['SN'][status][test][sn]:
                                if isinstance(DATA['SN'][status][test][sn][checkkey], dict):
                                    if displaykey in DATA['SN'][status][test][sn][checkkey]:
                                        if not "data" in displaydict:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]
                                        elif DATA['SN'][status][test][sn]["checktime"] > displaydict["checktime"]:
                                            displaydict["data"] = DATA['SN'][status][test][sn][checkkey][displaykey]
                                            displaydict["checktime"] = DATA['SN'][status][test][sn]["checktime"]                                            
                                        #finddisplay = True
                                        #wirtedata.append(DATA['SN'][status][test][sn][checkkey][displaykey])
                                        #break
                if "data" in displaydict:
                    wirtedata.append(displaydict["data"])
                else:
                    wirtedata.append('NODATA')
        # if status == 'LAST':
        #     last_dateandtime = overall['LAST']['date']

        #     last_datetimearray = last_dateandtime.split('_')
        #     reporttestandresult = "{} ({})".format(overall['LAST']['test'],overall['LAST']['result'])
        #     wirtedata.append(last_datetimearray[0])
        #     wirtedata.append(reporttestandresult)
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'QSPI' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['QSPI'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #     #'CPLD1'
        #     if sn in DATA['SN']['LAST'][DATA['SN']['TEST'][0]]:
        #         if 'CPLD1' in DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]:
        #             wirtedata.append(DATA['SN']['LAST'][DATA['SN']['TEST'][0]][sn]['CPLD1'])
        #         else:
        #             wirtedata.append('NO INFORMATION')
        #     else:
        #         wirtedata.append('NO INFORMATION')
        #wirtedata.append('')

        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            #sys.exit()
            if sn in DATA['SN'][status][test]:
                # print("TEST: {} | STATUS: {}".format(test, status))
                # print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                # sys.exit()
                #time.sleep(5)

                if "result" in DATA['SN'][status][test][sn]:
                    #DATA['SN']['ERROR'][teststep][sn][reportdetailstep]
                    dateandtime = DATA['SN'][status][test][sn]["checktime"]
                    datetimearray = dateandtime.split('_')
                    reportresut = "{} ({})".format(DATA['SN'][status][test][sn]["result"],datetimearray[0])
                    #DATA['SN']['LAST'][teststep][sn]['count']
                    if "count" in DATA['SN'][status][test][sn]:
                        reportresut = "{} <{}> ".format(reportresut,DATA['SN'][status][test][sn]["count"])
                    if Withallerror:
                        if sn in DATA['SN']['ERROR'][test]:
                            for eachfailstep in DATA['SN']['ERROR'][test][sn]['LIST']:
                                reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                    else:
                        if "FAIL" in DATA['SN'][status][test][sn]["result"]:
                            for eachfailstep in DATA['SN'][status][test][sn]['ERROR']['LIST']:
                                reportresut = "{}\r{} <{}>".format(reportresut,eachfailstep,DATA['SN']['ERROR'][test][sn]['DETAIL'][eachfailstep])
                    wirtedata.append(reportresut)
                else:
                    wirtedata.append('NO TEST DATA')
            else:
                wirtedata.append('NO TEST DATA')
            #time.sleep(2)
        ws2.append(wirtedata)
        
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')
    highlightinOrange(ws2,'INCOMPLETE')
    wraptest(ws2)
    freezePosition(ws2,'C2')
    
    return 0

def createlastfailurelogfolder(pr,DATA,workingonSNlist,inputconfig):
    lastfailurefolder = "{}LastFail_{}".format(inputconfig['DIR']["reportpath"],date_time)
    pr['modules'].createdirinserver(lastfailurefolder)
    status = 'LAST'
    countforcreateunzipfile = 0
    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            if sn in DATA['SN'][status][test]:
                #print(json.dumps(DATA['SN'][status][test][sn], indent = 4))
                #time.sleep(5)
                if "result" in DATA['SN'][status][test][sn]:
                    if "FAIL" in DATA['SN'][status][test][sn]["result"].upper():
                        dateandtime = DATA['SN'][status][test][sn]["checktime"]
                        datetimearray = dateandtime.split('_')
                        eachfile = DATA['SN'][status][test][sn]["testfile"]
                        #shutil.copy2(eachfile, lastfailurefolder)
                        unzipfolder = "{}/{}_{}_{}".format(lastfailurefolder,test,sn,datetimearray[0])
                        pr['modules'].createdirinserver(unzipfolder)
                        shutil.unpack_archive(eachfile, unzipfolder)
                        countforcreateunzipfile += 1

    print("createlastfailurelogfolder: <{}>".format(countforcreateunzipfile))
    return 0

def copyallfailurelogfolder(pr,DATA,workingonSNlist,inputconfig):
    lastfailurefolder = "{}LastFail_{}".format(inputconfig['DIR']["reportpath"],date_time)
    pr['modules'].createdirinserver(lastfailurefolder)
    status = 'LAST'
    countforcreateunzipfile = 0
    for sn in workingonSNlist:
        wirtedata = list()
        wirtedata.append(sn)
        for test in DATA['SN']['TEST']:
            #print("{} : {}".format(sn,test))
            #print(json.dumps(DATA['SN'][status][test], indent = 4))
            if sn in DATA['SN'][status][test]:
                if "result" in DATA['SN'][status][test][sn]:
                    if "FAIL" in DATA['SN'][status][test][sn]["result"].upper():
                        if sn in DATA['SN'][status][test]:
                            copyfolder = "{}/{}_{}".format(lastfailurefolder,test,sn)
                            pr['modules'].createdirinserver(copyfolder)
                            for eachfile in DATA['teststep'][test]['FILELISTS']:
                                if sn.upper() in eachfile.upper():

                                    shutil.copy2(eachfile, copyfolder)
                                    countforcreateunzipfile += 1
                            change_permissions_recursive(copyfolder, 0o777)
    linux_cmd_change_permissions = "chmod -R 777 {}".format(lastfailurefolder)
    os.system(linux_cmd_change_permissions)

    print("copyallfailurelogfolder: <{}>".format(countforcreateunzipfile))
    return 0

def change_permissions_recursive(path, mode):
    for root, dirs, files in os.walk(path, topdown=False):
        for dir in [os.path.join(root,d) for d in dirs]:
            os.chmod(dir, mode)
    for file in [os.path.join(root, f) for f in files]:
            os.chmod(file, mode)

def generatedailydataintxtfile(DATA,teststep):
    
    ## [2021-04-01_03-51-47] LOG: [MTP-213]: NIC-10 ORTANO2 FLM2110004F NIC_DIAG_REGRESSION_TEST_PASS
    ## [2021-04-01_03-51-47] ERR: [MTP-213]: NIC-07 ORTANO2 FLM2110005B NIC_DIAG_REGRESSION_TEST_FAIL (LV_ASIC L1)
    ## [2021-04-16_12-30-18] LOG: [MTPS-402]: NIC-02 ORTANO2 FLM21100003 NIC_DIAG_REGRESSION_TEST_PASS
    print("")
    print("teststep: {}".format(teststep))
    print("")
    print("")
    checkdatasetup = "2021-07-11"
    BYCHASSIS = dict()
    for sn in DATA["SN"]:
        for chassis in DATA["SN"][sn]:
            if not chassis in BYCHASSIS:
                BYCHASSIS[chassis] = list()
            for testdate in DATA["SN"][sn][chassis]:
                if not testdate == checkdatasetup:
                    continue
                for testetime in DATA["SN"][sn][chassis][testdate]:
                    
                    wirteline = "## [{}_{}] ".format(testdate, testetime)
                    if 'PASS' in DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']:
                        wirteline += "LOG: "
                    else:
                        wirteline += "ERR: "
                    wirteline += "[{}]: NIC-{} {} {} NIC_DIAG_REGRESSION_TEST_{}".format(chassis,DATA["SN"][sn][chassis][testdate][testetime]['SLOT'],DATA["SN"][sn][chassis][testdate][testetime]['CARDTYPE'],sn,DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                    if 'FAIL' in DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']:
                        wirteline += " ("
                        for eachdeatilteststep in DATA["DETAILTESTSTEP"]:
                            if eachdeatilteststep in DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]:
                                if 'FAIL' in DATA["SN"][sn][chassis][testdate][testetime]['DETAILTESTSTEP'][eachdeatilteststep]:
                                    wirteline += "{},".format(eachdeatilteststep)
                        if wirteline[-1] == ",":
                            wirteline = wirteline[:-1]
                        wirteline += ")"
                            
                    BYCHASSIS[chassis].append(wirteline)
                    print(wirteline)
                      
                            
                    
    
    #print(json.dumps(BYCHASSIS, indent = 4))
    print("teststep: {}".format(teststep))
    for eachchassis in BYCHASSIS:
        print("CHASSIS: {}".format(eachchassis))
        for eachline in BYCHASSIS[eachchassis]:
            print(eachline)
    print("")
    print("")
    print("")
    print("")
    print("")
    print("")
    print("")
    print("")
    print("")

    
    
    return 0

def GetTestTimedictbyweek(workingonSNlist,DATA,teststep,FULLDATA):

    DatabyWeek = dict()
    DatabyWeek['WEEKLIST'] = list()
    DatabyWeek['DATA'] = dict()

    TimeData = dict()
    TimeData["DATA"] = dict()
    TimeData["LIST"] = list()

    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        endtime = "{}_{}".format(testdate,testetime)
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            unittesttime = str(int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME'])))
                            #print("{}: {}".format(unittesttime,endtime))
                            #sys.exit()
                            begintime = findbegintesttime(endtime,unittesttime)
                            recordchassistime = "{}_{}".format(chassis,endtime)
                            if not int(unittesttime) in TimeData["LIST"]:
                                TimeData["LIST"].append(int(unittesttime))
                                TimeData["LIST"].sort(reverse=True)
                            if not unittesttime in TimeData["DATA"]:
                                TimeData["DATA"][unittesttime] = dict()
                                TimeData["DATA"][unittesttime]["MTP"] = dict()
                                TimeData["DATA"][unittesttime]["TESTTIME"] = int(unittesttime)
                            if not recordchassistime in TimeData["DATA"][unittesttime]["MTP"]:
                                TimeData["DATA"][unittesttime]["MTP"][recordchassistime] = dict()
                            TimeData["DATA"][unittesttime]["MTP"][recordchassistime]["start"] = begintime
                            TimeData["DATA"][unittesttime]["MTP"][recordchassistime]["end"] = endtime
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        weeknumber = findworkweek("{}_{}".format(testdate,testetime))
                        if not weeknumber in DatabyWeek['WEEKLIST']:
                            DatabyWeek['WEEKLIST'].append(weeknumber)
                            DatabyWeek['WEEKLIST'].sort()
                        if not weeknumber in DatabyWeek['DATA']:
                            DatabyWeek['DATA'][weeknumber] = dict()
                            DatabyWeek['DATA'][weeknumber]["numberofSLOTlist"] = list()
                            DatabyWeek['DATA'][weeknumber]["DATAofNumber"] = dict()
                        timestamp = "{}_{}".format(testdate,testetime)
                        numberofslotfortest = findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp)
                        if isinstance(numberofslotfortest, int):
                            numberofslotfortest = "{:02d}".format(numberofslotfortest)
                        if not numberofslotfortest in DatabyWeek['DATA'][weeknumber]["numberofSLOTlist"]:
                            if not 'NO' in numberofslotfortest.upper():
                                DatabyWeek['DATA'][weeknumber]["numberofSLOTlist"].append(numberofslotfortest)
                                DatabyWeek['DATA'][weeknumber]["numberofSLOTlist"].sort()
                        if not numberofslotfortest in DatabyWeek['DATA'][weeknumber]["DATAofNumber"]:
                            DatabyWeek['DATA'][weeknumber]["DATAofNumber"][numberofslotfortest] = list()

                        if '2C' in teststep or '4C' in teststep:
                            endtestime = "{}_{}".format(testdate,testetime)
                            MaxTestTime = findmaxtesttimeinChamber(endtestime,TimeData)
                            DatabyWeek['DATA'][weeknumber]["DATAofNumber"][numberofslotfortest].append(int(MaxTestTime))
                        else:
                            if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                                DatabyWeek['DATA'][weeknumber]["DATAofNumber"][numberofslotfortest].append(int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME'])))
    
    return DatabyWeek

def generateexecltestbyNon4Ctesttime(workingonSNlist,DATA,teststep,wb,FULLDATA):

    titlename = "{}_{}".format("Testtime", teststep)
    ws2 = wb.create_sheet(title=titlename)
    print("{}: {}".format("generateexecltestbyNon4Ctesttime", titlename))
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('WEEK')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('TESTTIME(HH:MM:SS)')
    wirtedata.append('TIME_RANGE')
    wirtedata.append('TESTSTATUS')
    wirtedata.append('# of Card in MTP')
    TimeData = dict()
    TimeData["DATA"] = dict()
    TimeData["LIST"] = list()

    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        slot = DATA["SN"][sn][chassis][testdate][testetime]['SLOT']
                        wirtedata = list()
                        wirtedata.append(teststep)
                        wirtedata.append(sn)
                        wirtedata.append(chassis)
                        timestamp = "{}_{}".format(testdate,testetime)

                        wirtedata.append(findworkweek("{}_{}".format(testdate,testetime)))
                        endtestime = "{}_{}".format(testdate,testetime)
                        wirtedata.append(testdate)
                        wirtedata.append(testetime.replace("-",":"))
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(timedelta(seconds=int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                        else:
                            wirtedata.append('NO TESTTIME')
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(converttoHourtoHour(int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                        else:
                            wirtedata.append("No TESTTIME")
                        #sys.exit()
                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                        timestamp = "{}_{}".format(testdate,testetime)
                        wirtedata.append(findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp))
                        ws2.append(wirtedata)
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'K2')
    return 0

def generateexecltestby4Ctesttime(workingonSNlist,DATA,teststep,wb,FULLDATA):

    titlename = "{}_{}".format("Testtime", teststep)
    ws2 = wb.create_sheet(title=titlename)
    print("{}: {}".format("generateexecltestby4Ctesttime", titlename))
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('WEEK')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('TESTTIME(HH:MM:SS)')
    wirtedata.append('GROUP_TESTTIME(HH:MM:SS)')
    wirtedata.append('TIME_RANGE')
    wirtedata.append('TESTSTATUS')
    wirtedata.append('# of Card in MTP')
    TimeData = dict()
    TimeData["DATA"] = dict()
    TimeData["LIST"] = list()

    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        endtime = "{}_{}".format(testdate,testetime)
                        unittesttime = str(int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME'])))
                        #print("{}: {}".format(unittesttime,endtime))
                        #sys.exit()
                        begintime = findbegintesttime(endtime,unittesttime)
                        recordchassistime = "{}_{}".format(chassis,endtime)
                        if not int(unittesttime) in TimeData["LIST"]:
                            TimeData["LIST"].append(int(unittesttime))
                            TimeData["LIST"].sort(reverse=True)
                        if not unittesttime in TimeData["DATA"]:
                            TimeData["DATA"][unittesttime] = dict()
                            TimeData["DATA"][unittesttime]["MTP"] = dict()
                            TimeData["DATA"][unittesttime]["TESTTIME"] = int(unittesttime)
                        if not recordchassistime in TimeData["DATA"][unittesttime]["MTP"]:
                            TimeData["DATA"][unittesttime]["MTP"][recordchassistime] = dict()
                        TimeData["DATA"][unittesttime]["MTP"][recordchassistime]["start"] = begintime
                        TimeData["DATA"][unittesttime]["MTP"][recordchassistime]["end"] = endtime
    #print(json.dumps(TimeData, indent = 4))
    #sys.exit()
    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        slot = DATA["SN"][sn][chassis][testdate][testetime]['SLOT']
                        wirtedata = list()
                        wirtedata.append(teststep)
                        wirtedata.append(sn)
                        wirtedata.append(chassis)
                        
                        #sys.exit()
                        wirtedata.append(findworkweek("{}_{}".format(testdate,testetime)))
                        endtestime = "{}_{}".format(testdate,testetime)
                        wirtedata.append(testdate)
                        wirtedata.append(testetime.replace("-",":"))
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(timedelta(seconds=int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                        else:
                            wirtedata.append('NO TESTTIME')
                        MaxTestTime = findmaxtesttimeinChamber(endtestime,TimeData)
                        wirtedata.append(timedelta(seconds=int(MaxTestTime)))
                        wirtedata.append(converttoHourtoHour(MaxTestTime))
                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                        #sys.exit()
                        timestamp = "{}_{}".format(testdate,testetime)
                        wirtedata.append(findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp))
                        ws2.append(wirtedata)
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'K2')
    return 0

def converttoHourtoHour(MaxTestTime):
    hours = int(MaxTestTime / 3600)
    displaystr = "{:02d} ~ {:02d}".format(hours, hours + 1)
    return displaystr


def findmaxtesttimeinChamber(endtesttime,TimeData):
    count = 1
    for unittesttime in TimeData["LIST"]:
        #print("{}: {}".format(count,unittesttime))
        #print(json.dumps(TimeData["DATA"][str(unittesttime)], indent = 4))
        for mtp in TimeData["DATA"][str(unittesttime)]["MTP"]:
            if endtesttime >= TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["start"] and endtesttime <= TimeData["DATA"][str(unittesttime)]["MTP"][mtp]["end"]:
                
                #print("Return {}: {} by END TIME: {}".format(count,unittesttime,endtesttime))
                #sys.exit()
                return unittesttime
        
        count += 1

def GetSNlistbyPNfromDLtest(workingonSNlist,DATA,teststep='DL'):

    SNbyPN = dict()
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        PN = None
                        if 'PN' in DATA["SN"][sn][chassis][testdate][testetime]:
                            PN = str(DATA["SN"][sn][chassis][testdate][testetime]['PN'])
                        else:
                            PN = 'CANNOT FIND'
                        
                        teststatus = DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']

                        if 'PASS' in teststatus.upper():
                            if not PN in SNbyPN:
                                SNbyPN[PN] = list()
                            if not sn in SNbyPN[PN]:
                                SNbyPN[PN].append(sn)
                                SNbyPN[PN].sort(reverse=True)
                        
    return SNbyPN


def generateexecltest(workingonSNlist,DATA,teststep,wb,FULLDATA):

    ws2 = wb.create_sheet(title=teststep)
    print("{}: {}".format("generateexecltest", teststep))
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('WEEK')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('TESTTIME(HH:MM:SS)')
    wirtedata.append('# of Card in MTP')
    wirtedata.append('CARDTYPE')
    wirtedata.append('SLOT')
    wirtedata.append('FINALRESULT')
    if teststep == 'DL':
        wirtedata.append('PN')
        wirtedata.append('MACADDRESS')
        #wirtedata.append('CPLD1')
        #wirtedata.append('CPLD2')
        #wirtedata.append('QSPI')
    chassisinfo = list()
    imageinfo = list()
    nicinfo = list()

    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        if "MTPINFO" in DATA["SN"][sn][chassis][testdate][testetime]:
                            if type(DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"]) is dict:
                                for chassiskey in DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"]:
                                    if not chassiskey in chassisinfo:
                                        chassisinfo.append(chassiskey)
                                        chassisinfo.sort()
                        if "IMAGE" in DATA["SN"][sn][chassis][testdate][testetime]:
                            if type(DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"]) is dict:
                                for imagekey in DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"]:
                                    if not imagekey in imageinfo:
                                        imageinfo.append(imagekey)
                                        imageinfo.sort()
                        if "NICINFO" in DATA["SN"][sn][chassis][testdate][testetime]:
                            if type(DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"]) is dict:
                                for nickey in DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"]:
                                    if not nickey in nicinfo:
                                        nicinfo.append(nickey)
                                        nicinfo.sort()

    for chassiskey in chassisinfo:
        wirtedata.append(chassiskey)
    for imagekey in imageinfo:
        wirtedata.append(imagekey)
    for nickey in nicinfo:
        wirtedata.append(nickey)

    for eachdeatilteststep in DATA["DETAILTESTSTEP"]:
         wirtedata.append(eachdeatilteststep)
    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        slot = DATA["SN"][sn][chassis][testdate][testetime]['SLOT']
                        wirtedata = list()
                        wirtedata.append(teststep)
                        wirtedata.append(sn)
                        wirtedata.append(chassis)
                        wirtedata.append(findworkweek("{}_{}".format(testdate,testetime)))
                        wirtedata.append(testdate)
                        wirtedata.append(testetime.replace("-",":"))
                        if 'TESTTIME' in DATA["SN"][sn][chassis][testdate][testetime]:
                            wirtedata.append(timedelta(seconds=int(float(DATA["SN"][sn][chassis][testdate][testetime]['TESTTIME']))))
                        else:
                            wirtedata.append('NO TESTTIME')
                        timestamp = "{}_{}".format(testdate,testetime)
                        wirtedata.append(findhowmanycardinthistestbymtp(FULLDATA,chassis,teststep,timestamp))
                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CARDTYPE'])
                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['SLOT'])
                        #DATA['teststep'][teststep]['SN'][sn][TESTCHASSIS][TESTDATE][TESTFINISHTIME]['CPLD']
                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                        if teststep == 'DL':
                            if 'PN' in DATA["SN"][sn][chassis][testdate][testetime]:
                                wirtedata.append(str(DATA["SN"][sn][chassis][testdate][testetime]['PN']))
                            else:
                                wirtedata.append('CANNOT FIND')
                                
                            if 'MACADDRESS' in DATA["SN"][sn][chassis][testdate][testetime]:
                                wirtedata.append(str(DATA["SN"][sn][chassis][testdate][testetime]['MACADDRESS']))
                            else:
                                wirtedata.append('CANNOT FIND')
                            #print(DATA["SN"][sn][chassis][testdate][testetime])
                            #sys.exit()
                            # if 'CPLD1' in DATA["SN"][sn][chassis][testdate][testetime]['CPLD']:
                            #     wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CPLD']['CPLD1'][slot])
                            # else:
                            #     wirtedata.append('CANNOT FIND')
                            # if 'CPLD2' in DATA["SN"][sn][chassis][testdate][testetime]['CPLD']:
                            #     wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CPLD']['CPLD2'][slot])
                            # else:
                            #     wirtedata.append('CANNOT FIND')
                            # if 'QSPI' in DATA["SN"][sn][chassis][testdate][testetime]:
                            #     wirtedata.append(str(DATA["SN"][sn][chassis][testdate][testetime]['QSPI'][slot]))
                            # else:
                            #     wirtedata.append('CANNOT FIND')
                        
                        for chassiskey in chassisinfo:
                            #wirtedata.append(chassiskey)
                            if "MTPINFO" in DATA["SN"][sn][chassis][testdate][testetime]:
                                if type(DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"]) is dict:
                                    if chassiskey in DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"]:
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]["MTPINFO"][chassiskey])
                                    else:
                                        wirtedata.append('CANNOT FIND')
                                else:
                                    wirtedata.append('CANNOT FIND')
                            else:
                                wirtedata.append('CANNOT FIND')
                        for imagekey in imageinfo:
                            #wirtedata.append(imagekey)
                            if "IMAGE" in DATA["SN"][sn][chassis][testdate][testetime]:
                                if type(DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"]) is dict:
                                    if imagekey in DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"]:
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]["IMAGE"][imagekey])
                                    else:
                                        wirtedata.append('CANNOT FIND')
                                else:
                                    wirtedata.append('CANNOT FIND')
                            else:
                                wirtedata.append('CANNOT FIND')
                        for nickey in nicinfo:
                            #wirtedata.append(nickey)
                            if "NICINFO" in DATA["SN"][sn][chassis][testdate][testetime]:
                                if type(DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"]) is dict:
                                    if nickey in DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"]:
                                        wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]["NICINFO"][nickey])
                                    else:
                                        wirtedata.append('CANNOT FIND')
                                else:
                                    wirtedata.append('CANNOT FIND')
                            else:
                                wirtedata.append('CANNOT FIND')    

                        for eachdeatilteststep in DATA["DETAILTESTSTEP"]:
                            if eachdeatilteststep in DATA["SN"][sn][chassis][testdate][testetime]["DETAILTESTSTEP"]:
                                wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['DETAILTESTSTEP'][eachdeatilteststep])
                            else:
                                wirtedata.append("NO TEST DATA")
                        
                        ws2.append(wirtedata)
    
    fixcolumnssize(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    freezePosition(ws2,'L2')
    return 0

def generateexeclerrdata(workingonSNlist,DATA,teststep,wb,FULLDATA):

    teststeperrortitle = "{}_ERROR".format(teststep)
    print("{}: {}".format("generateexeclerrdata", teststeperrortitle))
    ws2 = wb.create_sheet(title=teststeperrortitle)
    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN')
    wirtedata.append('CHASSIS')
    wirtedata.append('DATE')
    wirtedata.append('TIME')
    wirtedata.append('CARDTYPE')
    wirtedata.append('SLOT')

    wirtedata.append('FINALRESULT')
    wirtedata.append('ERROR_DATA')
    ws2.append(wirtedata)
    
    for sn in workingonSNlist:
        if sn in DATA["SN"]:
            for chassis in DATA["SN"][sn]:
                for testdate in DATA["SN"][sn][chassis]:
                    for testetime in DATA["SN"][sn][chassis][testdate]:
                        if "FAIL" in DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT']:
                            slot = DATA["SN"][sn][chassis][testdate][testetime]['SLOT']
                            howmanyerror = len(DATA["SN"][sn][chassis][testdate][testetime]['ERRORDETAIL'])
                            if howmanyerror:
                                for eacherror in DATA["SN"][sn][chassis][testdate][testetime]['ERRORDETAIL']:
                                    eacherrorlist = eacherror.split('\n')
                                    neweacherror = ''
                                    for checkeacherror in eacherrorlist:

                                        checkeacherror = _removeIllegalCharacterError(checkeacherror)
                                        neweacherror += checkeacherror
                                        neweacherror += "\r"
                                    wirtedata = list()
                                    wirtedata.append(teststep)
                                    wirtedata.append(sn)
                                    wirtedata.append(chassis)
                                    wirtedata.append(testdate)
                                    wirtedata.append(testetime)
                                    wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['CARDTYPE'])
                                    wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['SLOT'])

                                    wirtedata.append(DATA["SN"][sn][chassis][testdate][testetime]['FINALRESULT'])
                                    #print(eacherror)
                                    #if len(eacherror) > 200:
                                        #eacherror = eacherror[:200]
                                    wirtedata.append(neweacherror)
                                    #print(wirtedata)
                                    ws2.append(wirtedata)
                                    #print(wirtedata)
    
    fixcolumnssize2(ws2)
    highlightinyellow(ws2,'TIMEOUT')
    highlightinyellow(ws2,'NO TEST DATA')
    highlightingreen(ws2,'PASS')
    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    wraptest(ws2)
    return 0

def fixcolumnssize(ws,enablefilter=True):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
    for col, value in dims.items():
        if value > 30:
            value = 30
        ws.column_dimensions[col].width = value + 2
    if enablefilter:
        ws.auto_filter.ref = ws.dimensions

def fixcolumnssize2(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
    for col, value in dims.items():
        if value > 100:
            value = 100
        ws.column_dimensions[col].width = value + 2

    ws.auto_filter.ref = ws.dimensions

def wraptest(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True,vertical='top') 

def freezePosition(ws,keyword):
    ws.freeze_panes = keyword

def highlightinOrange(ws,keyword):
    from openpyxl.styles import Color, PatternFill, Font, Border
    maxRow = ws.max_row
    maxCol = ws.max_column
    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
    for rowNum in range(1, maxRow + 1):
        fillcolor = 0
        for colNum in range(1, maxCol + 1):
            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
            if keyword in str(ws.cell(row=rowNum, column=colNum).value):
                ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type = 'solid')

def highlightinyellow(ws,keyword):
	from openpyxl.styles import Color, PatternFill, Font, Border
	maxRow = ws.max_row
	maxCol = ws.max_column
	#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	for rowNum in range(1, maxRow + 1):
		fillcolor = 0
		for colNum in range(1, maxCol + 1):
			#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
			if keyword in str(ws.cell(row=rowNum, column=colNum).value):
				ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')

def highlightnumberincell(ws):
    import numbers
    from openpyxl.styles import Color, PatternFill, Font, Border
    maxRow = ws.max_row
    maxCol = ws.max_column
    #print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
    for rowNum in range(1, maxRow + 1):
        fillcolor = 0
        for colNum in range(1, maxCol + 1):
            #print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
            #if keyword in str(ws.cell(row=rowNum, column=colNum).value):
            if isinstance(ws.cell(row=rowNum, column=colNum).value, numbers.Number):
                if ws.cell(row=rowNum, column=colNum).value < 3:
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
                elif ws.cell(row=rowNum, column=colNum).value >= 3 and ws.cell(row=rowNum, column=colNum).value < 8:
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FFBF00', end_color='FFBF00', fill_type = 'solid')
                else:
                    ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')

def highlightingreen(ws,keyword):
	from openpyxl.styles import Color, PatternFill, Font, Border
	maxRow = ws.max_row
	maxCol = ws.max_column
	#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	for rowNum in range(1, maxRow + 1):
		fillcolor = 0
		for colNum in range(1, maxCol + 1):
			#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
			if keyword in str(ws.cell(row=rowNum, column=colNum).value):
				fillcolor = 1
				ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='66f86a', end_color='66f86a', fill_type = 'solid')
                
def highlightinred(ws,keyword):
	from openpyxl.styles import Color, PatternFill, Font, Border
	maxRow = ws.max_row
	maxCol = ws.max_column
	#print('highlightinyellow: ' + keyword + ' ' + str(maxRow) + ' ' + str(maxCol))
	for rowNum in range(1, maxRow + 1):
		fillcolor = 0
		for colNum in range(1, maxCol + 1):
			#print(str(rowNum) + ' ' + str(colNum) + ' ' + str(ws.cell(row=rowNum, column=colNum).value))
			if keyword in str(ws.cell(row=rowNum, column=colNum).value):
				ws.cell(row=rowNum, column=colNum).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')                

def findhowmanycardinthistestbymtp(DATA,mtp,test,timestamp):
    
    if not mtp in DATA["MTPCHASSIS"]:
        return "No Data"

    if not test in DATA["MTPCHASSIS"][mtp]:
        return "No Data"

    if not timestamp in DATA["MTPCHASSIS"][mtp][test]:
        return "No Data"

    mtpdata = DATA["MTPCHASSIS"][mtp][test][timestamp]

    #print(json.dumps(mtpdata, indent = 4))

    if not "NICRESULT" in mtpdata:
        return "No Data"

    return len(mtpdata["NICRESULT"].keys())

def generateexeclmtpstatusbyeachmtpreport(DATA,wb,inputconfig,mtp):
    print("generateexeclmtpstatusbyeachmtpreport: {}".format(mtp))
    titlename = mtp
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append("TEST")
    wirtedata.append("DATE_TIME")
    for slot in inputconfig["MTP_STATUS"]:
        wirtedata.append(slot)
    wirtedata.append("FAILURE_TYPE")
    for slot in inputconfig["MTP_STATUS"]:
        wirtedata.append(slot)
    ws2.append(wirtedata)

    for test in DATA["MTPCHASSIS"][mtp]:
        for datetime in sorted(DATA["MTPCHASSIS"][mtp][test].keys(), reverse=True):
            wirtedata = list()
            wirtedataFailure = list()
            wirtedata.append(test)
            wirtedata.append(datetime)
            for slot in inputconfig["MTP_STATUS"]:

                if slot in DATA["MTPCHASSIS"][mtp][test][datetime]["NICRESULT"]:
                    if "FAIL" in DATA["MTPCHASSIS"][mtp][test][datetime]["NICRESULT"][slot]["RESULT"]:
                        #print(json.dumps(DATA["MTPCHASSIS"][mtp][test][datetime], indent = 4))
                        testdateandtime = datetime.split("_")
                        SN = DATA["MTPCHASSIS"][mtp][test][datetime]["NICRESULT"][slot]["SN"]
                        if SN in DATA["teststep"][test]["SN"]:
                            #print(json.dumps(DATA["teststep"][test]["SN"][SN][mtp][testdateandtime[0]][testdateandtime[1]], indent = 4))
                            firstFailurestep = None 
                            if SN in DATA["teststep"][test]["SN"]:
                                if mtp in DATA["teststep"][test]["SN"][SN]:
                                    if testdateandtime[0] in DATA["teststep"][test]["SN"][SN][mtp]:
                                        if testdateandtime[1] in DATA["teststep"][test]["SN"][SN][mtp][testdateandtime[0]]:
                                            for teststepdata in DATA["teststep"][test]["SN"][SN][mtp][testdateandtime[0]][testdateandtime[1]]["TESTSTEPLIST"]:
                                                for teststep in teststepdata:
                                                    if "FAIL" in teststepdata[teststep]:
                                                        firstFailurestep = "{} <{}>".format(teststep,teststepdata[teststep])
                                                if firstFailurestep:
                                                    break
                            print(firstFailurestep)
                            if not firstFailurestep:
                                firstFailurestep = "CANNOT FIND FAILURE STEP"
                            wirtedataFailure.append(firstFailurestep)
                            #sys.exit()
                        else:
                            wirtedataFailure.append("CANNOT FIND SN: {}".format(SN))
                    else:
                        wirtedataFailure.append("PASS")
                    wirtedata.append(DATA["MTPCHASSIS"][mtp][test][datetime]["NICRESULT"][slot]["RESULT"])
                else:
                    wirtedata.append("None")
                    wirtedataFailure.append("None")
            wirtedata.append("|||")
            for eachdata in wirtedataFailure:
                wirtedata.append(eachdata)

            ws2.append(wirtedata)

    highlightinred(ws2, 'FAIL')
    highlightinred(ws2, 'FAILED')
    highlightingreen(ws2,'PASS')      
    fixcolumnssize(ws2)

    return True

def generateexeclmtpstatussummaryreport(DATA,wb,inputconfig):
    ws1 = wb.active
    ws1.title = "SUMMARY"

    mtpchassisusecountbyslot = dict()
    for MTP in DATA["MTPCHASSIS"]:
        #print(MTP)
        if not MTP in mtpchassisusecountbyslot:
            mtpchassisusecountbyslot[MTP] = dict()
            mtpchassisusecountbyslot[MTP]["TEST"] = list()
            for slot in inputconfig["MTP_STATUS"]:
                mtpchassisusecountbyslot[MTP][slot] = dict()
                mtpchassisusecountbyslot[MTP][slot]["TOTAL"] = 0
                mtpchassisusecountbyslot[MTP][slot]["PASS"] = 0
                mtpchassisusecountbyslot[MTP][slot]["FAIL"] = 0
        for test in DATA["MTPCHASSIS"][MTP]:
            #print(test)
            if not test in mtpchassisusecountbyslot[MTP]["TEST"]:
                mtpchassisusecountbyslot[MTP]["TEST"].append(test)
            for datetime in DATA["MTPCHASSIS"][MTP][test]:
                #print(datetime)
                #print(json.dumps(DATA["MTPCHASSIS"][MTP][test][datetime], indent = 4))
                for eachslot in DATA["MTPCHASSIS"][MTP][test][datetime]["NICRESULT"]:
                    if "PASS" in DATA["MTPCHASSIS"][MTP][test][datetime]["NICRESULT"][eachslot]["RESULT"]:
                        mtpchassisusecountbyslot[MTP][eachslot]["TOTAL"] += 1
                        mtpchassisusecountbyslot[MTP][eachslot]["PASS"] += 1
                    else:
                        mtpchassisusecountbyslot[MTP][eachslot]["TOTAL"] += 1
                        mtpchassisusecountbyslot[MTP][eachslot]["FAIL"] += 1                        
            #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
            #sys.exit()
        #break

    
    #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
    #sys.exit()
    wirtedata = list()
    wirtedata.append('MTP')
    wirtedata.append('TEST')
    wirtedata.append('STATUS')
    for slot in inputconfig["MTP_STATUS"]:
        wirtedata.append(slot)
    ws1.append(wirtedata)
    for mtp in mtpchassisusecountbyslot:
        wirtedatatotal = list()
        wirtedatapass = list()
        wirtedatafail = list()
        wirtedatatotal.append(mtp)
        wirtedatapass.append(mtp)
        wirtedatafail.append(mtp)
        teststep = mtpchassisusecountbyslot[mtp]["TEST"][0]
        if len(mtpchassisusecountbyslot[mtp]["TEST"]) > 1:
            for eachtest in mtpchassisusecountbyslot[mtp]["TEST"][1:]:
                teststep = "{},{}".format(teststep,eachtest)
        wirtedatatotal.append(teststep)
        wirtedatapass.append(teststep)
        wirtedatafail.append(teststep)  
        wirtedatatotal.append("TOTAL")
        wirtedatapass.append("PASS")
        wirtedatafail.append("FAIL")      
        for slot in inputconfig["MTP_STATUS"]:
            wirtedatatotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            wirtedatapass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            wirtedatafail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])
        ws1.append(wirtedatatotal)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatafail)           
    fixcolumnssize(ws1)

    return True

def generateexecldieidreport(DATA,wb,inputconfig,start=None):
    ws1 = wb.active
    ws1.title = "DIE_ID_DATA"

    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('DIE_ID')
    ws1.append(wirtedata)

    alllistofSN = list()
    for test in DATA['SN']['TEST']:
        wirtedata = list()
        wirtedata.append(test)
        listofSN = list()
        for SN in DATA['teststep'][test]["SN"]:
            recordSN = True
            if "SKIPPrefix" in inputconfig:
                if inputconfig["SKIPPrefix"] in SN:
                    recordSN = False
            if "STARTCountSN" in inputconfig:
                if SN < inputconfig["STARTCountSN"]:
                    recordSN = False
            if recordSN:
                listofSN.append(SN)
                if not SN in alllistofSN:
                    alllistofSN.append(SN)   

    for SN in alllistofSN:
        #DATA["SN"][sn][chassis][testdate][testetime]
        die_id = None
        for test in DATA['teststep']:
            if SN in DATA['teststep'][test]["SN"]:
                for chassis in DATA['teststep'][test]["SN"][SN]:
                    for testdate in DATA['teststep'][test]["SN"][SN][chassis]:
                        for testetime in DATA['teststep'][test]["SN"][SN][chassis][testdate]:
                            if "ELBA_DIE_ID" in DATA['teststep'][test]["SN"][SN][chassis][testdate][testetime]:
                                die_id = DATA['teststep'][test]["SN"][SN][chassis][testdate][testetime]["ELBA_DIE_ID"]
        wirtedata = list()
        wirtedata.append(SN)        
        if die_id:
            wirtedata.append(die_id)
        else:
            wirtedata.append("Cannot find")
        ws1.append(wirtedata)

    fixcolumnssize(ws1)

    return True


def generateexeclDatabyPN(DATA,wb,inputconfig,workingonSNlist,PN,start=None):

    titlename = "{}_SUMMARY".format(PN)
    ws2 = wb.create_sheet(title=titlename)
    print("{}: {}".format("generateexeclDatabyPN", titlename))

    SummaryReportDetail(DATA,ws2,inputconfig,workingonSNlist,start=None)

    return None    

def generateexeclreport(DATA,wb,inputconfig,workingonSNlist,start=None):

    # 1st Sheet keep active
    ws1 = wb.active
    
    # Will creat extra "sheet" using below
    #ws1 = wb.create_sheet(title='SUMMARY')
    ws1.title = "SUMMARY"

    SummaryReportDetail(DATA,ws1,inputconfig,workingonSNlist,start=None)

    return None


def SummaryReportDetail(DATA,ws1,inputconfig,workingonSNlist,start=None):

    #print(json.dumps(inputconfig, indent = 4))
    WeekTimedata = dict()
    for eachteststep in DATA['SN']['TEST']:
        WeekTimedata[eachteststep] = GetTestTimedictbyweek(workingonSNlist,DATA["teststep"][eachteststep],eachteststep,DATA)
        #print(json.dumps(WeekTimedata, indent = 4))
        #sys.exit()

    wirtedata = list()
    wirtedata.append('TEST')
    wirtedata.append('SN_QTY')
    wirtedata.append('DETAIL_STEP_QTY')
    wirtedata.append('First_PASS')
    wirtedata.append('First_FAIL')
    wirtedata.append('First_YIELD')
    wirtedata.append('Final_PASS')
    wirtedata.append('Final_FAIL')
    wirtedata.append('Final_YIELD')
    wirtedata.append('')
    wirtedata.append('Final_Failure\rCount (NoDupSN)')
    ws1.append(wirtedata)

    LastfailureremoveDupSN = dict()
    for SN in workingonSNlist:

        for test in DATA['SN']['TEST']:
            if not test in LastfailureremoveDupSN:
                LastfailureremoveDupSN[test] = dict()
                LastfailureremoveDupSN[test]["count"] = 0
                LastfailureremoveDupSN[test]["SN"] = list()
            if SN in DATA['SN']['LAST'][test]:
                if "result" in DATA['SN']['LAST'][test][SN]:
                    if "FAIL" in DATA['SN']['LAST'][test][SN]["result"]:
                        LastfailureremoveDupSN[test]["count"] += 1
                        LastfailureremoveDupSN[test]["SN"].append(SN)
                        break
                else:
                    print("LAST <{}> SN: {}".format(test, SN))
                    #print(json.dumps(DATA['SN']['LAST'][test][SN], indent = 4))
                    del DATA['SN']['LAST'][test][SN]
                    #sys.exit()
            if SN in DATA['SN']['FIRST'][test]:
                if not "result" in DATA['SN']['FIRST'][test][SN]:
                    print("FIRST <{}> SN: {}".format(test, SN))
                    #print(json.dumps(DATA['SN']['FIRST'][test][SN], indent = 4))
                    del DATA['SN']['FIRST'][test][SN]

    #print(json.dumps(LastfailureremoveDupSN, indent = 4))
    #sys.exit()

    resultlist = ["PASS", "FAIL"]
    alllistofSN = list()
    for test in DATA['SN']['TEST']:
        wirtedata = list()
        wirtedata.append(test)
        listofSN = list()
        for SN in DATA['teststep'][test]["SN"]:
            recordSN = True
            if "SKIPPrefix" in inputconfig:
                if inputconfig["SKIPPrefix"] in SN:
                    recordSN = False
            if "STARTCountSN" in inputconfig:
                if SN < inputconfig["STARTCountSN"]:
                    recordSN = False
            if recordSN:
                if SN in workingonSNlist:
                    listofSN.append(SN)
                    if not SN in alllistofSN:
                        alllistofSN.append(SN)    

        wirtedata.append(len(listofSN))
        wirtedata.append(len(DATA['teststep'][test]["DETAILTESTSTEP"]))
        countResult = dict()
        countResult["PASS"] = 0
        for sn in listofSN:
            if "result" in DATA['SN']['FIRST'][test][sn]:
                testresult = DATA['SN']['FIRST'][test][sn]["result"]
                if not testresult in countResult:
                    countResult[testresult] = 1
                else:
                    countResult[testresult] += 1
        for testresult in resultlist:
            if testresult in countResult:
                wirtedata.append(countResult[testresult])
            else:
                 wirtedata.append(0)
        if len(listofSN):
            firstyeild = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
        else:
            firstyeild = "{:.2f}%".format(0)
        wirtedata.append(firstyeild)
        countResult = dict()
        countResult["PASS"] = 0
        for sn in listofSN:
            testresult = DATA['SN']['LAST'][test][sn]["result"]
            if not testresult in countResult:
                countResult[testresult] = 1
            else:
                countResult[testresult] += 1
        for testresult in resultlist:
            if testresult in countResult:
                wirtedata.append(countResult[testresult])
            else:
                 wirtedata.append(0)
        if len(listofSN):
            firstyeild = "{:.2f}%".format(countResult["PASS"]/len(listofSN) * 100)
        else:
            firstyeild = "{:.2f}%".format(0)
        wirtedata.append(firstyeild)
        wirtedata.append('')
        wirtedata.append(LastfailureremoveDupSN[test]["count"])
        ws1.append(wirtedata)

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    ws1.append(wirtedata)

    testyeildbyworkweek(DATA, ws1, alllistofSN, resultlist, status='FIRST')
    testyeildbyworkweek(DATA, ws1, alllistofSN, resultlist, status='LAST')
    testtimebyworkweek(DATA, ws1, alllistofSN, WeekTimedata)

    testyeildbySNperfixworkweek(DATA, ws1, alllistofSN, resultlist, status='FIRST')
    testyeildbySNperfixworkweek(DATA, ws1, alllistofSN, resultlist, status='LAST')

    fixcolumnssize(ws1)

    return 0

def findworkweek(checktime):
    checktime_object = datetime.strptime(checktime, '%Y-%m-%d_%H-%M-%S')
    #print(checktime)
    #print(checktime_object.isocalendar()[:2])
    #hereweekcode = week_from_date(checktime_object)
    hereweekcode = checktime_object.isocalendar()[:2]
    recordweekcode = "{}wk{:02d}".format(hereweekcode[0],hereweekcode[1])
    return recordweekcode

def findbegintesttime(checktime,testime):
    checktime_object = datetime.strptime(checktime, '%Y-%m-%d_%H-%M-%S')
    begintime_object = checktime_object - timedelta(seconds=float(testime))
    #print(checktime_object)
    #print(begintime_object)
    begintimestr = date_time = begintime_object.strftime("%Y-%m-%d_%H-%M-%S")
    #print(begintimestr)
    return begintimestr

def testtimebyworkweek(DATA, ws1, alllistofSN, WeekTimedata):
    myweeknumber = dict()
    myweeknumber['LIST'] = list()

    for teststep in WeekTimedata:
        for weeknumber in WeekTimedata[teststep]["WEEKLIST"]:
            if not weeknumber in myweeknumber['LIST']:
                myweeknumber['LIST'].append(weeknumber)
                myweeknumber['LIST'].sort()
    
    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    wirtedata.append('TEST TIME FOR MAX CARD IN MTP BY WEEKCODE')

    ws1.append(wirtedata)
    wirtedata = list()
    wirtedata.append('TEST')
    for weekcode in myweeknumber['LIST']:
        wirtedata.append(weekcode)
    ws1.append(wirtedata)
    for test in DATA['SN']['TEST']:
        wirtedataslot = list()
        wirtedatatestunits = list()
        wirtedataMAX = list()
        wirtedataAVG = list()
        wirtedataslot.append("{}_UNIT_in_MTP".format(test))
        wirtedatatestunits.append("{}_Units".format(test))
        wirtedataMAX.append("{}_MAX_TIME".format(test))
        wirtedataAVG.append("{}_AVG_TIME".format(test))
        for weekcode in myweeknumber['LIST']:
            if weekcode in WeekTimedata[test]["DATA"]:
                MaxSlotinData = WeekTimedata[test]["DATA"][weekcode]["numberofSLOTlist"][-1]
                TestunitinMaxSlotData = len(WeekTimedata[test]["DATA"][weekcode]["DATAofNumber"][MaxSlotinData])
                MaxTimeinData = max(WeekTimedata[test]["DATA"][weekcode]["DATAofNumber"][MaxSlotinData])
                from statistics import mean
                AveTimeinData = mean(WeekTimedata[test]["DATA"][weekcode]["DATAofNumber"][MaxSlotinData])
                wirtedataslot.append(MaxSlotinData)
                wirtedatatestunits.append(TestunitinMaxSlotData)
                wirtedataMAX.append(timedelta(seconds=int(MaxTimeinData)))
                wirtedataAVG.append(timedelta(seconds=int(AveTimeinData)))
                
            else:
                wirtedataslot.append("None")
                wirtedatatestunits.append("None")
                wirtedataMAX.append("None")
                wirtedataAVG.append("None")
        ws1.append(wirtedataslot)
        ws1.append(wirtedatatestunits)
        ws1.append(wirtedataMAX) 
        ws1.append(wirtedataAVG) 

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)

def testyeildbyworkweek(DATA, ws1, alllistofSN, resultlist, status='FIRST'):
    myweeknumber = dict()
    myweeknumber['LIST'] = list()
    myweeknumber['DETAIL'] = dict()

    for sn in alllistofSN:
        for test in DATA['SN']['TEST']:
            if not test in myweeknumber['DETAIL']:
                myweeknumber['DETAIL'][test] = dict()
            if sn in DATA['SN'][status][test]:
                checktime = DATA['SN'][status][test][sn]["checktime"]
                checktime_object = datetime.strptime(checktime, '%Y-%m-%d_%H-%M-%S')
                #print(checktime)
                #print(checktime_object.isocalendar()[:2])
                #hereweekcode = week_from_date(checktime_object)
                hereweekcode = checktime_object.isocalendar()[:2]
                recordweekcode = "{}wk{}".format(hereweekcode[0],hereweekcode[1])
                #print(recordweekcode)
                if not recordweekcode in myweeknumber['LIST']:
                    myweeknumber['LIST'].append(recordweekcode)
                    myweeknumber['LIST'].sort()
                if not recordweekcode in myweeknumber['DETAIL'][test]:
                    myweeknumber['DETAIL'][test][recordweekcode] = dict()
                    myweeknumber['DETAIL'][test][recordweekcode]['TOTAL'] = 0
                    myweeknumber['DETAIL'][test][recordweekcode]['PASS'] = 0
                myweeknumber['DETAIL'][test][recordweekcode]['TOTAL'] += 1
                if 'PASS' == DATA['SN'][status][test][sn]["result"]:
                    myweeknumber['DETAIL'][test][recordweekcode]['PASS'] += 1

    #print(json.dumps(myweeknumber, indent = 4))

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)
    if status == 'FIRST':
        wirtedata.append('FIRST PASS YIELD BY WEEKCODE')
    else:
        wirtedata.append('FINAL YIELD BY WEEKCODE')
    ws1.append(wirtedata)
    wirtedata = list()
    wirtedata.append('TEST')
    testendtoendyeildlist = dict()
    wirtedataendtoendyeild = list()
    wirtedataendtoendyeild.append("TEST_ENDtoEND_YIELD")
    #print(myweeknumber['LIST'])
    for weekcode in myweeknumber['LIST']:
        #wirtedata.append("{}_PASS".format(SNPerfix))
        #wirtedata.append("{}_FAIL".format(SNPerfix))
        #wirtedata.append("{}_%".format(SNPerfix))
        wirtedata.append(weekcode)
    ws1.append(wirtedata)
    for test in DATA['SN']['TEST']:
        wirtedata = list()
        wirtedatapass = list()
        wirtedatayield = list()
        #wirtedata.append(test)
        wirtedata.append("{}_TOTAL".format(test))
        wirtedatapass.append("{}_PASS".format(test))
        wirtedatayield.append("{}_YIELD".format(test))
        for weekcode in myweeknumber['LIST']:
            if not weekcode in testendtoendyeildlist:
                testendtoendyeildlist[weekcode] = list()
            if weekcode in myweeknumber['DETAIL'][test]:
                wirtedata.append(myweeknumber['DETAIL'][test][weekcode]['TOTAL'])
                if 'PASS' in myweeknumber['DETAIL'][test][weekcode]:
                    wirtedatapass.append(myweeknumber['DETAIL'][test][weekcode]['PASS'])
                    firstyield = "{:.2f}%".format(myweeknumber['DETAIL'][test][weekcode]['PASS']/myweeknumber['DETAIL'][test][weekcode]['TOTAL'] * 100)
                    firstyieldsave = myweeknumber['DETAIL'][test][weekcode]['PASS']/myweeknumber['DETAIL'][test][weekcode]['TOTAL']
                else:
                    wirtedatapass.append(0)
                    firstyield = "{:.2f}%".format(0)
                    firstyieldsave = 0                    
                wirtedatayield.append(firstyield)
                if myweeknumber['DETAIL'][test][weekcode]['TOTAL']:
                    testendtoendyeildlist[weekcode].append(firstyieldsave)
                else:
                    testendtoendyeildlist[weekcode].append(1)
            else:
                wirtedata.append(0)
                wirtedatapass.append(0)
                wirtedatayield.append("0%")  
                testendtoendyeildlist[weekcode].append(1)
        ws1.append(wirtedata)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatayield) 
    for weekcode in myweeknumber['LIST']:
        endtoendyeild = numpy.prod(testendtoendyeildlist[weekcode])
        endtoendyeilddisplay = "{:.2f}%".format(endtoendyeild * 100)
        wirtedataendtoendyeild.append(endtoendyeilddisplay)    
    ws1.append(wirtedataendtoendyeild)

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)

def testyeildbySNperfixworkweek(DATA, ws1, alllistofSN, resultlist, status='FIRST'):
    checkeachwkcodeSNlist = dict()
    checkeachwkcodeSNlist['LIST'] = list()
    checkeachwkcodeSNlist['DETAIL'] = dict()
    for sn in alllistofSN:
        if not sn[:7] in checkeachwkcodeSNlist['LIST']:
            checkeachwkcodeSNlist['LIST'].append(sn[:7])
            checkeachwkcodeSNlist['LIST'].sort()
            checkeachwkcodeSNlist['DETAIL'][sn[:7]] = list()

        if not sn in checkeachwkcodeSNlist['DETAIL'][sn[:7]]:
            checkeachwkcodeSNlist['DETAIL'][sn[:7]].append(sn)
    wirtedata = list()
    if status == 'FIRST':
        wirtedata.append('FIRST PASS YIELD BY SN WEEKCODE')
    else:
        wirtedata.append('FINAL YIELD BY SN WEEKCODE')
    ws1.append(wirtedata)
    wirtedata = list()
    wirtedata.append('TEST')
    print(checkeachwkcodeSNlist['LIST'])
    for SNPerfix in checkeachwkcodeSNlist['LIST']:
        #wirtedata.append("{}_PASS".format(SNPerfix))
        #wirtedata.append("{}_FAIL".format(SNPerfix))
        #wirtedata.append("{}_%".format(SNPerfix))
        wirtedata.append(SNPerfix)
    ws1.append(wirtedata)
    testendtoendyeildlist = dict()
    wirtedataendtoendyeild = list()
    wirtedataendtoendyeild.append("TEST_ENDtoEND_YIELD")
    for test in DATA['SN']['TEST']:
        wirtedata = list()
        wirtedatapass = list()
        wirtedatayield = list()
        #wirtedata.append(test)
        wirtedata.append("{}_TOTAL".format(test))
        wirtedatapass.append("{}_PASS".format(test))
        wirtedatayield.append("{}_YIELD".format(test))
        for SNPerfix in checkeachwkcodeSNlist['LIST']:
            if not SNPerfix in testendtoendyeildlist:
                testendtoendyeildlist[SNPerfix] = list()
            countResult = dict()
            for sn in checkeachwkcodeSNlist['DETAIL'][SNPerfix]:
                if sn in DATA['SN'][status][test]:
                    testresult = DATA['SN'][status][test][sn]["result"]
                    if not testresult in countResult:
                        countResult[testresult] = 1
                    else:
                        countResult[testresult] += 1
            total = 0
            for testresult in resultlist:
                if testresult in countResult:
                    #wirtedata.append(countResult[testresult])
                    total += countResult[testresult]
            print(countResult)
            if 'PASS' in countResult:
                wirtedatapass.append(countResult["PASS"])
                firstyieldsave = countResult["PASS"]/total
                firstyield = "{:.2f}%".format(countResult["PASS"]/total * 100)
            else:
                wirtedatapass.append(0)
                firstyieldsave = 0
                firstyield = "{:.2f}%".format(0)
            testendtoendyeildlist[SNPerfix].append(firstyieldsave)
            wirtedatayield.append(firstyield)
            wirtedata.append(total)

        ws1.append(wirtedata)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatayield)
    for SNPerfix in checkeachwkcodeSNlist['LIST']:
        endtoendyeild = numpy.prod(testendtoendyeildlist[SNPerfix])
        endtoendyeilddisplay = "{:.2f}%".format(endtoendyeild * 100)
        wirtedataendtoendyeild.append(endtoendyeilddisplay)    
    ws1.append(wirtedataendtoendyeild)

    wirtedata = list()
    ws1.append(wirtedata)
    ws1.append(wirtedata)


def week_from_date(date_object):
    date_ordinal = date_object.toordinal()
    year = date_object.year
    week = ((date_ordinal - _week1_start_ordinal(year)) // 7) + 1
    if week >= 52:
        if date_ordinal >= _week1_start_ordinal(year + 1):
            year += 1
            week = 1
    return year, week

def _week1_start_ordinal(year):
    from datetime import date
    jan1 = date(year, 1, 1)
    jan1_ordinal = jan1.toordinal()
    jan1_weekday = jan1.weekday()
    week1_start_ordinal = jan1_ordinal - ((jan1_weekday + 1) % 7)
    return week1_start_ordinal

def getallfilebyfolder(rootdir,keyword=None,level=None):

    listofdirs = getallfolder(rootdir,keyword=keyword,level=level)

    listoffiles = list()

    for eachdir in listofdirs:
        somelists = glob.glob(eachdir + '/*')
        for something in somelists:
            if os.path.isfile(something):  
                basename = os.path.basename(something)
                if keyword:
                    if not keyword.upper() in basename.upper():
                        continue
                if not something in listoffiles:
                    listoffiles.append(something)

    listoffiles.sort()

    #print(json.dumps(listoffiles, indent = 4))

    return listoffiles

def getallfolder(rootdir,keyword=None,level=None):

    listofdirs = list()

    listofdirs.append(rootdir)

    stillhavedir = True

    count = 1
    #print(level)
    #sys.exit()

    while stillhavedir:
        
        if level:
            #print(json.dumps(listofdirs, indent = 4))
            if count >= level:
                break
        stillhavedir = False
        resultlist = list()         

        for checkdir in listofdirs:
            somelists = glob.glob(checkdir + '/*')
            #print("COUNT: {} DIR: {} STILLHAVE: {}".format(count,checkdir,stillhavedir))
            for something in somelists:
                if os.path.isdir(something):
                    if not something in listofdirs:
                        stillhavedir = True
                        resultlist.append(something)
        #print("COUNT: {} TOTAL resultlist: {} STILLHAVE: {}".format(count,len(resultlist), stillhavedir))
        for eachdir in resultlist:
            if not eachdir in listofdirs:
                listofdirs.append(eachdir)

        #print("COUNT: {} TOTAL listofdirs: {} STILLHAVE: {}".format(count,len(listofdirs), stillhavedir))
        count += 1


    #print("LEVEL: {}".format(level))

    return listofdirs

def _removeIllegalCharacterError(s):
    s = re.sub('[^A-Za-z0-9 =\[\]!\(\)_,-:\.]+', '', s)
    return s    

def _removeNonAscii(s): return "".join(i for i in s if ord(i)<128)

def _removespace(s): return "".join(i for i in s if ord(i)!=32)

def _onlykeepword(s): return "".join(i for i in s if (ord(i)>48 and ord(i)<57) or (ord(i)>97 and ord(i)<122) or (ord(i)>65 and ord(i)<90))

def _remove47(s): return "".join(i for i in s if ord(i)!=47 and ord(i)!=43 and ord(i)!=58 and ord(i)!=33)

if __name__ == "__main__":
    sys.exit(main())