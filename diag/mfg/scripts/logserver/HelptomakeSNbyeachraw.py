#!/usr/bin/python3

import shutil
import json
import sys
import glob
import os
import re
import mpu.io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
import timeit
from datetime import datetime
from datetime import timedelta
import pandas as pd
import math
import numpy
from os import listdir
import time
from logdef import KEY_WORD
import modules
import csv

now = datetime.now() # current date and time
date_time = now.strftime("%Y%m%d_%H%M%S")
print("date and time:",date_time)

#TEST on KEY_WORD
#print(KEY_WORD.NIC_DIAG_TEST_RSLT_RE)

xlsxfile = "ProductDatabySerialNumberv2Results734.xlsx"

def main():

    start=datetime.now()
    pr = dict()
    pr['modules'] = modules.modules()

    sn_list = dict()
    sn_list['DATA'] = dict()
    sn_list['HEADER'] = list()
    from openpyxl import load_workbook
    wb = load_workbook(filename = xlsxfile)
    for tabname in wb.sheetnames:
        pr['modules'].debug_print("tabname: {}".format(tabname))
        eachsheet = wb[tabname]
        headers = [c.value for c in next(eachsheet.iter_rows(min_row=1, max_row=1))]
        for eachheader in headers:
            if not eachheader in sn_list['HEADER']:
                sn_list['HEADER'].append(eachheader)
    #         headerraw, headercolunm = pr['modules'].findposition(eachsheet,eachheader)
    #         colunm = eachsheet[get_column_letter(headercolunm)]
    #         for countnumberinrow in range(2,len(colunm)+1):
    #             if not str(countnumberinrow) in sn_list['DATA']:
    #                 sn_list['DATA'][str(countnumberinrow)] = dict()
    #             eachdata = str(eachsheet.cell(row=countnumberinrow, column=headercolunm).value)
    #             eachdata = eachdata.replace('\n','')
    #             if eachheader == 'Serial/Lot Numbers':
    #                 listofsn = eachdata.split("_x000D_")
    #                 eachdata = listofsn
    #             sn_list['DATA'][str(countnumberinrow)][eachheader] = eachdata

    print(sn_list['HEADER'])
    #print(sn_list['DATA']['4'])
    #sys.exit()

    filename = 'ProductDatabySerialNumberv2Results734.csv'
    with open(filename, newline='', encoding = "ISO-8859-1") as csvfile:
        reader = csv.DictReader(csvfile)
        count = 0
        for row in reader:
            count += 1 
            if not str(count) in sn_list['DATA']:
                sn_list['DATA'][str(count)] = dict()
            for eachheader in sn_list['HEADER']:
                eachdata = row[eachheader]
                eachdata = eachdata.replace('\r\n',' ')
                if eachheader == 'Serial/Lot Numbers':
                    listofsn = eachdata.split()
                    eachdata = listofsn
                
                sn_list['DATA'][str(count)][eachheader] = eachdata

    print(sn_list['HEADER'])
    #print(sn_list['DATA'])
    #sys.exit()
    ws2 = wb.create_sheet(title='NetSuite SN Ship To_2')

    wirtedata = list()
    for anotherheader in sn_list['HEADER']:
        wirtedata.append(anotherheader)
    wirtedata.append("QTYofexeclData")
    ws2.append(wirtedata)
    for eachrow in sorted(sn_list['DATA']):
        for eachsn in sorted(sn_list['DATA'][eachrow]['Serial/Lot Numbers']):
            if eachsn == 'None':
                continue
            if len(eachsn) == 0:
                continue
            wirtedata = list()
            wirtedata.append(eachsn)
            for anotherheader in sn_list['HEADER'][1:]:
                wirtedata.append(sn_list['DATA'][eachrow][anotherheader])
            wirtedata.append(len(sn_list['DATA'][eachrow]['Serial/Lot Numbers']))
            ws2.append(wirtedata)
    pr['modules'].fixcolumnssize(ws2)
    pr['modules'].freezePosition(ws2,'B2')
    wb.save(filename = "ProductDatabySerialNumberv2Results734_2.xlsx")
    difftime = datetime.now()-start
    print('Done Time: ', difftime)       
    print("How many seconds use?: {} seconds".format(difftime.total_seconds()))       
    return None


if __name__ == "__main__":
    sys.exit(main())