#!/usr/bin/python3

import shutil
import json
import sys
import glob
import os
import re
import mpu.io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
from openpyxl.styles import numbers
from openpyxl.chart import BarChart, Reference, Series, LineChart
from openpyxl.chart.label import DataLabelList
from datetime import datetime

import modules

now = datetime.now() # current date and time
date_time = now.strftime("%Y%m%d_%H%M%S")
print("date and time:",date_time)

reportpath = "/samba/public/REPORT/MAC_SN/"
historyreportpath = "/samba/public/HISTORY/MAC_SN/"
MTPslot = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10"]

def main():

    start=datetime.now()
    pr = dict()
    pr['modules'] = modules.modules()
    pr['modules'].createdirinserver(reportpath)
    pr['modules'].createdirinserver(historyreportpath)
    pr['modules'].movereporttohistorydir(reportpath,historyreportpath)
    cwd = os.getcwd()
    
    listofJson = pr['modules'].getallfilebyfolder(cwd, keyword="input.json")

    #pr['modules'].print_anyinformation(listofJson)

    jsonfiledata = dict()
    listofcurrentusingdatabase = list()

    for eachjsonfile in listofJson:
        pr['modules'].debug_print("FILE: {}".format(eachjsonfile))
        eachjsondict = pr['modules'].readjsonfile(eachjsonfile)
        jsonfiledata[eachjsonfile] = eachjsondict
        thisdatabasefile = getcurrentlymtpdatabasejsonfile(eachjsondict)
        if thisdatabasefile:
            if not thisdatabasefile in listofcurrentusingdatabase:
                listofcurrentusingdatabase.append(thisdatabasefile)

    pr['modules'].print_anyinformation(jsonfiledata)
    pr['modules'].print_anyinformation(listofcurrentusingdatabase)
    #sys.exit()

    if len(listofcurrentusingdatabase) != 1:
        print("ERROR! Have {} file for MTP DATABASE, it is not correct!".format(len(listofcurrentusingdatabase)))
        sys.exit()

    SNMACDATA = pr['modules'].readjsonfile(listofcurrentusingdatabase[0])
    #pr['modules'].print_anyinformation(MTPDATA)

    for key in SNMACDATA:
        print(key)

    makemacaddresswithsnlist(pr,SNMACDATA['macsnlist'])



    difftime = datetime.now()-start
    print('Done Time: ', difftime)       
    print("How many seconds use?: {} seconds".format(difftime.total_seconds()))       
    return None

def makemacaddresswithsnlist(pr,MACDATA):
    start2=datetime.now()
    MACDATA2 = dict()
    for family in MACDATA:
        print(family)
        for eachmac in MACDATA[family]:
            #print(eachmac)
            if not eachmac in MACDATA2:
                MACDATA2[eachmac] = dict()
                MACDATA2[eachmac]['family'] = list()
                MACDATA2[eachmac]['PN'] = list()
                MACDATA2[eachmac]['SN'] = list()
            #pr['modules'].print_anyinformation(MACDATA[family][eachmac])
            if not family in MACDATA2[eachmac]['family']:
                MACDATA2[eachmac]['family'].append(family)
            for eachPN in MACDATA[family][eachmac]['PN']:
                if not eachPN in MACDATA2[eachmac]['PN']:
                    MACDATA2[eachmac]['PN'].append(eachPN)
            for eachSN in MACDATA[family][eachmac]['SN']:
                if not eachSN in MACDATA2[eachmac]['SN']:
                    MACDATA2[eachmac]['SN'].append(eachSN)
            #pr['modules'].print_anyinformation(MACDATA2)
            #sys.exit()
        difftime2 = datetime.now()-start2
        print("How many seconds use for family <{}> with mac <{}>?: {} seconds".format(family,len(MACDATA[family]),difftime2.total_seconds())) 

    generateMACreport(pr,MACDATA2)
        
    difftime3 = datetime.now()-start2
    print("How many seconds use for makemacaddresswithsnlist?: {} seconds".format(difftime3.total_seconds())) 

def generateMACreport(pr,MACDATA):
    dest_filename3 = "{}MACvsSN_{}_DATA.xlsx".format(reportpath,date_time)
    wsandc = Workbook()
    generateexeclsummaryreport(pr,wsandc,MACDATA)
    generateexeclmacandsnreport3(pr,wsandc,MACDATA)
    wsandc.save(filename = dest_filename3)

def generateexeclsummaryreport(pr,wb,snmaclist):
    ws1 = wb.active
    ws1.title = "SUMMARY"

    MACSUMMARY = dict()
    for mac in snmaclist:
        numberofSN = str(len(snmaclist[mac]['SN']))
        if not numberofSN in MACSUMMARY:
            MACSUMMARY[numberofSN] = list()
        MACSUMMARY[numberofSN].append(mac)

    wirtedata = list()
    wirtedata.append('Number of SN in one MAC')
    wirtedata.append('How many MAC?')
    ws1.append(wirtedata)

    for numberofSN in sorted(MACSUMMARY,reverse=True):
        print(numberofSN)
        wirtedata = list()
        wirtedata.append(numberofSN)
        wirtedata.append(len(MACSUMMARY[numberofSN]))
        ws1.append(wirtedata)

    pr['modules'].fixcolumnssize(ws1)

    for numberofSN in sorted(MACSUMMARY,reverse=True):
        if numberofSN == '1':
            continue
        generateexeclmacandsnreport4(pr,wb,numberofSN,MACSUMMARY[numberofSN],snmaclist)

    pr['modules'].fixcolumnssize(ws1)

    return True

def generateexeclmacandsnreport4(pr,wb,numberofSN,maclist,snmaclist):
    ws1 = wb.create_sheet(title="{} in One MAC".format(numberofSN))

    wirtedata = list()
    wirtedata.append('MAC')
    wirtedata.append('PN')
    wirtedata.append('# of PN')
    wirtedata.append('SN')
    ws1.append(wirtedata)

    for mac in maclist:
        wirtedata = list()
        wirtedata.append(mac.replace('-',''))        
        if len(snmaclist[mac]['PN']):
            PN = snmaclist[mac]['PN'][0]
            if len(snmaclist[mac]['PN']) > 1:
                for eachPN in snmaclist[mac]['PN'][1:]:
                    PN += '\r'
                    PN += eachPN
            wirtedata.append(PN)
        wirtedata.append(len(snmaclist[mac]['PN']))
        if len(snmaclist[mac]['SN']):
            for sn in snmaclist[mac]['SN']:
                wirtedata.append(sn)   
        ws1.append(wirtedata)

    pr['modules'].wraptest(ws1)
    pr['modules'].fixcolumnssize(ws1)


    return True

def generateexeclmacandsnreport3(pr,wb,snmaclist):
    ws1 = wb.create_sheet(title="ALLDATA")

    wirtedata = list()
    wirtedata.append('MAC')
    wirtedata.append('PN')
    wirtedata.append('# of PN')
    wirtedata.append('SN')
    ws1.append(wirtedata)

    for mac in snmaclist:
        wirtedata = list()
        wirtedata.append(mac.replace('-',''))        
        if len(snmaclist[mac]['PN']):
            PN = snmaclist[mac]['PN'][0]
            if len(snmaclist[mac]['PN']) > 1:
                for eachPN in snmaclist[mac]['PN'][1:]:
                    PN += '\r'
                    PN += eachPN
            wirtedata.append(PN)
        wirtedata.append(len(snmaclist[mac]['PN']))
        if len(snmaclist[mac]['SN']):
            for sn in snmaclist[mac]['SN']:
                wirtedata.append(sn)   
        ws1.append(wirtedata)

    pr['modules'].wraptest(ws1)
    pr['modules'].fixcolumnssize(ws1)

    return True


def getcurrentlymtpdatabasejsonfile(eachjsondict):
    if 'FILE' in eachjsondict:
        if 'snandmacjsonfile' in eachjsondict['FILE']:
            return eachjsondict['FILE']["snandmacjsonfile"]

    return None

if __name__ == "__main__":
    sys.exit(main())