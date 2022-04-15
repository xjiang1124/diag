#!/usr/bin/python3

import shutil
import json
import sys
import glob
import os
import re
import mpu.io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
from openpyxl.styles import numbers
from openpyxl.chart import BarChart, Reference, Series, LineChart
from openpyxl.chart.label import DataLabelList
from datetime import datetime

import modules

now = datetime.now() # current date and time
date_time = now.strftime("%Y%m%d_%H%M%S")
print("date and time:",date_time)

reportpath = "/samba/public/REPORT/MTP/"
historyreportpath = "/samba/public/HISTORY/MTP/"
MTPslot = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10"]

def main():

    start=datetime.now()
    pr = dict()
    pr['modules'] = modules.modules()
    pr['modules'].createdirinserver(reportpath)
    pr['modules'].createdirinserver(historyreportpath)
    pr['modules'].movereporttohistorydir(reportpath,historyreportpath)
    cwd = os.getcwd()
    
    listofJson = pr['modules'].getallfilebyfolder(cwd, keyword="input.json")

    #pr['modules'].print_anyinformation(listofJson)

    jsonfiledata = dict()
    listofcurrentusingdatabase = list()

    for eachjsonfile in listofJson:
        pr['modules'].debug_print("FILE: {}".format(eachjsonfile))
        eachjsondict = pr['modules'].readjsonfile(eachjsonfile)
        jsonfiledata[eachjsonfile] = eachjsondict
        thisdatabasefile = getcurrentlymtpdatabasejsonfile(eachjsondict)
        if thisdatabasefile:
            if not thisdatabasefile in listofcurrentusingdatabase:
                listofcurrentusingdatabase.append(thisdatabasefile)

    pr['modules'].print_anyinformation(jsonfiledata)
    pr['modules'].print_anyinformation(listofcurrentusingdatabase)
    #sys.exit()

    if len(listofcurrentusingdatabase) != 1:
        print("ERROR! Have {} file for MTP DATABASE, it is not correct!".format(len(listofcurrentusingdatabase)))
        sys.exit()

    MTPDATA = pr['modules'].readjsonfile(listofcurrentusingdatabase[0])
    #pr['modules'].print_anyinformation(MTPDATA)

    generateMTPallreport(pr,MTPDATA,startdate=None)
    generateMTPallreport(pr,MTPDATA,startdate=pr['modules'].getbeforedayinformation(checkday=31))
    generateMTPallreport(pr,MTPDATA,startdate=pr['modules'].getbeforedayinformation(checkday=7))
    generateMTPallreport2(pr,MTPDATA,startdate=None)

    difftime = datetime.now()-start
    print('Done Time: ', difftime)       
    print("How many seconds use?: {} seconds".format(difftime.total_seconds()))       
    return None

def generateMTPallreport(pr,MTPDATA,startdate=None):
    wb3 = Workbook()
    dest_filename3 = "{}MTP_STATUS_{}_DATA.xlsx".format(reportpath,date_time)
    if startdate:
        dest_filename3 = "{}MTP_STATUS_{}_DATA_StartWith_{}.xlsx".format(reportpath,date_time,startdate)
    print(dest_filename3)

    mtpchassisusecountbyslot = generateexeclallmtpstatussummaryreport(pr,MTPDATA,wb3,startdate)

    for mtp in sorted(mtpchassisusecountbyslot):
        generateexeclallmtpstatusbyeachmtpreport(pr,mtpchassisusecountbyslot,MTPDATA,wb3,mtp,startdate)

    print(dest_filename3)
    wb3.save(filename = dest_filename3) 

    return None

def generateMTPallreport2(pr,MTPDATA,startdate=None):
    wb3 = Workbook()
    dest_filename3 = "{}MTP_STATUS_{}_DATA2.xlsx".format(reportpath,date_time)
    if startdate:
        dest_filename3 = "{}MTP_STATUS_{}_DATA2_StartWith_{}.xlsx".format(reportpath,date_time,startdate)
    print(dest_filename3)

    mtpchassisusecountbyslot = generateexeclallmtpstatussummaryreport(pr,MTPDATA,wb3,startdate)

    for mtp in sorted(mtpchassisusecountbyslot):
        generateexeclallmtpstatusbyeachmtpreport2(pr,mtpchassisusecountbyslot,MTPDATA,wb3,mtp,startdate)

    print(dest_filename3)
    wb3.save(filename = dest_filename3) 

    return None

def generateexeclallmtpstatusbyeachmtpreport(pr,mtpchassisusecountbyslot,DATA,wb,mtp,startdate=None):
    print("generateexeclallmtpstatusbyeachmtpreport: {}".format(mtp))
    titlename = mtp
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append("FAMILY")
    wirtedata.append("TEST")
    wirtedata.append("DATE_TIME")
    for slot in MTPslot:
        wirtedata.append(slot)
    wirtedata.append("FAILURE_TYPE")
    for slot in MTPslot:
        wirtedata.append(slot)
    wirtedata.append("SN")
    for slot in MTPslot:
        wirtedata.append(slot)
    ws2.append(wirtedata)

    for datetime in mtpchassisusecountbyslot[mtp]["datetimelist"]:
        for family in DATA:
            if mtp in DATA[family]:
                for test in DATA[family][mtp]:
                    if datetime in DATA[family][mtp][test]:
                        wirtedata = list()
                        wirtedata.append(family)
                        wirtedata.append(test)
                        wirtedata.append(datetime)
                        #pr['modules'].print_anyinformation(DATA[family][mtp][test][datetime])
                        #sys.exit()
                        for slot in MTPslot:

                            if slot in DATA[family][mtp][test][datetime]["NICRESULT"]:
                                wirtedata.append(DATA[family][mtp][test][datetime]["NICRESULT"][slot]["RESULT"])
                            else:
                                wirtedata.append("None")

                        wirtedata.append("||||||||||")
                        for slot in MTPslot:

                            if not "FAILURESTEP" in DATA[family][mtp][test][datetime]:
                                wirtedata.append("No Data")
                                continue
                                
                            if slot in DATA[family][mtp][test][datetime]["FAILURESTEP"]:
                                wirtedata.append(DATA[family][mtp][test][datetime]["FAILURESTEP"][slot])
                            else:
                                wirtedata.append("None")
                        wirtedata.append("||||||||||")
                        for slot in MTPslot:

                            if slot in DATA[family][mtp][test][datetime]["NICRESULT"]:
                                if "SN" in DATA[family][mtp][test][datetime]["NICRESULT"][slot]:
                                    wirtedata.append(DATA[family][mtp][test][datetime]["NICRESULT"][slot]["SN"])
                                else:
                                    wirtedata.append("None")
                            else:
                                wirtedata.append("None")
                        ws2.append(wirtedata)

    pr['modules'].highlightinred(ws2, 'FAIL')
    pr['modules'].highlightinred(ws2, 'FAILED')
    pr['modules'].highlightingreen(ws2,'PASS')      
    pr['modules'].fixcolumnssize(ws2)
    pr['modules'].freezePosition(ws2,'D2')

    return True


def generateexeclallmtpstatusbyeachmtpreport2(pr,mtpchassisusecountbyslot,DATA,wb,mtp,startdate=None):
    print("generateexeclallmtpstatusbyeachmtpreport2: {}".format(mtp))
    titlename = mtp
    ws2 = wb.create_sheet(title=titlename)
    wirtedata = list()
    wirtedata.append("FAMILY")
    wirtedata.append("TEST")
    wirtedata.append("DATE_TIME")
    for slot in MTPslot:
        wirtedata.append(slot)
        wirtedata.append(slot)
        wirtedata.append(slot)
        wirtedata.append("|||")

    ws2.append(wirtedata)

    for datetime in mtpchassisusecountbyslot[mtp]["datetimelist"]:
        for family in DATA:
            if mtp in DATA[family]:
                for test in DATA[family][mtp]:
                    if datetime in DATA[family][mtp][test]:
                        wirtedata = list()
                        wirtedata.append(family)
                        wirtedata.append(test)
                        wirtedata.append(datetime)
                        #pr['modules'].print_anyinformation(DATA[family][mtp][test][datetime])
                        #sys.exit()
                        for slot in MTPslot:

                            if slot in DATA[family][mtp][test][datetime]["NICRESULT"]:
                                wirtedata.append(DATA[family][mtp][test][datetime]["NICRESULT"][slot]["RESULT"])
                            else:
                                wirtedata.append("None")

                            if not "FAILURESTEP" in DATA[family][mtp][test][datetime]:
                                wirtedata.append("No Data")
                                continue
                                
                            if slot in DATA[family][mtp][test][datetime]["FAILURESTEP"]:
                                wirtedata.append(DATA[family][mtp][test][datetime]["FAILURESTEP"][slot])
                            else:
                                wirtedata.append("None")

                            if slot in DATA[family][mtp][test][datetime]["NICRESULT"]:
                                if "SN" in DATA[family][mtp][test][datetime]["NICRESULT"][slot]:
                                    wirtedata.append(DATA[family][mtp][test][datetime]["NICRESULT"][slot]["SN"])
                                else:
                                    wirtedata.append("None")
                            else:
                                wirtedata.append("None")

                            wirtedata.append("|||")

                        ws2.append(wirtedata)

    pr['modules'].highlightinred(ws2, 'FAIL')
    pr['modules'].highlightinred(ws2, 'FAILED')
    pr['modules'].highlightingreen(ws2,'PASS')      
    pr['modules'].fixcolumnssize(ws2)
    pr['modules'].freezePosition(ws2,'D2')

    return True


def generateexeclallmtpstatussummaryreport(pr,DATA,wb,startdate=None):
    ws1 = wb.active
    ws1.title = "SUMMARY"

    mtpchassisusecountbyslot = dict()
    mtpchassisusecountbytest = dict()
    mtpchassisusecountbytest['TESTLIST'] = list()
    mtpchassisusecountbytest['STATUS'] = dict()
    mtpchassisusecountbyteststep = dict()
    mtpchassisusecountbyteststep['TESTSTEPLIST'] = list()
    mtpchassisusecountbyteststep['STATUS'] = dict()

    for family in DATA:
        for MTP in DATA[family]:
            if not "MTP" in MTP:
                continue
            for test in DATA[family][MTP]:
                for datetime in DATA[family][MTP][test]:
                    if startdate:
                        datelist = datetime.split("_")
                        if datelist[0] < startdate:
                            continue
                    for eachslot in DATA[family][MTP][test][datetime]["NICRESULT"]:
                        if eachslot in DATA[family][MTP][test][datetime]["FAILURESTEP"]:
                            if 'CANNOT FIND' in DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]:
                                continue
                            if 'PASS' in DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]:
                                continue
                            if not DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot] in mtpchassisusecountbyteststep['TESTSTEPLIST']:
                                mtpchassisusecountbyteststep['TESTSTEPLIST'].append(DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot])
                                mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]] = dict()
                            if not MTP in mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]]:
                                mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP] = dict()
                                mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["TEST"] = list()
                                mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["FAMILY"] = list()
                                mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["datetimelist"] = list()
                                for slot in MTPslot:
                                    mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP][slot] = dict()
                                    mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP][slot]["TOTAL"] = 0
                                    mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP][slot]["PASS"] = 0
                                    mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP][slot]["FAIL"] = 0
    #pr['modules'].print_anyinformation(mtpchassisusecountbyteststep)    

    #sys.exit()

    for family in DATA:
        print(family)
        #print(DATA[family])
        if len(DATA[family]) == 0:
            continue

        for MTP in DATA[family]:
            print(MTP)
            if not "MTP" in MTP:
                continue
            if not MTP in mtpchassisusecountbyslot:
                mtpchassisusecountbyslot[MTP] = dict()
                mtpchassisusecountbyslot[MTP]["TEST"] = list()
                mtpchassisusecountbyslot[MTP]["FAMILY"] = list()
                mtpchassisusecountbyslot[MTP]["datetimelist"] = list()
                for slot in MTPslot:
                    mtpchassisusecountbyslot[MTP][slot] = dict()
                    mtpchassisusecountbyslot[MTP][slot]["TOTAL"] = 0
                    mtpchassisusecountbyslot[MTP][slot]["PASS"] = 0
                    mtpchassisusecountbyslot[MTP][slot]["FAIL"] = 0
            for test in DATA[family][MTP]:
                #print(test)
                if not test in mtpchassisusecountbytest['TESTLIST']:
                    mtpchassisusecountbytest['TESTLIST'].append(test)
                if not test in mtpchassisusecountbytest['STATUS']:
                    mtpchassisusecountbytest['STATUS'][test] = dict()
                if not MTP in mtpchassisusecountbytest['STATUS'][test]:
                    mtpchassisusecountbytest['STATUS'][test][MTP] = dict()
                    mtpchassisusecountbytest['STATUS'][test][MTP]["TEST"] = list()
                    mtpchassisusecountbytest['STATUS'][test][MTP]["FAMILY"] = list()
                    mtpchassisusecountbytest['STATUS'][test][MTP]["datetimelist"] = list()
                    for slot in MTPslot:
                        mtpchassisusecountbytest['STATUS'][test][MTP][slot] = dict()
                        mtpchassisusecountbytest['STATUS'][test][MTP][slot]["TOTAL"] = 0
                        mtpchassisusecountbytest['STATUS'][test][MTP][slot]["PASS"] = 0
                        mtpchassisusecountbytest['STATUS'][test][MTP][slot]["FAIL"] = 0
                for datetime in DATA[family][MTP][test]:
                    #print(datetime)
                    #print(json.dumps(DATA["MTPCHASSIS"][MTP][test][datetime], indent = 4))
                    if startdate:
                        datelist = datetime.split("_")
                        if datelist[0] < startdate:
                            continue
                    if not test in mtpchassisusecountbyslot[MTP]["TEST"]:
                        mtpchassisusecountbyslot[MTP]["TEST"].append(test)
                        mtpchassisusecountbyslot[MTP]["TEST"].sort()
                    if not family in mtpchassisusecountbyslot[MTP]["FAMILY"]:
                        mtpchassisusecountbyslot[MTP]["FAMILY"].append(family)
                        mtpchassisusecountbyslot[MTP]["FAMILY"].sort()
                    if not datetime in mtpchassisusecountbyslot[MTP]["datetimelist"]:
                        mtpchassisusecountbyslot[MTP]["datetimelist"].append(datetime)
                        mtpchassisusecountbyslot[MTP]["datetimelist"].sort(reverse=True)
                    if not test in mtpchassisusecountbytest['STATUS'][test][MTP]["TEST"]:
                        mtpchassisusecountbytest['STATUS'][test][MTP]["TEST"].append(test)
                        mtpchassisusecountbytest['STATUS'][test][MTP]["TEST"].sort()
                    if not family in mtpchassisusecountbytest['STATUS'][test][MTP]["FAMILY"]:
                        mtpchassisusecountbytest['STATUS'][test][MTP]["FAMILY"].append(family)
                        mtpchassisusecountbytest['STATUS'][test][MTP]["FAMILY"].sort()
                    if not datetime in mtpchassisusecountbytest['STATUS'][test][MTP]["datetimelist"]:
                        mtpchassisusecountbytest['STATUS'][test][MTP]["datetimelist"].append(datetime)
                        mtpchassisusecountbytest['STATUS'][test][MTP]["datetimelist"].sort(reverse=True)
                    for eachslot in DATA[family][MTP][test][datetime]["NICRESULT"]:
                        if "PASS" in DATA[family][MTP][test][datetime]["NICRESULT"][eachslot]["RESULT"]:
                            mtpchassisusecountbyslot[MTP][eachslot]["TOTAL"] += 1
                            mtpchassisusecountbyslot[MTP][eachslot]["PASS"] += 1
                        else:
                            mtpchassisusecountbyslot[MTP][eachslot]["TOTAL"] += 1
                            mtpchassisusecountbyslot[MTP][eachslot]["FAIL"] += 1
                        if "PASS" in DATA[family][MTP][test][datetime]["NICRESULT"][eachslot]["RESULT"]:
                            mtpchassisusecountbytest['STATUS'][test][MTP][eachslot]["TOTAL"] += 1
                            mtpchassisusecountbytest['STATUS'][test][MTP][eachslot]["PASS"] += 1
                        else:
                            mtpchassisusecountbytest['STATUS'][test][MTP][eachslot]["TOTAL"] += 1
                            mtpchassisusecountbytest['STATUS'][test][MTP][eachslot]["FAIL"] += 1
                        if eachslot in DATA[family][MTP][test][datetime]["FAILURESTEP"]:
                            if 'CANNOT FIND' in DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]:
                                continue
                            if 'PASS' in DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]:
                                continue
                        if not test in mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["TEST"]:
                            mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["TEST"].append(test)
                            mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["TEST"].sort()
                        if not family in mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["FAMILY"]:
                            mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["FAMILY"].append(family)
                            mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["FAMILY"].sort()
                        if not datetime in mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["datetimelist"]:
                            mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["datetimelist"].append(datetime)
                            mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP]["datetimelist"].sort(reverse=True)
                        mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP][eachslot]["TOTAL"] = mtpchassisusecountbyslot[MTP][eachslot]["TOTAL"]
                        mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP][eachslot]["PASS"] = mtpchassisusecountbyslot[MTP][eachslot]["PASS"]
                        mtpchassisusecountbyteststep['STATUS'][DATA[family][MTP][test][datetime]["FAILURESTEP"][eachslot]][MTP][eachslot]["FAIL"] += 1                                

            #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
            #sys.exit()
        #break
    findzeroMTP = list()

    for mtp in sorted(mtpchassisusecountbyslot):
        listoftotal = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
        if sum(listoftotal) == 0:
            findzeroMTP.append(mtp)
    
    for mtp in findzeroMTP:
        del mtpchassisusecountbyslot[mtp]
    #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
    #sys.exit()
    wirtedata = list()
    wirtedata.append('MTP')
    wirtedata.append('TEST')
    wirtedata.append('STATUS')
    wirtedata.append('TOTAL')
    for slot in MTPslot:
        wirtedata.append(slot)
    wirtedata.append('FAMILY')
    ws1.append(wirtedata)
    for mtp in sorted(mtpchassisusecountbyslot):
        wirtedatatotal = list()
        wirtedatapass = list()
        wirtedatafail = list()
        wirtedatapassrate = list()
        wirtedatatotal.append(mtp)
        wirtedatapass.append(mtp)
        wirtedatafail.append(mtp)
        wirtedatapassrate.append(mtp)
        teststep = mtpchassisusecountbyslot[mtp]["TEST"][0]
        if len(mtpchassisusecountbyslot[mtp]["TEST"]) > 1:
            for eachtest in mtpchassisusecountbyslot[mtp]["TEST"][1:]:
                teststep = "{},{}".format(teststep,eachtest)
        allfamily = mtpchassisusecountbyslot[mtp]["FAMILY"][0]
        if len(mtpchassisusecountbyslot[mtp]["FAMILY"]) > 1:
            for eachfamily in mtpchassisusecountbyslot[mtp]["FAMILY"][1:]:
                allfamily = "{}, {}".format(allfamily,eachfamily)
        wirtedatatotal.append(teststep)
        wirtedatapass.append(teststep)
        wirtedatafail.append(teststep)
        wirtedatapassrate.append(teststep)
        listoftotal = list()
        listofpass = list()
        listoffail = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            listofpass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            listoffail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])  
        wirtedatatotal.append("TOTAL")
        wirtedatapass.append("PASS")
        wirtedatafail.append("FAIL")
        wirtedatapassrate.append("PASS_RATE")
        wirtedatatotal.append(sum(listoftotal))
        wirtedatapass.append(sum(listofpass))
        wirtedatafail.append(sum(listoffail))
        wirtedatapassrate.append(pr['modules'].calculePass_rate(sum(listoftotal),sum(listofpass)))           
        for slot in MTPslot:
            wirtedatatotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            wirtedatapass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            wirtedatafail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])
            wirtedatapassrate.append(pr['modules'].calculePass_rate(mtpchassisusecountbyslot[mtp][slot]["TOTAL"],mtpchassisusecountbyslot[mtp][slot]["PASS"]))
        wirtedatatotal.append(allfamily)
        wirtedatapass.append(allfamily)
        wirtedatafail.append(allfamily)
        wirtedatapassrate.append(allfamily)
        ws1.append(wirtedatatotal)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatafail)
        ws1.append(wirtedatapassrate)          
    pr['modules'].fixcolumnssize3(ws1)
    pr['modules'].highlightinlightredrow(ws1,'FAIL')
    pr['modules'].converttoPERCENTAGEnumber(ws1)
    pr['modules'].freezePosition(ws1,'D2')

    for teststep in mtpchassisusecountbytest['TESTLIST']:
        generateexeclallmtpstatussummaryreportbytest(pr,DATA,wb,teststep,mtpchassisusecountbytest['STATUS'][teststep],startdate)

    createMTPteststepfailurereport(pr,DATA,mtpchassisusecountbyteststep,startdate)

    return mtpchassisusecountbyslot

def createMTPteststepfailurereport(pr,DATA,mtpchassisusecountbyteststep,startdate=None):
    wb3 = Workbook()
    dest_filename5 = "{}MTP_FAILSTEP_{}_DATA.xlsx".format(reportpath,date_time)
    if startdate:
        dest_filename5 = "{}MTP_FAILSTEP_{}_DATA_StartWith_{}.xlsx".format(reportpath,date_time,startdate)
    print(dest_filename5)

    mtpchassisusecountbyslot = generateexeclallmtpfailurestepsummaryreport(pr,mtpchassisusecountbyteststep,wb3,startdate)

    for eachfailurestep in mtpchassisusecountbyteststep['TESTSTEPLIST']:
        generateexeclallmtpstatussummaryreportbyfailurestep(pr,DATA,wb3,eachfailurestep,mtpchassisusecountbyteststep['STATUS'][eachfailurestep],startdate)

    print(dest_filename5)
    wb3.save(filename = dest_filename5) 

def generateexeclallmtpfailurestepsummaryreport(pr,mtpchassisusecountbyteststep,wb,startdate=None):
    ws1 = wb.active
    ws1.title = "SUMMARY"
    for eachfailurestep in mtpchassisusecountbyteststep['TESTSTEPLIST']:
        wirtedata = list()
        wirtedata.append(eachfailurestep)
        ws1.append(wirtedata)

    return None

def generateexeclallmtpstatussummaryreportbyfailurestep(pr,DATA,wb,teststep,mtpchassisusecountbyslot,startdate=None):
    print("generateexeclallmtpstatussummaryreportbyfailurestep: {}".format(teststep))
    titlename = teststep
    ws1 = wb.create_sheet(title=titlename[:30])
    findzeroMTP = list()

    #pr['modules'].print_anyinformation(mtpchassisusecountbyslot)  

    #sys.exit()

    for mtp in sorted(mtpchassisusecountbyslot):
        listoftotal = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
        if sum(listoftotal) == 0:
            findzeroMTP.append(mtp)
    
    for mtp in findzeroMTP:
        del mtpchassisusecountbyslot[mtp]
    #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
    #sys.exit()
    wirtedata = list()
    wirtedata.append('MTP')
    wirtedata.append('TEST')
    wirtedata.append('STATUS')
    wirtedata.append('TOTAL')
    for slot in MTPslot:
        wirtedata.append(slot)
    wirtedata.append('FAMILY')
    ws1.append(wirtedata)
    for mtp in sorted(mtpchassisusecountbyslot):
        wirtedatatotal = list()
        wirtedatapass = list()
        wirtedatafail = list()
        wirtedatapassrate = list()
        wirtedatatotal.append(mtp)
        wirtedatapass.append(mtp)
        wirtedatafail.append(mtp)
        wirtedatapassrate.append(mtp)
        teststep = mtpchassisusecountbyslot[mtp]["TEST"][0]
        if len(mtpchassisusecountbyslot[mtp]["TEST"]) > 1:
            for eachtest in mtpchassisusecountbyslot[mtp]["TEST"][1:]:
                teststep = "{},{}".format(teststep,eachtest)
        allfamily = mtpchassisusecountbyslot[mtp]["FAMILY"][0]
        if len(mtpchassisusecountbyslot[mtp]["FAMILY"]) > 1:
            for eachfamily in mtpchassisusecountbyslot[mtp]["FAMILY"][1:]:
                allfamily = "{}, {}".format(allfamily,eachfamily)
        wirtedatatotal.append(teststep)
        wirtedatapass.append(teststep)
        wirtedatafail.append(teststep)
        wirtedatapassrate.append(teststep)
        listoftotal = list()
        listofpass = list()
        listoffail = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            listofpass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            listoffail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])  
        wirtedatatotal.append("TOTAL")
        wirtedatapass.append("PASS")
        wirtedatafail.append("FAIL")
        wirtedatapassrate.append("FAIL_RATE")
        wirtedatatotal.append(sum(listoftotal))
        wirtedatapass.append(sum(listofpass))
        wirtedatafail.append(sum(listoffail))
        wirtedatapassrate.append(pr['modules'].calculeFail_rate(sum(listoftotal),sum(listoffail)))           
        for slot in MTPslot:
            wirtedatatotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            wirtedatapass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            wirtedatafail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])
            wirtedatapassrate.append(pr['modules'].calculeFail_rate(mtpchassisusecountbyslot[mtp][slot]["TOTAL"],mtpchassisusecountbyslot[mtp][slot]["FAIL"]))
        wirtedatatotal.append(allfamily)
        wirtedatapass.append(allfamily)
        wirtedatafail.append(allfamily)
        wirtedatapassrate.append(allfamily)
        ws1.append(wirtedatatotal)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatafail)
        ws1.append(wirtedatapassrate)          
    pr['modules'].fixcolumnssize3(ws1)
    pr['modules'].highlightinlightredrow(ws1,'FAIL')
    pr['modules'].converttoPERCENTAGEnumberbyFailurerange(ws1)
    pr['modules'].freezePosition(ws1,'D2')

    return None


def generateexeclallmtpstatussummaryreportbytest(pr,DATA,wb,teststep,mtpchassisusecountbyslot,startdate=None):
    print("generateexeclallmtpstatussummaryreport: {}".format(teststep))
    titlename = teststep
    ws1 = wb.create_sheet(title=titlename)
    findzeroMTP = list()

    for mtp in sorted(mtpchassisusecountbyslot):
        listoftotal = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
        if sum(listoftotal) == 0:
            findzeroMTP.append(mtp)
    
    for mtp in findzeroMTP:
        del mtpchassisusecountbyslot[mtp]
    #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
    #sys.exit()
    wirtedata = list()
    wirtedata.append('MTP')
    wirtedata.append('TEST')
    wirtedata.append('STATUS')
    wirtedata.append('TOTAL')
    for slot in MTPslot:
        wirtedata.append(slot)
    wirtedata.append('FAMILY')
    ws1.append(wirtedata)
    for mtp in sorted(mtpchassisusecountbyslot):
        wirtedatatotal = list()
        wirtedatapass = list()
        wirtedatafail = list()
        wirtedatapassrate = list()
        wirtedatatotal.append(mtp)
        wirtedatapass.append(mtp)
        wirtedatafail.append(mtp)
        wirtedatapassrate.append(mtp)
        teststep = mtpchassisusecountbyslot[mtp]["TEST"][0]
        if len(mtpchassisusecountbyslot[mtp]["TEST"]) > 1:
            for eachtest in mtpchassisusecountbyslot[mtp]["TEST"][1:]:
                teststep = "{},{}".format(teststep,eachtest)
        allfamily = mtpchassisusecountbyslot[mtp]["FAMILY"][0]
        if len(mtpchassisusecountbyslot[mtp]["FAMILY"]) > 1:
            for eachfamily in mtpchassisusecountbyslot[mtp]["FAMILY"][1:]:
                allfamily = "{}, {}".format(allfamily,eachfamily)
        wirtedatatotal.append(teststep)
        wirtedatapass.append(teststep)
        wirtedatafail.append(teststep)
        wirtedatapassrate.append(teststep)
        listoftotal = list()
        listofpass = list()
        listoffail = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            listofpass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            listoffail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])  
        wirtedatatotal.append("TOTAL")
        wirtedatapass.append("PASS")
        wirtedatafail.append("FAIL")
        wirtedatapassrate.append("PASS_RATE")
        wirtedatatotal.append(sum(listoftotal))
        wirtedatapass.append(sum(listofpass))
        wirtedatafail.append(sum(listoffail))
        wirtedatapassrate.append(pr['modules'].calculePass_rate(sum(listoftotal),sum(listofpass)))           
        for slot in MTPslot:
            wirtedatatotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            wirtedatapass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            wirtedatafail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])
            wirtedatapassrate.append(pr['modules'].calculePass_rate(mtpchassisusecountbyslot[mtp][slot]["TOTAL"],mtpchassisusecountbyslot[mtp][slot]["PASS"]))
        wirtedatatotal.append(allfamily)
        wirtedatapass.append(allfamily)
        wirtedatafail.append(allfamily)
        wirtedatapassrate.append(allfamily)
        ws1.append(wirtedatatotal)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatafail)
        ws1.append(wirtedatapassrate)          
    pr['modules'].fixcolumnssize3(ws1)
    pr['modules'].converttoPERCENTAGEnumber(ws1)
    pr['modules'].highlightinlightredrow(ws1,'FAIL')
    pr['modules'].freezePosition(ws1,'D2')

    return None

def getcurrentlymtpdatabasejsonfile(eachjsondict):
    if 'FILE' in eachjsondict:
        if 'mtpdatabasejsonfile' in eachjsondict['FILE']:
            return eachjsondict['FILE']["mtpdatabasejsonfile"]

    return None

if __name__ == "__main__":
    sys.exit(main())