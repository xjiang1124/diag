#!/usr/bin/python3

import shutil
import json
import sys
import glob
import os
import re
import mpu.io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
from openpyxl.styles import numbers
from openpyxl.chart import BarChart, Reference, Series, LineChart
from openpyxl.chart.label import DataLabelList
from datetime import datetime

import modules

now = datetime.now() # current date and time
date_time = now.strftime("%Y%m%d_%H%M%S")
print("date and time:",date_time)

reportpath = "/samba/public/REPORT/SW/"
historyreportpath = "/samba/public/HISTORY/SW/"
MTPslot = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10"]

def main():

    start=datetime.now()
    pr = dict()
    pr['modules'] = modules.modules()
    pr['modules'].createdirinserver(reportpath)
    pr['modules'].createdirinserver(historyreportpath)
    pr['modules'].movereporttohistorydir(reportpath,historyreportpath)
    cwd = os.getcwd()
    
    listofJson = pr['modules'].getallfilebyfolder(cwd, keyword="input.json")

    #pr['modules'].print_anyinformation(listofJson)

    jsonfiledata = dict()
    listofcurrentusingdatabase = list()

    for eachjsonfile in listofJson:
        pr['modules'].debug_print("FILE: {}".format(eachjsonfile))
        eachjsondict = pr['modules'].readjsonfile(eachjsonfile)
        jsonfiledata[eachjsonfile] = eachjsondict
        thisdatabasefile = getcurrentlyswdatabasejsonfile(eachjsondict)
        if thisdatabasefile:
            if not thisdatabasefile in listofcurrentusingdatabase:
                listofcurrentusingdatabase.append(thisdatabasefile)

    pr['modules'].print_anyinformation(jsonfiledata)
    pr['modules'].print_anyinformation(listofcurrentusingdatabase)
    #sys.exit()

    if len(listofcurrentusingdatabase) != 1:
        print("ERROR! Have {} file for SW DATABASE, it is not correct!".format(len(listofcurrentusingdatabase)))
        sys.exit()

    SWDATA = pr['modules'].readjsonfile(listofcurrentusingdatabase[0])
    #pr['modules'].print_anyinformation(MTPDATA)

    generateSWallreport(pr,SWDATA,startdate=None)

    difftime = datetime.now()-start
    print('Done Time: ', difftime)       
    print("How many seconds use?: {} seconds".format(difftime.total_seconds()))       
    return None

def generateSWallreport(pr,SWDATA,startdate=None):
    wb3 = Workbook()
    dest_filename3 = "{}SW_{}_DATA.xlsx".format(reportpath,date_time)
    if startdate:
        dest_filename3 = "{}SW_{}_DATA_StartWith_{}.xlsx".format(reportpath,date_time,startdate)
    print(dest_filename3)

    generateexeclallswstatussummaryreport(pr,SWDATA,wb3,startdate)

    for family in sorted(SWDATA["status"]["family"]):
        generateexeclallswstatusbyeachfamilyreport(pr,SWDATA["data"][family],wb3,family,startdate)

    print(dest_filename3)
    wb3.save(filename = dest_filename3) 

    return None

def generateexeclallswstatusbyeachfamilyreport(pr,lastdata,wb3,family,startdate=None):
    print("generateexeclallswstatusbyeachfamilyreport: {}".format(family))
    titlename = family
    ws2 = wb3.create_sheet(title=family)

    imagenamelist = list()
    for SN in sorted(lastdata,reverse=True):
        result = ''
        if "result" in lastdata[SN]:
            result = lastdata[SN]["result"]
        if not "PASS" in result.upper():
            continue
        if "IMAGE" in lastdata[SN]:
            if lastdata[SN]["IMAGE"]:
                for imagename in lastdata[SN]["IMAGE"]:
                    if not imagename in imagenamelist:
                        imagenamelist.append(imagename)
    wirtedata = list()
    wirtedata.append('SN')
    wirtedata.append('PN')
    wirtedata.append('MAC')
    for imagename in imagenamelist:
        wirtedata.append(imagename)
    ws2.append(wirtedata)

    for SN in sorted(lastdata,reverse=True):
        result = ''
        if "result" in lastdata[SN]:
            result = lastdata[SN]["result"]
        if not "PASS" in result.upper():
            continue
        if not "NICINFO" in lastdata[SN]:
            continue
        if not lastdata[SN]["NICINFO"]:
            continue
        FRU = lastdata[SN]["NICINFO"]["FRU"]
        #FRU = FRU.replace(' ','')
        #print(FRU)
        frulist = FRU.split(', ')
        #print(frulist)
        #sys.exit()
        wirtedata = list()
        wirtedata.append(SN)
        wirtedata.append(frulist[2])
        wirtedata.append(frulist[1])
        if "IMAGE" in lastdata[SN]:
            if lastdata[SN]["IMAGE"]:
                for imagename in imagenamelist:
                    if imagename in lastdata[SN]["IMAGE"]:
                        wirtedata.append(lastdata[SN]["IMAGE"][imagename])
                    else:
                        wirtedata.append("NO DATA")

        ws2.append(wirtedata)

    pr['modules'].fixcolumnssize(ws2)
    pr['modules'].freezePosition(ws2, "B2")

    return True

def generateexeclallswstatussummaryreport(pr,SWDATA,wb,startdate=None):
    ws1 = wb.active
    ws1.title = "SUMMARY"
    

    wirtedata = list()
    wirtedata.append('FAMILY')
    ws1.append(wirtedata)

    for family in sorted(SWDATA["status"]["family"]):
        wirtedata = list()
        wirtedata.append(family)
        ws1.append(wirtedata)

    pr['modules'].fixcolumnssize3(ws1)
    pr['modules'].freezePosition(ws1,'A2')

    return None

def createMTPteststepfailurereport(pr,DATA,mtpchassisusecountbyteststep,startdate=None):
    wb3 = Workbook()
    dest_filename5 = "{}MTP_FAILSTEP_{}_DATA.xlsx".format(reportpath,date_time)
    if startdate:
        dest_filename5 = "{}MTP_FAILSTEP_{}_DATA_StartWith_{}.xlsx".format(reportpath,date_time,startdate)
    print(dest_filename5)

    mtpchassisusecountbyslot = generateexeclallmtpfailurestepsummaryreport(pr,mtpchassisusecountbyteststep,wb3,startdate)

    for eachfailurestep in mtpchassisusecountbyteststep['TESTSTEPLIST']:
        generateexeclallmtpstatussummaryreportbyfailurestep(pr,DATA,wb3,eachfailurestep,mtpchassisusecountbyteststep['STATUS'][eachfailurestep],startdate)

    print(dest_filename5)
    wb3.save(filename = dest_filename5) 

def generateexeclallmtpfailurestepsummaryreport(pr,mtpchassisusecountbyteststep,wb,startdate=None):
    ws1 = wb.active
    ws1.title = "SUMMARY"
    for eachfailurestep in mtpchassisusecountbyteststep['TESTSTEPLIST']:
        wirtedata = list()
        wirtedata.append(eachfailurestep)
        ws1.append(wirtedata)

    return None

def generateexeclallmtpstatussummaryreportbyfailurestep(pr,DATA,wb,teststep,mtpchassisusecountbyslot,startdate=None):
    print("generateexeclallmtpstatussummaryreportbyfailurestep: {}".format(teststep))
    titlename = teststep
    ws1 = wb.create_sheet(title=titlename[:30])
    findzeroMTP = list()

    #pr['modules'].print_anyinformation(mtpchassisusecountbyslot)  

    #sys.exit()

    for mtp in sorted(mtpchassisusecountbyslot):
        listoftotal = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
        if sum(listoftotal) == 0:
            findzeroMTP.append(mtp)
    
    for mtp in findzeroMTP:
        del mtpchassisusecountbyslot[mtp]
    #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
    #sys.exit()
    wirtedata = list()
    wirtedata.append('MTP')
    wirtedata.append('TEST')
    wirtedata.append('STATUS')
    wirtedata.append('TOTAL')
    for slot in MTPslot:
        wirtedata.append(slot)
    wirtedata.append('FAMILY')
    ws1.append(wirtedata)
    for mtp in sorted(mtpchassisusecountbyslot):
        wirtedatatotal = list()
        wirtedatapass = list()
        wirtedatafail = list()
        wirtedatapassrate = list()
        wirtedatatotal.append(mtp)
        wirtedatapass.append(mtp)
        wirtedatafail.append(mtp)
        wirtedatapassrate.append(mtp)
        teststep = mtpchassisusecountbyslot[mtp]["TEST"][0]
        if len(mtpchassisusecountbyslot[mtp]["TEST"]) > 1:
            for eachtest in mtpchassisusecountbyslot[mtp]["TEST"][1:]:
                teststep = "{},{}".format(teststep,eachtest)
        allfamily = mtpchassisusecountbyslot[mtp]["FAMILY"][0]
        if len(mtpchassisusecountbyslot[mtp]["FAMILY"]) > 1:
            for eachfamily in mtpchassisusecountbyslot[mtp]["FAMILY"][1:]:
                allfamily = "{}, {}".format(allfamily,eachfamily)
        wirtedatatotal.append(teststep)
        wirtedatapass.append(teststep)
        wirtedatafail.append(teststep)
        wirtedatapassrate.append(teststep)
        listoftotal = list()
        listofpass = list()
        listoffail = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            listofpass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            listoffail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])  
        wirtedatatotal.append("TOTAL")
        wirtedatapass.append("PASS")
        wirtedatafail.append("FAIL")
        wirtedatapassrate.append("FAIL_RATE")
        wirtedatatotal.append(sum(listoftotal))
        wirtedatapass.append(sum(listofpass))
        wirtedatafail.append(sum(listoffail))
        wirtedatapassrate.append(pr['modules'].calculeFail_rate(sum(listoftotal),sum(listoffail)))           
        for slot in MTPslot:
            wirtedatatotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            wirtedatapass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            wirtedatafail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])
            wirtedatapassrate.append(pr['modules'].calculeFail_rate(mtpchassisusecountbyslot[mtp][slot]["TOTAL"],mtpchassisusecountbyslot[mtp][slot]["FAIL"]))
        wirtedatatotal.append(allfamily)
        wirtedatapass.append(allfamily)
        wirtedatafail.append(allfamily)
        wirtedatapassrate.append(allfamily)
        ws1.append(wirtedatatotal)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatafail)
        ws1.append(wirtedatapassrate)          
    pr['modules'].fixcolumnssize3(ws1)
    pr['modules'].highlightinlightredrow(ws1,'FAIL')
    pr['modules'].converttoPERCENTAGEnumberbyFailurerange(ws1)
    pr['modules'].freezePosition(ws1,'D2')

    return None


def generateexeclallmtpstatussummaryreportbytest(pr,DATA,wb,teststep,mtpchassisusecountbyslot,startdate=None):
    print("generateexeclallmtpstatussummaryreport: {}".format(teststep))
    titlename = teststep
    ws1 = wb.create_sheet(title=titlename)
    findzeroMTP = list()

    for mtp in sorted(mtpchassisusecountbyslot):
        listoftotal = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
        if sum(listoftotal) == 0:
            findzeroMTP.append(mtp)
    
    for mtp in findzeroMTP:
        del mtpchassisusecountbyslot[mtp]
    #print(json.dumps(mtpchassisusecountbyslot, indent = 4))
    #sys.exit()
    wirtedata = list()
    wirtedata.append('MTP')
    wirtedata.append('TEST')
    wirtedata.append('STATUS')
    wirtedata.append('TOTAL')
    for slot in MTPslot:
        wirtedata.append(slot)
    wirtedata.append('FAMILY')
    ws1.append(wirtedata)
    for mtp in sorted(mtpchassisusecountbyslot):
        wirtedatatotal = list()
        wirtedatapass = list()
        wirtedatafail = list()
        wirtedatapassrate = list()
        wirtedatatotal.append(mtp)
        wirtedatapass.append(mtp)
        wirtedatafail.append(mtp)
        wirtedatapassrate.append(mtp)
        teststep = mtpchassisusecountbyslot[mtp]["TEST"][0]
        if len(mtpchassisusecountbyslot[mtp]["TEST"]) > 1:
            for eachtest in mtpchassisusecountbyslot[mtp]["TEST"][1:]:
                teststep = "{},{}".format(teststep,eachtest)
        allfamily = mtpchassisusecountbyslot[mtp]["FAMILY"][0]
        if len(mtpchassisusecountbyslot[mtp]["FAMILY"]) > 1:
            for eachfamily in mtpchassisusecountbyslot[mtp]["FAMILY"][1:]:
                allfamily = "{}, {}".format(allfamily,eachfamily)
        wirtedatatotal.append(teststep)
        wirtedatapass.append(teststep)
        wirtedatafail.append(teststep)
        wirtedatapassrate.append(teststep)
        listoftotal = list()
        listofpass = list()
        listoffail = list()
        for slot in MTPslot:
            listoftotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            listofpass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            listoffail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])  
        wirtedatatotal.append("TOTAL")
        wirtedatapass.append("PASS")
        wirtedatafail.append("FAIL")
        wirtedatapassrate.append("PASS_RATE")
        wirtedatatotal.append(sum(listoftotal))
        wirtedatapass.append(sum(listofpass))
        wirtedatafail.append(sum(listoffail))
        wirtedatapassrate.append(pr['modules'].calculePass_rate(sum(listoftotal),sum(listofpass)))           
        for slot in MTPslot:
            wirtedatatotal.append(mtpchassisusecountbyslot[mtp][slot]["TOTAL"])
            wirtedatapass.append(mtpchassisusecountbyslot[mtp][slot]["PASS"])
            wirtedatafail.append(mtpchassisusecountbyslot[mtp][slot]["FAIL"])
            wirtedatapassrate.append(pr['modules'].calculePass_rate(mtpchassisusecountbyslot[mtp][slot]["TOTAL"],mtpchassisusecountbyslot[mtp][slot]["PASS"]))
        wirtedatatotal.append(allfamily)
        wirtedatapass.append(allfamily)
        wirtedatafail.append(allfamily)
        wirtedatapassrate.append(allfamily)
        ws1.append(wirtedatatotal)
        ws1.append(wirtedatapass)
        ws1.append(wirtedatafail)
        ws1.append(wirtedatapassrate)          
    pr['modules'].fixcolumnssize3(ws1)
    pr['modules'].converttoPERCENTAGEnumber(ws1)
    pr['modules'].highlightinlightredrow(ws1,'FAIL')
    pr['modules'].freezePosition(ws1,'D2')

    return None

def getcurrentlyswdatabasejsonfile(eachjsondict):
    if 'FILE' in eachjsondict:
        if 'softwarejsonfile' in eachjsondict['FILE']:
            return eachjsondict['FILE']["softwarejsonfile"]

    return None

if __name__ == "__main__":
    sys.exit(main())