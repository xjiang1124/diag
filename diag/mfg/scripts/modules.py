#!/usr/bin/python3

import warnings
import openpyxl
import os
import sys
import time
import glob
import os.path
from os import path
import datetime
import re
from openpyxl import Workbook
from datetime import datetime
import yaml

class modules(object):
	def __init__(self, logfile=None):
		self.debug = True
		self.logfile = None
		if logfile:
			self.logfile = logfile


	def readyamltodict(self, yamlfile):
		with open(yamlfile) as file:
		    # The FullLoader parameter handles the conversion from YAML
		    # scalar values to Python the dictionary format
		    yamldict = yaml.load(file, Loader=yaml.FullLoader)

		    #self.print_anyinformation(yamldict)

		return yamldict

	#USE
	def wirtedicttoyaml(self, yamldict, yamlfile):
		with open(yamlfile, 'w') as file:
		    documents = yaml.dump(yamldict, file)
		return None

	#USE
	def getallfilebyfolder(self,rootdir,keyword=None,level=None):
	    start=datetime.now()
	    listofdirs = self.getallfolder(rootdir,keyword=keyword,level=level)

	    listoffiles = list()

	    for eachdir in listofdirs:
	        somelists = glob.glob(eachdir + '/*')
	        for something in somelists:
	            if os.path.isfile(something):  
	                basename = os.path.basename(something)
	                if keyword:
	                    if not keyword.upper() in basename.upper():
	                        continue
	                if not something in listoffiles:
	                    listoffiles.append(something)

	    listoffiles.sort()
	    difftime = datetime.now()-start
	    return listoffiles
	#USE
	def getallfolder(self,rootdir,keyword=None,level=None):
		start=datetime.now()

		listofdirs = list()

		listofdirs.append(rootdir)

		stillhavedir = True

		count = 1

		if not level:
		    for checkdir in listofdirs:
		        somelists = glob.glob(checkdir + '/*')
		        for something in somelists:
		            if os.path.isdir(something):
		                if not something in listofdirs:
		                    stillhavedir = True
		                    listofdirs.append(something)
		else:
		    while stillhavedir:
		        
		        if level:
		            if count >= level:
		                break
		        stillhavedir = False
		        resultlist = list()         

		        for checkdir in listofdirs:
		            somelists = glob.glob(checkdir + '/*')
		            for something in somelists:
		                if os.path.isdir(something):
		                    if not something in listofdirs:
		                        stillhavedir = True
		                        resultlist.append(something)
		        for eachdir in resultlist:
		            if not eachdir in listofdirs:
		                listofdirs.append(eachdir)
		        count += 1

		difftime = datetime.now()-start
		self.debug_print("DIR: {} have {} folder and running {} seconds".format(rootdir,len(listofdirs),difftime.total_seconds()))

		return listofdirs

	#USE
	def readSNxlsxfile(self,xlsxfile,SNvsPCBA):
		from openpyxl import load_workbook
		self.debug_print("FILE: {}".format(xlsxfile))
		wb_sn = load_workbook(filename = xlsxfile)
		for tabname in wb_sn.sheetnames:
			self.debug_print("tabname: {}".format(tabname))
			eachsheet = wb_sn[tabname]
			SNraw, SNcolunm = self.findposition(eachsheet,'SN')
			PCBAsnraw, PCBAsncolunm = self.findposition(eachsheet,'PCBA_SN')

			self.debug_print("SN: Raw {} Colunm {} | PCBA_SN: Raw {} Colunm {}".format(SNraw, SNcolunm, PCBAsnraw, PCBAsncolunm))

			if SNraw and PCBAsnraw:
				colunm = eachsheet['A']
				for countnumberinrow in range(2,len(colunm)+1):
					SN = str(eachsheet.cell(row=countnumberinrow, column=SNcolunm).value).strip()
					PCBA_SN = str(eachsheet.cell(row=countnumberinrow, column=PCBAsncolunm).value).strip()
					self.debug_print("SN: {} | PCBA_SN: {}".format(SN, PCBA_SN))
					SNvsPCBA[SN] = PCBA_SN		

		return None

	#USE
	def findposition(self,sheet,keyword):

		colunm = sheet[1]
		raw = sheet['A']

		for rawnumber in range(1,len(raw)+1):
			for colunmnumber in range(1,len(colunm)+1):
				if keyword == sheet.cell(row=rawnumber, column=colunmnumber).value:
					return rawnumber, colunmnumber
		return None, None

	#USE
	def debug_print(self,msg):
		if self.debug:
			
			if self.logfile:
				self.logfile.WriteLine(str(msg))
			else:
				print(msg)
		return None